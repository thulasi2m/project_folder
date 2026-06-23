# ---------------- Standard Library ----------------
import os
import json
import sqlite3
import time
import queue
import threading
import datetime
import sys
import re
from PIL import Image, ImageDraw, ImageOps, ImageEnhance

# ---------------- Tkinter / CustomTkinter ----------------
import tkinter as tk
from tkinter import messagebox, filedialog
from tkinter import ttk
import customtkinter as ctk
from tkcalendar import DateEntry

# ---------------- Serial Communication ----------------
import serial
import serial.tools.list_ports

# ---------------- Plotting ----------------
import matplotlib
matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg


# ---------------- Table / Sheet ----------------
import tksheet
import tempfile

from ui_animations import Easing, AnimationHelper, HoverAnimation

class ModernCardFrame(ctk.CTkFrame):
    """
    Premium card container frame matching UI/UX specifications.
    Features corner_radius=12, fg_color="#FAFAFA", border_width=1, border_color="#E0E0E0".
    """
    def __init__(self, parent, **kwargs):
        # Override default CustomTkinter styles with design token values
        kwargs["corner_radius"] = kwargs.get("corner_radius", 12)
        kwargs["fg_color"] = kwargs.get("fg_color", "#FAFAFA")
        kwargs["border_width"] = kwargs.get("border_width", 1)
        kwargs["border_color"] = kwargs.get("border_color", "#E0E0E0")
        super().__init__(parent, **kwargs)

class ModernButton(ctk.CTkButton):
    def __init__(self, parent, **kwargs):
        # Standardize font to Segoe UI
        if "font" not in kwargs:
            kwargs["font"] = ("Segoe UI", 11, "normal")
        else:
            font_val = kwargs["font"]
            if isinstance(font_val, tuple) and len(font_val) >= 2:
                kwargs["font"] = ("Segoe UI", font_val[1], font_val[2] if len(font_val) > 2 else "normal")
                
        takefocus_val = kwargs.pop("takefocus", None)
        super().__init__(parent, **kwargs)
        if takefocus_val is not None:
            try:
                self._canvas.configure(takefocus=takefocus_val)
            except:
                pass
        
        # Resolve normal_color and hover_color
        normal_color = self.cget("fg_color")
        hover_color = self.cget("hover_color")
        
        if isinstance(normal_color, (list, tuple)):
            normal_color = normal_color[0]
        if isinstance(hover_color, (list, tuple)):
            hover_color = hover_color[0]
            
        if (kwargs.get("hover") is not False and
            isinstance(normal_color, str) and normal_color.startswith("#") and 
            isinstance(hover_color, str) and hover_color.startswith("#")):
            try:
                self.hover_anim = HoverAnimation(self, hover_color, normal_color, duration=150)
            except Exception as e:
                pass

    def configure(self, **kwargs):
        super().configure(**kwargs)
        if "fg_color" in kwargs or "hover_color" in kwargs:
            normal_color = self.cget("fg_color")
            hover_color = self.cget("hover_color")
            if isinstance(normal_color, (list, tuple)): normal_color = normal_color[0]
            if isinstance(hover_color, (list, tuple)): hover_color = hover_color[0]
            if hasattr(self, "hover_anim") and self.hover_anim:
                if isinstance(normal_color, str) and normal_color.startswith("#"):
                    self.hover_anim.normal_color = normal_color
                if isinstance(hover_color, str) and hover_color.startswith("#"):
                    self.hover_anim.hover_color = hover_color

class SidebarButton(ctk.CTkFrame):
    def __init__(self, parent, symbol, word, command, **kwargs):
        super().__init__(parent, fg_color="transparent", corner_radius=8, height=48, **kwargs)
        self.pack_propagate(False)
        self.command = command
        self.symbol = symbol
        self.word = word
        
        self.symbol_lbl = ctk.CTkLabel(
            self,
            text=symbol,
            font=("Segoe UI", 16),
            text_color="#007B43",
            anchor="center",
            width=40
        )
        self.symbol_lbl.pack(side="left", padx=(10, 5), fill="y")
        
        self.word_lbl = ctk.CTkLabel(
            self,
            text=word,
            font=("Segoe UI", 13, "bold"),
            text_color="#1A1A1A",
            anchor="w"
        )
        self.word_lbl.pack(side="left", fill="both", expand=True)
        
        # Bind events
        for widget in (self, self.symbol_lbl, self.word_lbl):
            widget.bind("<Enter>", self.on_enter)
            widget.bind("<Leave>", self.on_leave)
            widget.bind("<Button-1>", self.on_click)
            
        self.active = False
        
    def on_enter(self, event):
        if not self.active:
            self.configure(fg_color="#F1F8E9") # Light green hover
            
    def on_leave(self, event):
        if not self.active:
            self.configure(fg_color="transparent")
            
    def on_click(self, event):
        self.command()
        
    def set_active(self, active):
        self.active = active
        if active:
            self.configure(fg_color="#007B43")
            self.symbol_lbl.configure(text_color="white")
            self.word_lbl.configure(text_color="white")
        else:
            self.configure(fg_color="transparent")
            self.symbol_lbl.configure(text_color="#007B43")
            self.word_lbl.configure(text_color="#1A1A1A")

def make_sheet_auto_resize(sheet_obj, frame_obj, cols):
    """Binds configure event of frame_obj to resize columns of sheet_obj to fit 100%."""
    def do_resize(event=None):
        w = frame_obj.winfo_width()
        if w > 100:
            cols_count = len(cols)
            avg_w = (w - 20) // cols_count if cols_count > 0 else 100
            for i in range(cols_count):
                try:
                    sheet_obj.column_width(column=i, width=avg_w, only_set_if_too_small=False)
                except:
                    try: sheet_obj.column_width(i, avg_w)
                    except: pass
            try: sheet_obj.refresh()
            except:
                try: sheet_obj.redraw()
                except: pass
    frame_obj.bind("<Configure>", do_resize)
    # Trigger initial size
    frame_obj.after(100, do_resize)
    return do_resize





def resource_path(relative_path):
    """Return absolute path for EXE and normal run."""
    try:
        base_path = sys._MEIPASS    # EXE temp folder
    except Exception:
        # When running normally (not EXE)
        base_path = os.path.dirname(os.path.abspath(__file__))

    return os.path.join(base_path, relative_path)

# --- WINDOWS 10/11 HIGH DPI FIX ---
try:
    import ctypes
    ctypes.windll.shcore.SetProcessDpiAwareness(1)
except Exception:
    pass

def add_corners(im, rad):
    """Add rounded corners to an image"""
    circle = Image.new('L', (rad * 2, rad * 2), 0)
    draw = ImageDraw.Draw(circle)
    draw.ellipse((0, 0, rad * 2, rad * 2), fill=255)
    alpha = Image.new('L', im.size, 255)
    w, h = im.size
    alpha.paste(circle.crop((0, 0, rad, rad)), (0, 0))
    alpha.paste(circle.crop((0, rad, rad, rad * 2)), (0, h - rad))
    alpha.paste(circle.crop((rad, 0, rad * 2, rad)), (w - rad, 0))
    alpha.paste(circle.crop((rad, rad, rad * 2, rad * 2)), (w - rad, h - rad))
    im.putalpha(alpha)
    return im


# --- GLOBAL STATE FOR REPORT PAGE PERSISTENCE ---
REPORT_FILTER_STATE = {}

# --- EXPORT LIBRARIES ---
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    openpyxl = None

try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage, PageBreak, KeepTogether
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfgen import canvas
except ImportError:
    # If reportlab is missing, we will handle it gracefully in the export method
    pass

class DataManager:
    """Central buffer for all parsed AirGauge data."""
    def __init__(self):
        self.buffer = []          # store all parsed readings
        self.subscribers = []     # pages that want to receive updates (LiveData, RunChat)

    def subscribe(self, page):
        """Register a page (like LiveDataPage or RunChatPage) for live updates."""
        if page not in self.subscribers:
            self.subscribers.append(page)

    def add_parsed(self, parsed):
        """Add new parsed data and broadcast to all subscribers."""
        if not parsed:
            return
        self.buffer.append(parsed)
        # notify all pages that care
        for page in list(self.subscribers):
            try:
                page.update_with_parsed(parsed)
            except Exception as e:
                print(f"Update error in {page.__class__.__name__}: {e}")

LIVE_DATA_FILE = "live_data_log.json"
try:
    if os.path.exists(LIVE_DATA_FILE):
        os.remove(LIVE_DATA_FILE)
        print("Old live data cleared (fresh start).")
except Exception as e:
    print("Warning: couldn't clear old data:", e)


COMP_FILE = "airgauge_component_config.json"

def load_component_map():
    """Load component configuration from SQLite."""
    try:
        conn = sqlite3.connect(resource_path("component_setup.db"))
        cursor = conn.cursor()
        cursor.execute("SELECT airgauge_id, channel, properties FROM component_setup")
        rows = cursor.fetchall()
        conn.close()
        comp_map = {}
        for ag_id, ch, props_str in rows:
            if ag_id not in comp_map:
                comp_map[ag_id] = {}
            try:
                comp_map[ag_id][ch] = json.loads(props_str)
            except:
                comp_map[ag_id][ch] = {}
        return comp_map
    except Exception:
        return {}

# Load once globally
COMP_MAP = load_component_map()


def parse_packet(packet):
    try:
        # basic validation
        if not (packet.startswith("*") and packet.endswith("#") and len(packet) >= 54):
            return None
        data = packet

        hours = data[1:3]
        minutes = data[3:5]
        seconds = data[5:7]

        # --- date ---
        datestr = data[8:14]
        formatted_date = f"{datestr[0:2]}/{datestr[2:4]}/{datestr[4:6]}"

        # --- numeric fields ---
        # Robustly handle AirGauge ID (can be digits or like "AG1")
        ag_raw = data[14:17].strip()
        if ag_raw.startswith("AG"):
            try: airgauge_id = int(ag_raw[2:])
            except: airgauge_id = 0
        else:
            try: airgauge_id = int(ag_raw) % 1000
            except: airgauge_id = 0

        channel_no = int(data[17:19])
        offset_value = int(data[19:21])

        # --- drawing value ---
        draw = data[22:29]
        if not draw.isdigit():
            cleaned = "".join(ch for ch in draw if ch.isdigit())
            cleaned = cleaned.ljust(7, "0")
            draw = cleaned
        drawing_val = f"{draw[:3]}.{draw[3:]}"

        # --- Reading ---
        calc_raw = data[31:38]
        if not calc_raw.isdigit():
            cleaned = "".join(ch for ch in calc_raw if ch.isdigit())
            cleaned = cleaned.ljust(7, "0")
            calc_raw = cleaned
        Reading = f"{calc_raw[:3]}.{calc_raw[3:]}"

        # --- status conversion ---
        status_raw = data[38].upper()

        if status_raw == "A":
            status = "ACCEPTED"
        elif status_raw == "B":
            status = "REJECTED"
        elif status_raw == "C":
            status = "REWORK"
        else:
            status = "UNKNOWN"

        # --- IDs ---
        compID = data[39:49].strip()
        userID = data[49:52]
        cncID = data[52:54]

        # =================================================================
        #  NEW LOGIC → Attach item_name/item_code from config JSON
        # =================================================================
        item_name = ""
        item_code = ""

        ag_str = str(airgauge_id)
        ch_key = f"CH{channel_no}"

        # Try matching with and without "AG" prefix
        lookups = [ag_str, f"AG{airgauge_id}", f"AG{airgauge_id:01d}", f"AG{airgauge_id:02d}", f"AG{airgauge_id:03d}"]
        for l_key in lookups:
            if l_key in COMP_MAP:
                if ch_key in COMP_MAP[l_key]:
                    item_name = COMP_MAP[l_key][ch_key].get("item_name", "")
                    item_code = COMP_MAP[l_key][ch_key].get("item_code", "")
                    break

        # =================================================================
        #  Return updated tuple WITH item_name + item_code
        # =================================================================
        return (hours, minutes, seconds, formatted_date,Reading,offset_value,status, airgauge_id, channel_no,
                drawing_val, userID,compID,item_name,item_code, cncID)

    except (ValueError, IndexError):
        return None






PASSWORD = "cherry@123"
CONFIG_FILE = "mac_airgauge_config.json"



class InstallerLoginPage(ctk.CTkFrame):
    """
    Pre-setup login: Hardcoded to 'cherry' / 'cherry@123'.
    Required before accessing FirstTimeSetupPage.
    Redesigned as a modern centered single-card with shadow and circular corners.
    """
    def __init__(self, parent, controller):
        super().__init__(parent)
        self.controller = controller

        # ── Background: soft neutral gradient ──────────────────────────────
        self.configure(fg_color="#F0F2F5")

        # ── Shadow layer: proper drop shadow offset down-right ────────────
        # We use a Canvas to draw a true soft blurred shadow behind the card
        self._shadow_canvas = tk.Canvas(self, bg="#F0F2F5", highlightthickness=0,
                                        width=420, height=550)
        self._shadow_canvas.place(relx=0.5, rely=0.5, anchor="center")

        # Draw multiple translucent rounded rectangles offset slightly for depth
        _shadow_colors = ["#CBCDD4", "#D3D5DC", "#DBDDE4", "#E3E5EB", "#EBECF0"]
        for i, col in enumerate(_shadow_colors):
            pad = i * 2
            off = 5 - i  # offset from center
            self._shadow_canvas.create_rectangle(
                10 + pad + off, 12 + pad + off,
                410 - pad + off, 542 - pad + off,
                fill=col, outline="", tags="shadow"
            )

        # ── Main Card ──────────────────────────────────────────────────────
        self.card = ctk.CTkFrame(
            self,
            fg_color="white",
            corner_radius=28,
            width=400,
            height=530
        )
        self.card.place(relx=0.5, rely=0.5, anchor="center")
        self.card.pack_propagate(False)
        self.card.lift()

        # ── Logo ───────────────────────────────────────────────────────────
        try:
            logo_path = resource_path("settings/cherry_signup_logo.png")
            if os.path.exists(logo_path):
                pil_img = Image.open(logo_path)
                self.cherry_logo_img = ctk.CTkImage(pil_img, size=(200, 62))
                logo_lbl = ctk.CTkLabel(self.card, text="", image=self.cherry_logo_img,
                                        fg_color="transparent")
                logo_lbl.pack(pady=(38, 6))
        except Exception as e:
            print("Error loading installer logo:", e)

        # ── Title ──────────────────────────────────────────────────────────
        title_frame = ctk.CTkFrame(self.card, fg_color="transparent")
        title_frame.pack(pady=(4, 2))
        ctk.CTkLabel(title_frame, text="Welcome ", font=("Segoe UI", 22, "bold"),
                     text_color="#1A1A1A").pack(side="left")
        ctk.CTkLabel(title_frame, text="Back", font=("Segoe UI", 22, "bold"),
                     text_color="#B0050E").pack(side="left")

        ctk.CTkLabel(self.card, text="Please sign in to configure the system",
                     font=("Segoe UI", 11), text_color="#757575").pack(pady=(0, 22))

        # ── Username Field ─────────────────────────────────────────────────
        user_container = ctk.CTkFrame(self.card, fg_color="white",
                                      border_color="#E0E0E0", border_width=1.5,
                                      corner_radius=12, height=50, width=310)
        user_container.pack(pady=6)
        user_container.pack_propagate(False)

        ctk.CTkLabel(user_container, text="👤", font=("Segoe UI", 15),
                     text_color="#9E9E9E", fg_color="transparent").pack(side="left", padx=(14, 6))

        self.user_entry = tk.Entry(user_container, relief="flat", bd=0,
                                   bg="white", fg="#333",
                                   font=("Segoe UI", 13), insertbackground="#333")
        self.user_entry.pack(side="left", fill="x", expand=True, padx=(0, 14), pady=14)

        user_ph = ctk.CTkLabel(user_container, text="Username", font=("Segoe UI", 13),
                               text_color="#BDBDBD", fg_color="transparent")
        user_ph.place(x=42, y=25, anchor="w")
        user_ph.bind("<Button-1>", lambda e: self.user_entry.focus())

        def _check_user_ph(*_):
            self.after(10, lambda: user_ph.place_forget() if self.user_entry.get()
                       else user_ph.place(x=42, y=25, anchor="w"))
        self.user_entry.bind("<KeyRelease>", _check_user_ph)
        self.user_entry.bind("<FocusIn>",
            lambda e: [user_container.configure(border_color="#1A1A1A"), user_ph.place_forget()])
        self.user_entry.bind("<FocusOut>",
            lambda e: [user_container.configure(border_color="#E0E0E0"), _check_user_ph()])
        self.user_entry.focus()

        # ── Password Field ─────────────────────────────────────────────────
        pass_container = ctk.CTkFrame(self.card, fg_color="white",
                                      border_color="#E0E0E0", border_width=1.5,
                                      corner_radius=12, height=50, width=310)
        pass_container.pack(pady=6)
        pass_container.pack_propagate(False)

        ctk.CTkLabel(pass_container, text="🔒", font=("Segoe UI", 15),
                     text_color="#9E9E9E", fg_color="transparent").pack(side="left", padx=(14, 6))

        self._show_pass = False
        self._actual_pass = ""

        self.pass_entry = tk.Entry(pass_container, relief="flat", bd=0,
                                   bg="white", fg="#333",
                                   font=("Segoe UI", 13), insertbackground="#333")
        self.pass_entry.pack(side="left", fill="x", expand=True, padx=(0, 6), pady=14)

        pass_ph = ctk.CTkLabel(pass_container, text="Password", font=("Segoe UI", 13),
                               text_color="#BDBDBD", fg_color="transparent")
        pass_ph.place(x=42, y=25, anchor="w")
        pass_ph.bind("<Button-1>", lambda e: self.pass_entry.focus())

        def _check_pass_ph(*_):
            self.after(10, lambda: pass_ph.place_forget() if self.pass_entry.get()
                       else pass_ph.place(x=42, y=25, anchor="w"))

        def _on_pass_key(event=None):
            raw = self.pass_entry.get()
            self._actual_pass = raw
            _check_pass_ph()

        self.pass_entry.bind("<KeyRelease>", _on_pass_key)
        self.pass_entry.bind("<FocusIn>",
            lambda e: [pass_container.configure(border_color="#1A1A1A"), pass_ph.place_forget()])
        self.pass_entry.bind("<FocusOut>",
            lambda e: [pass_container.configure(border_color="#E0E0E0"), _check_pass_ph()])

        def _toggle_pass():
            self._show_pass = not self._show_pass
            self.pass_entry.config(show="" if self._show_pass else "●")
            eye_btn.configure(text="👁" if self._show_pass else "👁‍🗨")

        eye_btn = ctk.CTkLabel(pass_container, text="👁‍🗨", font=("Segoe UI", 15),
                               text_color="#9E9E9E", fg_color="transparent", cursor="hand2")
        eye_btn.pack(side="right", padx=(0, 12))
        eye_btn.bind("<Button-1>", lambda e: _toggle_pass())

        # ── Return bindings ────────────────────────────────────────────────
        self.user_entry.bind("<Return>", lambda e: self.pass_entry.focus())
        self.pass_entry.bind("<Return>", lambda e: self.check_login())

        # ── Error Label ────────────────────────────────────────────────────
        self.error_label = ctk.CTkLabel(self.card, text="",
                                        text_color="#D32F2F", font=("Segoe UI", 11))
        self.error_label.pack(pady=(8, 2))

        # ── LOGIN Button (black, slightly smaller) ─────────────────────────
        self.login_btn = ModernButton(
            self.card,
            text="⬤  LOGIN",
            font=("Segoe UI", 12, "bold"),
            height=42, width=270,
            fg_color="#1A1A1A", hover_color="#333333",
            text_color="white",
            corner_radius=10,
            command=self.check_login
        )
        self.login_btn.pack(pady=(6, 4))

        # ── Footer ─────────────────────────────────────────────────────────
        ctk.CTkLabel(self.card, text="Authorized Personnel Only",
                     font=("Segoe UI", 10), text_color="#BDBDBD").pack(pady=(10, 0))

    def check_login(self):
        u = self.user_entry.get().strip()
        p = self.pass_entry.get().strip()

        # Hardcoded installer login
        if u == "cherry" and p == "cherry@123":
            self.controller.show_license_verification()
        else:
            self.error_label.configure(text="Invalid credentials")




class LicenseVerificationPage(ctk.CTkFrame):
    """
    Step 2 of Installer setup: License File upload.
    Redesigned to match the same single-card centered style as InstallerLoginPage.
    """
    def __init__(self, parent, controller):
        super().__init__(parent)
        self.controller = controller

        # ── Background ────────────────────────────────────────────────────
        self.configure(fg_color="#F0F2F5")

        # ── Shadow canvas (same technique as InstallerLoginPage) ──────────
        self._shadow_canvas = tk.Canvas(self, bg="#F0F2F5", highlightthickness=0,
                                        width=480, height=570)
        self._shadow_canvas.place(relx=0.5, rely=0.5, anchor="center")
        _shadow_colors = ["#CBCDD4", "#D3D5DC", "#DBDDE4", "#E3E5EB", "#EBECF0"]
        for i, col in enumerate(_shadow_colors):
            pad = i * 2
            off = 5 - i
            self._shadow_canvas.create_rectangle(
                10 + pad + off, 12 + pad + off,
                470 - pad + off, 562 - pad + off,
                fill=col, outline="", tags="shadow"
            )

        # ── Main Card ─────────────────────────────────────────────────────
        self.card = ctk.CTkFrame(
            self,
            fg_color="white",
            corner_radius=28,
            width=460,
            height=550
        )
        self.card.place(relx=0.5, rely=0.5, anchor="center")
        self.card.pack_propagate(False)
        self.card.lift()

        # ── Logo ──────────────────────────────────────────────────────────
        try:
            logo_path = resource_path("settings/cherry_signup_logo.png")
            if os.path.exists(logo_path):
                pil_img = Image.open(logo_path)
                self.cherry_logo_img = ctk.CTkImage(pil_img, size=(200, 62))
                logo_lbl = ctk.CTkLabel(self.card, text="", image=self.cherry_logo_img,
                                        fg_color="transparent")
                logo_lbl.pack(pady=(32, 6))
        except Exception as e:
            print("Error loading license page logo:", e)

        # ── Title ─────────────────────────────────────────────────────────
        ctk.CTkLabel(self.card, text="🔐  License Verification",
                     font=("Segoe UI", 20, "bold"), text_color="#1A1A1A").pack(pady=(4, 2))

        ctk.CTkLabel(self.card, text="Select your license key file (.key)",
                     font=("Segoe UI", 11), text_color="#757575").pack(pady=(0, 16))

        # ── Drop-zone: Canvas with dashed border + PIL upload-tray icon ──────
        dz_w, dz_h = 380, 210
        self._dz_canvas = tk.Canvas(
            self.card, bg="#F8FAFB",
            highlightthickness=0, width=dz_w, height=dz_h
        )
        self._dz_canvas.pack(pady=(0, 14))

        # Dashed border rectangle
        self._dz_canvas.create_rectangle(
            6, 6, dz_w - 6, dz_h - 6,
            outline="#AAAAAA", width=2, dash=(5, 4)
        )

        # Upload-tray icon via PIL → PhotoImage on canvas
        try:
            from PIL import ImageTk as _ITK
            _icon_pil = LicenseVerificationPage._make_cloud_icon(
                size=64, color="#1A1A1A", bg="#F8FAFB")
            self._cloud_photo = _ITK.PhotoImage(_icon_pil)
            self._dz_canvas.create_image(dz_w // 2, 52, image=self._cloud_photo)
        except Exception as _ce:
            self._dz_canvas.create_text(dz_w // 2, 46, text="⬆",
                                         font=("Segoe UI", 36), fill="#555")

        # "Drag & drop" label
        self._dz_canvas.create_text(
            dz_w // 2, 106,
            text="Drag & drop your license file here",
            font=("Segoe UI", 13, "bold"), fill="#1A1A1A"
        )

        # "or" divider
        self._dz_canvas.create_text(
            dz_w // 2, 128,
            text="or", font=("Segoe UI", 10), fill="#AAAAAA"
        )

        # Upload button embedded in canvas
        self.license_btn = ModernButton(
            self._dz_canvas,
            text="⬆  CLICK TO UPLOAD LICENSE",
            font=("Segoe UI", 11, "bold"),
            height=38, width=300,
            fg_color="#1A1A1A", hover_color="#333333",
            text_color="white",
            corner_radius=8,
            command=self.upload_license
        )
        self._dz_canvas.create_window(dz_w // 2, 170, window=self.license_btn)


        # ── Info hint ─────────────────────────────────────────────────────
        hint_frame = ctk.CTkFrame(self.card, fg_color="#F0F4F0",
                                  corner_radius=10, width=380, height=56)
        hint_frame.pack(pady=(0, 12))
        hint_frame.pack_propagate(False)

        ctk.CTkLabel(hint_frame,
                     text="ℹ  License file should be in .key format\n"
                          "    Contact your system administrator if you don't have one.",
                     font=("Segoe UI", 10), text_color="#4E6B50",
                     fg_color="transparent", justify="left").pack(
                         anchor="w", padx=12, pady=8)

        # ── Error Label ───────────────────────────────────────────────────
        self.error_label = ctk.CTkLabel(self.card, text="",
                                        text_color="#D32F2F", font=("Segoe UI", 11))
        self.error_label.pack(pady=(0, 6))

        # ── VERIFY & PROCEED Button (black) ───────────────────────────────
        self.proceed_btn = ModernButton(
            self.card,
            text="🔒  VERIFY & PROCEED",
            font=("Segoe UI", 12, "bold"),
            height=42, width=280,
            fg_color="#1A1A1A", hover_color="#333333",
            text_color="white",
            corner_radius=10,
            command=self.upload_license
        )
        self.proceed_btn.pack(pady=(0, 6))

        # ── Footer ────────────────────────────────────────────────────────
        ctk.CTkLabel(self.card, text="🔒  License Required for System Activation",
                     font=("Segoe UI", 10), text_color="#BDBDBD").pack(pady=(2, 0))

    def upload_license(self):
        filename = filedialog.askopenfilename(
            title="Select License Key",
            filetypes=(("License Files", "*.key *.dat *.lic"), ("All Files", "*.*"))
        )
        
        if filename:
            try:
                with open(filename, 'r') as f:
                    content = f.read().strip()
                
                # Valid License Keys
                valid_keys = [
                    "CHERRY-LIC-ALPHA-8821",
                    "CHERRY-LIC-BETA-9932",
                    "CHERRY-LIC-GAMMA-7743",
                    "CHERRY-LIC-DELTA-5514",
                    "CHERRY-LIC-EPSILON-2295",
                    "CHERRY-LIC-ZETA-1106"
                ]
                
                if content in valid_keys:
                    self.license_btn.configure(text="License Verified ✅",
                                               fg_color="#2E7D32", text_color="white")
                    self.proceed_btn.configure(text="✅  Verified — Proceeding...",
                                               fg_color="#2E7D32")
                    self.error_label.configure(text="")
                    self.update()
                    time.sleep(0.5)
                    self.controller.show_setup_wizard()
                else:
                    self.error_label.configure(text="❌  Invalid License File. Please try again.")
            except Exception as e:
                self.error_label.configure(text=f"Error reading file: {str(e)}")

    @staticmethod
    def _make_cloud_icon(size=64, color="#1A1A1A", bg="#F8FAFB"):
        """Draw an upload-tray icon (arrow + tray base) matching the reference UI."""
        from PIL import Image, ImageDraw
        ic = tuple(int(color.lstrip('#')[i:i+2], 16) for i in (0, 2, 4)) + (255,)
        img = Image.new("RGBA", (size, size), (0, 0, 0, 0))
        draw = ImageDraw.Draw(img)
        lw = max(2, round(size / 20))
        s   = float(size)
        cx  = size // 2

        # ─ Upward arrow ─
        tip_y      = int(0.05 * s)   # very top of arrowhead
        head_half  = int(0.18 * s)   # half-width of arrowhead base
        shaft_top  = tip_y + head_half
        shaft_btm  = int(0.55 * s)   # where shaft meets the tray

        # Filled arrowhead triangle
        draw.polygon([
            (cx,            tip_y),
            (cx - head_half, shaft_top),
            (cx + head_half, shaft_top),
        ], fill=ic)

        # Arrow shaft
        draw.line([cx, shaft_top, cx, shaft_btm], fill=ic, width=lw)

        # ─ Tray base ─
        tray_arm_w = int(0.38 * s)   # how far arms spread sideways from center
        tray_top_y = int(0.55 * s)   # top of tray (= shaft bottom)
        tray_btm_y = int(0.82 * s)   # bottom bar y

        # Left arm: diagonal from shaft-bottom going down-left
        draw.line([cx, tray_top_y, cx - tray_arm_w, tray_btm_y], fill=ic, width=lw)
        # Right arm: diagonal from shaft-bottom going down-right
        draw.line([cx, tray_top_y, cx + tray_arm_w, tray_btm_y], fill=ic, width=lw)
        # Horizontal base bar
        draw.line([cx - tray_arm_w, tray_btm_y, cx + tray_arm_w, tray_btm_y],
                  fill=ic, width=lw)

        return img


def bezier_curve(p0, p1, p2, p3, num_points=100):
    points = []
    for i in range(num_points + 1):
        t = i / num_points
        x = (1-t)**3 * p0[0] + 3*(1-t)**2 * t * p1[0] + 3*(1-t)*t**2 * p2[0] + t**3 * p3[0]
        y = (1-t)**3 * p0[1] + 3*(1-t)**2 * t * p1[1] + 3*(1-t)*t**2 * p2[1] + t**3 * p3[1]
        points.append((x, y))
    return points

def generate_panel_image(width, height):
    from PIL import ImageFilter
    scale = 4
    sw = width * scale
    sh = height * scale
    
    red_layer = Image.new("RGBA", (sw, sh), (0, 0, 0, 0))
    green_layer = Image.new("RGBA", (sw, sh), (0, 0, 0, 0))
    shadow_layer = Image.new("RGBA", (sw, sh), (0, 0, 0, 0))
    
    draw_red = ImageDraw.Draw(red_layer)
    draw_green = ImageDraw.Draw(green_layer)
    draw_shadow = ImageDraw.Draw(shadow_layer)
    
    # Red Shape (Triangular, overlapping under green)
    red_p0 = (0, 250 * scale)
    red_p1 = (80 * scale, 270 * scale)
    red_p2 = (160 * scale, 300 * scale)
    red_p3 = (215 * scale, 340 * scale) # Right vertex
    red_curve1 = bezier_curve(red_p0, red_p1, red_p2, red_p3)
    
    red_p4 = (215 * scale, 340 * scale)
    red_p5 = (180 * scale, 420 * scale)
    red_p6 = (160 * scale, 520 * scale)
    red_p7 = (130 * scale, sh) # Bottom edge
    red_curve2 = bezier_curve(red_p4, red_p5, red_p6, red_p7)
    
    red_poly = [(0, 250 * scale)] + red_curve1 + red_curve2 + [(0, sh)]
    
    # Green Shape (Triangular fold)
    green_p0 = (110 * scale, 0)
    green_p1 = (140 * scale, 60 * scale)
    green_p2 = (180 * scale, 140 * scale)
    green_p3 = (200 * scale, 240 * scale) # Right vertex
    green_curve1 = bezier_curve(green_p0, green_p1, green_p2, green_p3)
    
    green_p4 = (200 * scale, 240 * scale)
    green_p5 = (150 * scale, 270 * scale)
    green_p6 = (80 * scale, 310 * scale)
    green_p7 = (0, 310 * scale) # Left edge
    green_curve2 = bezier_curve(green_p4, green_p5, green_p6, green_p7)
    
    green_poly = [(0, 0)] + [(110 * scale, 0)] + green_curve1 + green_curve2 + [(0, 310 * scale)]
    
    # Shadow for Red Shape
    draw_shadow.polygon(red_poly, fill=(0, 0, 0, 40))
    
    # Shadow for Green Shape (casts shadow onto the red shape)
    draw_shadow.polygon(green_poly, fill=(0, 0, 0, 50))
    
    # Blur the shadow layer
    blurred_shadow = shadow_layer.filter(ImageFilter.GaussianBlur(radius=6 * scale))
    
    # Offset shadow slightly
    shadow_offset = Image.new("RGBA", (sw, sh), (0, 0, 0, 0))
    shadow_offset.paste(blurred_shadow, (int(3 * scale), int(3 * scale)))
    
    # Draw Red Shape
    red_color = (176, 5, 14, 255) # #B0050E
    draw_red.polygon(red_poly, fill=red_color)
    
    # Draw Green Shape with white outline highlight along both curves
    green_color = (32, 81, 36, 255) # #205124
    
    green_outline_points = green_curve1 + green_curve2
    draw_green.line(green_outline_points, fill=(255, 255, 255, 255), width=8 * scale)
    draw_green.polygon(green_poly, fill=green_color)
    
    # Composite
    final_img = Image.new("RGBA", (sw, sh), (255, 255, 255, 255))
    final_img.alpha_composite(shadow_offset)
    final_img.alpha_composite(red_layer)
    final_img.alpha_composite(green_layer)
    
    # Downsample
    final_img = final_img.resize((width, height), Image.Resampling.LANCZOS)
    
    # Round left corners
    rad = 20
    circle = Image.new('L', (rad * 2, rad * 2), 0)
    draw_circle = ImageDraw.Draw(circle)
    draw_circle.ellipse((0, 0, rad * 2, rad * 2), fill=255)
    
    alpha = Image.new('L', final_img.size, 255)
    alpha.paste(circle.crop((0, 0, rad, rad)), (0, 0))
    alpha.paste(circle.crop((0, rad, rad, rad * 2)), (0, height - rad))
    final_img.putalpha(alpha)
    
    return final_img


def generate_login_background(width, height, app=None):
    from PIL import Image, ImageDraw, ImageFilter
    
    # 1. Try to load the preloaded welcome background image
    raw_img = getattr(app, "_cached_bg_image", None)
    if raw_img is None:
        try:
            img_path = resource_path(os.path.join("settings", "welcome_bg.png"))
            raw_img = Image.open(img_path)
        except Exception as e:
            print("Failed to open welcome_bg.png in generate_login_background:", e)
            raw_img = Image.new("RGB", (width, height), "#F5F5F5")
            
    # 2. Resize to screen dimensions
    bg = raw_img.resize((width, height), Image.Resampling.LANCZOS)
    
    # 3. Draw the waves on the right side of the image (scale = 1 for instant CPU rendering)
    sw = width
    sh = height
    
    red_layer = Image.new("RGBA", (sw, sh), (0, 0, 0, 0))
    green_layer = Image.new("RGBA", (sw, sh), (0, 0, 0, 0))
    shadow_layer = Image.new("RGBA", (sw, sh), (0, 0, 0, 0))
    
    draw_red = ImageDraw.Draw(red_layer)
    draw_green = ImageDraw.Draw(green_layer)
    draw_shadow = ImageDraw.Draw(shadow_layer)
    
    # Red Shape: curves from top-right, bulges to the left, sweeps down to bottom-right
    red_p0 = (sw - 320, 0)
    red_p1 = (sw - 420, sh * 0.25)
    red_p2 = (sw - 380, sh * 0.7)
    red_p3 = (sw - 120, sh)
    red_curve1 = bezier_curve(red_p0, red_p1, red_p2, red_p3)
    red_poly = [red_p0] + red_curve1 + [(sw, sh), (sw, 0)]
    
    # Green Shape: crescent at bottom-right, sweeps from bottom edge to right edge
    green_p0 = (sw - 200, sh)
    green_p1 = (sw - 380, sh * 0.8)
    green_p2 = (sw - 420, sh * 0.45)
    green_p3 = (sw, sh * 0.35)
    green_curve1 = bezier_curve(green_p0, green_p1, green_p2, green_p3)
    green_poly = [green_p0] + green_curve1 + [(sw, sh * 0.35), (sw, sh)]
    
    # Draw shadows onto shadow layer
    draw_shadow.polygon(red_poly, fill=(0, 0, 0, 30))
    draw_shadow.polygon(green_poly, fill=(0, 0, 0, 40))
    
    # Blur shadows with a small fast radius
    blurred_shadow = shadow_layer.filter(ImageFilter.GaussianBlur(radius=8))
    
    # Offset shadow slightly left and down
    shadow_offset = Image.new("RGBA", (sw, sh), (0, 0, 0, 0))
    shadow_offset.paste(blurred_shadow, (-3, 3))
    
    # Draw solid colors
    red_color = (176, 5, 14, 255)  # #B0050E
    draw_red.polygon(red_poly, fill=red_color)
    
    green_color = (32, 81, 36, 255)  # #205124
    green_outline_points = green_curve1
    draw_green.line(green_outline_points, fill=(255, 255, 255, 255), width=2)
    draw_green.polygon(green_poly, fill=green_color)
    
    # Composite layers
    bg_rgba = bg.convert("RGBA")
    bg_rgba.alpha_composite(shadow_offset)
    bg_rgba.alpha_composite(red_layer)
    bg_rgba.alpha_composite(green_layer)
    
    return bg_rgba.convert("RGB")


class AdminLoginPage(ctk.CTkFrame):
    """
    Post-setup login: Uses Admin Name & Password from DB.
    Required every time the app launches if setup is complete.
    """
    def __init__(self, parent, controller):
        super().__init__(parent)
        self.controller = controller

        # Momentary password masking setup
        self.actual_password = ""
        self.mask_char = "●"
        self.mask_job = None
        self.show_pass = False

        self.configure(fg_color="#F5F5F5")
        self._photo = None
        self._resize_job = None

        # Full-window background Canvas
        self._canvas = tk.Canvas(self, bg="#F5F5F5", highlightthickness=0)
        self._canvas.pack(fill="both", expand=True)
        self._img_item = self._canvas.create_image(0, 0, anchor="nw")

        # Bind resize configuration
        self.bind("<Configure>", self._on_resize)

        # White Login Card Container
        self.login_card = ctk.CTkFrame(
            self,
            fg_color="white",
            corner_radius=24,
            border_width=1,
            border_color="#E0E0E0",
            width=420,
            height=610
        )
        self.login_card.place(relx=0.75, rely=0.5, anchor="center")
        self.login_card.pack_propagate(False)

        # ── 1. Logo ──
        try:
            logo_path = resource_path("settings/cherry_signup_logo.png")
            if os.path.exists(logo_path):
                pil_img = Image.open(logo_path)
                self.cherry_logo_img = ctk.CTkImage(pil_img, size=(220, 68))
                logo_lbl = ctk.CTkLabel(self.login_card, text="", image=self.cherry_logo_img, fg_color="transparent")
                logo_lbl.pack(pady=(30, 8))
        except Exception as e:
            print("Error loading cherry logo in login:", e)

        # ── 1.5. Air Gauge Management Label ──
        ctk.CTkLabel(self.login_card, text="Air Gauge Management",
                     font=("Segoe UI", 13, "bold"), text_color="#555555").pack(pady=(0, 10))

        # ── 2. Welcome Back Title ──
        welcome_frame = ctk.CTkFrame(self.login_card, fg_color="transparent")
        welcome_frame.pack(pady=(0, 2))
        
        welcome_lbl1 = ctk.CTkLabel(welcome_frame, text="Welcome ", font=("Segoe UI", 24, "bold"), text_color="#1A1A1A")
        welcome_lbl1.pack(side="left")
        welcome_lbl2 = ctk.CTkLabel(welcome_frame, text="Back!", font=("Segoe UI", 24, "bold"), text_color="#B0050E")
        welcome_lbl2.pack(side="left")

        # ── 3. Subtitle ──
        ctk.CTkLabel(self.login_card, text="Sign in to continue",
                     font=("Segoe UI", 11), text_color="#757575").pack(pady=(0, 12))

        # ── 4. Username Field Container ──
        user_container = ctk.CTkFrame(self.login_card, fg_color="white", border_color="#E0E0E0", border_width=1.5, corner_radius=8, height=48, width=320)
        user_container.pack(pady=6)
        user_container.pack_propagate(False)

        user_icon = ctk.CTkLabel(user_container, text="👤", font=("Segoe UI", 16), text_color="#757575", fg_color="transparent")
        user_icon.pack(side="left", padx=(12, 5))

        self.user_entry = tk.Entry(user_container, relief="flat", bd=0, bg="white", fg="#333",
                                   selectbackground="#B0050E", selectforeground="white",
                                   font=("Segoe UI", 13), insertbackground="#333")
        self.user_entry.pack(side="left", fill="x", expand=True, padx=(2, 12), pady=12)
        self.user_entry.focus()

        # Username Placeholder label
        user_placeholder = ctk.CTkLabel(user_container, text="Username", font=("Segoe UI", 13), text_color="#9E9E9E", fg_color="transparent")
        user_placeholder.place(x=40, y=24, anchor="w")
        user_placeholder.bind("<Button-1>", lambda e: self.user_entry.focus())

        def on_user_key(event=None):
            self.after(10, check_user_placeholder)

        def check_user_placeholder():
            if self.user_entry.get():
                user_placeholder.place_forget()
            else:
                user_placeholder.place(x=40, y=24, anchor="w")

        self.user_entry.bind("<KeyRelease>", on_user_key)

        # ── 5. Password Field Container ──
        pass_container = ctk.CTkFrame(self.login_card, fg_color="white", border_color="#E0E0E0", border_width=1.5, corner_radius=8, height=48, width=320)
        pass_container.pack(pady=6)
        pass_container.pack_propagate(False)

        pass_icon = ctk.CTkLabel(pass_container, text="🔒", font=("Segoe UI", 16), text_color="#757575", fg_color="transparent")
        pass_icon.pack(side="left", padx=(12, 5))

        self.pass_entry = tk.Entry(pass_container, relief="flat", bd=0, bg="white", fg="#333",
                                   selectbackground="#B0050E", selectforeground="white",
                                   font=("Segoe UI", 13), insertbackground="#333")
        self.pass_entry.pack(side="left", fill="x", expand=True, padx=(2, 5), pady=12)

        # Password Placeholder label
        pass_placeholder = ctk.CTkLabel(pass_container, text="Password", font=("Segoe UI", 13), text_color="#9E9E9E", fg_color="transparent")
        pass_placeholder.place(x=40, y=24, anchor="w")
        pass_placeholder.bind("<Button-1>", lambda e: self.pass_entry.focus())

        def on_pass_key(event=None):
            self.after(10, check_pass_placeholder)

        def check_pass_placeholder():
            if self.pass_entry.get():
                pass_placeholder.place_forget()
            else:
                pass_placeholder.place(x=40, y=24, anchor="w")

        # Hook validation for momentary password masking
        pass_vcmd = (self.register(self._validate_pass_input), '%d', '%i', '%P', '%s', '%S')
        self.pass_entry.config(validate="key", validatecommand=pass_vcmd)

        self.pass_entry.bind("<KeyRelease>", on_pass_key)

        # Show/Hide eye button
        def toggle_pass_visibility():
            self.show_pass = not self.show_pass
            if self.show_pass:
                self._set_pass_display_text(self.actual_password)
                eye_btn.configure(text="👁")
            else:
                self._set_pass_display_text(self.mask_char * len(self.actual_password))
                eye_btn.configure(text="👁‍🗨")

        eye_btn = ctk.CTkLabel(pass_container, text="👁‍🗨", font=("Segoe UI", 16), text_color="#757575", fg_color="transparent", cursor="hand2")
        eye_btn.pack(side="right", padx=(5, 12))
        eye_btn.bind("<Button-1>", lambda e: toggle_pass_visibility())

        # Focus ring bindings
        self.user_entry.bind("<FocusIn>", lambda e: [user_container.configure(border_color="#B0050E"), user_placeholder.place_forget()])
        self.user_entry.bind("<FocusOut>", lambda e: [user_container.configure(border_color="#E0E0E0"), check_user_placeholder()])
        
        self.pass_entry.bind("<FocusIn>", lambda e: [pass_container.configure(border_color="#B0050E"), pass_placeholder.place_forget()])
        self.pass_entry.bind("<FocusOut>", lambda e: [pass_container.configure(border_color="#E0E0E0"), check_pass_placeholder()])

        # Return bindings
        self.user_entry.bind("<Return>", lambda event: self.pass_entry.focus())
        self.pass_entry.bind("<Return>", lambda event: self.check_login())

        # ── 6. Forgot Password Row ──
        forgot_frame = ctk.CTkFrame(self.login_card, fg_color="transparent", width=320, height=25)
        forgot_frame.pack(pady=(4, 10))
        forgot_frame.pack_propagate(False)

        forgot_lbl = ctk.CTkLabel(
            forgot_frame,
            text="Forgot Password?",
            font=("Segoe UI", 11, "bold"),
            text_color="#205124",
            cursor="hand2"
        )
        forgot_lbl.pack(side="right")
        forgot_lbl.bind("<Button-1>", lambda e: self.open_forgot_password())

        # ── 7. Error Label ──
        self.error_label = ctk.CTkLabel(self.login_card, text="", text_color="#D32F2F",
                                        font=("Segoe UI", 11), fg_color="transparent")
        self.error_label.pack(pady=(2, 2))

        # ── 8. Login Button ──
        self.login_btn = ModernButton(self.login_card, text="🔒  LOGIN",
                                       font=("Segoe UI", 13, "bold"),
                                       height=48, width=320,
                                       fg_color="#B0050E", hover_color="#90030A",
                                       corner_radius=8,
                                       command=self.check_login)
        self.login_btn.pack(pady=10)

        # ── 9. OR Separator ──
        or_frame = ctk.CTkFrame(self.login_card, fg_color="transparent", width=320, height=20)
        or_frame.pack(pady=(0, 4))
        or_frame.pack_propagate(False)

        or_left = ctk.CTkFrame(or_frame, fg_color="#E0E0E0", height=1)
        or_left.place(relx=0.0, rely=0.5, relwidth=0.38, anchor="w")

        ctk.CTkLabel(or_frame, text="OR", font=("Segoe UI", 10, "bold"),
                     text_color="#9E9E9E", fg_color="transparent").place(relx=0.5, rely=0.5, anchor="center")

        or_right = ctk.CTkFrame(or_frame, fg_color="#E0E0E0", height=1)
        or_right.place(relx=1.0, rely=0.5, relwidth=0.38, anchor="e")

        # ── 10. Create Account Button ──
        self.create_btn = ModernButton(
            self.login_card,
            text="👤  CREATE ACCOUNT",
            font=("Segoe UI", 13, "bold"),
            height=48, width=320,
            fg_color="white",
            hover_color="#F1F8E9",
            text_color="#205124",
            border_color="#205124",
            border_width=1.5,
            corner_radius=8,
            command=self.start_registration
        )
        self.create_btn.pack(pady=(0, 10))

        # ── 11. Secure Terminal v1.0 label ──
        ctk.CTkLabel(
            self.login_card,
            text="Secure Terminal v1.0",
            font=("Segoe UI", 11),
            text_color="#9E9E9E"
        ).place(relx=0.5, rely=0.95, anchor="center")

    def _on_resize(self, event=None):
        if self._resize_job:
            self.after_cancel(self._resize_job)
        self._resize_job = self.after(50, self._do_resize)

    def _do_resize(self):
        self._resize_job = None
        w = self.winfo_width()
        h = self.winfo_height()
        if w < 100 or h < 100:
            return
        try:
            bg_img = generate_login_background(w, h, self.controller)
            from PIL import ImageTk
            self._photo = ImageTk.PhotoImage(bg_img)
            self._canvas.itemconfig(self._img_item, image=self._photo)
        except Exception as e:
            print("AdminLoginPage background resize error:", e)

    def _validate_pass_input(self, action, index, value_if_allowed, prior_value, text_inserted_deleted):
        idx = int(index)
        action = int(action)

        if action == 1: # Insert
            self.actual_password = self.actual_password[:idx] + text_inserted_deleted + self.actual_password[idx:]
        elif action == 0: # Delete
            self.actual_password = self.actual_password[:idx] + self.actual_password[idx + len(text_inserted_deleted):]

        # Schedule the display update
        if self.show_pass:
            self.after_idle(lambda: self._set_pass_display_text(self.actual_password))
        else:
            if action == 1 and len(text_inserted_deleted) == 1:
                masked = "".join(self.mask_char if i != idx else c for i, c in enumerate(self.actual_password))
                self.after_idle(lambda: self._set_pass_display_text(masked))
                self._schedule_pass_masking()
            else:
                masked = self.mask_char * len(self.actual_password)
                self.after_idle(lambda: self._set_pass_display_text(masked))

        return True

    def _set_pass_display_text(self, text):
        self.pass_entry.config(validate="none")
        cursor_pos = self.pass_entry.index(tk.INSERT)
        self.pass_entry.delete(0, tk.END)
        self.pass_entry.insert(0, text)
        self.pass_entry.icursor(cursor_pos)
        self.pass_entry.config(validate="key")

    def _schedule_pass_masking(self):
        if self.mask_job:
            self.after_cancel(self.mask_job)
        self.mask_job = self.after(1500, self._mask_all_pass)

    def _mask_all_pass(self):
        self.mask_job = None
        if not self.show_pass:
            masked = self.mask_char * len(self.actual_password)
            self._set_pass_display_text(masked)

    def check_login(self):
        admin_name_input = self.user_entry.get().strip()
        password_input   = self.actual_password.strip()

        stored_admin = SetupDatabase.get_admin_name()
        stored_pass  = SetupDatabase.get_admin_password()

        is_user_valid   = (admin_name_input == stored_admin and password_input == stored_pass)
        is_master_valid = (admin_name_input == "cherry"    and password_input == "cherry@123")

        if is_user_valid or is_master_valid:
            # No remember me saving needed

            self.controller.complete_login()
        else:
            self.error_label.configure(text="Invalid Admin Name or Password")

    def open_forgot_password(self):
        ForgotPasswordDialog(self, self.controller)

    def start_registration(self):
        """Transition from login page to the setup/registration wizard."""
        self.controller.start_setup_from_login()



# ─────────────────────────────────────────────────────────────────────────────
#  Forgot Password Dialog
# ─────────────────────────────────────────────────────────────────────────────
class ForgotPasswordDialog(ctk.CTkToplevel):
    """
    Two-panel modal:
      Left  – Cherry service contact information
      Right – Cherry Service ID & Password entry
    """
    def __init__(self, parent, controller):
        super().__init__(parent)
        self.withdraw()   # Hide window during construction to prevent layout flash glitch
        self.controller  = controller
        self.parent_page = parent

        # Momentary password masking setup
        self.pw_actual_password = ""
        self.pw_mask_char = "●"
        self.pw_mask_job = None
        self.pw_show_pass = False



        self.title("Forgot Password — Cherry Precision Service")
        self.configure(fg_color="white")
        self.transient(parent)
        self.grab_set()

        # Set window icon bitmap
        try:
            icon_path = resource_path("cherry_precision_transparent.ico")
            if os.path.exists(icon_path):
                self.after(200, lambda: self.iconbitmap(icon_path))
        except:
            pass

        # ── Outer card ───────────────────────────────────────
        card = ctk.CTkFrame(self, fg_color="white", corner_radius=0)
        card.pack(fill="both", expand=True)

        # ── LEFT: Unified contact info & wavy separator ─────────────────
        left_canvas = tk.Canvas(card, width=370, height=500, bg="white", highlightthickness=0)
        left_canvas.pack(side="left", fill="y")

        # Draw green background area with S-curve boundary on the right
        import math
        points = [(0, 0)]
        for y in range(0, 505, 5):
            x = 340 + 18 + 12 * math.cos(y * 2 * math.pi / 500)
            points.append((x, y))
        points.append((0, 500))

        flat_points = []
        for pt in points:
            flat_points.extend(pt)

        left_canvas.create_polygon(flat_points, fill="#1B5E20", outline="#1B5E20")

        # Draw red ribbon line
        line_points = []
        for y in range(0, 505, 5):
            x = 340 + 18 + 12 * math.cos(y * 2 * math.pi / 500)
            line_points.extend((x, y))
        left_canvas.create_line(line_points, fill="#D32F2F", width=3, smooth=True)

        # Content frame inside canvas
        content_frame = tk.Frame(left_canvas, bg="#1B5E20")

        # Lifebuoy Canvas Badge (White circle with red lifebuoy inside)
        badge_canvas = tk.Canvas(content_frame, width=70, height=70, bg="#1B5E20", highlightthickness=0)
        badge_canvas.pack(pady=(28, 4))
        # White circle background
        badge_canvas.create_oval(5, 5, 65, 65, fill="white", outline="")
        # Red ring of lifebuoy
        badge_canvas.create_oval(15, 15, 55, 55, outline="#C62828", width=8, fill="white")
        # Center white donut hole
        badge_canvas.create_oval(24, 24, 46, 46, fill="white", outline="#C62828", width=2)
        # Four red bands
        badge_canvas.create_rectangle(32, 15, 38, 25, fill="#C62828", outline="")
        badge_canvas.create_rectangle(32, 45, 38, 55, fill="#C62828", outline="")
        badge_canvas.create_rectangle(15, 32, 25, 38, fill="#C62828", outline="")
        badge_canvas.create_rectangle(45, 32, 55, 38, fill="#C62828", outline="")

        ctk.CTkLabel(content_frame, text="Need Help?",
                     font=("Segoe UI", 18, "bold"), text_color="white", fg_color="transparent").pack(pady=(4, 4))
        ctk.CTkLabel(content_frame, text="Cherry Precision Service Support",
                     font=("Segoe UI", 11), text_color="#A5D6A7", fg_color="transparent").pack(pady=(0, 14))
        ctk.CTkFrame(content_frame, fg_color="#4CAF50", height=1, width=280, corner_radius=0).pack(pady=(0, 16))

        body = ctk.CTkFrame(content_frame, fg_color="transparent")
        body.pack(fill="x", padx=22)

        intro = ("If you have forgotten your admin password,\n"
                 "please contact your Cherry Precision\n"
                 "service representative. They will provide\n"
                 "you with a Cherry Service ID & Password\n"
                 "to reset your admin credentials.")
        ctk.CTkLabel(body, text=intro, font=("Segoe UI", 11),
                     text_color="#C8E6C9", fg_color="transparent", justify="left",
                     wraplength=280).pack(anchor="w", pady=(0, 18))

        # Contacts list
        contacts = [
            ("📞", "+91-98765-43210"),
            ("✉️", "service@cherryprecision.com"),
            ("🕒", "Mon-Sat | 9:00 AM - 6:00 PM"),
            ("🏢", "Cherry Precision Products"),
        ]
        for icon, value in contacts:
            row = ctk.CTkFrame(body, fg_color="transparent")
            row.pack(fill="x", pady=5)
            ctk.CTkLabel(row, text=icon, font=("Segoe UI", 14),
                         text_color="#A5D6A7", fg_color="transparent", width=30, anchor="w").pack(side="left")
            ctk.CTkLabel(row, text=value, font=("Segoe UI", 11),
                         text_color="white", fg_color="transparent", anchor="w").pack(side="left")

        # Place Content Frame on Canvas
        left_canvas.create_window(170, 250, window=content_frame, anchor="center")

        # ── RIGHT: credentials entry ─────────────────────────
        right = ctk.CTkFrame(card, fg_color="white", corner_radius=0)
        right.pack(side="right", fill="both", expand=True)

        # Top Header Bar
        header_bar = ctk.CTkFrame(right, fg_color="white", height=45, corner_radius=0)
        header_bar.pack(fill="x", side="top")
        
        # Header Label (no logo)
        hdr_title = ctk.CTkLabel(header_bar, text="Forgot Password — Cherry Precision Service",
                                  font=("Segoe UI", 11, "bold"), text_color="#333", fg_color="transparent")
        hdr_title.pack(side="left", padx=25, pady=8)
            
        border_line = ctk.CTkFrame(right, fg_color="#E0E0E0", height=1, corner_radius=0)
        border_line.pack(fill="x")

        # Form container
        inner = ctk.CTkFrame(right, fg_color="white", corner_radius=0)
        inner.pack(expand=True, padx=40, pady=(10, 20))

        # Title: Enter Cherry Service Credentials
        title_frame = ctk.CTkFrame(inner, fg_color="transparent")
        title_frame.pack(pady=(0, 2))
        ctk.CTkLabel(title_frame, text="Enter ", font=("Segoe UI", 18, "bold"), text_color="#1B5E20").pack(side="left")
        ctk.CTkLabel(title_frame, text="Cherry ", font=("Segoe UI", 18, "bold"), text_color="#C62828").pack(side="left")
        ctk.CTkLabel(title_frame, text="Service Credentials", font=("Segoe UI", 18, "bold"), text_color="#1B5E20").pack(side="left")

        ctk.CTkLabel(inner,
                     text="Use the Cherry Service ID & Password\nprovided by your service representative.",
                     font=("Segoe UI", 11), text_color="#757575", fg_color="transparent",
                     justify="center").pack(pady=(0, 22))

        # ID Field
        ctk.CTkLabel(inner, text="Cherry Service ID", font=("Segoe UI", 11, "bold"),
                     text_color="#333", fg_color="transparent", anchor="w").pack(fill="x")
                     
        id_container = ctk.CTkFrame(inner, fg_color="white", border_color="#B0BEC5", border_width=1, corner_radius=8, height=44, width=320)
        id_container.pack(pady=(4, 12))
        id_container.pack_propagate(False)

        id_icon = ctk.CTkLabel(id_container, text="👤", font=("Segoe UI", 16), text_color="#757575", fg_color="transparent")
        id_icon.pack(side="left", padx=(12, 5))

        self.id_entry = tk.Entry(id_container, relief="flat", bd=0, bg="white", fg="#333",
                                 selectbackground="#1B5E20", selectforeground="white",
                                 font=("Segoe UI", 13), insertbackground="#333")
        self.id_entry.pack(side="left", fill="x", expand=True, padx=(2, 12), pady=10)
        self.id_entry.focus()

        id_placeholder = ctk.CTkLabel(id_container, text="Enter service ID", font=("Segoe UI", 13), text_color="#9E9E9E", fg_color="transparent")
        id_placeholder.place(x=40, y=22, anchor="w")
        id_placeholder.bind("<Button-1>", lambda e: self.id_entry.focus())

        def on_id_key(event=None):
            self.after(10, check_id_placeholder)

        def check_id_placeholder():
            if self.id_entry.get():
                id_placeholder.place_forget()
            else:
                id_placeholder.place(x=40, y=22, anchor="w")

        self.id_entry.bind("<KeyRelease>", on_id_key)

        # PW Field
        ctk.CTkLabel(inner, text="Cherry Service Password", font=("Segoe UI", 11, "bold"),
                     text_color="#333", fg_color="transparent", anchor="w").pack(fill="x")
                     
        pw_container = ctk.CTkFrame(inner, fg_color="white", border_color="#B0BEC5", border_width=1, corner_radius=8, height=44, width=320)
        pw_container.pack(pady=(4, 12))
        pw_container.pack_propagate(False)

        pw_icon = ctk.CTkLabel(pw_container, text="🔒", font=("Segoe UI", 16), text_color="#757575", fg_color="transparent")
        pw_icon.pack(side="left", padx=(12, 5))

        self.pw_entry = tk.Entry(pw_container, relief="flat", bd=0, bg="white", fg="#333",
                                 selectbackground="#1B5E20", selectforeground="white",
                                 font=("Segoe UI", 13), insertbackground="#333")
        self.pw_entry.pack(side="left", fill="x", expand=True, padx=(2, 5), pady=10)

        pw_placeholder = ctk.CTkLabel(pw_container, text="Enter service password", font=("Segoe UI", 13), text_color="#9E9E9E", fg_color="transparent")
        pw_placeholder.place(x=40, y=22, anchor="w")
        pw_placeholder.bind("<Button-1>", lambda e: self.pw_entry.focus())

        def on_pw_key(event=None):
            self.after(10, check_pw_placeholder)

        def check_pw_placeholder():
            if self.pw_entry.get():
                pw_placeholder.place_forget()
            else:
                pw_placeholder.place(x=40, y=22, anchor="w")

        # Hook validation for password masking
        pw_vcmd = (self.register(self._validate_forgot_pw_input), '%d', '%i', '%P', '%s', '%S')
        self.pw_entry.config(validate="key", validatecommand=pw_vcmd)
        self.pw_entry.bind("<KeyRelease>", on_pw_key)

        # Show/Hide eye button
        def toggle_pw_visibility():
            self.pw_show_pass = not self.pw_show_pass
            if self.pw_show_pass:
                self._set_forgot_pw_display_text(self.pw_actual_password)
                eye_btn.configure(text="👁")
            else:
                self._set_forgot_pw_display_text(self.pw_mask_char * len(self.pw_actual_password))
                eye_btn.configure(text="👁‍🗨")

        eye_btn = ctk.CTkLabel(pw_container, text="👁‍🗨", font=("Segoe UI", 16), text_color="#757575", fg_color="transparent", cursor="hand2")
        eye_btn.pack(side="right", padx=(5, 12))
        eye_btn.bind("<Button-1>", lambda e: toggle_pw_visibility())

        # Focus ring bindings
        self.id_entry.bind("<FocusIn>", lambda e: [id_container.configure(border_color="#1B5E20"), id_placeholder.place_forget()])
        self.id_entry.bind("<FocusOut>", lambda e: [id_container.configure(border_color="#B0BEC5"), check_id_placeholder()])
        
        self.pw_entry.bind("<FocusIn>", lambda e: [pw_container.configure(border_color="#1B5E20"), pw_placeholder.place_forget()])
        self.pw_entry.bind("<FocusOut>", lambda e: [pw_container.configure(border_color="#B0BEC5"), check_pw_placeholder()])

        self.id_entry.bind("<Return>", lambda event: self.pw_entry.focus())
        self.pw_entry.bind("<Return>", lambda event: self.on_verify())

        self.err_lbl = ctk.CTkLabel(inner, text="", font=("Segoe UI", 11),
                                    text_color="#C62828", fg_color="transparent", wraplength=320, justify="center")
        self.err_lbl.pack(pady=(0, 8))

        # Solid red verify button
        ModernButton(inner, text="🛡️  VERIFY & RESET PASSWORD",
                      font=("Segoe UI", 13, "bold"),
                      height=46, width=320,
                      fg_color="#B71C1C", hover_color="#990000",
                      corner_radius=8,
                      command=self.on_verify).pack(pady=4)

        # Green outlined cancel button
        ModernButton(inner, text="Cancel",
                      font=("Segoe UI", 11), height=36, width=320,
                      fg_color="transparent", hover_color="#E8F5E9",
                      text_color="#1B5E20", border_width=1, border_color="#1B5E20",
                      corner_radius=8,
                      command=self.destroy).pack(pady=(6, 0))

        # Set explicit size and centre on screen
        W, H = 820, 500
        x = (self.winfo_screenwidth()  - W) // 2
        y = (self.winfo_screenheight() - H) // 2
        self.geometry(f"{W}x{H}+{x}+{y}")
        self.update_idletasks()
        self.resizable(False, False)
        self.deiconify()  # Reveal the window fully centered
        self.lift()
        self.focus_force()

    def _validate_forgot_pw_input(self, action, index, value_if_allowed, prior_value, text_inserted_deleted):
        idx = int(index)
        action = int(action)

        if action == 1: # Insert
            self.pw_actual_password = self.pw_actual_password[:idx] + text_inserted_deleted + self.pw_actual_password[idx:]
        elif action == 0: # Delete
            self.pw_actual_password = self.pw_actual_password[:idx] + self.pw_actual_password[idx + len(text_inserted_deleted):]

        # Schedule the display update
        if self.pw_show_pass:
            self.after_idle(lambda: self._set_forgot_pw_display_text(self.pw_actual_password))
        else:
            if action == 1 and len(text_inserted_deleted) == 1:
                masked = "".join(self.pw_mask_char if i != idx else c for i, c in enumerate(self.pw_actual_password))
                self.after_idle(lambda: self._set_forgot_pw_display_text(masked))
                self._schedule_forgot_pw_masking()
            else:
                masked = self.pw_mask_char * len(self.pw_actual_password)
                self.after_idle(lambda: self._set_forgot_pw_display_text(masked))

        return True

    def _set_forgot_pw_display_text(self, text):
        self.pw_entry.config(validate="none")
        cursor_pos = self.pw_entry.index(tk.INSERT)
        self.pw_entry.delete(0, tk.END)
        self.pw_entry.insert(0, text)
        self.pw_entry.icursor(cursor_pos)
        self.pw_entry.config(validate="key")

    def _schedule_forgot_pw_masking(self):
        if self.pw_mask_job:
            self.after_cancel(self.pw_mask_job)
        self.pw_mask_job = self.after(1500, self._mask_all_forgot_pw)

    def _mask_all_forgot_pw(self):
        self.pw_mask_job = None
        if not self.pw_show_pass:
            masked = self.pw_mask_char * len(self.pw_actual_password)
            self._set_forgot_pw_display_text(masked)

    def on_verify(self):
        cherry_id = self.id_entry.get().strip()
        cherry_pw = self.pw_actual_password.strip()

        if not cherry_id or not cherry_pw:
            self.err_lbl.configure(text="Please enter both Cherry Service ID and Password.")
            return

        if not SetupDatabase.validate_cherry_credentials(cherry_id, cherry_pw):
            self.err_lbl.configure(text="Invalid Cherry Service ID or Password.\nPlease check and try again.")
            return

        use_count = SetupDatabase.get_cherry_use_count(cherry_id)
        if use_count >= SetupDatabase.MAX_CHERRY_USES:
            self.err_lbl.configure(
                text=f"This Cherry Service ID has already been\n"
                     f"used {SetupDatabase.MAX_CHERRY_USES} times and is no longer valid.\n"
                     f"Please contact Cherry service for a new ID."
            )
            return

        # Valid — open reset dialog
        self.destroy()
        ResetPasswordDialog(self.parent_page, self.controller, cherry_id)


# ─────────────────────────────────────────────────────────────────────────────
#  Reset Password Dialog
# ─────────────────────────────────────────────────────────────────────────────
class ResetPasswordDialog(ctk.CTkToplevel):
    """
    Allows setting a new Admin Username + Password after presenting
    a valid Cherry Service ID. Increments the Cherry ID use-count on save.
    """
    def __init__(self, parent, controller, cherry_id):
        super().__init__(parent)
        self.withdraw()   # Hide window during construction to prevent layout flash glitch
        self.controller = controller
        self.cherry_id  = cherry_id

        self.title("Reset Admin Password — Cherry Precision")
        self.configure(fg_color="white")
        self.transient(parent)
        self.grab_set()

        # Set window icon bitmap
        try:
            icon_path = resource_path("cherry_precision_transparent.ico")
            if os.path.exists(icon_path):
                self.after(200, lambda: self.iconbitmap(icon_path))
        except:
            pass

        card = ctk.CTkFrame(self, fg_color="white", corner_radius=0)
        card.pack(fill="both", expand=True, padx=44, pady=30)

        ctk.CTkLabel(card, text="Set New Admin Credentials",
                     font=("Segoe UI", 18, "bold"), text_color="#1B5E20", fg_color="transparent").pack(pady=(0, 6))
        ctk.CTkLabel(card,
                     text="Set a new admin username and a strong password.\n"
                          "Your previous credentials will be deactivated after saving.",
                     font=("Segoe UI", 11), text_color="#555", fg_color="transparent",
                     justify="center").pack(pady=(0, 20))

        ctk.CTkLabel(card, text="New Admin Username", font=("Segoe UI", 11, "bold"),
                     text_color="#333", fg_color="transparent", anchor="w").pack(fill="x")
        self.name_ent = ctk.CTkEntry(card, placeholder_text="Enter new admin username",
                                     height=44, width=360, font=("Segoe UI", 13),
                                     border_color="#B0BEC5", corner_radius=8)
        self.name_ent.pack(pady=(4, 10))
        self.name_ent.focus()

        ctk.CTkLabel(card, text="New Password", font=("Segoe UI", 11, "bold"),
                     text_color="#333", fg_color="transparent", anchor="w").pack(fill="x")
        self.pass_ent = ctk.CTkEntry(card, placeholder_text="Enter new password",
                                     show="●", height=44, width=360,
                                     font=("Segoe UI", 13),
                                     border_color="#B0BEC5", corner_radius=8)
        self.pass_ent.pack(pady=(4, 10))
        self.pass_ent.bind("<KeyRelease>", self._validate)

        ctk.CTkLabel(card, text="Confirm Password", font=("Segoe UI", 11, "bold"),
                     text_color="#333", fg_color="transparent", anchor="w").pack(fill="x")
        self.conf_ent = ctk.CTkEntry(card, placeholder_text="Re-enter new password",
                                     show="●", height=44, width=360,
                                     font=("Segoe UI", 13),
                                     border_color="#B0BEC5", corner_radius=8)
        self.conf_ent.pack(pady=(4, 10))
        self.conf_ent.bind("<KeyRelease>", self._validate)

        self.name_ent.bind("<Return>", lambda event: self.pass_ent.focus())
        self.pass_ent.bind("<Return>", lambda event: self.conf_ent.focus())
        self.conf_ent.bind("<Return>", lambda event: self.save_btn.invoke() if self.save_btn.cget('state') == 'normal' else None)

        # Password rules checklist
        rules_frame = ctk.CTkFrame(card, fg_color="white", corner_radius=0)
        rules_frame.pack(fill="x", pady=(0, 8))
        self._rule_labels = {}
        for key, text in [("upper",   "One uppercase letter"),
                          ("lower",   "One lowercase letter"),
                          ("number",  "One number"),
                          ("special", "One special character"),
                          ("match",   "Passwords match")]:
            lbl = ctk.CTkLabel(rules_frame, text=f"❌  {text}",
                               font=("Segoe UI", 11), text_color="#757575", fg_color="transparent", anchor="w")
            lbl.pack(fill="x")
            self._rule_labels[key] = lbl

        self.err_lbl = ctk.CTkLabel(card, text="", font=("Segoe UI", 11),
                                    text_color="#C62828", fg_color="transparent")
        self.err_lbl.pack(pady=(0, 6))

        self.save_btn = ModernButton(card, text="SAVE NEW CREDENTIALS",
                                      font=("Segoe UI", 13, "bold"),
                                      height=46, width=360,
                                      fg_color="#2E7D32", hover_color="#1B5E20",
                                      corner_radius=8,
                                      state="disabled",
                                      command=self.on_save)
        self.save_btn.pack(pady=4)

        ModernButton(card, text="Cancel",
                      font=("Segoe UI", 11), height=36, width=360,
                      fg_color="transparent", hover_color="#E8F5E9",
                      text_color="#555", border_width=1, border_color="#BDBDBD",
                      corner_radius=8,
                      command=self.destroy).pack(pady=(6, 0))

        self.update_idletasks()
        # Use explicit size to avoid winfo_reqwidth returning 0 before render
        W, H = 480, 620
        x = (self.winfo_screenwidth()  - W) // 2
        y = (self.winfo_screenheight() - H) // 2
        self.geometry(f"{W}x{H}+{x}+{y}")
        self.update_idletasks()
        self.resizable(False, False)
        self.deiconify()  # Reveal the window fully centered
        self.lift()
        self.focus_force()


    def _validate(self, event=None):
        pw   = self.pass_ent.get()
        conf = self.conf_ent.get()

        checks = [
            ("upper",   bool(re.search(r'[A-Z]',        pw)), "One uppercase letter"),
            ("lower",   bool(re.search(r'[a-z]',        pw)), "One lowercase letter"),
            ("number",  bool(re.search(r'\d',           pw)), "One number"),
            ("special", bool(re.search(r'[^A-Za-z0-9]', pw)), "One special character"),
            ("match",   (pw == conf) and len(pw) > 0,          "Passwords match"),
        ]
        for key, ok, text in checks:
            self._rule_labels[key].configure(
                text=f"{'✅' if ok else '❌'}  {text}",
                text_color="#2E7D32" if ok else "#757575"
            )
        self.save_btn.configure(state="normal" if all(ok for _, ok, _ in checks) else "disabled")

    def on_save(self):
        new_name = self.name_ent.get().strip()
        new_pass = self.pass_ent.get().strip()

        if not new_name:
            self.err_lbl.configure(text="Admin username cannot be empty.")
            return
        try:
            company = SetupDatabase.get_company_name()
            conn    = sqlite3.connect(resource_path(SetupDatabase.DB_FILE))
            cursor  = conn.cursor()
            cursor.execute("SELECT address FROM setup_info LIMIT 1")
            row = cursor.fetchone()
            conn.close()
            address = row[0] if row and row[0] else ""

            SetupDatabase.save_setup_info(company, address, new_pass, new_name)
            SetupDatabase.increment_cherry_use(self.cherry_id)

            messagebox.showinfo(
                "Password Reset Successful",
                "Your admin credentials have been updated successfully.\n\n"
                "Please log in with your new username and password.",
                parent=self
            )
            self.destroy()
        except Exception as e:
            self.err_lbl.configure(text=f"Error saving credentials: {e}")


class SetupDatabase:
    DB_FILE = "app_data.db"

    # --- 10 Cherry Service Accounts (hardcoded, give one to each customer) ---
    CHERRY_SERVICE_ACCOUNTS = {
        "CHERRY-SVC-001": "Svc@Cherry#01",
        "CHERRY-SVC-002": "Svc@Cherry#02",
        "CHERRY-SVC-003": "Svc@Cherry#03",
        "CHERRY-SVC-004": "Svc@Cherry#04",
        "CHERRY-SVC-005": "Svc@Cherry#05",
        "CHERRY-SVC-006": "Svc@Cherry#06",
        "CHERRY-SVC-007": "Svc@Cherry#07",
        "CHERRY-SVC-008": "Svc@Cherry#08",
        "CHERRY-SVC-009": "Svc@Cherry#09",
        "CHERRY-SVC-010": "Svc@Cherry#10",
    }
    MAX_CHERRY_USES = 3  # Each Cherry ID can be used at most 3 times

    @staticmethod
    def init_db():
        conn = sqlite3.connect(resource_path(SetupDatabase.DB_FILE))
        cursor = conn.cursor()
        # Create table with admin_name
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS setup_info (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_name TEXT,
                address TEXT,
                admin_password TEXT,
                admin_name TEXT
            )
        """)

        # Cherry service usage tracking table
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS cherry_service_usage (
                cherry_id TEXT PRIMARY KEY,
                use_count INTEGER DEFAULT 0
            )
        """)

        # New Master Tables
        conn_emp = sqlite3.connect(resource_path("employee_master.db"))
        conn_emp.cursor().execute("""
            CREATE TABLE IF NOT EXISTS employee_master (
                id TEXT PRIMARY KEY,
                name TEXT,
                description TEXT,
                phone TEXT
            )
        """)
        conn_emp.commit()
        conn_emp.close()

        conn_mac = sqlite3.connect(resource_path("machine_master.db"))
        conn_mac.cursor().execute("""
            CREATE TABLE IF NOT EXISTS machine_master (
                code TEXT PRIMARY KEY,
                name TEXT,
                description TEXT
            )
        """)
        conn_mac.commit()
        conn_mac.close()

        conn_cust = sqlite3.connect(resource_path("customers.db"))
        conn_cust.cursor().execute("""
            CREATE TABLE IF NOT EXISTS customers (
                code TEXT PRIMARY KEY,
                name TEXT,
                description TEXT,
                email TEXT,
                phone TEXT
            )
        """)
        conn_cust.commit()
        conn_cust.close()

        conn_item = sqlite3.connect(resource_path("items.db"))
        conn_item.cursor().execute("""
            CREATE TABLE IF NOT EXISTS items (
                code TEXT PRIMARY KEY,
                drawing_no TEXT,
                name TEXT,
                revision_no TEXT
            )
        """)
        conn_item.commit()
        conn_item.close()

        conn_proc = sqlite3.connect(resource_path("process_master.db"))
        conn_proc.cursor().execute("""
            CREATE TABLE IF NOT EXISTS process_master (
                id TEXT PRIMARY KEY,
                name TEXT,
                description TEXT
            )
        """)
        conn_proc.commit()
        conn_proc.close()

        conn_comp = sqlite3.connect(resource_path("component_setup.db"))
        conn_comp.cursor().execute("""
            CREATE TABLE IF NOT EXISTS component_setup (
                airgauge_id TEXT,
                channel TEXT,
                properties TEXT,
                PRIMARY KEY(airgauge_id, channel)
            )
        """)
        conn_comp.commit()
        conn_comp.close()

        conn_ag = sqlite3.connect(resource_path("airgauge_master.db"))
        conn_ag.cursor().execute("""
            CREATE TABLE IF NOT EXISTS airgauges (
                id TEXT PRIMARY KEY,
                name TEXT,
                description TEXT
            )
        """)
        conn_ag.commit()
        conn_ag.close()
        
        # Unify production_data.db creation here too
        conn_prod = sqlite3.connect(resource_path("production_data.db"))
        conn_prod.cursor().execute('''
            CREATE TABLE IF NOT EXISTS measurements (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                source TEXT,
                date TEXT,
                time TEXT,
                reading TEXT,
                offset TEXT,
                status TEXT,
                airgauge_id TEXT,
                channel TEXT,
                drawing TEXT,
                user_id TEXT,
                component_id TEXT,
                item TEXT,
                machine_id TEXT,
                customer TEXT,
                utl TEXT,
                ltl TEXT,
                upload_timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Add columns if they do not exist
        try:
            conn_prod.cursor().execute("SELECT utl, ltl FROM measurements LIMIT 1")
        except sqlite3.OperationalError:
            try:
                conn_prod.cursor().execute("ALTER TABLE measurements ADD COLUMN utl TEXT")
                conn_prod.cursor().execute("ALTER TABLE measurements ADD COLUMN ltl TEXT")
            except Exception:
                pass

        conn_prod.commit()
        conn_prod.close()
                
        # Migration: Check if admin_name column exists (for existing DBs)
        try:
            cursor.execute("SELECT admin_name FROM setup_info LIMIT 1")
        except sqlite3.OperationalError:
            # Column missing, add it
            try:
                cursor.execute("ALTER TABLE setup_info ADD COLUMN admin_name TEXT")
            except:
                pass # Already exists or other error
                
        conn.commit()
        conn.close()
        # Port existing tables from app_data.db to new separate DBs seamlessly
        SetupDatabase._port_v1_tables_to_separate()
        SetupDatabase._port_v1_airgauge_config()
        
        # Trigger one-time migration from JSON files to SQLite (for older JSON users)
        SetupDatabase.migrate_json_to_sqlite()
        SetupDatabase.migrate_airgauge_config_json()

    @staticmethod
    def _port_v1_airgauge_config():
        """Safely transfers airgauge config from monolith app_data.db if it was there, 
           or just ensures the new db is reachable."""
        monolith_path = resource_path(SetupDatabase.DB_FILE)
        if not os.path.exists(monolith_path):
            return
        try:
            old_conn = sqlite3.connect(monolith_path)
            old_cursor = old_conn.cursor()
            try:
                old_cursor.execute("SELECT * FROM airgauge_master")
                rows = old_cursor.fetchall()
                if rows:
                    conn = sqlite3.connect(resource_path("airgauge_master.db"))
                    conn.executemany("INSERT OR IGNORE INTO airgauge_master VALUES (?, ?)", rows)
                    conn.commit()
                    conn.close()
                old_cursor.execute("DROP TABLE airgauge_master")
                old_conn.commit()
            except: pass
            old_conn.close()
        except: pass

    @staticmethod
    def _port_v1_tables_to_separate():
        """Safely transfers data from monolithic app_data.db out to the independent DBs."""
        monolith_path = resource_path(SetupDatabase.DB_FILE)
        if not os.path.exists(monolith_path):
            return
        try:
            old_conn = sqlite3.connect(monolith_path)
            old_cursor = old_conn.cursor()

            # Employee Master
            try:
                old_cursor.execute("SELECT * FROM employee_master")
                rows = old_cursor.fetchall()
                if rows:
                    conn = sqlite3.connect(resource_path("employee_master.db"))
                    conn.executemany("INSERT OR IGNORE INTO employee_master VALUES (?, ?, ?, ?)", rows)
                    conn.commit()
                    conn.close()
                old_cursor.execute("DROP TABLE employee_master")
            except: pass

            # Machine Master
            try:
                old_cursor.execute("SELECT * FROM machine_master")
                rows = old_cursor.fetchall()
                if rows:
                    conn = sqlite3.connect(resource_path("machine_master.db"))
                    conn.executemany("INSERT OR IGNORE INTO machine_master VALUES (?, ?, ?)", rows)
                    conn.commit()
                    conn.close()
                old_cursor.execute("DROP TABLE machine_master")
            except: pass

            # Customers
            try:
                old_cursor.execute("SELECT * FROM customers")
                rows = old_cursor.fetchall()
                if rows:
                    conn = sqlite3.connect(resource_path("customers.db"))
                    conn.executemany("INSERT OR IGNORE INTO customers VALUES (?, ?, ?, ?, ?)", rows)
                    conn.commit()
                    conn.close()
                old_cursor.execute("DROP TABLE customers")
            except: pass

            # Items
            try:
                old_cursor.execute("SELECT * FROM items")
                rows = old_cursor.fetchall()
                if rows:
                    conn = sqlite3.connect(resource_path("items.db"))
                    conn.executemany("INSERT OR IGNORE INTO items VALUES (?, ?, ?, ?)", rows)
                    conn.commit()
                    conn.close()
                old_cursor.execute("DROP TABLE items")
            except: pass

            # Process
            try:
                old_cursor.execute("SELECT * FROM process_master")
                rows = old_cursor.fetchall()
                if rows:
                    conn = sqlite3.connect(resource_path("process_master.db"))
                    conn.executemany("INSERT OR IGNORE INTO process_master VALUES (?, ?, ?)", rows)
                    conn.commit()
                    conn.close()
                old_cursor.execute("DROP TABLE process_master")
            except: pass

            # Component Setup
            try:
                old_cursor.execute("SELECT * FROM component_setup")
                rows = old_cursor.fetchall()
                if rows:
                    conn = sqlite3.connect("component_setup.db")
                    conn.executemany("INSERT OR IGNORE INTO component_setup VALUES (?, ?, ?)", rows)
                    conn.commit()
                    conn.close()
                old_cursor.execute("DROP TABLE component_setup")
            except: pass

            old_conn.commit()
            old_conn.close()
        except Exception as e:
            print("V1 Data Port Error:", e)

    @staticmethod
    def migrate_json_to_sqlite():
        """One-time migration script to move existing data from .json files into the new separate DBs"""
        
        # --- 1. migrate operators.json (Employee Master) ---
        if os.path.exists("operators.json"):
            try:
                conn = sqlite3.connect(resource_path("employee_master.db"))
                cursor = conn.cursor()
                with open("operators.json", "r", encoding="utf-8") as f:
                    ops = json.load(f)
                for op in ops:
                    cursor.execute("INSERT OR IGNORE INTO employee_master (id, name, description, phone) VALUES (?, ?, ?, ?)", 
                                   (op.get('id',''), op.get('name',''), op.get('description',''), op.get('phone','')))
                conn.commit()
                conn.close()
                os.rename("operators.json", "operators.json.bak")
            except Exception as e: print("Error migrating operators:", e)

        # --- 2. migrate machines.json (Machine Master) ---
        if os.path.exists("machines.json"):
            try:
                conn = sqlite3.connect(resource_path("machine_master.db"))
                cursor = conn.cursor()
                with open("machines.json", "r", encoding="utf-8") as f:
                    macs = json.load(f)
                for mac in macs:
                    cursor.execute("INSERT OR IGNORE INTO machine_master (code, name, description) VALUES (?, ?, ?)", 
                                   (mac.get('code',''), mac.get('name',''), mac.get('description','')))
                conn.commit()
                conn.close()
                os.rename("machines.json", "machines.json.bak")
            except Exception as e: print("Error migrating machines:", e)

        # --- 3. migrate customers.json (Customer Master) ---
        if os.path.exists("customers.json"):
            try:
                conn = sqlite3.connect(resource_path("customers.db"))
                cursor = conn.cursor()
                with open("customers.json", "r", encoding="utf-8") as f:
                    custs = json.load(f)
                for c in custs:
                    cursor.execute("INSERT OR IGNORE INTO customers (code, name, description, email, phone) VALUES (?, ?, ?, ?, ?)",
                                   (c.get('code',''), c.get('name',''), c.get('description',''), c.get('email',''), c.get('phone','')))
                conn.commit()
                conn.close()
                os.rename("customers.json", "customers.json.bak")
            except Exception as e: print("Error migrating customers:", e)
            
        # --- 4. migrate items.json (Item Master) ---
        if os.path.exists("items.json"):
            try:
                conn = sqlite3.connect(resource_path("items.db"))
                cursor = conn.cursor()
                with open("items.json", "r", encoding="utf-8") as f:
                    items = json.load(f)
                for item in items:
                    cursor.execute("INSERT OR IGNORE INTO items (code, drawing_no, name, revision_no) VALUES (?, ?, ?, ?)",
                                   (item.get('code',''), item.get('drawing_no',''), item.get('name',''), item.get('revision_no','')))
                conn.commit()
                conn.close()
                os.rename("items.json", "items.json.bak")
            except Exception as e: print("Error migrating items:", e)
            
        # --- 5. migrate process.json (Process Master) ---
        if os.path.exists("process.json"):
            try:
                conn = sqlite3.connect(resource_path("process_master.db"))
                cursor = conn.cursor()
                with open("process.json", "r", encoding="utf-8") as f:
                    procs = json.load(f)
                for proc in procs:
                    cursor.execute("INSERT OR IGNORE INTO process_master (id, name, description) VALUES (?, ?, ?)", 
                                   (proc.get('code',''), proc.get('name',''), proc.get('desc','')))
                conn.commit()
                conn.close()
                os.rename("process.json", "process.json.bak")
            except Exception as e: print("Error migrating process:", e)
            
        # --- 6. migrate airgauge_component_config.json (Component Setup) ---
        if os.path.exists("airgauge_component_config.json"):
            try:
                conn = sqlite3.connect(resource_path("component_setup.db"))
                cursor = conn.cursor()
                with open("airgauge_component_config.json", "r", encoding="utf-8") as f:
                    config = json.load(f)
                # config format: {"1": {"CH1": {"item_code": "...", "lsl": "...", ...}, "CH2": {...}}, "2": {...}}
                for ag_id, channels in config.items():
                    if isinstance(channels, dict):
                        for ch_id, props in channels.items():
                            cursor.execute("INSERT OR IGNORE INTO component_setup (airgauge_id, channel, properties) VALUES (?, ?, ?)",
                                           (ag_id, ch_id, json.dumps(props)))
                conn.commit()
                conn.close()
                os.rename("airgauge_component_config.json", "airgauge_component_config.json.bak")
            except Exception as e: print("Error migrating component setup:", e)

    @staticmethod
    def is_setup_complete():
        if not os.path.exists(SetupDatabase.DB_FILE):
            return False
        try:
            conn = sqlite3.connect(resource_path(SetupDatabase.DB_FILE))
            cursor = conn.cursor()
            # Ensure table exists (and migrate if needed by calling Init implicitly or just checking)
            # Better to call init_db to ensure structure is correct before checking count
            SetupDatabase.init_db() 
            
            cursor.execute("SELECT COUNT(*) FROM setup_info")
            count = cursor.fetchone()[0]
            conn.close()
            return count > 0
        except:
            return False

    @staticmethod
    def save_setup_info(company_name, address, admin_password, admin_name):
        SetupDatabase.init_db()
        conn = sqlite3.connect(resource_path(SetupDatabase.DB_FILE))
        cursor = conn.cursor()
        # Clear old info just in case
        cursor.execute("DELETE FROM setup_info")
        cursor.execute("INSERT INTO setup_info (company_name, address, admin_password, admin_name) VALUES (?, ?, ?, ?)",
                       (company_name, address, admin_password, admin_name))
        conn.commit()
        conn.close()

    @staticmethod
    def get_admin_password():
        if not SetupDatabase.is_setup_complete():
            return "cherry@123" # Default fallback
        try:
            conn = sqlite3.connect(SetupDatabase.DB_FILE)
            cursor = conn.cursor()
            cursor.execute("SELECT admin_password FROM setup_info LIMIT 1")
            row = cursor.fetchone()
            conn.close()
            if row and row[0]:
                return row[0]
        except:
            pass
        return "cherry@123"

    @staticmethod
    def get_admin_name():
        if not SetupDatabase.is_setup_complete():
            return "admin" # Default fallback
        try:
            conn = sqlite3.connect(SetupDatabase.DB_FILE)
            cursor = conn.cursor()
            cursor.execute("SELECT admin_name FROM setup_info LIMIT 1")
            row = cursor.fetchone()
            conn.close()
            if row and row[0]:
                return row[0]
        except:
            pass
        return "admin"

    @staticmethod
    def get_company_name():
        if not SetupDatabase.is_setup_complete():
            return "Cherry Precision" # Default fallback
        try:
            conn = sqlite3.connect(SetupDatabase.DB_FILE)
            cursor = conn.cursor()
            cursor.execute("SELECT company_name FROM setup_info LIMIT 1")
            row = cursor.fetchone()
            conn.close()
            if row and row[0]:
                return row[0]
        except:
            pass
        return "Cherry Precision"

    @staticmethod
    def migrate_airgauge_config_json():
        """Migrate legacy airgauge_config.json to airgauge_master.db."""
        if os.path.exists("airgauge_config.json"):
            try:
                with open("airgauge_config.json", "r") as f:
                    data = json.load(f)
                conn = sqlite3.connect("airgauge_master.db")
                cursor = conn.cursor()
                for ip, ag_id in data.items():
                    cursor.execute("INSERT OR IGNORE INTO airgauge_master (ip, airgauge_id) VALUES (?, ?)", (ip, str(ag_id)))
                conn.commit()
                conn.close()
                os.rename("airgauge_config.json", "airgauge_config.json.bak")
            except Exception as e:
                print("AirGauge migration error:", e)

    @staticmethod
    def validate_cherry_credentials(cherry_id, cherry_password):
        """Returns True if cherry_id and cherry_password match CHERRY_SERVICE_ACCOUNTS."""
        expected = SetupDatabase.CHERRY_SERVICE_ACCOUNTS.get(cherry_id)
        return expected is not None and expected == cherry_password

    @staticmethod
    def get_cherry_use_count(cherry_id):
        """Returns how many times this cherry_id has been used for password reset."""
        try:
            SetupDatabase.init_db()
            conn = sqlite3.connect(resource_path(SetupDatabase.DB_FILE))
            cursor = conn.cursor()
            cursor.execute("SELECT use_count FROM cherry_service_usage WHERE cherry_id = ?", (cherry_id,))
            row = cursor.fetchone()
            conn.close()
            return row[0] if row else 0
        except:
            return 0

    @staticmethod
    def increment_cherry_use(cherry_id):
        """Increments the use count for the given cherry_id."""
        try:
            SetupDatabase.init_db()
            conn = sqlite3.connect(resource_path(SetupDatabase.DB_FILE))
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO cherry_service_usage (cherry_id, use_count) VALUES (?, 1)
                ON CONFLICT(cherry_id) DO UPDATE SET use_count = use_count + 1
            """, (cherry_id,))
            conn.commit()
            conn.close()
        except Exception as e:
            print("Error incrementing cherry use:", e)

class OrganizationSetupPage(ctk.CTkFrame):
    def __init__(self, parent, controller):
        super().__init__(parent)
        self.controller = controller
        
        # Background
        self.configure(fg_color="#F5F7FA") # Light nice background
        
        # Main Container - Split Card Style
        self.main_card = ctk.CTkFrame(self, fg_color="white", corner_radius=20, width=1000, height=600)
        self.main_card.place(relx=0.5, rely=0.5, anchor="center")
        self.main_card.pack_propagate(False)

        # --- LEFT SIDE: INFO PANEL ---
        self.left_panel = ctk.CTkFrame(self.main_card, fg_color="#1565C0", corner_radius=20, width=400, height=600)
        self.left_panel.place(x=0, y=0)
        self.left_panel.pack_propagate(False)
        
        # Load cherry logo image
        try:
            logo_path = resource_path("settings/cherry_signup_logo.png")
            if os.path.exists(logo_path):
                pil_img = Image.open(logo_path)
                self.cherry_logo_img = ctk.CTkImage(pil_img, size=(180, 86))
                logo_lbl = ctk.CTkLabel(self.left_panel, text="", image=self.cherry_logo_img, fg_color="transparent")
                logo_lbl.place(relx=0.5, rely=0.15, anchor="center")
        except Exception as e:
            print("Error loading setup cherry logo:", e)
        
        # Content Left
        ctk.CTkLabel(self.left_panel, text="Setup Wizard", 
                     font=("Segoe UI", 24, "bold"), text_color="white").place(relx=0.5, rely=0.30, anchor="center")
        
        ctk.CTkLabel(self.left_panel, text="Step 1: Organization", 
                     font=("Segoe UI", 18, "bold"), text_color="#BBDEFB").place(relx=0.5, rely=0.38, anchor="center")
        
        info_text = ("Welcome to your Air Gauge System.\n\n"
                     "Please provide the company details.\n\n"
                     "These settings can be changed later\n"
                     "in the application settings.")
        
        ctk.CTkLabel(self.left_panel, text=info_text, 
                     font=("Segoe UI", 13), text_color="#E3F2FD", justify="center").place(relx=0.5, rely=0.58, anchor="center")

        # --- RIGHT SIDE: FORM ---
        self.right_panel = ctk.CTkFrame(self.main_card, fg_color="white", corner_radius=20, width=600, height=600)
        self.right_panel.place(x=400, y=0)
        
        self.form_container = ctk.CTkFrame(self.right_panel, fg_color="transparent")
        self.form_container.place(relx=0.5, rely=0.5, anchor="center")

        # Title
        ctk.CTkLabel(self.form_container, text="Organization Details", 
                     font=("Segoe UI", 24, "bold"), text_color="#333").pack(pady=(0, 20))

        # Form Fields
        self.fields = {}
        
        self.create_field("Company Name", "company_name")
        self.create_field("Address", "address")
        
        try:
            self.fields["company_name"].focus()
            self.fields["company_name"].bind("<Return>", lambda event: self.fields["address"].focus())
            self.fields["address"].bind("<Return>", lambda event: self.save_btn.invoke())
        except Exception as e:
            print(f"Error binding OrganizationSetupPage: {e}")

        # Error Label
        self.error_label = ctk.CTkLabel(self.form_container, text="", text_color="red", font=("Segoe UI", 11))
        self.error_label.pack(pady=(10, 5))

        # Save Button
        self.save_btn = ModernButton(self.form_container, text="Next: Admin Setup ➔", 
                                      font=("Segoe UI", 13, "bold"),
                                      height=45, width=250,
                                      fg_color="#1976D2", hover_color="#1565C0",
                                      corner_radius=8,
                                      command=self.on_save)
        self.save_btn.pack(pady=20)

    def create_field(self, label_text, key):
        frame = ctk.CTkFrame(self.form_container, fg_color="transparent")
        frame.pack(pady=10, fill="x")
        
        ctk.CTkLabel(frame, text=label_text, font=("Segoe UI", 11, "bold"), text_color="#555", anchor="w").pack(fill="x", pady=(0, 2))
        
        entry = ctk.CTkEntry(frame, height=40, width=320, font=("Segoe UI", 13), border_color="#E0E0E0", corner_radius=6)
        entry.pack(fill="x")
        
        self.fields[key] = entry
        
    def on_save(self):
        company = self.fields["company_name"].get().strip()
        address = self.fields["address"].get().strip()

        if not company or not address:
            self.error_label.configure(text="All fields are mandatory.")
            return

        self.controller.show_admin_setup(company, address)


class AdminSetupPage(ctk.CTkFrame):
    def __init__(self, parent, controller, company, address):
        super().__init__(parent)
        self.controller = controller
        self.company = company
        self.address = address
        
        # Background
        self.configure(fg_color="#F5F7FA")
        
        # Main Container
        self.main_card = ctk.CTkFrame(self, fg_color="white", corner_radius=20, width=1000, height=600)
        self.main_card.place(relx=0.5, rely=0.5, anchor="center")
        self.main_card.pack_propagate(False)

        # --- LEFT SIDE: INFO PANEL ---
        self.left_panel = ctk.CTkFrame(self.main_card, fg_color="#1565C0", corner_radius=20, width=400, height=600)
        self.left_panel.place(x=0, y=0)
        self.left_panel.pack_propagate(False)
        
        # Load cherry logo image
        try:
            logo_path = resource_path("settings/cherry_signup_logo.png")
            if os.path.exists(logo_path):
                pil_img = Image.open(logo_path)
                self.cherry_logo_img = ctk.CTkImage(pil_img, size=(180, 86))
                logo_lbl = ctk.CTkLabel(self.left_panel, text="", image=self.cherry_logo_img, fg_color="transparent")
                logo_lbl.place(relx=0.5, rely=0.15, anchor="center")
        except Exception as e:
            print("Error loading setup cherry logo:", e)
        
        ctk.CTkLabel(self.left_panel, text="Setup Wizard", 
                     font=("Segoe UI", 24, "bold"), text_color="white").place(relx=0.5, rely=0.30, anchor="center")
        
        ctk.CTkLabel(self.left_panel, text="Step 2: Administrator", 
                     font=("Segoe UI", 18, "bold"), text_color="#BBDEFB").place(relx=0.5, rely=0.38, anchor="center")
        
        info_text = ("Set up the primary admin account.\n\n"
                     "This account will be used to access\n"
                     "system settings and configurations.")
        
        ctk.CTkLabel(self.left_panel, text=info_text, 
                     font=("Segoe UI", 13), text_color="#E3F2FD", justify="center").place(relx=0.5, rely=0.58, anchor="center")

        # --- RIGHT SIDE: FORM ---
        self.right_panel = ctk.CTkFrame(self.main_card, fg_color="white", corner_radius=20, width=600, height=600)
        self.right_panel.place(x=400, y=0)
        
        self.form_container = ctk.CTkFrame(self.right_panel, fg_color="transparent")
        self.form_container.place(relx=0.5, rely=0.5, anchor="center")

        ctk.CTkLabel(self.form_container, text="Admin Credentials", 
                     font=("Segoe UI", 24, "bold"), text_color="#333").pack(pady=(0, 10))

        # Admin Name
        name_frame = ctk.CTkFrame(self.form_container, fg_color="transparent")
        name_frame.pack(pady=5, fill="x")
        ctk.CTkLabel(name_frame, text="Admin Name", font=("Segoe UI", 11, "bold"), text_color="#555", anchor="w").pack(fill="x", pady=(0, 2))
        self.name_entry = ctk.CTkEntry(name_frame, height=40, width=320, font=("Segoe UI", 13), border_color="#E0E0E0", corner_radius=6)
        self.name_entry.pack(fill="x")
        self.name_entry.focus()

        # Password
        pass_frame = ctk.CTkFrame(self.form_container, fg_color="transparent")
        pass_frame.pack(pady=5, fill="x")
        ctk.CTkLabel(pass_frame, text="Password", font=("Segoe UI", 11, "bold"), text_color="#555", anchor="w").pack(fill="x", pady=(0, 2))
        self.pass_entry = ctk.CTkEntry(pass_frame, height=40, width=320, font=("Segoe UI", 13), show="●", border_color="#E0E0E0", corner_radius=6)
        self.pass_entry.pack(fill="x")
        self.pass_entry.bind("<KeyRelease>", self.validate_password)

        # Confirm Password
        confirm_frame = ctk.CTkFrame(self.form_container, fg_color="transparent")
        confirm_frame.pack(pady=5, fill="x")
        ctk.CTkLabel(confirm_frame, text="Confirm Password", font=("Segoe UI", 11, "bold"), text_color="#555", anchor="w").pack(fill="x", pady=(0, 2))
        self.confirm_entry = ctk.CTkEntry(confirm_frame, height=40, width=320, font=("Segoe UI", 13), show="●", border_color="#E0E0E0", corner_radius=6)
        self.confirm_entry.pack(fill="x")
        self.confirm_entry.bind("<KeyRelease>", self.validate_password)

        self.name_entry.bind("<Return>", lambda event: self.pass_entry.focus())
        self.pass_entry.bind("<Return>", lambda event: self.confirm_entry.focus())
        self.confirm_entry.bind("<Return>", lambda event: self.save_btn.invoke() if self.save_btn.cget('state') == 'normal' else None)

        # Password Rules Checklist
        self.rules_frame = ctk.CTkFrame(self.form_container, fg_color="transparent")
        self.rules_frame.pack(pady=5, fill="x")
        
        self.rule_labels = {}
        rules = [("upper", "One uppercase letter"), ("lower", "One lowercase letter"),
                 ("number", "One number"), ("special", "One special character"),
                 ("match", "Passwords match")]
        
        for key, text in rules:
            lbl = ctk.CTkLabel(self.rules_frame, text="❌ " + text, font=("Segoe UI", 11), text_color="#757575", anchor="w")
            lbl.pack(fill="x")
            self.rule_labels[key] = lbl

        # Error
        self.error_label = ctk.CTkLabel(self.form_container, text="", text_color="red", font=("Segoe UI", 11))
        self.error_label.pack(pady=(5, 5))

        # Save
        self.save_btn = ModernButton(self.form_container, text="Save & Launch", 
                                      font=("Segoe UI", 13, "bold"),
                                      height=45, width=250,
                                      fg_color="#1976D2", hover_color="#1565C0",
                                      state="disabled",
                                      corner_radius=8,
                                      command=self.on_save)
        self.save_btn.pack(pady=10)

    def validate_password(self, event=None):
        password = self.pass_entry.get()
        confirm = self.confirm_entry.get()
        
        has_upper = bool(re.search(r'[A-Z]', password))
        has_lower = bool(re.search(r'[a-z]', password))
        has_number = bool(re.search(r'\d', password))
        has_special = bool(re.search(r'[^A-Za-z0-9]', password))
        is_match = (password == confirm) and len(password) > 0
        
        # Update UI Ticks
        self._update_tick("upper", has_upper, "One uppercase letter")
        self._update_tick("lower", has_lower, "One lowercase letter")
        self._update_tick("number", has_number, "One number")
        self._update_tick("special", has_special, "One special character")
        self._update_tick("match", is_match, "Passwords match")
        
        # Enable/Disable Save button
        if all([has_upper, has_lower, has_number, has_special, is_match]):
            self.save_btn.configure(state="normal")
        else:
            self.save_btn.configure(state="disabled")

    def _update_tick(self, key, is_valid, text):
        if is_valid:
            self.rule_labels[key].configure(text="✅ " + text, text_color="#2E7D32")
        else:
            self.rule_labels[key].configure(text="❌ " + text, text_color="#757575")

    def on_save(self):
        admin_name = self.name_entry.get().strip()
        password = self.pass_entry.get().strip()
        
        if not admin_name:
            self.error_label.configure(text="Admin Name is required.")
            return

        try:
            SetupDatabase.save_setup_info(self.company, self.address, password, admin_name)
            self.controller.complete_setup()
        except Exception as e:
            self.error_label.configure(text=f"Error saving: {e}")


# ─────────────────────────────────────────────────────────────────────────────
#  Startup Splash Screen
# ─────────────────────────────────────────────────────────────────────────────
class SplashScreen(tk.Toplevel):
    """
    Full-screen branded splash shown while the app initialises.
    Auto-closes after ~2.5 s with a smooth fade-out.
    """
    _DOTS = ["   ", ".  ", ".. ", "..."]

    def __init__(self, parent):
        super().__init__(parent)
        self.overrideredirect(True)          # No title bar
        self.attributes("-topmost", True)

        # --- Full screen size ---
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        self.geometry(f"{sw}x{sh}+0+0")
        self.configure(bg="#0D1B2A")

        # ── Background gradient canvas ──────────────────────────
        canvas = tk.Canvas(self, bg="#0D1B2A", highlightthickness=0)
        canvas.pack(fill="both", expand=True)

        # Dark gradient via horizontal bands
        for i in range(sh):
            ratio = i / sh
            r = int(13  + ratio * (30  - 13))
            g = int(27  + ratio * (60  - 27))
            b = int(42  + ratio * (90  - 42))
            canvas.create_line(0, i, sw, i, fill=f"#{r:02x}{g:02x}{b:02x}")

        # ── Decorative accent bar at top ───────────────────────
        canvas.create_rectangle(0, 0, sw, 6, fill="#1976D2", outline="")

        # ── Cherry brand circle ────────────────────────────────
        cx, cy = sw // 2, sh // 2 - 80
        canvas.create_oval(cx-60, cy-60, cx+60, cy+60,
                           fill="#1565C0", outline="#42A5F5", width=2)
        canvas.create_text(cx, cy, text="🍒", font=("Segoe UI", 24, "bold"))

        # ── App title ──────────────────────────────────────────
        canvas.create_text(sw // 2, sh // 2 + 10,
                           text="Cherry Precision Products",
                           font=("Segoe UI", 24, "bold"),
                           fill="white")
        canvas.create_text(sw // 2, sh // 2 + 56,
                           text="Air Gauge  •  SPC System  •  v2.0",
                           font=("Segoe UI", 13),
                           fill="#90CAF9")

        # Divider line
        canvas.create_line(sw // 2 - 200, sh // 2 + 80,
                           sw // 2 + 200, sh // 2 + 80,
                           fill="#1976D2", width=1)

        # ── Loading dots label ─────────────────────────────────
        self._dot_label = canvas.create_text(
            sw // 2, sh // 2 + 110,
            text="Loading   ",
            font=("Segoe UI", 13),
            fill="#64B5F6"
        )
        self._canvas   = canvas
        self._dot_idx  = 0
        self._dot_job  = None
        self._fade_job = None

        # ── Progress bar (animated) ────────────────────────────
        bar_y   = sh // 2 + 145
        bar_w   = 360
        bar_x   = sw // 2 - bar_w // 2
        bar_h   = 6
        canvas.create_rectangle(bar_x, bar_y,
                                 bar_x + bar_w, bar_y + bar_h,
                                 fill="#1A2E40", outline="")
        self._progress_rect = canvas.create_rectangle(
            bar_x, bar_y, bar_x, bar_y + bar_h,
            fill="#1976D2", outline=""
        )
        self._bar_x    = bar_x
        self._bar_w    = bar_w
        self._bar_y    = bar_y
        self._bar_h    = bar_h
        self._progress = 0.0

        # ── Copyright footer ───────────────────────────────────
        canvas.create_text(sw // 2, sh - 30,
                           text="© 2025 Cherry Precision Products. All rights reserved.",
                           font=("Segoe UI", 9),
                           fill="#37474F")

        # Start animations
        self._animate_dots()
        self._animate_progress()

    # ── Dot animation ──────────────────────────────────────────
    def _animate_dots(self):
        if not self.winfo_exists():
            return
        self._dot_idx = (self._dot_idx + 1) % len(self._DOTS)
        try:
            self._canvas.itemconfig(
                self._dot_label,
                text=f"Loading{self._DOTS[self._dot_idx]}"
            )
        except Exception:
            return
        self._dot_job = self.after(350, self._animate_dots)

    # ── Progress bar animation ─────────────────────────────────
    def _animate_progress(self):
        if not self.winfo_exists():
            return
        if self._progress < 1.0:
            self._progress = min(self._progress + 0.008, 1.0)
            filled_w = int(self._bar_w * self._progress)
            try:
                self._canvas.coords(
                    self._progress_rect,
                    self._bar_x, self._bar_y,
                    self._bar_x + filled_w, self._bar_y + self._bar_h
                )
                # Colour shift: blue → green as it fills
                r = int(25  + self._progress * 0)
                g = int(118 + self._progress * (168 - 118))
                b = int(210 + self._progress * (107 - 210))
                self._canvas.itemconfig(
                    self._progress_rect, fill=f"#{r:02x}{g:02x}{b:02x}"
                )
            except Exception:
                return
            self.after(18, self._animate_progress)

    # ── Real-progress update (driven by background preload thread) ─
    def _set_real_progress(self, fraction):
        """
        Advance the progress bar to reflect actual background-task
        completion.  fraction is 0.0 – 1.0.  The bar only moves
        forward, never backward.  Called from the main thread via
        app.after(0, ...) from the worker thread.
        """
        if not self.winfo_exists():
            return
        self._progress = max(self._progress, min(fraction, 1.0))
        filled_w = int(self._bar_w * self._progress)
        try:
            self._canvas.coords(
                self._progress_rect,
                self._bar_x, self._bar_y,
                self._bar_x + filled_w, self._bar_y + self._bar_h
            )
            r = int(25  + self._progress * 0)
            g = int(118 + self._progress * (168 - 118))
            b = int(210 + self._progress * (107 - 210))
            self._canvas.itemconfig(
                self._progress_rect, fill=f"#{r:02x}{g:02x}{b:02x}")
        except Exception:
            pass

    def set_status(self, text):
        """Replace the animated loading dots with a real status string."""
        if not self.winfo_exists():
            return
        try:
            self._canvas.itemconfig(self._dot_label, text=text)
        except Exception:
            pass

    # ── Fade-out and destroy ───────────────────────────────────
    def fade_out(self, alpha=1.0, on_done=None):
        if not self.winfo_exists():
            if on_done:
                on_done()
            return
        if alpha > 0:
            self.attributes("-alpha", alpha)
            self.after(20, lambda: self.fade_out(alpha - 0.06, on_done))
        else:
            # Cancel pending jobs
            for job_attr in ("_dot_job", "_fade_job"):
                job = getattr(self, job_attr, None)
                if job:
                    try:
                        self.after_cancel(job)
                    except Exception:
                        pass
            try:
                self.destroy()
            except Exception:
                pass
            if on_done:
                on_done()


class CherryApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        # ── Set maximised state FIRST, then hide the window ────────────
        # Critical order: state("zoomed") must be called before withdraw().
        # The old approach scheduled maximize via self.after(200, …) which
        # briefly deiconified the window 200 ms after mainloop started,
        # causing the 1-second login-page flash the user reported.
        try:
            self.state("zoomed")
        except Exception:
            try:
                sw = self.winfo_screenwidth()
                sh = self.winfo_screenheight()
                self.geometry(f"{sw}x{sh}+0+0")
            except Exception:
                self.geometry("1250x720")
        self.withdraw()   # ← hide AFTER state is set — zero flash

        # ---------------- Window setup ----------------
        self.title("Cherry Precision Product – Air Gauge System")
        self.minsize(1100, 650)

        # ---------------- Force Light Theme ----------------
        ctk.set_appearance_mode("light")

        # ---------------- App state ----------------
        self.active_page = None
        self.sidebar_user_visible = False
        self.run_chat_active = False
        self._closing = False

        self.component_page = None
        self.runchat_page = None
        self.bg_resize_job = None  # For debouncing

        # ---------------- Core systems ----------------
        self.data_manager = DataManager()
        self.data_queue = queue.Queue()
        self.rssi_queue = queue.Queue(maxsize=5)  # Only keep latest RSSI readings


        # ---------------- Setup Check ----------------
        # ---------------- Setup Check ----------------
        if not SetupDatabase.is_setup_complete():
            # Step 1: Installer Login
            self.show_installer_login()
        else:
            # Step 2: Normal Admin Login
            self.show_admin_login()

    def show_installer_login(self):
        """Show the Installer Login Page (Hardcoded)."""
        self.installer_login = InstallerLoginPage(self, self)
        self.installer_login.pack(fill="both", expand=True)

    def show_license_verification(self):
        """Step 2: Show License Verification Page."""
        if hasattr(self, "installer_login"):
            self.installer_login.pack_forget()
            self.installer_login.destroy()
        
        self.license_verify = LicenseVerificationPage(self, self)
        self.license_verify.pack(fill="both", expand=True)

    def show_setup_wizard(self):
        """Called after successful license verification. Now shows Organization Details."""
        if hasattr(self, "license_verify"):
            self.license_verify.pack_forget()
            self.license_verify.destroy()
        
        self.setup_page = OrganizationSetupPage(self, self)
        self.setup_page.pack(fill="both", expand=True)

    def show_admin_setup(self, company, address):
        """Transition from Organization details to Admin details."""
        if hasattr(self, "setup_page"):
            self.setup_page.pack_forget()
            self.setup_page.destroy()
        
        self.admin_setup = AdminSetupPage(self, self, company, address)
        self.admin_setup.pack(fill="both", expand=True)

    def complete_setup(self):
        """Called when admin setup is done."""
        if hasattr(self, "admin_setup"):
            self.admin_setup.pack_forget()
            self.admin_setup.destroy()
        self.init_main_ui()

    def show_admin_login(self):
        """Show the Admin Login Page (DB-based)."""
        self.admin_login = AdminLoginPage(self, self)
        self.admin_login.pack(fill="both", expand=True)

    def start_setup_from_login(self):
        """Called when user clicks CREATE ACCOUNT on the login page.
        Destroys the login page and starts the installer/setup wizard flow."""
        if hasattr(self, "admin_login"):
            self.admin_login.pack_forget()
            self.admin_login.destroy()
        self.show_installer_login()

    def complete_login(self):
        """Called by AdminLoginPage on success."""
        if hasattr(self, "admin_login"):
            self.admin_login.pack_forget()
            self.admin_login.destroy()
        self.init_main_ui()

    def destroy(self):
        """Graceful shutdown: stop all recurring timers before destroying."""
        self._closing = True
        super().destroy()

    def init_main_ui(self):
        """Initialize the main application UI."""
        # Load component setup data (shared across pages for tolerance lookups)
        self.comp_json = self._load_comp_json()

        # ---------------- UI creation ----------------
        self.configure(fg_color="#F4F6F8")
        self.create_header()
        self.create_content_area()
        self.create_sidebar()
        
        self.sidebar_user_visible = True

        # ---------------- Resize optimization ----------------
        self._install_resize_throttle()

        # ---------------- Live data loop ----------------
        self.after(10, self.update_live_data)

        # ---------------- Pages ----------------
        self.live_data_page = LiveDataPage(self.content_frame, self)
        self.live_data_page.pack(fill="both", expand=True)
        self.live_data_page.pack_forget()  # hidden initially

        # ---------------- Home Page (default) ----------------
        self.home_page = HomePage(self.content_frame, self)
        self.home_page.pack(fill="both", expand=True)
        # Mark Home as active in sidebar after sidebar is ready
        self.after(50, lambda: self._sidebar_nav(lambda: None, "home"))

        # ---------------- Global Keyboard ----------------
        try:
            self.bind("<BackSpace>", self.global_go_back)
        except Exception:
            pass

    # ─────────────────────────────────────────────────────────────────
    #  PAGE-SWITCH OVERLAY  –  prevents the white-flash between pages
    # ─────────────────────────────────────────────────────────────────
    def _switch_page_with_overlay(self, load_fn, delay_ms=50, manual_dismiss=False,
                                    page_attr=None):
        # ── 1. Smart Cache Bypass ────────────────────────────────────
        # If the page already exists, load it instantly with zero delay
        if page_attr is not None:
            existing = getattr(self, page_attr, None)
            if existing is not None and hasattr(existing, "winfo_exists"):
                try:
                    if existing.winfo_exists():
                        load_fn()
                        return
                except Exception:
                    pass

        # ── 2. Instant Switch (No Slide Animation) ───────────────────
        # Runs load_fn directly to load/switch the page instantly, avoiding lag.
        try:
            load_fn()
            
            # Setup callback if manual dismiss is required (e.g. for AnalysisPage thread completion)
            if manual_dismiss:
                self._dismiss_overlay_callback = lambda: None
        except Exception as e:
            print("Page load error:", e)

    def _load_comp_json(self):
        """Load component setup JSON from SQLite for tolerance lookups."""
        try:
            conn = sqlite3.connect(resource_path("component_setup.db"))
            cursor = conn.cursor()
            cursor.execute("SELECT airgauge_id, channel, properties FROM component_setup")
            rows = cursor.fetchall()
            conn.close()
            comp_map = {}
            for ag_id, ch, props_str in rows:
                if ag_id not in comp_map:
                    comp_map[ag_id] = {}
                try:
                    comp_map[ag_id][ch] = json.loads(props_str)
                except:
                    comp_map[ag_id][ch] = {}
            return comp_map
        except Exception:
            return {}

    def global_go_back(self, event=None):
        """Global handler for Backspace key."""
        # 1. Focus check: don't navigate if typing in any entry/text widget
        focused = self.focus_get()
        if isinstance(focused, (ctk.CTkEntry, tk.Entry, tk.Text, ctk.CTkTextbox)):
            return
        
        # Also check internal tk.Entry if ctk wrapper focus isn't direct
        try:
            # Some ctk widgets have internal names that focus_get returns
            if "entry" in str(focused).lower() or "text" in str(focused).lower():
                return
        except:
            pass

        # 2. delegate to currently visible page in content_frame
        for widget in self.content_frame.winfo_children():
            if widget.winfo_viewable() and hasattr(widget, "go_back"):
                # Call the page's back logic
                widget.go_back(event)
                return


    def _install_resize_throttle(self):
        """
        Call once in App.__init__ after creating UI.
        Debounces <Configure> events so heavy widgets update only after resize stops.
        """
        # small token to schedule final update
        self._resize_after_id = None
        # choose debounce delay in ms (120–200 is a good compromise)
        self._resize_debounce_ms = 140

        def _on_configure(event):
            # cancel previous schedule
            if self._resize_after_id:
                try: self.after_cancel(self._resize_after_id)
                except: pass
            # schedule final resize
            self._resize_after_id = self.after(self._resize_debounce_ms, self._do_final_resize)

        def _do_final_resize():
            self._resize_after_id = None
            try:
                # If you have a content_frame where pages are placed, update it
                if hasattr(self, "content_frame"):
                    # update layout once
                    self.content_frame.update_idletasks()
                    
                    # --- DYNAMIC BACKGROUND RESIZE ---
                    try:
                        if hasattr(self, "bg_pil_original") and hasattr(self, "bg_label"):
                            w = self.content_frame.winfo_width()
                            h = self.content_frame.winfo_height()
                            
                            # Skip resize if dimensions unchanged (avoid redundant work)
                            if w > 10 and h > 10 and (w, h) != getattr(self, '_last_bg_size', None):
                                self._last_bg_size = (w, h)
                                # NEAREST is ~10x faster than BILINEAR, imperceptible on a bg image
                                new_img = self.bg_pil_original.resize((w, h), Image.Resampling.NEAREST)
                                new_ctk_img = ctk.CTkImage(new_img, size=(w, h))
                                
                                self.bg_label.configure(image=new_ctk_img)
                                self.bg_label.image = new_ctk_img  # keep reference
                    except Exception as e:
                        print(f"Resize bg error: {e}")

                    # if current child has a refresh method, call it (optional)
                    cur = None
                    try:
                        children = self.content_frame.winfo_children()
                        if children:
                            cur = children[-1]
                    except:
                        cur = None
                    # If the current page implements on_parent_resized(), call it
                    try:
                        if cur and hasattr(cur, "on_parent_resized"):
                            cur.on_parent_resized()
                    except:
                        pass
            except Exception:
                pass

        # bind to configure on the root window
        try:
            self.bind("<Configure>", _on_configure)
        except Exception:
            pass

        # store helpers so ReportPage could call them if needed
        self._do_final_resize = _do_final_resize

       
    def load_operator_list(self):
        """Show full Operator Manager page."""
        self._switch_page_with_overlay(self._do_load_operator_list,
                                       page_attr="operator_page")

    def _do_load_operator_list(self):
        for widget in self.content_frame.winfo_children():
            widget.pack_forget()
        if hasattr(self, "operator_page") and getattr(self, "operator_page", None) and self.operator_page.winfo_exists():
            self.operator_page.pack(fill="both", expand=True)
        else:
            self.operator_page = OperatorManagerPage(self.content_frame, self)
            self.operator_page.pack(fill="both", expand=True)
        self.status_label.configure(text="Operator Manager loaded ✅")

    def load_customer_master(self):
        """Show the Customer Master page."""
        self._switch_page_with_overlay(self._do_load_customer_master,
                                       page_attr="customer_page")

    def _do_load_customer_master(self):
        for widget in self.content_frame.winfo_children():
            widget.pack_forget()
        if hasattr(self, "customer_page") and self.customer_page.winfo_exists():
            self.customer_page.pack(fill="both", expand=True)
        else:
            self.customer_page = CustomerMasterPage(self.content_frame, self)
            self.customer_page.pack(fill="both", expand=True)
        try:
            self.status_label.configure(text="Customer Master loaded ✅")
        except Exception:
            pass

    def update_live_data(self):
        """Continuously read parsed data from queue and update DataManager."""
        if getattr(self, '_closing', False):
            return
        # Process up to 20 items per tick; 80 ms gives smooth UI without queue buildup
        count = 0
        while not self.data_queue.empty() and count < 20:
            try:
                parsed = self.data_queue.get_nowait()
            except Exception:
                break
            self.data_manager.add_parsed(parsed)
            count += 1
        
        self.after(80, self.update_live_data)


    def verify_password(self):
        """Show professional password overlay with smooth fade-in and fade-out animations."""
        # --- Create overlay window (dark background) ---
        overlay = ctk.CTkToplevel(self)
        scale = self._get_window_scaling()
        w = int(self.winfo_width() / scale)
        h = int(self.winfo_height() / scale)
        x = int(self.winfo_rootx() / scale)
        y = int(self.winfo_rooty() / scale)
        overlay.geometry(f"{w}x{h}+{x}+{y}")
        overlay.overrideredirect(True)
        overlay.configure(fg_color="black")
        overlay.attributes("-alpha", 0.95) # instantly fully black backdrop to prevent Windows DWM lag
        overlay.grab_set()

        # --- Center white frame ---
        box_width = 360
        box_height = 220

        # --- Center shadow frame (drop shadow) ---
        shadow = ctk.CTkFrame(overlay, fg_color="#101010", corner_radius=16, border_width=0)
        shadow.place(relx=0.5, rely=0.5, anchor="center", x=4, y=4)
        shadow.pack_propagate(False)
        shadow.configure(width=box_width, height=box_height)

        box = ctk.CTkFrame(overlay, fg_color="#FFFFFF", corner_radius=16)
        box.place(relx=0.5, rely=0.5, anchor="center")
        box.pack_propagate(False)
        box.configure(width=box_width, height=box_height)

        # --- Title and labels ---
        title = ctk.CTkLabel(box, text="Admin Authorization Required",
                             font=("Segoe UI", 16, "bold"), text_color="black")
        title.pack(pady=(22, 2))

        subtitle = ctk.CTkLabel(box, text="Enter password to access AirGauge Setup.",
                                font=("Segoe UI", 12), text_color="#555555")
        subtitle.pack(pady=(0, 10))

        # --- Custom Styled Password Entry directly on Box with Key Icon ---
        pwd_entry = ctk.CTkEntry(
            box,
            show="*",
            width=280,
            height=36,
            corner_radius=8,
            border_width=1,
            border_color="#CCCCCC",
            fg_color="#FFFFFF",
            text_color="black",
            placeholder_text="",
            font=("Segoe UI", 15)
        )
        pwd_entry.pack(pady=(10, 2))

        # Shift text inside entry to make room for eye button (right)
        try:
            pwd_entry._entry.configure(padx=8)
        except Exception:
            pass

        # Toggle password visibility helper
        self._show_pwd_setup = False
        def toggle_pwd():
            self._show_pwd_setup = not self._show_pwd_setup
            if self._show_pwd_setup:
                pwd_entry.configure(show="")
                eye_btn.configure(text="👁")
            else:
                pwd_entry.configure(show="*")
                eye_btn.configure(text="🔒")

        eye_btn = ctk.CTkButton(
            pwd_entry,
            text="🔒",
            width=22,
            height=22,
            corner_radius=11,
            fg_color="transparent",
            hover_color="#EEEEEE",
            text_color="#666666",
            font=("Segoe UI", 11),
            command=toggle_pwd
        )
        eye_btn.place(relx=1.0, rely=0.5, x=-10, anchor="e")

        def on_focus_in(event):
            pwd_entry.configure(border_color="black")
        def on_focus_out(event):
            pwd_entry.configure(border_color="#CCCCCC")
        pwd_entry.bind("<FocusIn>", on_focus_in)
        pwd_entry.bind("<FocusOut>", on_focus_out)

        status_lbl = ctk.CTkLabel(box, text="", text_color="red", font=("Segoe UI", 11), height=14)
        status_lbl.pack(pady=(2, 2))

        def fade_out_overlay():
            overlay.destroy()

        # --- Button Actions ---
        def check_pass():
            input_val = pwd_entry.get().strip()
            stored_pass = SetupDatabase.get_admin_password()
            if input_val == "cherry@123" or input_val == stored_pass:
                fade_out_overlay()
                self.load_machine_setup()
                self.status_label.configure(text="AirGauge Setup unlocked ✅")
            else:
                status_lbl.configure(text="Incorrect password ❌")
                pwd_entry.delete(0, "end")

        def cancel_login():
            fade_out_overlay()
            self.status_label.configure(text="Access cancelled")

        # --- Black & Small Submit Button ---
        submit_btn = ModernButton(
            box,
            text="✓ Submit",
            width=120,
            height=32,
            fg_color="black",
            hover_color="#333333",
            text_color="white",
            corner_radius=16,
            font=("Segoe UI", 12, "bold"),
            command=check_pass
        )
        submit_btn.pack(pady=(12, 15))

        # --- Circular Overlapping Close Button ---
        close_btn = ctk.CTkButton(
            box,
            text="✕",
            width=24,
            height=24,
            corner_radius=12,
            fg_color="black",
            hover_color="#333333",
            text_color="white",
            bg_color="#FFFFFF",
            font=("Segoe UI", 10, "bold"),
            command=cancel_login
        )
        close_btn.place(x=box_width - 15, y=15, anchor="ne")

        pwd_entry.bind("<Return>", lambda event: check_pass())
        pwd_entry.focus()


    def verify_settings_password(self):
        """Show professional password overlay with smooth fade-in and fade-out animations."""
        # --- Create overlay window (dark background) ---
        overlay = ctk.CTkToplevel(self)
        scale = self._get_window_scaling()
        w = int(self.winfo_width() / scale)
        h = int(self.winfo_height() / scale)
        x = int(self.winfo_rootx() / scale)
        y = int(self.winfo_rooty() / scale)
        overlay.geometry(f"{w}x{h}+{x}+{y}")
        overlay.overrideredirect(True)
        overlay.configure(fg_color="black")
        overlay.attributes("-alpha", 0.95) # instantly fully black backdrop to prevent Windows DWM lag
        overlay.grab_set()

        # --- Center white frame ---
        box_width = 360
        box_height = 220

        # --- Center shadow frame (drop shadow) ---
        shadow = ctk.CTkFrame(overlay, fg_color="#101010", corner_radius=16, border_width=0)
        shadow.place(relx=0.5, rely=0.5, anchor="center", x=4, y=4)
        shadow.pack_propagate(False)
        shadow.configure(width=box_width, height=box_height)

        box = ctk.CTkFrame(overlay, fg_color="#FFFFFF", corner_radius=16)
        box.place(relx=0.5, rely=0.5, anchor="center")
        box.pack_propagate(False)
        box.configure(width=box_width, height=box_height)

        # --- Title and labels ---
        title = ctk.CTkLabel(box, text="Admin Authorization Required",
                             font=("Segoe UI", 16, "bold"), text_color="black")
        title.pack(pady=(22, 2))

        subtitle = ctk.CTkLabel(box, text="Enter password to modify system settings.",
                                font=("Segoe UI", 12), text_color="#555555")
        subtitle.pack(pady=(0, 10))

        # --- Custom Styled Password Entry directly on Box with Key Icon ---
        pwd_entry = ctk.CTkEntry(
            box,
            show="*",
            width=280,
            height=36,
            corner_radius=8,
            border_width=1,
            border_color="#CCCCCC",
            fg_color="#FFFFFF",
            text_color="black",
            placeholder_text="",
            font=("Segoe UI", 15)
        )
        pwd_entry.pack(pady=(10, 2))

        # Shift text inside entry to make room for eye button (right)
        try:
            pwd_entry._entry.configure(padx=8)
        except Exception:
            pass

        # Toggle password visibility helper
        self._show_pwd_settings = False
        def toggle_pwd():
            self._show_pwd_settings = not self._show_pwd_settings
            if self._show_pwd_settings:
                pwd_entry.configure(show="")
                eye_btn.configure(text="👁")
            else:
                pwd_entry.configure(show="*")
                eye_btn.configure(text="🔒")

        eye_btn = ctk.CTkButton(
            pwd_entry,
            text="🔒",
            width=22,
            height=22,
            corner_radius=11,
            fg_color="transparent",
            hover_color="#EEEEEE",
            text_color="#666666",
            font=("Segoe UI", 11),
            command=toggle_pwd
        )
        eye_btn.place(relx=1.0, rely=0.5, x=-10, anchor="e")

        def on_focus_in(event):
            pwd_entry.configure(border_color="black")
        def on_focus_out(event):
            pwd_entry.configure(border_color="#CCCCCC")
        pwd_entry.bind("<FocusIn>", on_focus_in)
        pwd_entry.bind("<FocusOut>", on_focus_out)

        status_lbl = ctk.CTkLabel(box, text="", text_color="red", font=("Segoe UI", 11), height=14)
        status_lbl.pack(pady=(2, 2))

        def fade_out_overlay():
            overlay.destroy()

        # --- Button Actions ---
        def check_pass():
            input_val = pwd_entry.get().strip()
            admin_pass = SetupDatabase.get_admin_password()
            if input_val == admin_pass or input_val == "cherry@123":
                fade_out_overlay()
                self.load_settings_page()
                self.status_label.configure(text="Settings unlocked ✅")
            else:
                status_lbl.configure(text="Incorrect password ❌")
                pwd_entry.delete(0, "end")

        def cancel_login():
            fade_out_overlay()
            self.status_label.configure(text="Access cancelled")

        # --- Black & Small Submit Button ---
        submit_btn = ModernButton(
            box,
            text="✓ Submit",
            width=120,
            height=32,
            fg_color="black",
            hover_color="#333333",
            text_color="white",
            corner_radius=16,
            font=("Segoe UI", 12, "bold"),
            command=check_pass
        )
        submit_btn.pack(pady=(12, 15))

        # --- Circular Overlapping Close Button ---
        close_btn = ctk.CTkButton(
            box,
            text="✕",
            width=24,
            height=24,
            corner_radius=12,
            fg_color="black",
            hover_color="#333333",
            text_color="white",
            bg_color="#FFFFFF",
            font=("Segoe UI", 10, "bold"),
            command=cancel_login
        )
        close_btn.place(x=box_width - 15, y=15, anchor="ne")

        pwd_entry.bind("<Return>", lambda event: check_pass())
        pwd_entry.focus()
        
    def load_settings_page(self):
        """Show the Settings Page (called after password verification)."""
        self._switch_page_with_overlay(self._do_load_settings_page,
                                       page_attr="settings_page")

    def _do_load_settings_page(self):
        for widget in self.content_frame.winfo_children():
            widget.pack_forget()
        if hasattr(self, "settings_page") and self.settings_page.winfo_exists():
            self.settings_page.pack(fill="both", expand=True)
        else:
            self.settings_page = SettingsPage(self.content_frame, self)
            self.settings_page.pack(fill="both", expand=True)
        self.status_label.configure(text="Settings Page opened ✅")

    def load_process_master(self):
        """Show the Process Master page."""
        self._switch_page_with_overlay(self._do_load_process_master,
                                       page_attr="process_page")

    def _do_load_process_master(self):
        for widget in self.content_frame.winfo_children():
            widget.pack_forget()
        if hasattr(self, "process_page") and self.process_page.winfo_exists():
            self.process_page.pack(fill="both", expand=True)
        else:
            self.process_page = ProcessMasterPage(self.content_frame, self)
            self.process_page.pack(fill="both", expand=True)
        try:
            self.status_label.configure(text="Process Master loaded ✅")
        except Exception:
            pass

    def load_item_master(self):
        """Show the Item Master page."""
        self._switch_page_with_overlay(self._do_load_item_master,
                                       page_attr="item_page")

    def _do_load_item_master(self):
        for widget in self.content_frame.winfo_children():
            widget.pack_forget()
        if hasattr(self, "item_page") and self.item_page.winfo_exists():
            self.item_page.pack(fill="both", expand=True)
        else:
            self.item_page = ItemMasterPage(self.content_frame, self)
            self.item_page.pack(fill="both", expand=True)
        try:
            self.status_label.configure(text="Item Master loaded ✅")
        except Exception:
            pass

    def load_settings(self):
        """Show the Settings page."""
        self._switch_page_with_overlay(self._do_load_settings_page,
                                       page_attr="settings_page")

    def load_machine_master(self):
        """Show the Machine Master Page."""
        self._switch_page_with_overlay(self._do_load_machine_master,
                                       page_attr="machine_page")

    def _do_load_machine_master(self):
        for widget in self.content_frame.winfo_children():
            widget.pack_forget()
        if hasattr(self, "machine_page") and self.machine_page.winfo_exists():
            self.machine_page.pack(fill="both", expand=True)
        else:
            self.machine_page = MachineMasterPage(self.content_frame, self)
            self.machine_page.pack(fill="both", expand=True)
        try:
            self.status_label.configure(text="Machine Master loaded ✅")
        except Exception:
            pass

    def load_machine_setup(self):
        """Show the Machine Setup Page after password success."""
        self._switch_page_with_overlay(self._do_load_machine_setup)

    def _do_load_machine_setup(self):
        for widget in self.content_frame.winfo_children():
            widget.pack_forget()
        self.machine_setup_page = MachineSetupPage(self.content_frame, self)
        self.machine_setup_page.pack(fill="both", expand=True)

    def load_component_setup(self):
        self._switch_page_with_overlay(self._do_load_component_setup)

    def _do_load_component_setup(self):
        for widget in self.content_frame.winfo_children():
            widget.pack_forget()
        self.component_page = ComponentSetupPage(self.content_frame, self)
        self.component_page.pack(fill="both", expand=True)
       
    def load_home(self):
        """Show the Home page (welcome image). Default page after login."""
        self._switch_page_with_overlay(self._do_load_home,
                                       page_attr="home_page")

    def _do_load_home(self):
        for widget in self.content_frame.winfo_children():
            widget.pack_forget()
        if hasattr(self, "runchat_page") and hasattr(self.runchat_page, "on_hide"):
            try:
                self.runchat_page.on_hide()
            except Exception:
                pass
        if hasattr(self, "home_page") and self.home_page.winfo_exists():
            self.home_page.pack(fill="both", expand=True)
        else:
            self.home_page = HomePage(self.content_frame, self)
            self.home_page.pack(fill="both", expand=True)
        self.active_page = "Home"

    def load_run_chat(self):
        """Show the RunChat page and enable live updates only while visible."""
        self._switch_page_with_overlay(self._do_load_run_chat,
                                       page_attr="runchat_page")

    def _do_load_run_chat(self):
        # 1️⃣ Hide current visible pages
        for widget in self.content_frame.winfo_children():
            widget.pack_forget()
        # 2️⃣ Pause LiveData updates
        if hasattr(self, "live_data_page") and hasattr(self.live_data_page, "on_hide"):
            try:
                self.live_data_page.on_hide()
            except Exception:
                pass
        # 3️⃣ Reuse if exists; else create
        if hasattr(self, "runchat_page") and self.runchat_page and self.runchat_page.winfo_exists():
            self.runchat_page.pack(fill="both", expand=True)
        else:
            self.runchat_page = RunChatPage(self.content_frame, self)
            self.runchat_page.pack(fill="both", expand=True)
        # 4️⃣ Activate RunChat and subscribe to data
        if hasattr(self.runchat_page, "on_show"):
            self.runchat_page.on_show()
        try:
            self.data_manager.subscribe(self.runchat_page)
        except Exception:
            pass
        self.run_chat_active = True
        self.active_page = "RunChat"

    def load_live_data(self):
        """Show the Live Data page and pause RunChat redraws."""
        self._switch_page_with_overlay(self._do_load_live_data,
                                       page_attr="live_data_page")

    def _do_load_live_data(self):
        for widget in self.content_frame.winfo_children():
            widget.pack_forget()
        if hasattr(self, "runchat_page") and hasattr(self.runchat_page, "on_hide"):
            try:
                self.runchat_page.on_hide()
            except Exception:
                pass
        if hasattr(self, "live_data_page") and self.live_data_page.winfo_exists():
            self.live_data_page.pack(fill="both", expand=True)
        else:
            self.live_data_page = LiveDataPage(self.content_frame, self)
            self.live_data_page.pack(fill="both", expand=True)
        self.active_page = "LiveData"

    def load_usb_data(self):
        """Show the USB Data page."""
        self._switch_page_with_overlay(self._do_load_usb_data,
                                       page_attr="usb_data_page")

    def _do_load_usb_data(self):
        for widget in self.content_frame.winfo_children():
            widget.pack_forget()
        if hasattr(self, "runchat_page") and hasattr(self.runchat_page, "on_hide"):
            try:
                self.runchat_page.on_hide()
            except Exception:
                pass
        if hasattr(self, "usb_data_page") and self.usb_data_page.winfo_exists():
            self.usb_data_page.pack(fill="both", expand=True)
        else:
            self.usb_data_page = UsbDataPage(self.content_frame, self)
            self.usb_data_page.pack(fill="both", expand=True)
        self.active_page = "UsbData"

    def load_report_page(self):
        """Show the SPC Report page."""
        self._switch_page_with_overlay(self._do_load_report_page,
                                       page_attr="report_page")

    def _do_load_report_page(self):
        for widget in self.content_frame.winfo_children():
            widget.pack_forget()
        if hasattr(self, "report_page") and self.report_page.winfo_exists():
            self.report_page.pack(fill="both", expand=True)
            try:
                self.report_page.refresh_table_data()
            except Exception:
                pass
        else:
            self.report_page = ReportPage(self.content_frame, self)
            self.report_page.pack(fill="both", expand=True)
        self.active_page = "Report"

    def load_logout(self):
        """Show the Logout page."""
        self._switch_page_with_overlay(self._do_load_logout,
                                       page_attr="logout_page")

    def _do_load_logout(self):
        for widget in self.content_frame.winfo_children():
            widget.pack_forget()
        if hasattr(self, "runchat_page") and hasattr(self.runchat_page, "on_hide"):
            try:
                self.runchat_page.on_hide()
            except Exception:
                pass
        if hasattr(self, "logout_page") and self.logout_page.winfo_exists():
            self.logout_page.pack(fill="both", expand=True)
        else:
            self.logout_page = LogoutPage(self.content_frame, self)
            self.logout_page.pack(fill="both", expand=True)
        self.active_page = "Logout"


    def create_header(self):
        # === Modern Header Bar with White Background ===
        self.header = ctk.CTkFrame(self, fg_color="white", height=70, corner_radius=0)
        self.header.pack(side="top", fill="x")
        
        # Add subtle inner shadow effect with nested frame
        header_inner = ctk.CTkFrame(self.header, fg_color="transparent")
        header_inner.pack(fill="both", expand=True, padx=2, pady=2)
        
        # --- Title with modern styling ---
        try:
            # Use same logo as sidebar (ensure resource_path is working)
            logo_path = resource_path("settings/cherry_full_logo.png")
            
            # Load and round corners
            pil_img = Image.open(logo_path)
            pil_img = pil_img.resize((180, 50), Image.Resampling.LANCZOS) # Slightly smaller for header
            pil_img = add_corners(pil_img, 15) # Add 15px radius corners
            
            # Keep persistent reference
            self.logo_header = ctk.CTkImage(pil_img, size=(180, 50))
            
            logo_label = ctk.CTkLabel(header_inner, text="", image=self.logo_header)
            logo_label.pack(side="left", padx=(25, 10), pady=10)
            
            # Add Text Title Next to Logo (Red separator, Green text)
            separator_label = ctk.CTkLabel(
                header_inner,
                text="|",
                font=("Segoe UI", 18, "bold"),
                text_color="#D32F2F"
            )
            separator_label.pack(side="left", padx=(0, 10), pady=15)
            
            title_text = ctk.CTkLabel(
                header_inner, 
                text="Airgauge Monitoring",
                font=("Segoe UI", 18, "bold"),
                text_color="#1B5E20"
            )
            title_text.pack(side="left", padx=(0, 24), pady=15)
            
        except Exception as e:
            # Fallback text if logo fails
            print(f"Header logo error: {e}")
            title = ctk.CTkLabel(
                header_inner,
                text="Cherry Precision – Air Gauge Monitoring",
                font=("Segoe UI", 18, "bold"),
                text_color="#1B5E20"
            )
            title.pack(side="left", padx=24, pady=15)

        # --- Connection Status Indicator (Modern Glass Card) ---
        status_frame = ctk.CTkFrame(
            header_inner, 
            fg_color="white", 
            corner_radius=12,
            border_width=1,
            border_color="#E0E0E0"
        )
        status_frame.pack(side="right", padx=12, pady=12)

        self.status_light = ctk.CTkLabel(
            status_frame,
            text="●",
            text_color="#F44336",  # 🔴 default disconnected
            font=("Segoe UI", 18, "bold")
        )
        self.status_light.pack(side="left", padx=(12, 6), pady=8)

        self.status_label = ctk.CTkLabel(
            status_frame,
            text="Disconnected",
            font=("Segoe UI", 13, "bold"),
            text_color="black"
        )
        self.status_label.pack(side="left", padx=(0, 12), pady=8)
        
        # Start pulsing animation for status light
        self._pulse_status_light()

        # --- Modern Connect Button with Glass Effect ---
        self.connect_btn = ModernButton(
            header_inner,
            text="🔌 Connect",
            width=120,
            height=42,
            fg_color="#1B5E20",
            hover_color="#144A26",
            corner_radius=12,
            border_width=0,
            font=("Segoe UI", 13, "bold"),
            command=self.connect_esp32
        )
        self.connect_btn.pack(side="left", padx=8, pady=12)

        # --- Modern Disconnect Button ---
        self.disconnect_btn = ModernButton(
            header_inner,
            text="❌ Disconnect",
            width=130,
            height=42,
            fg_color="#C62828",
            hover_color="#A91E1E",
            corner_radius=12,
            border_width=0,
            font=("Segoe UI", 13, "bold"),
            command=self.disconnect_esp32
        )
        self.disconnect_btn.pack(side="left", padx=8, pady=12)

        # --- WiFi Signal Strength Widget ---
        wifi_frame = ctk.CTkFrame(
            header_inner,
            fg_color="white",
            corner_radius=12,
            border_width=1,
            border_color="#E0E0E0"
        )
        wifi_frame.pack(side="left", padx=(4, 8), pady=12)

        # Canvas for WiFi bars icon
        self._wifi_canvas = tk.Canvas(
            wifi_frame,
            width=36, height=28,
            bg="white",
            highlightthickness=0
        )
        self._wifi_canvas.pack(side="left", padx=(8, 2), pady=4)

        self._wifi_label = ctk.CTkLabel(
            wifi_frame,
            text="--",
            font=("Segoe UI", 11, "bold"),
            text_color="#757575",
            width=58
        )
        self._wifi_label.pack(side="left", padx=(2, 8), pady=4)

        # Draw initial empty bars
        self._draw_wifi_bars(0, "#9E9E9E")

        # Start polling rssi_queue
        self.after(1000, self._update_rssi_display)

    
    def _draw_wifi_bars(self, bars, color):
        """Draw 4-bar WiFi strength icon on the canvas. bars=0..4."""
        try:
            c = self._wifi_canvas
            c.delete("all")
            # 4 bars, each progressively taller, left to right
            bar_w = 5
            gap   = 3
            max_h = 22
            base_y = 26
            for i in range(4):
                x1 = 2 + i * (bar_w + gap)
                x2 = x1 + bar_w
                h  = int(max_h * (i + 1) / 4)
                y1 = base_y - h
                # Active bar if within signal level
                bar_color = color if (i < bars) else "#E0E0E0"
                c.create_rectangle(x1, y1, x2, base_y, fill=bar_color, outline="", tags="bar")
        except Exception:
            pass

    def _update_rssi_display(self):
        """Poll rssi_queue every 1 second and update WiFi bars widget."""
        if getattr(self, '_closing', False):
            return
        try:
            # Drain queue, keep only latest value
            val = None
            while not self.rssi_queue.empty():
                val = self.rssi_queue.get_nowait()

            if val is not None:
                if val == "DISC" or val == "DISCONNECTED":
                    self._draw_wifi_bars(0, "#EF4444")
                    self._wifi_label.configure(text="No WiFi", text_color="#EF4444")
                else:
                    try:
                        rssi = int(val)
                        # Map RSSI to bars and color
                        if rssi >= -50:
                            bars, color, txt_color = 4, "#22C55E", "#22C55E"
                        elif rssi >= -65:
                            bars, color, txt_color = 3, "#22C55E", "#22C55E"
                        elif rssi >= -75:
                            bars, color, txt_color = 2, "#F59E0B", "#F59E0B"
                        elif rssi >= -85:
                            bars, color, txt_color = 1, "#EF4444", "#EF4444"
                        else:
                            bars, color, txt_color = 0, "#EF4444", "#EF4444"
                        self._draw_wifi_bars(bars, color)
                        self._wifi_label.configure(
                            text=f"{rssi} dBm",
                            text_color=txt_color
                        )
                    except ValueError:
                        pass
        except Exception:
            pass
        # Schedule next poll
        self.after(1000, self._update_rssi_display)

    def _pulse_status_light(self):

        """Animate status light with pulsing effect"""
        try:
            if getattr(self, '_closing', False) or not self.winfo_exists():
                return
        except Exception:
            return
        if not hasattr(self, '_pulse_state'):
            self._pulse_state = 0
        
        # Only pulse when disconnected (red)
        try:
            current_color = self.status_light.cget("text_color")
            if current_color == "#F44336":  # Red (disconnected)
                # Alternate between red and darker red
                colors = ["#F44336", "#C62828"]
                self._pulse_state = (self._pulse_state + 1) % 2
                self.status_light.configure(text_color=colors[self._pulse_state])
        except Exception:
            pass
        
        # Continue pulsing (1500ms: less frequent = less UI thread overhead)
        self.after(1500, self._pulse_status_light)


    def toggle_sidebar(self):
        if self.sidebar.winfo_ismapped():
            self.sidebar.pack_forget()
            self.sidebar_user_visible = False
        else:
            self.sidebar.pack(side="left", fill="y", padx=(15, 0), pady=15, before=self.content_frame)
            self.sidebar_user_visible = True

    def create_sidebar(self):
        # === Modern Sidebar with Glassmorphism ===
        self.sidebar = ctk.CTkFrame(
            self, 
            fg_color="white",
            width=240,
            corner_radius=15,
            border_width=1,
            border_color="#E0E0E0"
        )
        # Packed by default
        self.sidebar.pack(side="left", fill="y", padx=(15, 0), pady=15, before=self.content_frame)
        self.sidebar_user_visible = True
        
        # Add subtle inner border effect
        sidebar_inner = ctk.CTkFrame(
            self.sidebar,
            fg_color="transparent",
            corner_radius=15
        )
        sidebar_inner.pack(fill="both", expand=True, padx=0, pady=0)

        # --- Modern Menu Items with Icons ---
        menu_items = [
            ("🏠", "Home", self.load_home, "home"),
            ("⚙️", "Settings", self.verify_settings_password, "settings"),
            ("📈", "Run Chart", self.load_run_chat, "runchat"),
            ("📡", "Live Data", self.load_live_data, "livedata"),
            ("🔌", "USB", self.load_usb_data, "usbdata"),
            ("📊", "Report", self.load_report_page, "report"),
        ]

        # Store buttons for active state management
        self.sidebar_buttons = {}
        
        for symbol, word, func, key in menu_items:
             # Create button container for hover effect
             btn_container = ctk.CTkFrame(
                 sidebar_inner,
                 fg_color="transparent"
             )
             # Add top margin for the first item (Home) to align it perfectly
             pady_val = (15, 4) if key == "home" else 4
             btn_container.pack(fill="x", padx=12, pady=pady_val)
             
             btn = SidebarButton(
                 btn_container,
                 symbol=symbol,
                 word=word,
                 command=lambda f=func, k=key: self._sidebar_nav(f, k)
             )
             btn.pack(fill="x", padx=2, pady=2)
             
             # Store button reference
             self.sidebar_buttons[key] = btn
        
        # --- Footer/Version Info ---
        footer_frame = ctk.CTkFrame(
            sidebar_inner,
            fg_color="transparent"
        )
        footer_frame.pack(side="bottom", fill="x", pady=(10, 20))
        
        ctk.CTkLabel(
            footer_frame,
            text="v2.0",
            font=("Segoe UI", 9),
            text_color="#757575"
        ).pack()

        # --- Logout Button at the bottom (above version info) ---
        logout_container = ctk.CTkFrame(
            sidebar_inner,
            fg_color="transparent"
        )
        logout_container.pack(side="bottom", fill="x", padx=12, pady=10)
        
        logout_btn = SidebarButton(
            logout_container,
            symbol="🚪",
            word="Logout",
            command=lambda: self._sidebar_nav(self.load_logout, "logout")
        )
        logout_btn.pack(fill="x", padx=2, pady=2)
        
        self.sidebar_buttons["logout"] = logout_btn
    
    def _sidebar_nav(self, func, key):
        """Handle sidebar navigation with active state"""
        # Reset all buttons to normal state
        for btn_key, btn in self.sidebar_buttons.items():
            if btn_key == key:
                btn.set_active(True)
            else:
                btn.set_active(False)
        func()



    def create_content_area(self):
        # === Modern Content Area with Subtle Background ===
        self.content_frame = ctk.CTkFrame(
            self, 
            fg_color="#F5F7FA",
            corner_radius=0
        )
        self.content_frame.pack(side="right", fill="both", expand=True)

    def _on_content_resize(self, event):
        """Debounced resize of background image."""
        if self.bg_resize_job:
            self.after_cancel(self.bg_resize_job)
        self.bg_resize_job = self.after(100, lambda: self._resize_background_action(event.width, event.height))

    def _resize_background_action(self, new_w, new_h):
        if not hasattr(self, "bg_pil_original") or not self.bg_pil_original:
            return
        # Safety for very small windows
        if new_w < 10 or new_h < 10:
            return
            
        try:
            # Skip if dimensions unchanged (avoid redundant CPU work)
            if (new_w, new_h) == getattr(self, '_last_bg_size', None):
                return
            self._last_bg_size = (new_w, new_h)
            # NEAREST is ~10x faster than BILINEAR, imperceptible on a blurred bg
            resized = self.bg_pil_original.resize((new_w, new_h), Image.Resampling.NEAREST)
            self.bg_image_ref = ctk.CTkImage(resized, size=(new_w, new_h))
            self.bg_label.configure(image=self.bg_image_ref)
        except Exception as e:
            print("Resize error:", e)
            
        # Welcome card removed as per user request
        # welcome_container was here

    # ---------------- UI Animations ----------------
    from ui_animations import AnimationHelper, Easing

    def load_page(self, page_class, name):
        """Load a page with smooth slide transition."""
        
        # 1. Prevent overlapping transitions
        if hasattr(self, "_transition_in_progress") and self._transition_in_progress:
            return
        
        # 2. Check if already active
        if self.active_page_name == name:
            return

        self._transition_in_progress = True
        
        # 3. Identify Old Page
        old_page = None
        for widget in self.content_frame.winfo_children():
            # Find the currently packed widget
            if widget.winfo_manager() == "pack":
                old_page = widget
                break

        # 4. Prepare New Page
        new_page = None
        
        # Reuse or Create
        if name == "Live Data" and hasattr(self, "live_data_page"):
            new_page = self.live_data_page
        elif name == "Component Setup" and hasattr(self, "component_page"):
            new_page = self.component_page
        elif name == "Run Chart" and hasattr(self, "runchat_page"):
            new_page = self.runchat_page
        elif name == "Report" and hasattr(self, "report_page"):
            new_page = self.report_page
        elif name == "USB Upload" and hasattr(self, "usb_page"):
            new_page = self.usb_page
        elif name == "Machine Setup":
            # For simpler pages, we might recreate or verify existence
             # (Machine setup logic is complex, keeping existing flow but adapting for animation)
            self._transition_in_progress = False # Abort transition for complex modal-like pages if preferred, OR implement fully.
            # actually, let's just stick to the main pages for animation to be safe.
            self.load_machine_setup() 
            return
        elif name == "Settings":
            self._transition_in_progress = False
            self.verify_settings_password()
            return
        else:
            # Create new page HIDDEN to prevent piece-by-piece rendering
            new_page = page_class(self.content_frame, self)
            new_page.update_idletasks()  # Force layout computation while hidden
            if name == "Component Setup": self.component_page = new_page
            elif name == "Run Chart": self.runchat_page = new_page
            elif name == "Live Data": self.live_data_page = new_page
            elif name == "Report": self.report_page = new_page
            elif name == "USB Upload": self.usb_page = new_page

        # Safety check
        if not new_page:
             self._transition_in_progress = False
             return

        self.active_page_name = name
        self.status_label.configure(text=f"{name} loaded ✅")

        # 5. Setup Animation Geometry
        container_width = self.content_frame.winfo_width()
        container_height = self.content_frame.winfo_height()
        
        # If no old page (first load), just pack
        if not old_page:
            new_page.pack(fill="both", expand=True)
            self._transition_in_progress = False
            return

        # Prepare for Animation:
        # Move old_page to 'place' at current position (0,0) so it doesn't vanish when we allow new_page to exist
        old_page.pack_forget()
        old_page.place(x=0, y=0, width=container_width, height=container_height)
        
        # Place new_page off-screen to the right
        new_page.place(x=container_width, y=0, width=container_width, height=container_height)
        new_page.lift() # Ensure it's on top

        # 6. Animate
        # We will manually animate both to slide: Old moves Left, New moves Left (into view)
        
        duration = 350 # ms
        start_time = time.time() * 1000
        
        def animate_step():
            current_time = time.time() * 1000
            elapsed = current_time - start_time
            progress = min(elapsed / duration, 1.0)
            
            # Easing
            eased = Easing.ease_out_cubic(progress)
            
            # Calculate positions
            # Old Page: 0 -> -width
            old_x = int(0 - (container_width * eased))
            
            # New Page: width -> 0
            new_x = int(container_width - (container_width * eased))
            
            try:
                old_page.place(x=old_x, y=0)
                new_page.place(x=new_x, y=0)
            except:
                # Page might be destroyed or error
                pass

            if progress < 1.0:
                self.after(16, animate_step)
            else:
                # 7. Cleanup
                try:
                    old_page.place_forget()
                    
                    # Finalize new page state
                    new_page.place_forget() # Remove exact positioning
                    new_page.pack(fill="both", expand=True) # Lock into layout
                    
                except Exception as e:
                    print(f"Animation cleanup error: {e}")
                
                self._transition_in_progress = False

        animate_step()



    import serial.tools.list_ports

    def auto_detect_esp32(self):
        """
        Return COM port of ANY ESP32/ESP8266 board.
        Detects based on VID/PID of all common USB-UART bridges.
        """

        # List of VID/PID pairs for all ESP boards
        ESP_USB_IDS = [
            ("1A86", "7523"),  # CH340
            ("1A86", "55D4"),  # CH9102F newer chip
            ("10C4", "EA60"),  # CP2102
            ("10C4", "EA70"),  # CP2104
            ("0403", "6001"),  # FT232R (FTDI)
            ("303A", "1001"),  # ESP32-S3 / C3 native USB CDC
            ("303A", "0002"),  # Espressif USB JTAG / UART
        ]

        for p in serial.tools.list_ports.comports():
            try:
                vid = f"{p.vid:04X}" if p.vid else None
                pid = f"{p.pid:04X}" if p.pid else None

                if (vid, pid) in ESP_USB_IDS:
                    return p.device  # COMx
            except Exception:
                continue

        return None  # No ESP detected


    def connect_esp32(self):
        port = self.auto_detect_esp32()
        if not port:
            self.status_label.configure(text="ESP32 not detected ⚠️")
            self.status_light.configure(text_color="#F44336")  # 🔴
            return

        try:
            self.serial_conn = serial.Serial(port, 115200, timeout=0.002)
            self.status_label.configure(text=f"Port {port} opened ✅ Waiting for ESP32...")
            self.status_light.configure(text_color="#FFC107")  # 🟡 connecting

            time.sleep(1.5)
            self.serial_conn.reset_input_buffer()

            self.serial_conn.write(b"PING\n")
            time.sleep(0.2)
            response = self.serial_conn.read(self.serial_conn.in_waiting or 1).decode(errors="ignore")

            self.start_serial_reader()
            self.serial_conn.write(b"START\n")

            self.status_label.configure(text=f"Connected to {port} ✅ (Streaming started)")
            self.status_light.configure(text_color="#4CAF50")  # 🟢 Connected
            self.connect_btn.configure(state="disabled", fg_color="#9E9E9E", hover_color="#9E9E9E")

        except Exception as e:
            self.status_label.configure(text=f"Failed to open {port}: {e}")
            self.status_light.configure(text_color="#F44336")  # 🔴
            self.serial_conn = None

    def disconnect_esp32(self):
        """Send STOP command before closing connection."""
        try:
            if hasattr(self, "serial_conn") and self.serial_conn and self.serial_conn.is_open:
                self.serial_conn.write(b"STOP\n")
                time.sleep(0.12)
                self.serial_conn.close()
                self.status_label.configure(text="Disconnected ✅")
                self.status_light.configure(text_color="#F44336")  # 🔴
                self.connect_btn.configure(state="normal", fg_color="#43A047", hover_color="#388E3C")

        except Exception as e:
            print(f"Error disconnecting: {e}")
        self.serial_conn = None

    def auto_disconnect_handler(self):
        """Automatically called when ESP32 cable is unplugged or serial connection fails."""
        try:
            # 1. Safety check: if app is closing/closed, do nothing
            try:
                if not self.winfo_exists():
                    return
            except Exception:
                return

            self.status_label.configure(text="Disconnected (Cable removed) ⚠️")
            self.status_light.configure(text_color="#F44336")  # 🔴 red

            # Re-enable Connect button
            self.connect_btn.configure(
                state="normal",
                fg_color="#43A047",
                hover_color="#388E3C"
            )

            # Close serial safely
            if hasattr(self, "serial_conn") and self.serial_conn:
                try:
                    self.serial_conn.close()
                except:
                    pass

            self.serial_conn = None

        except Exception as e:
            # Ignore errors if we are shutting down
            if "invalid command name" in str(e):
                return
            print("Auto-disconnect handler error:", e)


    def start_serial_reader(self):
        """High-speed serial reader with auto-disconnect detection."""
        
        def read_loop():
            buffer = ""
            first_data = True
            sync_expected_lines = 0  # Number of IP lines expected after LIST_IP command

            while True:
                try:
                    # If serial connection object missing or closed → disconnect
                    if not getattr(self, "serial_conn", None) or not self.serial_conn.is_open:
                        raise serial.SerialException("Serial closed")

                    # Read incoming bytes
                    incoming = self.serial_conn.read(self.serial_conn.in_waiting or 1).decode(errors="ignore")

                    if incoming:
                        buffer += incoming

                        # Process full packets
                        while "\n" in buffer:
                            line, buffer = buffer.split("\n", 1)
                            line = line.strip()
                            if not line:
                                continue

                            # --- SYNC: Check for Master ACK messages ---
                            if line.startswith("IP ADDED:") or line.startswith("IP REMOVED:") or \
                               line.startswith("IP ALREADY EXISTS") or line.startswith("IP NOT FOUND") or \
                               line.startswith("MAX SLAVES REACHED"):
                                # Dispatch to main thread safely
                                self.after(0, lambda m=line: self.handle_master_message(m))
                                continue

                            # --- LIST_IP Sync: Detect "COUNT: N" and raw IPs ---
                            # Firmware sends: "COUNT: N" then N lines of IPs
                            
                            # 1. Detect Header
                            if line.startswith("COUNT:"):
                                try:
                                    cnt = int(line.split(":")[1].strip())
                                    sync_expected_lines = cnt
                                except:
                                    sync_expected_lines = 0
                                
                                # Send the COUNT message itself to GUI for feedback
                                self.after(0, lambda m=line: self.handle_master_message(m))
                                continue

                            # 2. In Sync Mode? Capture anything as an IP
                            if sync_expected_lines > 0:
                                sync_expected_lines -= 1
                                # Treat this line as an IP, even if weird
                                self.after(0, lambda m=line: self.handle_master_message(m))
                                continue

                            # 3. Fallback: accidental sync catch (if count missed)
                            if line.replace(".", "").isdigit() and line.count(".") == 3:
                                self.after(0, lambda m=line: self.handle_master_message(m))
                                continue

                            # --- RSSI: intercept before data parsing ---
                            if line.startswith("RSSI:"):
                                try:
                                    val = line[5:].strip()
                                    if not self.rssi_queue.full():
                                        self.rssi_queue.put_nowait(val)
                                except Exception:
                                    pass
                                continue

                            parsed = parse_packet(line)
                            if parsed:
                                self.data_queue.put(parsed)


                                if first_data:
                                    self.status_label.configure(text="✅ Data streaming started")
                                    first_data = False

                    # prevent CPU overload (2ms sleep = max 500 iterations/sec)
                    time.sleep(0.002)

                except serial.SerialException:
                    # 🔥 ESP32 cable removed / port closed
                    self.auto_disconnect_handler()
                    break

                except OSError:
                    # 🔥 Some OS-level disconnection events
                    self.auto_disconnect_handler()
                    break

                except Exception:
                    # ignore parsing noise or junk
                    continue

        threading.Thread(target=read_loop, daemon=True).start()


    def refresh_id_dependent_pages(self):
        """Update dropdowns or tables in Component Setup & RunChat when IDs change."""
        if hasattr(self, "component_page") and self.component_page:
            self.component_page.refresh_airgauge_ids()
        if hasattr(self, "runchat_page") and self.runchat_page:
            self.runchat_page.refresh_airgauge_ids()

    def handle_master_message(self, msg):
        """Handle control messages from Master (IP ADDED/REMOVED)."""
        print(f"Master Message: {msg}")
        
        # Parse message
        if msg.startswith("IP ADDED:"):
            # Format: IP ADDED: 192.168.1.10
            ip = msg.split(":", 1)[1].strip()
            # We need the ID requested. Since this is async, we can check 
            # if the MachineSetupPage has a pending add, OR we just trust the user flow.
            # Ideally, the Master should return "IP ADDED: <IP> -> <ID>" if we sent both.
            # But the protocol only sends ADD_IP <IP>. 
            # So here we will just trigger a reload if we have the data locally or 
            # we can't fully sync the ID from the "IP ADDED" message alone if it's new.
            
            # WAIT: The Master stores IPs. It doesn't store IDs (IDs are local mapping in hh.py).
            # The firmware command was `ADD_IP <IP>`. It stores IP in a list.
            # The standard `hh.py` logic maps IP -> ID in `ConfigManager`.
            # If we don't save locally first, we lose the ID mapping.
            # FIX: We MUST save the ID mapping locally *before* confirming, OR 
            # we just confirm the IP is alive. 
            
            # The USER wants: "only in serial master send added cmd ... then only in gui show"
            # So we should probably cache the pending ID in MachineSetupPage and save it now.
            
            if hasattr(self, "machine_setup_page") and self.machine_setup_page and self.machine_setup_page.winfo_exists():
                self.machine_setup_page.on_master_ack_add(ip)
            
            self.status_label.configure(text=f"✅ ACKnowledged: {ip} added")
            
        elif msg.startswith("IP REMOVED:"):
            ip = msg.split(":", 1)[1].strip()
            
            # Remove from local config
            try:
                conn = sqlite3.connect("airgauge_master.db")
                conn.execute("DELETE FROM airgauge_master WHERE ip=?", (ip.strip(),))
                conn.commit()
                conn.close()
            except:
                pass
            
            if hasattr(self, "machine_setup_page") and self.machine_setup_page and self.machine_setup_page.winfo_exists():
                self.machine_setup_page.refresh_table()
                
            self.status_label.configure(text=f"✅ ACKnowledged: {ip} removed")

        elif msg.startswith("IP ALREADY EXISTS") or ("ALREADY EXISTS" in msg):
             # If user is trying to ADD an IP that the Master already has, we should 
             # treat this as a success for the local config (restoring the mapping).
             if hasattr(self, "machine_setup_page") and \
                self.machine_setup_page and \
                self.machine_setup_page.winfo_exists() and \
                hasattr(self.machine_setup_page, "pending_add_ip") and \
                self.machine_setup_page.pending_add_ip:
                 
                 # The Master has it, so just update our local map
                 ip_attempt = self.machine_setup_page.pending_add_ip
                 self.machine_setup_page.on_master_ack_add(ip_attempt)
                 self.status_label.configure(text=f"✅ Sync restored: {ip_attempt}")
             else:
                 messagebox.showinfo("Info", f"Master already has this IP: {msg}")

        elif msg.startswith("COUNT:"):
             # Start of a LIST_IP response
             count = msg.split(":")[1].strip()
             self.status_label.configure(text=f"Syncing... Master has {count} IPs")
        
        elif "MAX SLAVES" in msg:
             messagebox.showerror("Master Error", "Max Slave limit reached on Master.")
        
        else:
             # Assume any other message reaching here during sync is an IP
             # (Or matched the strict IP check)
             potential_ip = msg.strip()
             # Basic sanity check to avoid adding "Unknown command" etc
             if "Unknown" not in potential_ip and "CMD" not in potential_ip and len(potential_ip) > 3:
                 if hasattr(self, "machine_setup_page") and \
                    self.machine_setup_page and \
                    self.machine_setup_page.winfo_exists():
                     self.machine_setup_page.add_discovered_ip(potential_ip)


        
#================================================
#---------- MachineSetupPage ---------------------
#================================================
        
class MachineSetupPage(ctk.CTkFrame):
    def __init__(self, parent, app):
        super().__init__(parent, fg_color=("#F8F9FA", "#1C1C1C"))
        self.app = app
        self.build_ui()
        self.load_existing_data()

    def build_ui(self):
        # Header Frame for Back Button + Title
        header_frame = ctk.CTkFrame(self, fg_color="transparent")
        header_frame.pack(fill="x", pady=(15, 10), padx=15)

        # Back Button in Green
        btn_back = ModernButton(
            header_frame,
            text="← Back",
            width=80,
            height=32,
            font=("Segoe UI", 11, "bold"),
            fg_color="#007B43",
            hover_color="#005C32",
            text_color="white",
            corner_radius=6,
            command=self.go_back
        )
        btn_back.pack(side="left")

        # Branded Title: "AirGauge" (Green) + "Setup" (Red)
        title_container = ctk.CTkFrame(header_frame, fg_color="transparent")
        title_container.pack(side="left", padx=20)
        
        lbl_airgauge = ctk.CTkLabel(
            title_container,
            text="AirGauge",
            font=("Segoe UI", 18, "bold"),
            text_color="#007B43"
        )
        lbl_airgauge.pack(side="left")
        
        lbl_setup = ctk.CTkLabel(
            title_container,
            text=" Setup",
            font=("Segoe UI", 18, "bold"),
            text_color="#B0050E"
        )
        lbl_setup.pack(side="left")

        # Company Logo on the right
        try:
            logo_path = resource_path("settings/cherry_full_logo.png")
            pil_img = Image.open(logo_path)
            pil_img = pil_img.resize((144, 40), Image.Resampling.LANCZOS)
            pil_img = add_corners(pil_img, 10)
            self.logo_img = ctk.CTkImage(pil_img, size=(144, 40))
            logo_lbl = ctk.CTkLabel(header_frame, text="", image=self.logo_img)
            logo_lbl.pack(side="right", padx=(10, 20))
        except Exception as e:
            print(f"Error loading logo in MachineSetupPage: {e}")

        # Controls Card Frame
        entry_frame = ModernCardFrame(self)
        entry_frame.pack(pady=10, padx=15, fill="x")
        
        # Configure columns for grid layout
        entry_frame.grid_columnconfigure(0, weight=0)
        entry_frame.grid_columnconfigure(1, weight=1)
        entry_frame.grid_columnconfigure(2, weight=0)
        entry_frame.grid_columnconfigure(3, weight=1)
        for col_idx in range(4, 8):
            entry_frame.grid_columnconfigure(col_idx, weight=0)

        # --- IP Address entry ---
        ctk.CTkLabel(entry_frame, text="IP Address:", font=("Segoe UI", 11, "bold"), text_color="#202124").grid(row=0, column=0, padx=(15, 5), pady=8, sticky="e")
        self.mac_entry = ctk.CTkEntry(entry_frame, width=160, height=30, corner_radius=6, border_color="#E0E0E0")
        self.mac_entry.grid(row=0, column=1, padx=(5, 15), pady=8, sticky="ew")

        # --- AirGauge ID entry ---
        ctk.CTkLabel(entry_frame, text="AirGauge ID:", font=("Segoe UI", 11, "bold"), text_color="#202124").grid(row=0, column=2, padx=(15, 5), pady=8, sticky="e")
        self.id_entry = ctk.CTkEntry(entry_frame, width=160, height=30, corner_radius=6, border_color="#E0E0E0")
        self.id_entry.grid(row=0, column=3, padx=(5, 15), pady=8, sticky="ew")

        # --- Action Buttons ---
        add_btn = ModernButton(
            entry_frame,
            text="+ Add",
            height=32,
            font=("Segoe UI", 11, "bold"),
            fg_color="#007B43",
            hover_color="#005C32",
            text_color="white",
            corner_radius=6,
            command=self.add_entry
        )
        add_btn.grid(row=0, column=4, padx=5, pady=8)
        
        del_btn = ModernButton(
            entry_frame,
            text="🗑 Delete",
            height=32,
            font=("Segoe UI", 11, "bold"),
            fg_color="#007B43",
            hover_color="#005C32",
            text_color="white",
            corner_radius=6,
            command=self.delete_entry
        )
        del_btn.grid(row=0, column=5, padx=5, pady=8)
        
        sync_btn = ModernButton(
            entry_frame,
            text="☁ Sync from Master",
            height=32,
            font=("Segoe UI", 11, "bold"),
            fg_color="#007B43",
            hover_color="#005C32",
            text_color="white",
            corner_radius=6,
            command=self.sync_from_master
        )
        sync_btn.grid(row=0, column=6, padx=5, pady=8)
        
        excel_btn = ModernButton(
            entry_frame,
            text="📊 Export Excel",
            height=32,
            font=("Segoe UI", 11, "bold"),
            fg_color="#007B43",
            hover_color="#005C32",
            text_color="white",
            corner_radius=6,
            command=self.export_to_excel
        )
        excel_btn.grid(row=0, column=7, padx=(5, 15), pady=8)
        
        # --- Table Frame ---
        self.table_frame = ctk.CTkScrollableFrame(self, corner_radius=6, fg_color="white", border_width=1, border_color="#E0E0E0")
        self.table_frame.pack(padx=15, pady=10, fill="both", expand=True)
        
        # Configure green theme scrollbar
        try:
            self.table_frame._scrollbar.configure(button_color="#007B43", button_hover_color="#005C32")
        except Exception:
            pass

        # Build table header
        self.header_row = ctk.CTkFrame(self.table_frame, fg_color="white", height=40, corner_radius=0)
        self.header_row.pack(fill="x", pady=(0, 0))
        self.header_row.pack_propagate(False)
        
        # IP Address Cell
        cell_ip = ctk.CTkFrame(self.header_row, fg_color="white", corner_radius=0, border_width=1, border_color="#E0E0E0", width=190, height=40)
        cell_ip.grid(row=0, column=0, sticky="nsew")
        cell_ip.grid_propagate(False)
        
        content_ip = ctk.CTkFrame(cell_ip, fg_color="transparent")
        content_ip.pack(side="left", padx=20)
        ctk.CTkLabel(content_ip, text="🌐\ufe0e", font=("Segoe UI", 11), text_color="#007B43").pack(side="left", padx=(0, 4))
        ctk.CTkLabel(content_ip, text="IP Address", font=("Segoe UI", 11, "bold"), text_color="#1A1A1A").pack(side="left")

        # AirGauge ID Cell
        cell_ag = ctk.CTkFrame(self.header_row, fg_color="white", corner_radius=0, border_width=1, border_color="#E0E0E0", width=190, height=40)
        cell_ag.grid(row=0, column=1, sticky="nsew")
        cell_ag.grid_propagate(False)
        
        content_ag = ctk.CTkFrame(cell_ag, fg_color="transparent")
        content_ag.pack(side="left", padx=20)
        ctk.CTkLabel(content_ag, text="🏷\ufe0e", font=("Segoe UI", 11), text_color="#007B43").pack(side="left", padx=(0, 4))
        ctk.CTkLabel(content_ag, text="AirGauge ID", font=("Segoe UI", 11, "bold"), text_color="#1A1A1A").pack(side="left")

        # Green bottom border line
        self.header_border = ctk.CTkFrame(self.table_frame, fg_color="#007B43", height=2, corner_radius=0)
        self.header_border.pack(fill="x", pady=(0, 5))

    def go_back(self, event=None):
        try:
            self.app.load_settings_page()
        except:
            pass

    def export_to_excel(self):
        """Export the AirGauge Setup table data to Excel."""
        try:
            data = self.load_json()
            if not data:
                messagebox.showinfo("Export", "No data available to export.")
                return
            
            from tkinter import filedialog
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            xlsx_path = filedialog.asksaveasfilename(
                title="Save Excel Setup As",
                initialfile=f"AirGauge_Setup_{timestamp}.xlsx",
                defaultextension=".xlsx",
                filetypes=[("Excel File", "*.xlsx")]
            )
            if not xlsx_path:
                return

            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "AirGauge Setup"
            
            # Write headers
            headers = ["IP Address", "AirGauge ID"]
            ws.append(headers)
            
            # Style header
            header_font = Font(bold=True, color="FFFFFF")
            fill_header = PatternFill(start_color="007B43", end_color="007B43", fill_type="solid")
            for col_idx in range(1, 3):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = fill_header
                cell.alignment = Alignment(horizontal="center")
                
            # Write data
            for ip, ag_id in data.items():
                ws.append([ip, str(ag_id)])
                
            # Auto-adjust columns
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 15)
                
            wb.save(xlsx_path)
            messagebox.showinfo("Export Success", f"Successfully exported to:\n{xlsx_path}")
            
            try:
                os.startfile(xlsx_path)
            except:
                pass
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export: {e}")

    def load_existing_data(self):
        self.refresh_table()

    def refresh_table(self):
        for widget in self.table_frame.winfo_children():
            # Skip internal scrollbar widgets safely
            if hasattr(widget, "_parent_frame"):
                continue
            # Skip header row
            if widget == getattr(self, "header_row", None):
                continue
            if widget == getattr(self, "header_border", None):
                continue
            if isinstance(widget, ctk.CTkFrame):
                widget.destroy()


        data = self.load_json()
        for mac, ag_id in data.items():
            row = ctk.CTkFrame(self.table_frame, fg_color=("#FAFAFA", "#3a3a3a"), corner_radius=6)
            row.pack(fill="x", pady=2)
            
            # Use a transparent button or bind events to labels to make row clickable
            # Using labels with bindings is easier for layout control
            lbl_ip = ctk.CTkLabel(row, text=mac, font=("Segoe UI", 11), width=150, anchor="w")
            lbl_ip.grid(row=0, column=0, padx=20, sticky="w")
            
            lbl_id = ctk.CTkLabel(row, text=str(ag_id), font=("Segoe UI", 11), width=150, anchor="w")
            lbl_id.grid(row=0, column=1, padx=20, sticky="w")
            
            # Bind click to populate fields
            for widget in (row, lbl_ip, lbl_id):
                widget.bind("<Button-1>", lambda e, m=mac, i=ag_id: self.on_row_click(m, i))

    def on_row_click(self, mac, ag_id):
        self.mac_entry.delete(0, "end")
        self.mac_entry.insert(0, mac)
        
        self.id_entry.delete(0, "end")
        self.id_entry.insert(0, str(ag_id))

    def _ensure_db(self):
        """Create the airgauge_master table if it doesn't exist."""
        db_path = resource_path("airgauge_master.db")
        conn = sqlite3.connect(db_path)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS airgauge_master (
                ip TEXT PRIMARY KEY,
                airgauge_id TEXT
            )
        """)
        conn.commit()
        conn.close()
        return db_path

    def load_json(self):
        try:
            db_path = self._ensure_db()
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT ip, airgauge_id FROM airgauge_master")
            rows = cursor.fetchall()
            conn.close()
            return {r[0]: r[1] for r in rows}
        except Exception:
            return {}

    def save_json(self, data):
        try:
            db_path = self._ensure_db()
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            cursor.execute("DELETE FROM airgauge_master")
            for ip, ag_id in data.items():
                cursor.execute("INSERT INTO airgauge_master (ip, airgauge_id) VALUES (?, ?)", (ip, str(ag_id)))
            conn.commit()
            conn.close()
        except Exception as e:
            print("Save AirGauge error:", e)

    def add_discovered_ip(self, ip):
        """Called when a raw IP is received during Sync. Saves to DB and refreshes table."""
        try:
            ip = ip.strip()
            if not ip:
                return
            db_path = self._ensure_db()          # always create table first
            conn = sqlite3.connect(db_path)
            exists = conn.execute("SELECT 1 FROM airgauge_master WHERE ip=?", (ip,)).fetchone()
            if not exists:
                conn.execute("INSERT INTO airgauge_master (ip, airgauge_id) VALUES (?, ?)", (ip, "Synced"))
                conn.commit()
            conn.close()
        except Exception as e:
            print("add_discovered_ip error:", e)

        # Always reschedule a refresh (cancel old one first so it doesn't fire early)
        if hasattr(self, "_refresh_job"):
            try:
                self.after_cancel(self._refresh_job)
            except Exception:
                pass
        self._refresh_job = self.after(600, self._do_batched_refresh)

    def _do_batched_refresh(self):
        if hasattr(self, "_refresh_job"):
            del self._refresh_job
        self.refresh_table()

    def destroy(self):
        super().destroy()


    def add_entry(self):
        mac = self.mac_entry.get().strip()
        ag_id = self.id_entry.get().strip()

        if not mac or not ag_id:
            messagebox.showwarning("Invalid Input", "Please enter both IP Address and AirGauge ID.")
            return

        # Always save locally first so the list updates immediately
        data = self.load_json()
        data[mac] = ag_id
        self.save_json(data)
        self.refresh_table()

        # Also send to ESP32 master if connected
        if getattr(self.app, "serial_conn", None):
            cmd = f"ADD_IP {mac}\n"
            self.app.serial_conn.write(cmd.encode())
            self.app.status_label.configure(text=f"Sent ADD_IP {mac} to Master...")
            # Store pending for ACK tracking
            self.pending_add_id = ag_id
            self.pending_add_ip = mac
        else:
            self.app.status_label.configure(text=f"Saved locally: {mac} (Master not connected)")

        self.mac_entry.delete(0, "end")
        self.id_entry.delete(0, "end")

    def on_master_ack_add(self, ack_ip):
        """Called by CherryApp when 'IP ADDED: ...' is received."""
        # Check if this matches what we tried to add
        if hasattr(self, 'pending_add_ip') and self.pending_add_ip == ack_ip:
            # Now we save to local JSON
            data = self.load_json()
            data[ack_ip] = self.pending_add_id
            self.save_json(data)
            
            # NOW refresh table
            self.refresh_table()
            
            # Cleanup
            self.pending_add_ip = None
            self.pending_add_id = None
            
        else:
            # Maybe added from another source or mismatch
            pass # Just ignore or refresh if you want to see raw IPs (but we map IP->ID)

    def delete_entry(self):
        mac = self.mac_entry.get().strip()
        ag_id = self.id_entry.get().strip()

        if not mac and not ag_id:
             messagebox.showwarning("Invalid Input", "Select a row or enter IP/ID to delete.")
             return

        data = self.load_json()
        found = False
        key_to_del = None
        
        # Search by IP first (most reliable for Sync items)
        if mac in data:
            key_to_del = mac
            found = True
        else:
            # Search by ID
            for k, v in list(data.items()):
                if str(v) == ag_id:
                    key_to_del = k
                    found = True
                    break
        
        if found and key_to_del:
            # We don't delete locally yet, wait for ACK
            # But we need the IP to send the command
            ip_to_delete = key_to_del
            
            if getattr(self.app, "serial_conn", None):
                cmd = f"DEL_IP {ip_to_delete}\n"
                self.app.serial_conn.write(cmd.encode())
                self.app.status_label.configure(text=f"Requesting delete of {ip_to_delete}...")
            else:
                messagebox.showwarning("Connection Error", "Master not connected. Cannot verify deletion.")
        else:
            messagebox.showinfo("Not Found", f"No entry found for IP {mac} or ID {ag_id}")

        self.mac_entry.delete(0, "end")
        self.id_entry.delete(0, "end")

    def sync_from_master(self):
        """Send LIST_IP to get all IPs from Master."""
        if getattr(self.app, "serial_conn", None):
            self.app.serial_conn.write(b"LIST_IP\n")
            self.app.status_label.configure(text="Requesting IP list from Master...")
            # Schedule a forced final refresh 2.5s after sending,
            # to catch all IPs that arrive in the response burst
            if hasattr(self, "_sync_refresh_job"):
                try:
                    self.after_cancel(self._sync_refresh_job)
                except Exception:
                    pass
            self._sync_refresh_job = self.after(2500, self.refresh_table)
        else:
            messagebox.showwarning("Connection Error", "Master not connected.")



        
# ==========================================================
# SETTING PAGE
# ==========================================================        


import os
import sys
from PIL import Image
import customtkinter as ctk
from tkinter import messagebox

class SettingsPage(ctk.CTkFrame):
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app

        # ===========================================================
        #  FUNCTION TO LOAD FILES (WORKS IN EXE + NORMAL PYTHON)
        # ===========================================================
        def resource_path(relative_path):
            """Return absolute path for EXE and normal run."""
            try:
                base_path = sys._MEIPASS    # EXE temp folder
            except Exception:
                # When running normally (not EXE)
                base_path = os.path.dirname(os.path.abspath(__file__))

            return os.path.join(base_path, relative_path)

        # ===========================================================
        #  LOAD ICONS FROM settings/ FOLDER
        # ===========================================================

        def _load_icon(rel_path, size=(100, 100)):
            try:
                img = Image.open(resource_path(rel_path)).resize(size, Image.Resampling.LANCZOS)
                return ctk.CTkImage(dark_image=img, size=size)
            except Exception as e:
                print(f"Icon load error ({rel_path}):", e)
                return None

        self.machine_icon        = _load_icon("settings/machinesetup.png",    (120, 95))
        self.item_icon           = _load_icon("settings/item_master.png",      (100, 100))
        self.general_icon        = _load_icon("settings/componentsetup.png",   (100, 100))
        self.customer_icon       = _load_icon("settings/coustomer.png",        (100, 100))
        self.operator_icon       = _load_icon("settings/operatorlist.png",     (100, 100))
        self.process_icon        = _load_icon("settings/process_master.png",   (100, 100))
        self.machine_master_icon = _load_icon("settings/machine_master.png",   (120, 100))

        # ===========================================================
        #  PAGE BACKGROUND  (light gray like the mockup)
        # ===========================================================
        self.configure(fg_color="#F0F2F5")

        # ===========================================================
        #  HEADER BAR  (white bar with logo + title + green border)
        # ===========================================================
        header_bar = ctk.CTkFrame(self, fg_color="white", corner_radius=0, height=70)
        header_bar.pack(fill="x", side="top")
        header_bar.pack_propagate(False)

        # Cherry full logo on the LEFT
        try:
            logo_path = resource_path("settings/cherry_full_logo.png")
            pil_logo = Image.open(logo_path).resize((140, 42), Image.Resampling.LANCZOS)
            self._hdr_logo = ctk.CTkImage(pil_logo, size=(140, 42))
            logo_lbl = ctk.CTkLabel(header_bar, image=self._hdr_logo, text="")
            logo_lbl.pack(side="left", padx=(18, 0), pady=14)
        except Exception as e:
            print("Header logo error:", e)

        # Centered title container (place so it stays centered regardless of logo width)
        title_center = ctk.CTkFrame(header_bar, fg_color="transparent")
        title_center.place(relx=0.5, rely=0.5, anchor="center")

        gear_lbl = ctk.CTkLabel(
            title_center, text="\u2699",
            font=("Segoe UI", 28, "bold"),
            text_color="#007B43"
        )
        gear_lbl.pack(side="left", padx=(0, 6))

        settings_lbl = ctk.CTkLabel(
            title_center, text="Settings",
            font=("Segoe UI", 24, "bold"),
            text_color="#007B43"
        )
        settings_lbl.pack(side="left")

        # Thin green separator line below header
        sep = ctk.CTkFrame(self, fg_color="#007B43", height=3, corner_radius=0)
        sep.pack(fill="x", side="top")

        # ===========================================================
        #  SCROLLABLE CONTENT AREA
        # ===========================================================
        scroll_area = ctk.CTkScrollableFrame(
            self, fg_color="#F0F2F5",
            scrollbar_button_color="#C8C8C8",
            scrollbar_button_hover_color="#A0A0A0"
        )
        scroll_area.pack(fill="both", expand=True)

        # ===========================================================
        #  GRID CONTAINER  (3 columns)
        # ===========================================================
        grid = ctk.CTkFrame(scroll_area, fg_color="transparent")
        grid.pack(fill="both", expand=True, padx=30, pady=28)

        for col in range(3):
            grid.grid_columnconfigure(col, weight=1, uniform="card_col")

        self.text_labels = []  # store for resize

        # ===========================================================
        #  CARD TILE BUILDER
        # ===========================================================
        def make_card(icon_image, label_text, row, col):
            """White rounded card with icon, bold label, green underline bar."""
            outer = ctk.CTkFrame(grid, fg_color="transparent")
            outer.grid(row=row, column=col, padx=18, pady=18, sticky="nsew")

            card = ctk.CTkFrame(
                outer,
                fg_color="white",
                corner_radius=16,
                border_width=1,
                border_color="#E4E8EE"
            )
            card.pack(fill="both", expand=True)

            # Icon
            if icon_image is not None:
                icon_lbl = ctk.CTkLabel(card, image=icon_image, text="")
            else:
                icon_lbl = ctk.CTkLabel(
                    card, text="\u2699",
                    font=("Segoe UI", 40, "bold"),
                    text_color="#007B43"
                )
            icon_lbl.pack(pady=(28, 10))

            # Label
            text_lbl = ctk.CTkLabel(
                card, text=label_text,
                font=("Segoe UI", 14, "bold"),
                text_color="#1A1A1A"
            )
            text_lbl.pack(pady=(0, 8))

            # Green underline bar
            bar_wrap = ctk.CTkFrame(card, fg_color="transparent")
            bar_wrap.pack(pady=(0, 20))
            bar = ctk.CTkFrame(bar_wrap, fg_color="#007B43", height=4, corner_radius=2, width=36)
            bar.pack()

            # Hover / click bindings
            def on_enter(e, c=card):
                c.configure(border_color="#007B43", border_width=2)

            def on_leave(e, c=card):
                c.configure(border_color="#E4E8EE", border_width=1)

            click_cb = lambda e, t=label_text: self.on_option_click(t)
            for w in (outer, card, icon_lbl, text_lbl, bar, bar_wrap):
                w.bind("<Button-1>", click_cb)
                w.bind("<Enter>", on_enter)
                w.bind("<Leave>", on_leave)

            self.text_labels.append((card, text_lbl))
            return card

        # ===========================================================
        #  CREATE 7 CARDS  (matches mockup layout)
        # ===========================================================
        make_card(self.machine_icon,        "AirGauge Setup",   0, 0)
        make_card(self.item_icon,           "Item Master",      0, 1)
        make_card(self.general_icon,        "Component Setup",  0, 2)
        make_card(self.customer_icon,       "Customer Master",  1, 0)
        make_card(self.operator_icon,       "Employee Master",  1, 1)
        make_card(self.process_icon,        "Process Master",   1, 2)
        make_card(self.machine_master_icon, "Machine Master",   2, 0)

    def on_parent_resized(self):
        """Called by main app window resize to adjust layout if needed."""
        try:
            for card, label in self.text_labels:
                w = card.winfo_width()
                if w > 20:
                    label.configure(wraplength=w - 20)
        except Exception:
            pass

    # ===========================================================
    #  CLICK HANDLER
    # ===========================================================
    def on_option_click(self, name):
        if name == "Component Setup":
            self.app.load_component_setup()
        elif name == "AirGauge Setup":
            self.app.verify_password()
        elif name == "Settings":
            self.app.verify_settings_password()
        elif name == "Employee Master":
            self.app.load_operator_list()
        elif name == "Customer Master":
            self.app.load_customer_master()
        elif name == "Item Master":
            self.app.load_item_master()
        elif name == "Process Master":
            self.app.load_process_master()
        elif name == "Machine Master":
            self.app.load_machine_master()
        else:
            messagebox.showinfo("Settings", f"You selected: {name}")


# ==========================================================
# Component Setup Page
# ==========================================================
class ComponentSetupPage(ctk.CTkFrame):
    COMPONENT_FILE = "airgauge_component_config.json"
    ITEMS_FILE = "items.json"  # item master source (same file used by Page)
    CUSTOMERS_FILE = "customers.json"

    def __init__(self, parent, app):
            super().__init__(parent, fg_color=("white", "#1c1c1c"))
            self.app = app

            # memory structures
            self.comp_map = self.load_json()
            self.items_master = self._load_items_master()  # list of dicts from items.json
            # mapping helpers
            self._code_to_name = {str(it.get("code", "")).strip(): str(it.get("name", "")).strip() for it in self.items_master}
            self._display_list = self._build_display_list(self.items_master)  # e.g. ["Guide Plate (IC001)", ...]
            self.selected_channel = ctk.StringVar(value="CH1")

            # ---- Customers master ----
            self.customers_master = self._load_customers_master()
            self._cust_code_to_name = {
                str(c.get("code", "")).strip(): str(c.get("name", "")).strip()
                for c in self.customers_master
            }
            self._customer_display_list = self._build_customer_display_list(self.customers_master)

            self.comp_type = ctk.StringVar(value="Shaft")

            # UI state
            self.entry_vars = {}
            self.entries = []
            self.build_ui()



    # -------------------------
    # UI BUILD (ALL WIDGETS HERE)
    # -------------------------
    def build_ui(self):
        # ====== HEADER ======
        header = ctk.CTkFrame(self, fg_color="transparent")
        header.pack(fill="x", side="top", pady=(15, 10), padx=20)

        # Back Button in Green
        ModernButton(
            header,
            text="← Back",
            width=80,
            height=32,
            font=("Segoe UI", 11, "bold"),
            fg_color="#007B43",
            hover_color="#005C32",
            text_color="white",
            corner_radius=6,
            command=self.go_back
        ).pack(side="left")
        
        # Branded Title: "Component" (Green) + "Setup" (Red) next to Gear
        title_container = ctk.CTkFrame(header, fg_color="transparent")
        title_container.pack(side="left", padx=20)
        
        lbl_gear = ctk.CTkLabel(
            title_container,
            text="⚙\ufe0e",
            font=("Segoe UI", 18, "bold"),
            text_color="#007B43"
        )
        lbl_gear.pack(side="left")
        
        lbl_component = ctk.CTkLabel(
            title_container,
            text=" Component",
            font=("Segoe UI", 18, "bold"),
            text_color="#007B43"
        )
        lbl_component.pack(side="left")
        
        lbl_setup = ctk.CTkLabel(
            title_container,
            text=" Setup",
            font=("Segoe UI", 18, "bold"),
            text_color="#B0050E"
        )
        lbl_setup.pack(side="left")

        # Company Logo on the right
        try:
            logo_path = resource_path("settings/cherry_full_logo.png")
            pil_img = Image.open(logo_path)
            pil_img = pil_img.resize((144, 40), Image.Resampling.LANCZOS)
            pil_img = add_corners(pil_img, 10)
            self.logo_img = ctk.CTkImage(pil_img, size=(144, 40))
            logo_lbl = ctk.CTkLabel(header, text="", image=self.logo_img)
            logo_lbl.pack(side="right", padx=(10, 20))
        except Exception as e:
            print(f"Error loading logo in ComponentSetupPage: {e}")

        # ====== FORM CONTAINER ======
        form_card = ModernCardFrame(self)
        form_card.pack(padx=20, pady=(15, 10), fill="x")

        # Row 0: Horizontal layout for dropdowns
        dropdown_row = ctk.CTkFrame(form_card, fg_color="transparent")
        dropdown_row.grid(row=0, column=0, columnspan=2, sticky="ew", padx=10, pady=10)
        
        # --- AirGauge ID Selection ---
        ctk.CTkLabel(dropdown_row, text="AirGauge ID:", font=("Segoe UI", 11, "bold"), text_color="#202124").pack(side="left", padx=(5, 5))
        
        # Fetch IDs from airgauge_master.db
        try:
            conn = sqlite3.connect("airgauge_master.db")
            rows = conn.execute("SELECT airgauge_id FROM airgauge_master").fetchall()
            ag_ids = sorted(list(set(str(r[0]) for r in rows if r[0])))
            conn.close()
        except:
            ag_ids = []
        
        if not ag_ids:
            ag_ids = [f"AG{i}" for i in range(1, 11)] # fallback
        
        self.id_dropdown = ctk.CTkComboBox(
            dropdown_row,
            values=ag_ids,
            width=160,
            command=lambda _: self._on_airgauge_changed(),
            border_color="#007B43",
            button_color="#007B43",
            button_hover_color="#005C32",
            dropdown_fg_color="white",
            dropdown_hover_color="#E8F5EE",
            dropdown_text_color="#202124"
        )
        self.id_dropdown.pack(side="left", padx=(0, 25))
        
        # --- Item Selection ---
        ctk.CTkLabel(dropdown_row, text="Item:", font=("Segoe UI", 11, "bold"), text_color="#202124").pack(side="left", padx=(5, 5))
        
        self.item_dropdown = ctk.CTkComboBox(
            dropdown_row,
            values=self._display_list,
            width=180,
            hover=False,
            command=lambda _: self._on_item_selected_by_user(),
            border_color="#007B43",
            button_color="#007B43",
            button_hover_color="#005C32",
            dropdown_fg_color="white",
            dropdown_hover_color="#E8F5EE",
            dropdown_text_color="#202124"
        )
        self.item_dropdown.pack(side="left", padx=(0, 25))
        self.item_dropdown.set("")
        
        # Enable typing-search on item entry
        try:
            entry_widget = getattr(self.item_dropdown, "_entry", None)
            if entry_widget is not None:
                entry_widget.bind("<KeyRelease>", self._on_item_entry_typed)
        except Exception:
            pass
            
        # --- Customer Selection ---
        ctk.CTkLabel(dropdown_row, text="Customer:", font=("Segoe UI", 11, "bold"), text_color="#202124").pack(side="left", padx=(5, 5))
        
        self.customer_dropdown = ctk.CTkComboBox(
            dropdown_row,
            values=self._customer_display_list,
            width=180,
            hover=False,
            command=lambda _: self._on_customer_selected_by_user(),
            border_color="#007B43",
            button_color="#007B43",
            button_hover_color="#005C32",
            dropdown_fg_color="white",
            dropdown_hover_color="#E8F5EE",
            dropdown_text_color="#202124"
        )
        self.customer_dropdown.pack(side="left", padx=(0, 5))
        self.customer_dropdown.set("")
        
        # Enable typing-search on customer entry
        try:
            entry_widget = getattr(self.customer_dropdown, "_entry", None)
            if entry_widget:
                entry_widget.bind("<KeyRelease>", self._on_customer_entry_typed)
        except Exception:
            pass

        # --- Type Selection (Shaft/Hole) ---
        ctk.CTkLabel(form_card, text="Type:", font=("Segoe UI", 11, "bold"), text_color="#202124").grid(
            row=1, column=0, padx=10, pady=10, sticky="e"
        )
        type_inner_frame = ctk.CTkFrame(form_card, fg_color="transparent")
        type_inner_frame.grid(row=1, column=1, padx=4, pady=10, sticky="w")
        
        ctk.CTkRadioButton(
            type_inner_frame, text="Shaft", variable=self.comp_type,
            value="Shaft", fg_color="#007B43", hover_color="#005C32", text_color="#202124"
        ).pack(side="left", padx=5)
        ctk.CTkRadioButton(
            type_inner_frame, text="Hole", variable=self.comp_type,
            value="Hole", fg_color="#007B43", hover_color="#005C32", text_color="#202124"
        ).pack(side="left", padx=5)

        # --- Channel Selection ---
        ctk.CTkLabel(form_card, text="Select Channel:", font=("Segoe UI", 11, "bold"), text_color="#202124").grid(
            row=2, column=0, padx=10, pady=10, sticky="e"
        )
        ch_frame = ctk.CTkFrame(form_card, fg_color="transparent")
        ch_frame.grid(row=2, column=1, padx=4, pady=10, sticky="w")
        self.selected_channel = ctk.StringVar(value="CH1")
        for ch in ["CH1", "CH2", "CH3", "CH4", "CH5", "CH6", "CH7", "CH8"]:
            ctk.CTkRadioButton(
                ch_frame, text=ch, variable=self.selected_channel,
                value=ch, command=self.load_channel_data,
                fg_color="#007B43", hover_color="#005C32", text_color="#202124"
            ).pack(side="left", padx=5)

        # --- Separator ---
        sep = ctk.CTkFrame(form_card, height=1, fg_color="#E0E0E0")
        sep.grid(row=3, column=0, columnspan=2, pady=5, sticky="ew")

        # --- Entry Fields (Drawing & Tolerances) ---
        self.entry_vars = {}
        self.entries = []
        entries = [
            ("Drawing Value (mm):", "drawing_value"),
            ("Low Tolerance (mm):", "low_tolerance"),
            ("High Tolerance (mm):", "high_tolerance")
        ]

        for i, (label, key) in enumerate(entries, start=4):
            ctk.CTkLabel(form_card, text=label, font=("Segoe UI", 11, "bold"), text_color="#202124").grid(
                row=i, column=0, padx=10, pady=8, sticky="e"
            )
            var = ctk.StringVar()
            
            def validate_numeric(char, new_value):
                if new_value == "":
                    return True
                for c in new_value:
                    if c not in "0123456789.":
                        return False
                if new_value.count(".") > 1:
                    return False
                return True

            validate_cmd = self.register(validate_numeric)

            entry = ctk.CTkEntry(
                form_card,
                textvariable=var,
                width=220,
                height=30,
                corner_radius=6,
                border_color="#007B43",
                validate="key",
                validatecommand=(validate_cmd, "%S", "%P")
            )

            entry.grid(row=i, column=1, padx=10, pady=8, sticky="w")
            # Focus/key bindings
            entry.bind("<FocusIn>", lambda e, ent=entry: self.highlight_entry(ent, True))
            entry.bind("<FocusOut>", lambda e, ent=entry: self.highlight_entry(ent, False))
            entry.bind("<Return>", lambda e, idx=i-4: self.focus_next_field(idx, save_on_last=True))
            entry.bind("<Down>", lambda e, idx=i-4: self.focus_next_field(idx, save_on_last=False))
            entry.bind("<Up>", lambda e, idx=i-4: self.focus_prev_field(idx))

            self.entries.append(entry)
            self.entry_vars[key] = var

        # ====== BUTTON ROW ======
        btn_frame = ctk.CTkFrame(self, fg_color="transparent")
        btn_frame.pack(pady=(10, 5))
        
        ModernButton(
            btn_frame,
            text="💾 Save Setup",
            fg_color="#007B43",
            hover_color="#005C32",
            height=32,
            width=160,
            corner_radius=6,
            font=("Segoe UI", 11, "bold"),
            command=self.save_data
        ).pack(side="left", padx=10)
        
        self.delete_btn = ModernButton(
            btn_frame,
            text="🗑 Delete Selected",
            fg_color="#B0050E",
            hover_color="#90040B",
            height=32,
            width=160,
            corner_radius=6,
            font=("Segoe UI", 11, "bold"),
            state="disabled",
            command=self.delete_selected
        )
        self.delete_btn.pack(side="left", padx=10)

        # --- Temporary Status Label ---
        self.status_label = ctk.CTkLabel(self, text="", font=("Segoe UI", 11), text_color="#007B43")
        self.status_label.pack(pady=(0, 5))

        # ====== TABLE ======
        table_frame = ModernCardFrame(self)
        table_frame.pack(fill="both", expand=True, padx=20, pady=(5, 15))

        title_lbl = ctk.CTkLabel(
            table_frame,
            text="🗎 Saved Component Configurations",
            font=("Segoe UI", 13, "bold"),
            text_color="#007B43"
        )
        title_lbl.pack(pady=5)

        # ---------- Treeview Style ----------
        style = ttk.Style()
        style.theme_use("default")
        style.configure(
            "Component.Treeview",
            font=("Segoe UI", 11),
            rowheight=28,
            background="white",
            foreground="#212121",
            fieldbackground="white",
            borderwidth=0
        )
        style.configure(
            "Component.Treeview.Heading",
            font=("Segoe UI", 11, "bold"),
            background="#205124",
            foreground="white",
            borderwidth=0,
            relief="flat"
        )
        style.map("Component.Treeview.Heading", background=[("active", "#005C32")])
        style.layout("Component.Treeview", [('Treeview.treearea', {'sticky': 'nswe'})])

        # ---------- Scrollbar MUST be created first ----------
        y_scroll = ctk.CTkScrollbar(
            table_frame,
            orientation="vertical",
            button_color="#007B43",
            button_hover_color="#005C32"
        )
        y_scroll.pack(side="right", fill="y", padx=(0, 10), pady=10)

        # --- Custom Table Header Row ---
        cols_with_icons = [
            ("AirGauge ID", "🏷\ufe0e"),
            ("Channel", "📊\ufe0e"),
            ("Item", "📋\ufe0e"),
            ("Type", "⚙\ufe0e"),
            ("Drawing", "📄\ufe0e"),
            ("Low Tol", "⬇\ufe0e"),
            ("High Tol", "⬆\ufe0e"),
            ("Customer", "👥\ufe0e")
        ]

        self.header_row_frame = ctk.CTkFrame(table_frame, fg_color="white", height=35, corner_radius=0)
        self.header_row_frame.pack(fill="x", padx=(10, 10), pady=(5, 0))
        self.header_row_frame.pack_propagate(False)

        # Scrollbar spacer on the right of header
        self.header_scrollbar_spacer = ctk.CTkFrame(
            self.header_row_frame,
            fg_color="white",
            corner_radius=0,
            width=16,
            height=35
        )
        self.header_scrollbar_spacer.pack(side="right", fill="y")

        self.header_inner_frame = ctk.CTkFrame(self.header_row_frame, fg_color="white", corner_radius=0)
        self.header_inner_frame.pack(side="left", fill="both", expand=True)

        self.header_widgets = []
        for name, icon in cols_with_icons:
            cell = ctk.CTkFrame(
                self.header_inner_frame,
                fg_color="white",
                corner_radius=0,
                border_width=1,
                border_color="#E0E0E0",
                height=35
            )
            cell.pack(side="left", fill="y")
            cell.pack_propagate(False)
            
            content_frame = ctk.CTkFrame(cell, fg_color="transparent")
            content_frame.pack(expand=True)
            
            ctk.CTkLabel(
                content_frame,
                text=icon,
                font=("Segoe UI", 11),
                text_color="#007B43"
            ).pack(side="left", padx=(0, 4))
            
            lbl = ctk.CTkLabel(
                content_frame,
                text=name,
                font=("Segoe UI", 11, "bold"),
                text_color="#1A1A1A",
                anchor="center"
            )
            lbl.pack(side="left")
            self.header_widgets.append(cell)

        # Underline
        self.border_line = ctk.CTkFrame(table_frame, fg_color="#007B43", height=2, corner_radius=0)
        self.border_line.pack(fill="x", padx=(10, 26), pady=(0, 0))

        # ---------- Treeview ----------
        columns = ("AirGauge ID", "Channel", "Item", "Type", "Drawing", "Low Tol", "High Tol", "Customer")
        self.tree = ttk.Treeview(
            table_frame,
            columns=columns,
            show="",
            height=8,
            style="Component.Treeview",
            yscrollcommand=y_scroll.set
        )

        y_scroll.configure(command=self.tree.yview)
        self.tree.pack(side="left", fill="both", expand=True, padx=(10, 0), pady=(0, 10))

        # ---------- Column Setup ----------
        for col in columns:
            if col == "AirGauge ID":
                self.tree.column(col, anchor="center", width=120)
            elif col == "Channel":
                self.tree.column(col, anchor="center", width=100)
            elif col == "Item":
                self.tree.column(col, anchor="center", width=280)
            elif col == "Type":
                self.tree.column(col, anchor="center", width=80)
            elif col == "Customer":
                self.tree.column(col, anchor="center", width=220)
            else:
                self.tree.column(col, anchor="center", width=130)

        # Dynamic widths synchronization
        def sync_widths(event=None):
            for idx, (name, icon) in enumerate(cols_with_icons):
                if idx < len(self.header_widgets):
                    try:
                        w = self.tree.column(name, "width")
                        self.header_widgets[idx].configure(width=w)
                    except Exception:
                        pass
        
        self.tree.bind("<Configure>", sync_widths)
        self.tree.bind("<ButtonRelease-1>", sync_widths)
        self.after(100, sync_widths)

        # ---------- Row Coloring ----------
        self.tree.tag_configure("evenrow", background="#FAFAFA")
        self.tree.tag_configure("oddrow", background="#F4F6F5")
        self.tree.tag_configure("highlight", background="#E8F5EE")

        self.refresh_table()


    # ------------------------------------------------------------
    # ------------------- SMALL UI HELPERS -----------------------
    # ------------------------------------------------------------
    def highlight_entry(self, entry, active):
        try:
            if active:
                entry.configure(border_width=2, border_color="#007B43")
                self.animate_glow(entry, 0)
            else:
                entry.configure(border_width=1, border_color="#007B43")
        except Exception:
            pass

    def animate_glow(self, entry, step):
        try:
            if self.focus_get() != entry:
                return
        except Exception:
            return
        wave_colors = ["#007B43", "#005C32", "#43A047", "#388E3C", "#005C32", "#007B43"]
        entry.configure(border_color=wave_colors[step % len(wave_colors)])
        self.after(150, lambda: self.animate_glow(entry, step + 1))

    def focus_next_field(self, current_index, save_on_last=True):
        if current_index + 1 < len(self.entries):
            self.entries[current_index + 1].focus_set()
        else:
            if save_on_last:
                self.save_data()
            else:
                self.entries[0].focus_set()

    def focus_prev_field(self, current_index):
        if current_index > 0:
            self.entries[current_index - 1].focus_set()

    # ------------------------------------------------------------
    # ---------------- JSON LOAD / SAVE for COMPONENTS -----------
    # ------------------------------------------------------------
    def load_json(self):
        try:
            conn = sqlite3.connect(resource_path("component_setup.db"))
            cursor = conn.cursor()
            cursor.execute("SELECT airgauge_id, channel, properties FROM component_setup")
            rows = cursor.fetchall()
            conn.close()
            comp_map = {}
            for ag_id, ch, props_str in rows:
                if ag_id not in comp_map:
                    comp_map[ag_id] = {}
                try:
                    comp_map[ag_id][ch] = json.loads(props_str)
                except:
                    comp_map[ag_id][ch] = {}
            return comp_map
        except Exception:
            return {}

    def save_json(self):
        try:
            conn = sqlite3.connect(resource_path("component_setup.db"))
            cursor = conn.cursor()
            cursor.execute("DELETE FROM component_setup")
            for ag_id, ch_data in self.comp_map.items():
                for ch, props in ch_data.items():
                    cursor.execute("INSERT INTO component_setup (airgauge_id, channel, properties) VALUES (?, ?, ?)",
                                   (ag_id, ch, json.dumps(props)))
            conn.commit()
            conn.close()
        except Exception as e:
            messagebox.showerror("Save Failed", f"Could not save component setup: {e}")

    # ------------------------------------------------------------
    # ---------------- ITEM MASTER LOADING HELPERS --------------
    # ------------------------------------------------------------
    def _load_items_master(self):
        """Read items from DB (Item Master). Returns list of dicts."""
        try:
            conn = sqlite3.connect(resource_path("items.db"))
            cursor = conn.cursor()
            cursor.execute("SELECT code, drawing_no, name, revision_no FROM items")
            rows = cursor.fetchall()
            conn.close()
            return [{"code": r[0], "drawing": r[1] or "", "name": r[2] or "", "revision": r[3] or ""} for r in rows]
        except Exception:
            return []

    def _build_display_list(self, items):
        """Return list of 'Name (Code)' strings for dropdown."""
        out = []
        for it in items:
            code = str(it.get("code", "")).strip()
            name = str(it.get("name", "")).strip()
            if name and code:
                out.append(f"{name} ({code})")
            elif name:
                out.append(name)
            elif code:
                out.append(code)
        return out

    def _split_display_to_code_name(self, display_text):
        """Given 'Name (Code)' return (code, name). Best-effort parsing."""
        if not display_text:
            return None, None
        # If user selected display that matches exact pattern
        if "(" in display_text and display_text.endswith(")"):
            try:
                name_part, code_part = display_text.rsplit("(", 1)
                code = code_part[:-1].strip()
                name = name_part.strip()
                return code, name
            except Exception:
                pass
        # Fallback: try to find code by name lookup
        # If exact name found among master items, return matching code
        for it in self.items_master:
            if str(it.get("name", "")).strip().lower() == display_text.strip().lower():
                return str(it.get("code", "")).strip(), str(it.get("name", "")).strip()
        return None, display_text.strip()
    def _on_customer_selected_by_user(self):
        try:
            sel = self.customer_dropdown.get()
            if sel and sel not in self._customer_display_list:
                self.customer_dropdown.configure(
                    values=self._customer_display_list + [sel]
                )
        except Exception:
            pass

    def _on_customer_entry_typed(self, event=None):
        try:
            entry = getattr(self.customer_dropdown, "_entry", None)
            if not entry:
                return
            q = entry.get().strip().lower()
            if not q:
                self.customer_dropdown.configure(values=self._customer_display_list)
                return

            filtered = []
            for c in self.customers_master:
                code = str(c.get("code", "")).strip()
                name = str(c.get("name", "")).strip()
                if q in code.lower() or q in name.lower():
                    filtered.append(f"{name} ({code})")
            self.customer_dropdown.configure(values=filtered or self._customer_display_list)
        except Exception:
            pass
    # ------------------------------------------------------------
    # ---------------- CUSTOMER MASTER HELPERS -------------------
    # ------------------------------------------------------------
    def _load_customers_master(self):
        try:
            conn = sqlite3.connect(resource_path("customers.db"))
            cursor = conn.cursor()
            cursor.execute("SELECT code, name, description, email, phone FROM customers")
            rows = cursor.fetchall()
            conn.close()
            return [{"code": r[0], "name": r[1] or "", "description": r[2] or "", "email": r[3] or "", "phone": r[4] or ""} for r in rows]
        except Exception:
            return []

    def _build_customer_display_list(self, customers):
        out = []
        for c in customers:
            code = str(c.get("code", "")).strip()
            name = str(c.get("name", "")).strip()
            if name and code:
                out.append(f"{name} ({code})")
            elif name:
                out.append(name)
            elif code:
                out.append(code)
        return out

    def _split_customer_display(self, display_text):
        if not display_text:
            return None, None
        if "(" in display_text and display_text.endswith(")"):
            try:
                name, code = display_text.rsplit("(", 1)
                return code[:-1].strip(), name.strip()
            except Exception:
                pass
        for c in self.customers_master:
            if str(c.get("name", "")).strip().lower() == display_text.lower():
                return str(c.get("code", "")).strip(), str(c.get("name", "")).strip()
        return None, display_text.strip()

    # ------------------------------------------------------------
    # ------------------ CORE LOGIC: load/save -------------------
    # ------------------------------------------------------------
    def _on_airgauge_changed(self):
        # When AirGauge ID selection changes: reload master items (in case updated) and load channel data
        self.items_master = self._load_items_master()
        self._code_to_name = {str(it.get("code", "")).strip(): str(it.get("name", "")).strip() for it in self.items_master}
        self._display_list = self._build_display_list(self.items_master)
        try:
            self.item_dropdown.configure(values=self._display_list)
        except Exception:
            pass
        self.load_channel_data()

    def _on_item_selected_by_user(self):
        """User explicitly selected an item from dropdown — nothing automatic here, just keep value for save."""
        # No-op; selection will be read by save_data() or when switching channel.
        # But keep dropdown consistent (refresh list on selection)
        try:
            sel = self.item_dropdown.get()
            if sel and sel not in self._display_list:
                # refresh values to include typed selection if it matches pattern
                vals = list(self._display_list) + [sel]
                self.item_dropdown.configure(values=vals)
        except Exception:
            pass

    def _on_item_entry_typed(self, event=None):
        """Filter the item_dropdown list on typed substring (search both code & name). Best-effort using internal entry."""
        try:
            entry_widget = getattr(self.item_dropdown, "_entry", None)
            if entry_widget is None:
                return
            q = entry_widget.get().strip().lower()
            if not q:
                self.item_dropdown.configure(values=self._display_list)
                return
            filtered = []
            for it in self.items_master:
                code = str(it.get("code", "")).strip()
                name = str(it.get("name", "")).strip()
                if q in code.lower() or q in name.lower():
                    filtered.append(f"{name} ({code})" if name and code else (name or code))
            # avoid empty list (combo expects at least [])
            self.item_dropdown.configure(values=filtered or self._display_list)
        except Exception:
            pass

    def load_channel_data(self):
        """Keep entry fields empty even when AirGauge ID or Channel changes.
        There is no need for auto-entering values from stored specs.
        """
        for var in self.entry_vars.values():
            var.set("")
        try:
            self.item_dropdown.set("")
        except Exception:
            try:
                entry_widget = getattr(self.item_dropdown, "_entry", None)
                if entry_widget:
                    entry_widget.delete(0, "end")
            except Exception:
                pass
        try:
            self.customer_dropdown.configure(values=self._customer_display_list)
            self.customer_dropdown.set("")
        except Exception:
            try:
                entry_widget = getattr(self.customer_dropdown, "_entry", None)
                if entry_widget:
                    entry_widget.delete(0, "end")
            except Exception:
                pass
        self.comp_type.set("Shaft")


    def save_data(self):
        """Save all component data to JSON and update table."""
        ag_id = self.id_dropdown.get().strip()
        ch = self.selected_channel.get().strip()
        draw = self.entry_vars["drawing_value"].get().strip()
        low = self.entry_vars["low_tolerance"].get().strip()
        high = self.entry_vars["high_tolerance"].get().strip()

        if not ag_id:
            messagebox.showerror("Error", "Please select an AirGauge ID.")
            return
        if not draw or not low or not high:
            messagebox.showerror("Error", "All fields (Drawing, Low, High) are required.")
            return
        # strict number validation (must be convertible to float, not empty)
        def is_number(value):
            try:
                float(value)
                return True
            except:
                return False

        if not (is_number(draw) and is_number(low) and is_number(high)):
            messagebox.showerror(
                "Invalid Input",
                "Drawing value, Low tolerance, and High tolerance must be valid numbers."
            )
            return


        # read currently selected item (parse to code & name)
        selected_display = ""
        try:
            selected_display = self.item_dropdown.get().strip()
        except Exception:
            # fallback to internal entry
            try:
                entry_widget = getattr(self.item_dropdown, "_entry", None)
                selected_display = entry_widget.get().strip() if entry_widget else ""
            except Exception:
                selected_display = ""

        item_code, item_name = self._split_display_to_code_name(selected_display)
        # if code missing but name exists, try reverse lookup in master
        if not item_code and item_name:
            # find code by name (exact match)
            for it in self.items_master:
                if str(it.get("name", "")).strip().lower() == item_name.strip().lower():
                    item_code = str(it.get("code", "")).strip()
                    break

        # ---- Customer parsing ----
        customer_display = ""
        try:
            customer_display = self.customer_dropdown.get().strip()
        except Exception:
            pass

        cust_code, cust_name = self._split_customer_display(customer_display)


        # Prepare map for ag_id
        if ag_id not in self.comp_map:
            self.comp_map[ag_id] = {}

        # determine update or new
        is_update = False
        if ch in self.comp_map[ag_id]:
            is_update = True

        # store values (store both code & name as requested)
        self.comp_map[ag_id][ch] = {
            "drawing_value": draw,
            "low_tolerance": low,
            "high_tolerance": high,
            "item_code": item_code or "",
            "item_name": item_name or "",
            "customer_code": cust_code or "",
            "customer_name": cust_name or "",
            "type": self.comp_type.get()
        }


        # persist
        self.save_json()
        # refresh table and indicate which row to highlight
        self.refresh_table(highlight=(ag_id, ch))
        # refresh other pages that depend on IDs (unchanged)
        try:
            self.app.refresh_id_dependent_pages()
        except Exception:
            pass

        # show friendly message
        if is_update:
            self.show_temp_status("✏️ Updated existing config", "#FFA000")
        else:
            self.show_temp_status("✅ Saved new config", "#43A047")

        # reset inputs (keeps item selection behavior consistent)
        self.reset_form()

    def delete_selected(self):
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showwarning("No Selection", "Please select a row to delete.")
            return

        item = self.tree.item(selected_item)
        values = item["values"]
        if len(values) < 2:
            return

        air_id, channel = str(values[0]).strip(), str(values[1]).strip().upper()

        if not messagebox.askyesno(
            "Confirm Delete",
            f"Are you sure you want to delete config for\nAirGauge {air_id} ({channel})?"
        ):
            return

        try:
            if air_id in self.comp_map and channel in self.comp_map[air_id]:
                del self.comp_map[air_id][channel]
                if not self.comp_map[air_id]:
                    del self.comp_map[air_id]
                self.save_json()
                self.refresh_table()
                self.show_temp_status("🗑 Deleted Successfully", "#E53935")
                self.reset_form()
                # reload in-memory map fresh
                self.comp_map = self.load_json()
                return
            else:
                messagebox.showwarning("Not Found", f"No matching entry found for {air_id} ({channel}).")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to delete: {e}")

    # ------------------------------------------------------------
    # ------------------- Table interactions ---------------------
    # ------------------------------------------------------------
    def refresh_airgauge_ids(self):
        """Just refresh the AirGauge dropdown values (useful after ItemMaster updates)."""
        try:
            conn = sqlite3.connect("airgauge_master.db")
            ids = [str(r[0]) for r in conn.execute("SELECT airgauge_id FROM airgauge_master").fetchall()]
            conn.close()
            if not ids:
                ids = [f"AG{i}" for i in range(1, 11)]
            self.id_dropdown.configure(values=ids)
        except Exception:
            pass

    def refresh_table(self, highlight=None):
        """Show all saved component configurations including Item Name (only)."""
        for i in self.tree.get_children():
            self.tree.delete(i)

        row_num = 0
        for ag_id, ch_data in (self.comp_map.items()):
            for ch, vals in ch_data.items():
                tag = "evenrow" if row_num % 2 == 0 else "oddrow"
                item_name = vals.get("item_name", "")  # display only name
                customer_name = vals.get("customer_name", "")
                item_id = self.tree.insert("", "end", values=(
                    ag_id,
                    ch,
                    item_name,
                    vals.get("type", "Shaft"),
                    vals.get("drawing_value", ""),
                    vals.get("low_tolerance", ""),
                    vals.get("high_tolerance", ""),
                    customer_name
                ), tags=(tag,))
                if highlight and (ag_id, ch) == highlight:
                    self.tree.item(item_id, tags=("highlight",))
                    # schedule remove highlight
                    self.after(1500, lambda iid=item_id, tag=tag: self.tree.item(iid, tags=(tag,)))
                row_num += 1

    def on_row_select(self, event=None):
        if self.tree.selection():
            self.delete_btn.configure(state="normal")
        else:
            self.delete_btn.configure(state="disabled")

    def on_row_click(self, event=None):
        """When a row is clicked, load data into form (including item)."""
        selected_item = self.tree.selection()
        if not selected_item:
            return

        item = self.tree.item(selected_item)
        values = item["values"]
        if len(values) < 6:
            return

        air_id, ch, item_name, draw, low, high, customer_name = values

        # Update form fields
        try:
            self.id_dropdown.set(str(air_id))
        except Exception:
            pass
        try:
            self.selected_channel.set(str(ch).upper())
        except Exception:
            pass

        # find code if possible (reverse lookup by name)
        item_code = ""
        if item_name:
            for it in self.items_master:
                if str(it.get("name", "")).strip().lower() == str(item_name).strip().lower():
                    item_code = str(it.get("code", "")).strip()
                    break

        # set item dropdown to 'Name (Code)' if code present else name
        display_val = f"{item_name} ({item_code})" if item_code else item_name
        try:
            self.item_dropdown.configure(values=self._display_list)
            self.item_dropdown.set(display_val)
        except Exception:
            try:
                entry_widget = getattr(self.item_dropdown, "_entry", None)
                if entry_widget:
                    entry_widget.delete(0, "end")
                    entry_widget.insert(0, display_val)
            except Exception:
                pass
        # ---- Load customer into dropdown ----
        customer_code = ""
        if customer_name:
            for c in self.customers_master:
                if str(c.get("name", "")).strip().lower() == str(customer_name).strip().lower():
                    customer_code = str(c.get("code", "")).strip()
                    break

        customer_display = f"{customer_name} ({customer_code})" if customer_code else customer_name
        try:
            self.customer_dropdown.configure(values=self._customer_display_list)
            self.customer_dropdown.set(customer_display)
        except Exception:
            try:
                entry_widget = getattr(self.customer_dropdown, "_entry", None)
                if entry_widget:
                    entry_widget.delete(0, "end")
                    entry_widget.insert(0, customer_display)
            except Exception:
                pass


        self.entry_vars["drawing_value"].set(str(draw))
        self.entry_vars["low_tolerance"].set(str(low))
        self.entry_vars["high_tolerance"].set(str(high))

        self.delete_btn.configure(state="normal")

    def show_temp_status(self, text, color):
        self.status_label.configure(text=text, text_color=color)
        self.fade_status_opacity(1.0)

    def fade_status_opacity(self, opacity):
        if opacity <= 0:
            self.status_label.configure(text="")
            return
        # keep color as provided (no alpha blending complexity)
        self.after(100, lambda: self.fade_status_opacity(opacity - 0.1))

    def reset_form(self):
        for var in self.entry_vars.values():
            var.set("")
        self.selected_channel.set("CH1")
        # reload ids and master lists
        try:
            self.refresh_airgauge_ids()
        except Exception:
            pass
        # reload items master list (in case it changed)
        self.items_master = self._load_items_master()
        self._display_list = self._build_display_list(self.items_master)
        try:
            self.item_dropdown.configure(values=self._display_list)
            self.item_dropdown.set("")  # clear selection
        except Exception:
            pass

        try:
            self.customers_master = self._load_customers_master()
            self._customer_display_list = self._build_customer_display_list(self.customers_master)
        except Exception:
            pass
        
    def go_back(self, event=None):
        try:
            self.app.load_settings_page()
        except:
            try: self.app.load_settings()
            except: pass

    def destroy(self):
        super().destroy()

    def _update_dropdown_values(self):
        try:
            self.customer_dropdown.configure(values=self._customer_display_list)
            self.customer_dropdown.set("")
        except Exception:
            pass

#=============================================================
#------------------Machine Master-----------------------------
#=============================================================
class MachineMasterPage(ctk.CTkFrame):
    """
    Machine Master Page
    Fields:
      - Machine Code
      - Machine Name
      - Description
    Storage: machines.json
    """
    FNAME = "machines.json"

    def __init__(self, parent, app):
        super().__init__(parent, fg_color=("white", "#1c1c1c"))
        self.app = app

        self.machines = []
        self.search_var = tk.StringVar()
        self.use_tksheet = False
        self.sheet = None
        self.tree = None
        self.resize_sheet = lambda event=None: None

        # Fixed column widths (similar style)
        self._col_widths = [140, 420, 550]

        self.load_machine_data()
        self.build_ui()
        self.refresh_table()

        try:
            self.after(60, self.resize_sheet)
        except:
            pass

    # ----------------------------------------------------
    # JSON HANDLING
    # ----------------------------------------------------
    def _get_data_path(self):
        return os.path.join(os.getcwd(), self.FNAME)

    def load_machine_data(self):
        try:
            conn = sqlite3.connect(resource_path("machine_master.db"))
            cursor = conn.cursor()
            cursor.execute("SELECT code, name, description FROM machine_master")
            rows = cursor.fetchall()
            conn.close()
            self.machines = [{"code": r[0], "name": r[1], "desc": r[2] or ""} for r in rows]
        except Exception as e:
            messagebox.showerror("Load Failed", f"Could not load data from DB: {e}")
            self.machines = []

    # ----------------------------------------------------
    # BUILD UI
    # ----------------------------------------------------
    def build_ui(self):
        # HEADER
        header = ctk.CTkFrame(self, fg_color="transparent")
        header.pack(fill="x", side="top", pady=(15, 10), padx=20)
        
        # Back Button in Green
        ModernButton(
            header,
            text="← Back",
            width=80,
            height=32,
            font=("Segoe UI", 11, "bold"),
            fg_color="#007B43",
            hover_color="#005C32",
            text_color="white",
            corner_radius=6,
            command=self.go_back
        ).pack(side="left")

        # Branded Title: "Machine" (Green) + "Master" (Red)
        title_container = ctk.CTkFrame(header, fg_color="transparent")
        title_container.pack(side="left", padx=20)
        
        lbl_icon = ctk.CTkLabel(
            title_container,
            text="🏭\ufe0e ",
            font=("Segoe UI", 18, "bold"),
            text_color="#007B43"
        )
        lbl_icon.pack(side="left")
        
        lbl_machine = ctk.CTkLabel(
            title_container,
            text="Machine",
            font=("Segoe UI", 18, "bold"),
            text_color="#007B43"
        )
        lbl_machine.pack(side="left")
        
        lbl_master = ctk.CTkLabel(
            title_container,
            text=" Master",
            font=("Segoe UI", 18, "bold"),
            text_color="#B0050E"
        )
        lbl_master.pack(side="left")

        # Company Logo on the right
        try:
            logo_path = resource_path("settings/cherry_full_logo.png")
            pil_img = Image.open(logo_path)
            pil_img = pil_img.resize((144, 40), Image.Resampling.LANCZOS)
            pil_img = add_corners(pil_img, 10)
            self.logo_img = ctk.CTkImage(pil_img, size=(144, 40))
            logo_lbl = ctk.CTkLabel(header, text="", image=self.logo_img)
            logo_lbl.pack(side="right", padx=(10, 20))
        except Exception as e:
            print(f"Error loading logo in MachineMasterPage: {e}")

        # FORM
        add_frame = ModernCardFrame(self)
        add_frame.pack(fill="x", padx=20, pady=(10, 6))

        LABEL_WIDTH = 100
        ENTRY_WIDTH = 260

        # Row 0: Code + Name
        ctk.CTkLabel(add_frame, text="Code:", width=LABEL_WIDTH, anchor="w", text_color="#202124", font=("Segoe UI", 11, "bold")).grid(
            row=0, column=0, sticky="w", pady=6, padx=(10, 0)
        )
        self.code_e = ctk.CTkEntry(add_frame, width=ENTRY_WIDTH, height=30, corner_radius=6, border_color="#007B43")
        self.code_e.grid(row=0, column=1, sticky="w", padx=(10, 30))

        ctk.CTkLabel(add_frame, text="Name:", width=LABEL_WIDTH, anchor="e", text_color="#202124", font=("Segoe UI", 11, "bold")).grid(
            row=0, column=2, sticky="e", padx=(0, 10), pady=6
        )
        self.name_e = ctk.CTkEntry(add_frame, width=ENTRY_WIDTH, height=30, corner_radius=6, border_color="#007B43")
        self.name_e.grid(row=0, column=3, sticky="w", padx=(10, 10))

        # Row 1: Description
        ctk.CTkLabel(
            add_frame, text="Description:", width=LABEL_WIDTH, anchor="w", text_color="#202124", font=("Segoe UI", 11, "bold")
        ).grid(row=1, column=0, sticky="w", pady=6, padx=(10, 0))
        self.desc_e = ctk.CTkEntry(add_frame, width=ENTRY_WIDTH * 2 + 130, height=30, corner_radius=6, border_color="#007B43")
        self.desc_e.grid(row=1, column=1, columnspan=3, sticky="w", padx=(10, 10))

        # Buttons
        btn_frame = ctk.CTkFrame(add_frame, fg_color="transparent")
        btn_frame.grid(row=2, column=0, columnspan=4, sticky="w", pady=(6, 10), padx=(10, 0))

        self.action_btn = ModernButton(
            btn_frame,
            text="➕ Add Machine",
            width=150,
            height=32,
            font=("Segoe UI", 11, "bold"),
            fg_color="#007B43",
            hover_color="#005C32",
            text_color="white",
            corner_radius=6,
            command=self._on_action_clicked
        )
        self.action_btn.pack(side="left", padx=(0, 10))

        del_btn = ModernButton(
            btn_frame,
            text="🗑 Delete Selected",
            width=150,
            height=32,
            font=("Segoe UI", 11, "bold"),
            fg_color="#B0050E",
            hover_color="#90040B",
            text_color="white",
            corner_radius=6,
            command=self._delete_selected
        )
        del_btn.pack(side="left")

        # Bind focus highlights
        for entry in (self.code_e, self.name_e, self.desc_e):
            entry.bind("<FocusIn>", lambda e, ent=entry: self.highlight_entry(ent, True))
            entry.bind("<FocusOut>", lambda e, ent=entry: self.highlight_entry(ent, False))

        # ====== TABLE ======
        table_card = ModernCardFrame(self)
        table_card.pack(fill="both", expand=True, padx=20, pady=(6, 12))
        table_card.grid_rowconfigure(0, weight=1)
        table_card.grid_columnconfigure(0, weight=1)

        # Try import tksheet
        try:
            from tksheet import Sheet
            SheetClass = Sheet
        except Exception:
            try:
                SheetClass = tksheet.Sheet
            except Exception:
                SheetClass = None

        cols = ["Code", "Name", "Description"]
        header_info = [
            ("Code", "🔑\ufe0e"),
            ("Name", "🏭\ufe0e"),
            ("Description", "📄\ufe0e")
        ]

        if SheetClass is not None:
            self.use_tksheet = True
            
            # --- Custom Horizontal Scroll-Synced Header Frame ---
            self.header_row_frame = ctk.CTkFrame(table_card, fg_color="white", height=42, corner_radius=0)
            self.header_row_frame.pack(fill="x", padx=10, pady=(10, 0))
            self.header_row_frame.pack_propagate(False)
            
            # Canvas inside header frame for scrolling
            self.header_canvas = tk.Canvas(self.header_row_frame, bg="white", highlightthickness=0, height=42)
            self.header_canvas.pack(side="left", fill="both", expand=True)
            
            # Frame inside canvas for custom header cells
            self.header_inner_frame = tk.Frame(self.header_canvas, bg="white")
            self.header_canvas.create_window(0, 0, window=self.header_inner_frame, anchor="nw")
            
            # Vertical scrollbar spacer on the right of header
            self.header_scrollbar_spacer = ctk.CTkFrame(
                self.header_row_frame,
                fg_color="white",
                corner_radius=0,
                width=16,
                height=42
            )
            self.header_scrollbar_spacer.pack(side="right", fill="y")
            
            # Row index spacer cell on the far left displaying green up-triangle sorting indicator
            row_index_cell = ctk.CTkFrame(
                self.header_inner_frame,
                fg_color="white",
                corner_radius=0,
                border_width=1,
                border_color="#E0E0E0",
                width=40,
                height=42
            )
            row_index_cell.pack(side="left", fill="y")
            row_index_cell.pack_propagate(False)
            
            lbl_tri = ctk.CTkLabel(
                row_index_cell,
                text="▲",
                font=("Segoe UI", 10, "bold"),
                text_color="#007B43",
                anchor="center"
            )
            lbl_tri.pack(expand=True)

            self.header_widgets = []
            for name, icon in header_info:
                cell = ctk.CTkFrame(
                    self.header_inner_frame,
                    fg_color="white",
                    corner_radius=0,
                    border_width=1,
                    border_color="#E0E0E0",
                    height=42
                )
                cell.pack(side="left", fill="y")
                cell.pack_propagate(False)
                
                content_frame = ctk.CTkFrame(cell, fg_color="transparent")
                content_frame.pack(expand=True)
                
                ctk.CTkLabel(
                    content_frame,
                    text=icon,
                    font=("Segoe UI", 12, "normal"),
                    text_color="#007B43",
                    anchor="center"
                ).pack(side="left", padx=(0, 4))
                
                lbl = ctk.CTkLabel(
                    content_frame,
                    text=name,
                    font=("Segoe UI", 11, "bold"),
                    text_color="#1A1A1A", # BLACK text color
                    anchor="center"
                )
                lbl.pack(side="left")
                self.header_widgets.append(cell)

            # Green border line under headers
            self.border_line = ctk.CTkFrame(table_card, fg_color="#007B43", height=2, corner_radius=0)
            self.border_line.pack(fill="x", padx=10, pady=(0, 0))

            # Table sheet frame
            self.table_sheet_frame = ctk.CTkFrame(table_card, fg_color="white", corner_radius=0)
            self.table_sheet_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))
            self.table_sheet_frame.grid_rowconfigure(0, weight=1)
            self.table_sheet_frame.grid_columnconfigure(0, weight=1)

            data = [
                [
                    m.get("code", ""),
                    m.get("name", ""),
                    m.get("desc", ""),
                ]
                for m in self.machines
            ]

            self.sheet = SheetClass(
                self.table_sheet_frame,
                headers=cols,
                data=data,
                show_header=False,        # Hide native header
                show_row_index=True,       # Show row numbers index
                row_index_width=40,        # Width matches spacer (40px)
                show_x_scrollbar=True,
                show_y_scrollbar=True
            )
            self.sheet.grid(row=0, column=0, sticky="nsew")

            # Style tksheet cells to match
            try:
                self.sheet.set_options(
                    grid_fg="#E0E0E0",
                    table_bg="white",
                    table_fg="#202124",
                    frame_bg="white",
                    select_bg="#E8F5EE",
                    select_fg="#007B43",
                    font=("Segoe UI", 10, "normal")
                )
            except Exception:
                pass

            try:
                self.sheet.enable_bindings(("single_select", "row_select", "arrowkeys", "copy", "select_all","right_click_popup_menu"))
            except Exception:
                pass

            # Double click → edit popup
            def _dbl(event):
                idx = self._get_selected_index()
                if idx is not None:
                    self._open_editor_dialog("edit", idx)

            try:
                self.sheet.bind("<Double-1>", _dbl)
            except Exception:
                pass

            # Sync horizontal scrolling to custom header
            try:
                orig_xscroll = self.sheet.MT.cget("xscrollcommand")
                def sync_scroll(first, last):
                    if orig_xscroll:
                        try:
                            self.sheet.tk.eval(f"{orig_xscroll} {first} {last}")
                        except Exception:
                            pass
                    try:
                        self.header_canvas.xview_moveto(first)
                    except Exception:
                        pass
                self.sheet.MT.configure(xscrollcommand=sync_scroll)
            except Exception as e:
                print("Failed to sync Machine Master header scrollbar:", e)

            # Resize columns resizes both sheet and header cells
            def resize_sheet(event=None):
                try:
                    total_width = self.table_sheet_frame.winfo_width() or 1
                    if total_width < 120:
                        self.after(60, self.resize_sheet)
                        return
                    col_count = len(cols)
                    # Deduct spacing for the vertical scrollbar (16px) and row index column (40px)
                    available = max(500, total_width - 16 - 40)
                    col_width = max(80, int(available / col_count))
                    for c in range(col_count):
                        try:
                            self.sheet.column_width(column=c, width=col_width, only_set_if_too_small=False)
                        except TypeError:
                            self.sheet.column_width(c, col_width)
                        
                        # Apply width to header cells
                        if c < len(self.header_widgets):
                            self.header_widgets[c].configure(width=col_width)
                    try:
                        self.sheet.refresh()
                    except Exception:
                        pass
                    try:
                        self.header_inner_frame.update_idletasks()
                        self.header_canvas.configure(scrollregion=self.header_canvas.bbox("all"))
                    except Exception:
                        pass
                except Exception as e:
                    print("Error resizing Machine Master sheet:", e)

            self.resize_sheet = resize_sheet
            self.table_sheet_frame.bind("<Configure>", self.resize_sheet)
            try:
                self.after(80, self.resize_sheet)
            except Exception:
                pass
            
        else:
            self.use_tksheet = False
            self.sheet = None

            self.tree = ttk.Treeview(
                table_card,
                columns=("code", "name", "desc"),
                show="headings",
                selectmode="browse",
            )
            self.tree.heading("code", text="Code")
            self.tree.heading("name", text="Name")
            self.tree.heading("desc", text="Description")

            self.tree.column("code", width=260)
            self.tree.column("name", width=320)
            self.tree.column("desc", width=520)

            vsb = ttk.Scrollbar(table_card, orient="vertical", command=self.tree.yview)
            hsb = ttk.Scrollbar(table_card, orient="horizontal", command=self.tree.xview)

            self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

            self.tree.grid(row=0, column=0, sticky="nsew")
            vsb.grid(row=0, column=1, sticky="ns")
            hsb.grid(row=1, column=0, sticky="ew")

            self.tree.bind(
                "<Double-1>",
                lambda e: self._open_editor_dialog("edit", self._get_selected_index()),
            )

        self.after(80, self.refresh_table)


    # ----------------------------------------------------
    # UTILITIES
    # ----------------------------------------------------
    def _get_selected_index(self):
        try:
            if self.use_tksheet:
                cur = self.sheet.get_currently_selected()
                if cur and isinstance(cur, tuple):
                    return cur[0]

                rows = self.sheet.get_selected_rows()
                if rows:
                    return rows[0]

                return None

            else:
                sel = self.tree.selection()
                if not sel:
                    return None

                vals = self.tree.item(sel[0], "values")
                code = str(vals[0])

                for i, rec in enumerate(self.machines):
                    if str(rec.get("code", "")) == code:
                        return i
        except:
            return None

    def _clear_inputs(self):
        try:
            self.code_e.delete(0, "end")
            self.name_e.delete(0, "end")
            self.desc_e.delete(0, "end")
        except:
            pass

    # ----------------------------------------------------
    # ADD / UPDATE / DELETE
    # ----------------------------------------------------
    def _on_action_clicked(self):
        code = self.code_e.get().strip()
        name = self.name_e.get().strip()
        desc = self.desc_e.get().strip()

        if not code:
            messagebox.showwarning("Validation", "Code is required.")
            return

        if not name:
            messagebox.showwarning("Validation", "Name is required.")
            return

        if getattr(self, "edit_index", None) is None:
            # ADD MODE
            if any(str(m.get("code", "")) == code for m in self.machines):
                messagebox.showwarning("Validation", "Code already exists.")
                return

            try:
                conn = sqlite3.connect(resource_path("machine_master.db"))
                cursor = conn.cursor()
                cursor.execute("INSERT INTO machine_master (code, name, description) VALUES (?, ?, ?)",
                               (code, name, desc))
                conn.commit()
                conn.close()
            except Exception as e:
                messagebox.showerror("DB Error", f"Failed to add machine: {e}")
                return

            self.load_machine_data()
            self.refresh_table()

            new_idx = next((i for i, m in enumerate(self.machines) if str(m["code"]) == code), len(self.machines) - 1)

            try:
                self.after(120, lambda: self._blink_new_row(new_idx))
                self.after(180, lambda: self._fade_row(new_idx))
            except:
                pass

            self._clear_inputs()
            self.code_e.focus_set()

            try:
                self.app.status_label.configure(text=f"Machine Added: {code} ✅")
            except:
                pass

        else:
            # EDIT MODE
            idx = self.edit_index
            old_m = self.machines[idx]
            old_code = old_m.get("code")

            try:
                conn = sqlite3.connect(resource_path("machine_master.db"))
                cursor = conn.cursor()
                cursor.execute("UPDATE machine_master SET code=?, name=?, description=? WHERE code=?",
                               (code, name, desc, old_code))
                conn.commit()
                conn.close()
            except Exception as e:
                messagebox.showerror("DB Error", f"Failed to update machine: {e}")
                return

            self.load_machine_data()
            self.refresh_table()

            def _restore():
                try:
                    if self.use_tksheet:
                        self.sheet.select_row(idx)
                except:
                    pass

            self.after(80, _restore)

            self._clear_inputs()
            self.edit_index = None
            self.action_btn.configure(text="➕ Add Machine")

            try:
                self.app.status_label.configure(text=f"Machine Updated: {code} ✅")
            except:
                pass

    def _delete_selected(self):
        idx = self._get_selected_index()
        if idx is None:
            messagebox.showwarning("No Selection", "Select a row to delete.")
            return

        rec = self.machines[idx]

        if not messagebox.askyesno(
            "Confirm Delete", f"Delete machine '{rec.get('name')}' (Code: {rec.get('code')})?"
        ):
            return

        try:
            conn = sqlite3.connect(resource_path("machine_master.db"))
            cursor = conn.cursor()
            cursor.execute("DELETE FROM machine_master WHERE code=?", (rec.get('code'),))
            conn.commit()
            conn.close()
        except Exception as e:
            messagebox.showerror("DB Error", f"Failed to delete machine: {e}")
            return

        self.load_machine_data()
        self.refresh_table()

        try:
            self.app.status_label.configure(text=f"Deleted {rec.get('code')} ✅")
        except:
            pass

    # ----------------------------------------------------
    # POPUP EDITOR
    # ----------------------------------------------------
    def _open_editor_dialog(self, mode="edit", index=None):
        if mode not in ("edit", "add"):
            return

        win = ctk.CTkToplevel(self)
        win.title("Edit Machine" if mode == "edit" else "Add Machine")
        win.geometry("580x320")
        win.grab_set()

        frm = ctk.CTkFrame(win)
        frm.pack(fill="both", expand=True, padx=12, pady=12)

        ctk.CTkLabel(frm, text="Code:", font=("Segoe UI", 11)).grid(
            row=0, column=0, sticky="w", padx=6, pady=6
        )
        code_e = ctk.CTkEntry(frm, width=300)
        code_e.grid(row=0, column=1, sticky="w", padx=6, pady=6)

        ctk.CTkLabel(frm, text="Name:", font=("Segoe UI", 11)).grid(
            row=1, column=0, sticky="w", padx=6, pady=6
        )
        name_e = ctk.CTkEntry(frm, width=420)
        name_e.grid(row=1, column=1, sticky="w", padx=6, pady=6)

        ctk.CTkLabel(frm, text="Description:", font=("Segoe UI", 11)).grid(
            row=2, column=0, sticky="w", padx=6, pady=6
        )
        desc_e = ctk.CTkEntry(frm, width=420)
        desc_e.grid(row=2, column=1, sticky="w", padx=6, pady=6)

        if mode == "edit" and index is not None:
            rec = self.machines[index]
            code_e.insert(0, rec.get("code", ""))
            name_e.insert(0, rec.get("name", ""))
            desc_e.insert(0, rec.get("desc", ""))

        btnf = ctk.CTkFrame(frm, fg_color="transparent")
        btnf.grid(row=3, column=0, columnspan=2, pady=(12, 0))

        def save_btn():
            code = code_e.get().strip()
            name = name_e.get().strip()
            desc = desc_e.get().strip()

            if not code or not name:
                messagebox.showwarning("Validation", "Code and Name required.")
                return

            if mode == "edit":
                if not messagebox.askyesno(
                    "Confirm Update", f"Apply changes to {code}?"
                ):
                    return

                old_code = self.machines[index].get("code")
                try:
                    conn = sqlite3.connect(resource_path("machine_master.db"))
                    cursor = conn.cursor()
                    cursor.execute("UPDATE machine_master SET code=?, name=?, description=? WHERE code=?",
                                   (code, name, desc, old_code))
                    conn.commit()
                    conn.close()
                except Exception as e:
                    messagebox.showerror("DB Error", f"Failed to update machine: {e}")
                    return

            else:
                if any(str(m.get("code", "")) == code for m in self.machines):
                    messagebox.showwarning("Validation", "Code already exists.")
                    return
                try:
                    conn = sqlite3.connect(resource_path("machine_master.db"))
                    cursor = conn.cursor()
                    cursor.execute("INSERT INTO machine_master (code, name, description) VALUES (?, ?, ?)",
                                   (code, name, desc))
                    conn.commit()
                    conn.close()
                except Exception as e:
                    messagebox.showerror("DB Error", f"Failed to add machine: {e}")
                    return

            self.load_machine_data()
            self.refresh_table()
            win.destroy()

        ModernButton(btnf, text="Save", fg_color="#43A047", command=save_btn).pack(
            side="left", padx=8
        )
        ModernButton(btnf, text="Cancel", fg_color="#E53935", command=win.destroy).pack(
            side="left", padx=8
        )

    # ----------------------------------------------------
    # ANIMATIONS
    # ----------------------------------------------------
    def _blink_new_row(self, row_index):
        if not (self.use_tksheet and self.sheet):
            return
        try:
            self.sheet.highlight_rows(rows=[row_index], bg="#FFF176")
            self.after(500, lambda: self.sheet.highlight_rows(rows=[row_index], bg=None))
        except:
            pass

    def _fade_row(self, row_index):
        if not (self.use_tksheet and self.sheet):
            return

        try:
            start = (255, 245, 120)
            end = (255, 255, 255)
            steps = 12
            interval = 15

            def fade(i=0):
                if i > steps:
                    return
                r = int(start[0] + (end[0] - start[0]) * i / steps)
                g = int(start[1] + (end[1] - start[1]) * i / steps)
                b = int(start[2] + (end[2] - start[2]) * i / steps)
                self.sheet.highlight_rows(rows=[row_index], bg=f"#{r:02x}{g:02x}{b:02x}")
                self.after(interval, lambda: fade(i + 1))

            fade()

        except:
            pass

    # ----------------------------------------------------
    # REFRESH TABLE
    # ----------------------------------------------------
    def refresh_table(self):
        query = self.search_var.get().strip().lower()

        def match(rec):
            return (
                query in rec.get("code", "").lower()
                or query in rec.get("name", "").lower()
                or query in rec.get("desc", "").lower()
            )

        if self.use_tksheet:
            try:
                data = [
                    [m.get("code", ""), m.get("name", ""), m.get("desc", "")]
                    for m in self.machines
                    if match(m)
                ]
                self.sheet.set_sheet_data(data)

                for i, w in enumerate(self._col_widths):
                    self.sheet.column_width(i, w)

                self.sheet.refresh()

            except:
                pass

        else:
            try:
                for r in self.tree.get_children():
                    self.tree.delete(r)

                for rec in self.machines:
                    if not match(rec):
                        continue

                    self.tree.insert(
                        "",
                        "end",
                        values=(rec.get("code", ""), rec.get("name", ""), rec.get("desc", "")),
                    )
            except:
                pass

    # ----------------------------------------------------
    # RESET
    # ----------------------------------------------------
    def reset_form(self):
        self._clear_inputs()
        self.edit_index = None
        try:
            self.action_btn.configure(text="➕ Add Machine")
        except:
            pass

    def go_back(self, event=None):
        try:
            self.app.load_settings_page()
        except:
            try: self.app.load_settings()
            except: pass

    def highlight_entry(self, entry, active):
        try:
            if active:
                entry.configure(border_width=2, border_color="#007B43")
                self.animate_glow(entry, 0)
            else:
                entry.configure(border_width=1, border_color="#007B43")
        except Exception:
            pass

    def animate_glow(self, entry, step):
        try:
            if self.focus_get() != entry:
                return
        except Exception:
            return
        wave_colors = ["#007B43", "#005C32", "#43A047", "#388E3C", "#005C32", "#007B43"]
        entry.configure(border_color=wave_colors[step % len(wave_colors)])
        self.after(250, lambda: self.animate_glow(entry, step + 1))

    def destroy(self):
        super().destroy()

#=============================================================
#---------------ITEM MASTER-----------------------------------
#=============================================================
        
class ItemMasterPage(ctk.CTkFrame):
    """
    Item Master Page
    Fields:
      - Item Code
      - Item Name
      - drawing no
      - revision no
    Storage: items.json
    """
    FNAME = "items.json"

    def __init__(self, parent, app):
        super().__init__(parent, fg_color=("white", "#1c1c1c"))
        self.app = app
        self.items = []  # list of dicts
        self.search_var = tk.StringVar()
        self.use_tksheet = False
        self.sheet = None
        self.tree = None
        self.resize_sheet = lambda event=None: None
        self._col_widths = [130, 550, 212, 212]  # default widths
        self.load_item_data()
        self.build_ui()
        self.refresh_table()
        try:
            self.after(60, self.resize_sheet)
        except Exception:
            pass

    # -------------------------
    # JSON helpers
    # -------------------------
    def _get_data_path(self):
        return os.path.join(os.getcwd(), self.FNAME)

    def load_item_data(self):
        try:
            conn = sqlite3.connect(resource_path("items.db"))
            cursor = conn.cursor()
            cursor.execute("SELECT code, drawing_no, name, revision_no FROM items")
            rows = cursor.fetchall()
            conn.close()
            self.items = [
                {"code": r[0], "drawing": r[1] or "", "name": r[2] or "", "revision": r[3] or ""}
                for r in rows
            ]
        except Exception as e:
            messagebox.showerror("Load Error", f"Failed to load items from DB: {e}")
            self.items = []

    # -------------------------
    # UI
    # -------------------------
    def build_ui(self):
        # Header
        header = ctk.CTkFrame(self, fg_color="transparent")
        header.pack(fill="x", side="top", pady=(15, 10), padx=20)
        
        # Back Button in Green
        ModernButton(
            header,
            text="← Back",
            width=80,
            height=32,
            font=("Segoe UI", 11, "bold"),
            fg_color="#007B43",
            hover_color="#005C32",
            text_color="white",
            corner_radius=6,
            command=self.go_back
        ).pack(side="left")

        # Branded Title: "Item" (Green) + "Master" (Red)
        title_container = ctk.CTkFrame(header, fg_color="transparent")
        title_container.pack(side="left", padx=20)
        
        lbl_item = ctk.CTkLabel(
            title_container,
            text="Item",
            font=("Segoe UI", 18, "bold"),
            text_color="#007B43"
        )
        lbl_item.pack(side="left")
        
        lbl_master = ctk.CTkLabel(
            title_container,
            text=" Master",
            font=("Segoe UI", 18, "bold"),
            text_color="#B0050E"
        )
        lbl_master.pack(side="left")

        # Company Logo on the right
        try:
            logo_path = resource_path("settings/cherry_full_logo.png")
            pil_img = Image.open(logo_path)
            pil_img = pil_img.resize((144, 40), Image.Resampling.LANCZOS)
            pil_img = add_corners(pil_img, 10)
            self.logo_img = ctk.CTkImage(pil_img, size=(144, 40))
            logo_lbl = ctk.CTkLabel(header, text="", image=self.logo_img)
            logo_lbl.pack(side="right", padx=(10, 20))
        except Exception as e:
            print(f"Error loading logo in ItemMasterPage: {e}")

        # Form frame
        add_frame = ModernCardFrame(self)
        add_frame.pack(fill="x", padx=20, pady=(10, 6))

        LABEL_WIDTH = 100
        ENTRY_WIDTH = 260

        # Row 0: Item Code + drawing no
        ctk.CTkLabel(add_frame, text="Item Code:", width=LABEL_WIDTH, anchor="w", text_color="#202124", font=("Segoe UI", 11, "bold")).grid(row=0, column=0, sticky="w", pady=6, padx=(10, 0))
        self.code_e = ctk.CTkEntry(add_frame, width=ENTRY_WIDTH, height=30, corner_radius=6, border_color="#E0E0E0")
        self.code_e.grid(row=0, column=1, sticky="w", padx=(10, 30))

        ctk.CTkLabel(add_frame, text="Drawing No:", width=LABEL_WIDTH, anchor="e", text_color="#202124", font=("Segoe UI", 11, "bold")).grid(row=0, column=2, sticky="e", padx=(0, 10), pady=6)
        self.drawing_e = ctk.CTkEntry(add_frame, width=ENTRY_WIDTH, height=30, corner_radius=6, border_color="#E0E0E0")
        self.drawing_e.grid(row=0, column=3, sticky="w", padx=(10, 10))

        # Row 1: Item Name + revision no
        ctk.CTkLabel(add_frame, text="Item Name:", width=LABEL_WIDTH, anchor="w", text_color="#202124", font=("Segoe UI", 11, "bold")).grid(row=1, column=0, sticky="w", pady=6, padx=(10, 0))
        self.name_e = ctk.CTkEntry(add_frame, width=ENTRY_WIDTH, height=30, corner_radius=6, border_color="#E0E0E0")
        self.name_e.grid(row=1, column=1, sticky="w", padx=(10, 30))

        ctk.CTkLabel(add_frame, text="Revision No:", width=LABEL_WIDTH, anchor="e", text_color="#202124", font=("Segoe UI", 11, "bold")).grid(row=1, column=2, sticky="e", padx=(0, 10), pady=6)
        self.revision_e = ctk.CTkEntry(add_frame, width=ENTRY_WIDTH, height=30, corner_radius=6, border_color="#E0E0E0")
        self.revision_e.grid(row=1, column=3, sticky="w", padx=(10, 10))

        # Row 2: Buttons
        btn_frame = ctk.CTkFrame(add_frame, fg_color="transparent")
        btn_frame.grid(row=2, column=0, columnspan=4, sticky="w", pady=(6, 10), padx=(10, 0))

        self.action_btn = ModernButton(
            btn_frame,
            text="➕ Add Item",
            width=140,
            height=32,
            font=("Segoe UI", 11, "bold"),
            fg_color="#007B43",
            hover_color="#005C32",
            text_color="white",
            corner_radius=6,
            command=self._on_action_clicked
        )
        self.action_btn.pack(side="left", padx=(0, 10))

        del_btn = ModernButton(
            btn_frame,
            text="🗑 Delete Selected",
            width=150,
            height=32,
            font=("Segoe UI", 11, "bold"),
            fg_color="#B0050E",
            hover_color="#90040B",
            text_color="white",
            corner_radius=6,
            command=self._delete_selected
        )
        del_btn.pack(side="left")

        # ====== TABLE ======
        table_card = ModernCardFrame(self)
        table_card.pack(fill="both", expand=True, padx=20, pady=(6, 12))
        table_card.grid_rowconfigure(0, weight=1)
        table_card.grid_columnconfigure(0, weight=1)

        # Try import tksheet
        try:
            from tksheet import Sheet
            SheetClass = Sheet
        except Exception:
            try:
                SheetClass = tksheet.Sheet
            except Exception:
                SheetClass = None

        self._col_widths = [130, 550, 212, 212]  # Code, Name, Drawing, Revision
        cols = ["Item Code", "Item Name", "Drawing No", "Revision No"]
        header_info = [
            ("Item Code", "🔑\ufe0e"),
            ("Item Name", "📦\ufe0e"),
            ("Drawing No", "📄\ufe0e"),
            ("Revision No", "🔄\ufe0e")
        ]

        if SheetClass is not None:
            self.use_tksheet = True
            
            # --- Custom Horizontal Scroll-Synced Header Frame ---
            self.header_row_frame = ctk.CTkFrame(table_card, fg_color="white", height=42, corner_radius=0)
            self.header_row_frame.pack(fill="x", padx=10, pady=(10, 0))
            self.header_row_frame.pack_propagate(False)
            
            # Canvas inside header frame for scrolling
            self.header_canvas = tk.Canvas(self.header_row_frame, bg="white", highlightthickness=0, height=42)
            self.header_canvas.pack(side="left", fill="both", expand=True)
            
            # Frame inside canvas for custom header cells
            self.header_inner_frame = tk.Frame(self.header_canvas, bg="white")
            self.header_canvas.create_window(0, 0, window=self.header_inner_frame, anchor="nw")
            
            # Vertical scrollbar spacer on the right of header
            self.header_scrollbar_spacer = ctk.CTkFrame(
                self.header_row_frame,
                fg_color="white",
                corner_radius=0,
                width=16,
                height=42
            )
            self.header_scrollbar_spacer.pack(side="right", fill="y")
            
            # Row index spacer cell on the far left displaying green up-triangle sorting indicator
            row_index_cell = ctk.CTkFrame(
                self.header_inner_frame,
                fg_color="white",
                corner_radius=0,
                border_width=1,
                border_color="#E0E0E0",
                width=40,
                height=42
            )
            row_index_cell.pack(side="left", fill="y")
            row_index_cell.pack_propagate(False)
            
            lbl_tri = ctk.CTkLabel(
                row_index_cell,
                text="▲",
                font=("Segoe UI", 10, "bold"),
                text_color="#007B43",
                anchor="center"
            )
            lbl_tri.pack(expand=True)

            self.header_widgets = []
            for name, icon in header_info:
                cell = ctk.CTkFrame(
                    self.header_inner_frame,
                    fg_color="white",
                    corner_radius=0,
                    border_width=1,
                    border_color="#E0E0E0",
                    height=42
                )
                cell.pack(side="left", fill="y")
                cell.pack_propagate(False)
                
                content_frame = ctk.CTkFrame(cell, fg_color="transparent")
                content_frame.pack(expand=True)
                
                ctk.CTkLabel(
                    content_frame,
                    text=icon,
                    font=("Segoe UI", 12, "normal"),
                    text_color="#007B43",
                    anchor="center"
                ).pack(side="left", padx=(0, 4))
                
                lbl = ctk.CTkLabel(
                    content_frame,
                    text=name,
                    font=("Segoe UI", 11, "bold"),
                    text_color="#1A1A1A", # BLACK text color
                    anchor="center"
                )
                lbl.pack(side="left")
                self.header_widgets.append(cell)

            # Green border line under headers
            self.border_line = ctk.CTkFrame(table_card, fg_color="#007B43", height=2, corner_radius=0)
            self.border_line.pack(fill="x", padx=10, pady=(0, 0))

            # Table sheet frame
            self.table_sheet_frame = ctk.CTkFrame(table_card, fg_color="white", corner_radius=0)
            self.table_sheet_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))
            self.table_sheet_frame.grid_rowconfigure(0, weight=1)
            self.table_sheet_frame.grid_columnconfigure(0, weight=1)

            data = [[it.get("code", ""), it.get("name", ""), it.get("drawing", ""), it.get("revision", "")] for it in self.items]

            self.sheet = SheetClass(
                self.table_sheet_frame,
                headers=cols,
                data=data,
                show_header=False,        # Hide native header
                show_row_index=True,       # Show row numbers index
                row_index_width=40,        # Width matches spacer (40px)
                show_x_scrollbar=True,
                show_y_scrollbar=True
            )
            self.sheet.grid(row=0, column=0, sticky="nsew")

            # Style tksheet cells to match
            try:
                self.sheet.set_options(
                    grid_fg="#E0E0E0",
                    table_bg="white",
                    table_fg="#202124",
                    frame_bg="white",
                    select_bg="#E8F5EE",
                    select_fg="#007B43",
                    font=("Segoe UI", 10, "normal")
                )
            except Exception:
                pass

            try:
                self.sheet.enable_bindings(("single_select", "row_select", "arrowkeys", "copy", "select_all","right_click_popup_menu"))
            except Exception:
                pass

            # double-click -> popup editor
            def _on_sheet_double_click(event):
                idx = self._get_selected_index()
                if idx is None:
                    return
                try:
                    self._open_editor_dialog(mode="edit", index=idx)
                except Exception:
                    pass

            try:
                self.sheet.bind("<Double-1>", _on_sheet_double_click)
            except Exception:
                try:
                    self.sheet.bind("<Double-Button-1>", _on_sheet_double_click)
                except Exception:
                    pass

            # Resize columns resizes both sheet and header cells
            def resize_sheet(event=None):
                try:
                    total_width = self.table_sheet_frame.winfo_width() or 1
                    if total_width < 120:
                        self.after(60, self.resize_sheet)
                        return
                    col_count = len(cols)
                    # Deduct spacing for the vertical scrollbar (16px) and row index column (40px)
                    available = max(500, total_width - 16 - 40)
                    col_width = max(80, int(available / col_count))
                    for c in range(col_count):
                        try:
                            self.sheet.column_width(column=c, width=col_width, only_set_if_too_small=False)
                        except TypeError:
                            self.sheet.column_width(c, col_width)
                        
                        # Apply width to header cells
                        if c < len(self.header_widgets):
                            self.header_widgets[c].configure(width=col_width)
                    try:
                        self.sheet.refresh()
                    except Exception:
                        pass
                    try:
                        self.header_inner_frame.update_idletasks()
                        self.header_canvas.configure(scrollregion=self.header_canvas.bbox("all"))
                    except Exception:
                        pass
                except Exception as e:
                    print("Error resizing Item Master sheet:", e)

            self.resize_sheet = resize_sheet
            self.table_sheet_frame.bind("<Configure>", self.resize_sheet)
            try:
                self.after(80, self.resize_sheet)
            except Exception:
                pass
            
        else:
            # Treeview fallback
            self.use_tksheet = False
            self.sheet = None
            self.tree = ttk.Treeview(table_card, columns=("code", "name", "drawing", "revision"), show="headings", selectmode="browse")
            self.tree.heading("code", text="Item Code"); self.tree.column("code", width=390, anchor="center")
            self.tree.heading("name", text="Item Name"); self.tree.column("name", width=400, anchor="w")
            self.tree.heading("drawing", text="Drawing No"); self.tree.column("drawing", width=240, anchor="w")
            self.tree.heading("revision", text="Revision No"); self.tree.column("revision", width=140, anchor="center")
            vsb = ttk.Scrollbar(table_card, orient="vertical", command=self.tree.yview)
            hsb = ttk.Scrollbar(table_card, orient="horizontal", command=self.tree.xview)
            self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
            self.tree.grid(row=0, column=0, sticky="nsew")
            vsb.grid(row=0, column=1, sticky="ns")
            hsb.grid(row=1, column=0, sticky="ew")
            # bind double click to open popup
            self.tree.bind("<Double-1>", lambda e: self._open_editor_dialog(mode="edit", index=self._get_selected_index()))

        # initial refresh
        self.after(50, self.refresh_table)

    def go_back(self, event=None):
        try:
            self.app.load_settings_page()
        except:
            try: self.app.load_settings()
            except: pass


    # -------------------------
    # Utilities
    # -------------------------
    def _get_selected_index(self):
        try:
            if self.use_tksheet and self.sheet is not None:
                try:
                    cur = self.sheet.get_currently_selected()
                    if cur and isinstance(cur, tuple):
                        row = cur[0]
                        if isinstance(row, int) and row >= 0:
                            return row
                except Exception:
                    pass
                try:
                    rows = self.sheet.get_selected_rows()
                    if rows and isinstance(rows, list) and len(rows) > 0 and rows[0] >= 0:
                        return rows[0]
                except Exception:
                    pass
                return None
            else:
                sel = self.tree.selection()
                if not sel:
                    return None
                vals = self.tree.item(sel[0], "values")
                code = str(vals[0])
                for i, rec in enumerate(self.items):
                    if str(rec.get("code", "")) == code:
                        return i
        except Exception:
            return None

    def _clear_inputs(self):
        try:
            self.code_e.delete(0, "end")
            self.name_e.delete(0, "end")
            self.drawing_e.delete(0, "end")
            self.revision_e.delete(0, "end")
        except Exception:
            pass

    # -------------------------
    # Add / Update / Delete
    # -------------------------
    def _on_action_clicked(self):
        code = self.code_e.get().strip()
        name = self.name_e.get().strip()
        drawing = self.drawing_e.get().strip()
        revision = self.revision_e.get().strip()

        if not code:
            messagebox.showwarning("Validation", "Item Code required.")
            return
        if not name:
            messagebox.showwarning("Validation", "Item Name required.")
            return

        # Add mode
        if getattr(self, "edit_index", None) is None:
            if any(str(it.get("code", "")) == code for it in self.items):
                messagebox.showwarning("Validation", "Item Code already exists.")
                return
            try:
                conn = sqlite3.connect(resource_path("items.db"))
                cursor = conn.cursor()
                cursor.execute("INSERT INTO items (code, drawing_no, name, revision_no) VALUES (?, ?, ?, ?)",
                               (code, drawing, name, revision))
                conn.commit()
                conn.close()
            except Exception as e:
                messagebox.showerror("DB Error", f"Failed to add item: {e}")
                return
                
            self.load_item_data()
            self.refresh_table()

            new_index = next((i for i, it in enumerate(self.items) if str(it["code"]) == code), len(self.items) - 1)
            try:
                self.after(120, lambda idx=new_index: self._blink_new_row(idx))
                self.after(180, lambda idx=new_index: self._fade_row(idx))
            except Exception:
                pass
            self._clear_inputs()
            try:
                self.code_e.focus_set()
            except Exception:
                pass
            try:
                self.app.status_label.configure(text=f"Item added: {code} ✅")
            except Exception:
                pass
        else:
            # Update mode
            idx = self.edit_index
            old_code = self.items[idx].get("code")
            try:
                conn = sqlite3.connect(resource_path("items.db"))
                cursor = conn.cursor()
                cursor.execute("UPDATE items SET code=?, drawing_no=?, name=?, revision_no=? WHERE code=?",
                               (code, drawing, name, revision, old_code))
                conn.commit()
                conn.close()
            except Exception as e:
                messagebox.showerror("Update Error", f"Failed to update entry: {e}")
                return
                
            self.load_item_data()
            self.refresh_table()
            # restore selection
            def _restore():
                try:
                    if self.use_tksheet and self.sheet is not None:
                        try:
                            self.sheet.select_row(idx)
                        except Exception:
                            try:
                                self.sheet.set_selected_rows([idx])
                            except Exception:
                                pass
                    else:
                        for iid in self.tree.get_children():
                            vals = self.tree.item(iid, "values")
                            if str(vals[0]) == str(self.items[idx].get("code", "")):
                                self.tree.selection_set(iid)
                                self.tree.see(iid)
                                break
                except Exception:
                    pass
            try:
                self.after(80, _restore)
            except Exception:
                _restore()
            self._clear_inputs()
            self.edit_index = None
            self.action_btn.configure(text="➕ Add Item")
            try:
                self.app.status_label.configure(text=f"Item updated: {code} ✅")
            except Exception:
                pass

    def _delete_selected(self):
        idx = self._get_selected_index()
        if idx is None:
            messagebox.showwarning("No Selection", "Select a row to delete.")
            return
        rec = self.items[idx]
        if not messagebox.askyesno("Confirm Delete", f"Delete item '{rec.get('name','')}' (Code: {rec.get('code','')})?"):
            return
        try:
            conn = sqlite3.connect(resource_path("items.db"))
            cursor = conn.cursor()
            cursor.execute("DELETE FROM items WHERE code=?", (rec.get('code'),))
            conn.commit()
            conn.close()
            
            self.load_item_data()
            self.refresh_table()
            try:
                self.app.status_label.configure(text=f"Deleted item {rec.get('code')} ✅")
            except Exception:
                pass
        except Exception as e:
            messagebox.showerror("Delete Failed", f"Could not delete: {e}")

    # -------------------------
    # Popup editor
    # -------------------------
    def _open_editor_dialog(self, mode="edit", index=None):
        if mode not in ("edit", "add"):
            return
        win = ctk.CTkToplevel(self)
        win.title("Edit Item" if mode == "edit" else "Add Item")
        win.geometry("580x320")
        win.grab_set()

        frm = ctk.CTkFrame(win)
        frm.pack(fill="both", expand=True, padx=12, pady=12)

        ctk.CTkLabel(frm, text="Item Code:", font=("Segoe UI", 11)).grid(row=0, column=0, sticky="w", padx=6, pady=6)
        code_e = ctk.CTkEntry(frm, width=300); code_e.grid(row=0, column=1, sticky="w", padx=6, pady=6)

        ctk.CTkLabel(frm, text="Item Name:", font=("Segoe UI", 11)).grid(row=1, column=0, sticky="w", padx=6, pady=6)
        name_e = ctk.CTkEntry(frm, width=420); name_e.grid(row=1, column=1, sticky="w", padx=6, pady=6)

        ctk.CTkLabel(frm, text="Drawing No:", font=("Segoe UI", 11)).grid(row=2, column=0, sticky="w", padx=6, pady=6)
        drawing_e = ctk.CTkEntry(frm, width=300); drawing_e.grid(row=2, column=1, sticky="w", padx=6, pady=6)

        ctk.CTkLabel(frm, text="Revision No:", font=("Segoe UI", 11)).grid(row=3, column=0, sticky="w", padx=6, pady=6)
        revision_e = ctk.CTkEntry(frm, width=180); revision_e.grid(row=3, column=1, sticky="w", padx=6, pady=6)

        if mode == "edit" and index is not None:
            rec = self.items[index]
            code_e.insert(0, str(rec.get("code","")))
            name_e.insert(0, str(rec.get("name","")))
            drawing_e.insert(0, str(rec.get("drawing","")))
            revision_e.insert(0, str(rec.get("revision","")))

        btnf = ctk.CTkFrame(frm, fg_color="transparent")
        btnf.grid(row=4, column=0, columnspan=2, pady=(12, 0))

        def on_save():
            code = code_e.get().strip()
            name = name_e.get().strip()
            drawing = drawing_e.get().strip()
            revision = revision_e.get().strip()
            if not code or not name:
                messagebox.showwarning("Validation", "Item Code and Item Name required.")
                return
            if mode == "edit" and index is not None:
                if not messagebox.askyesno("Confirm Update", f"Apply changes to item {code}?"):
                    return
                old_code = self.items[index].get("code")
                try:
                    conn = sqlite3.connect(resource_path("items.db"))
                    cursor = conn.cursor()
                    cursor.execute("UPDATE items SET code=?, drawing_no=?, name=?, revision_no=? WHERE code=?",
                                   (code, drawing, name, revision, old_code))
                    conn.commit()
                    conn.close()
                except Exception as e:
                    messagebox.showerror("Update Error", f"Failed to update entry: {e}")
                    return
            else:
                if any(str(it.get("code","")) == code for it in self.items):
                    messagebox.showwarning("Validation", "Item Code already exists.")
                    return
                try:
                    conn = sqlite3.connect(resource_path("items.db"))
                    cursor = conn.cursor()
                    cursor.execute("INSERT INTO items (code, drawing_no, name, revision_no) VALUES (?, ?, ?, ?)",
                                   (code, drawing, name, revision))
                    conn.commit()
                    conn.close()
                except Exception as e:
                    messagebox.showerror("DB Error", f"Failed to add item: {e}")
                    return
                    
            self.load_item_data()
            self.refresh_table()
            win.destroy()

        ModernButton(btnf, text="Save", fg_color="#43A047", command=on_save).pack(side="left", padx=8)
        ModernButton(btnf, text="Cancel", fg_color="#E53935", command=win.destroy).pack(side="left", padx=8)

    # -------------------------
    # Animations (tksheet)
    # -------------------------
    def _blink_new_row(self, row_index):
        if not (self.use_tksheet and self.sheet):
            return
        try:
            self.sheet.see(row_index, 0)
            self.sheet.highlight_rows(rows=[row_index], bg="#FFF176")
            self.after(500, lambda: self.sheet.highlight_rows(rows=[row_index], bg=None))
        except Exception:
            pass

    def _fade_row(self, row_index):
        if not (self.use_tksheet and self.sheet):
            return
        try:
            start_r, start_g, start_b = (255, 245, 120)
            end_r, end_g, end_b = (255, 255, 255)
            steps = 12
            duration = 15
            def interp(t):
                r = int(start_r + (end_r - start_r) * t)
                g = int(start_g + (end_g - start_g) * t)
                b = int(start_b + (end_b - start_b) * t)
                return f"#{r:02x}{g:02x}{b:02x}"
            def animate(frame=0):
                try:
                    if frame <= steps:
                        self.sheet.highlight_rows(rows=[row_index], bg=interp(frame/steps))
                        self.after(duration, lambda: animate(frame+1))
                    else:
                        self.sheet.highlight_rows(rows=[row_index], bg=None)
                except Exception:
                    pass
            animate(0)
        except Exception:
            pass

    # -------------------------
    # Refresh table
    # -------------------------
    def refresh_table(self):
        q = ""
        try:
            q = self.search_var.get().strip().lower()
        except Exception:
            q = ""
        def matches(rec):
            if not q:
                return True
            return (q in str(rec.get("code","")).lower() or
                    q in str(rec.get("name","")).lower() or
                    q in str(rec.get("drawing","")).lower() or
                    q in str(rec.get("revision","")).lower())
        if getattr(self, "use_tksheet", False) and getattr(self, "sheet", None):
            try:
                data = [[rec.get("code",""), rec.get("name",""), rec.get("drawing",""), rec.get("revision","")] for rec in self.items if matches(rec)]
                self.sheet.set_sheet_data(data)
                # Reapply fixed widths every refresh
                try:
                    for i, w in enumerate(self._col_widths):
                        self.sheet.column_width(i, w)
                except Exception:
                    pass
                try:
                    self.resize_sheet()
                    self.sheet.refresh()
                except Exception:
                    pass
            except Exception:
                pass
        else:
            try:
                for r in self.tree.get_children():
                    self.tree.delete(r)
                for rec in self.items:
                    if not matches(rec):
                        continue
                    self.tree.insert("", "end", values=(rec.get("code",""), rec.get("name",""), rec.get("drawing",""), rec.get("revision","")))
            except Exception:
                pass

    # -------------------------
    # Reset form helper
    # -------------------------
    def reset_form(self):
        self._clear_inputs()


#=========================================================
#-----------------ProcessMasterPage-----------------------
#=========================================================
class ProcessMasterPage(ctk.CTkFrame):
    """
    Process Master Page
    Fields:
      - Process Code
      - Process Name
      - Description
    Storage: process.json
    """
    FNAME = "process.json"

    def __init__(self, parent, app):
        super().__init__(parent, fg_color=("white", "#1c1c1c"))
        self.app = app

        self.processes = []  # list of dict entries
        self.search_var = tk.StringVar()
        self.use_tksheet = False
        self.sheet = None
        self.tree = None
        self.resize_sheet = lambda event=None: None

        # Manual fixed column widths (same as Item page style)
        self._col_widths = [140, 420, 550]

        self.load_process_data()
        self.build_ui()
        self.refresh_table()

        try:
            self.after(60, self.resize_sheet)
        except Exception:
            pass

    # -------------------------------------
    # JSON Helpers
    # -------------------------------------
    def _get_data_path(self):
        return os.path.join(os.getcwd(), self.FNAME)

    def load_process_data(self):
        try:
            conn = sqlite3.connect(resource_path("process_master.db"))
            cursor = conn.cursor()
            cursor.execute("SELECT id, name, description FROM process_master")
            rows = cursor.fetchall()
            conn.close()
            self.processes = [
                {"code": r[0], "name": r[1] or "", "desc": r[2] or ""}
                for r in rows
            ]
        except Exception as e:
            messagebox.showerror("Load Error", f"Failed to load processes from DB: {e}")
            self.processes = []

    # -------------------------------------
    # UI Builder
    # -------------------------------------
    def build_ui(self):
        # HEADER
        header = ctk.CTkFrame(self, fg_color="transparent")
        header.pack(fill="x", side="top", pady=(15, 10), padx=20)
        
        # Back Button in Green
        ModernButton(
            header,
            text="← Back",
            width=80,
            height=32,
            font=("Segoe UI", 11, "bold"),
            fg_color="#007B43",
            hover_color="#005C32",
            text_color="white",
            corner_radius=6,
            command=self.go_back
        ).pack(side="left")

        # Branded Title: "Process" (Green) + "Master" (Red)
        title_container = ctk.CTkFrame(header, fg_color="transparent")
        title_container.pack(side="left", padx=20)
        
        lbl_icon = ctk.CTkLabel(
            title_container,
            text="⚙\ufe0e ",
            font=("Segoe UI", 18, "bold"),
            text_color="#007B43"
        )
        lbl_icon.pack(side="left")
        
        lbl_process = ctk.CTkLabel(
            title_container,
            text="Process",
            font=("Segoe UI", 18, "bold"),
            text_color="#007B43"
        )
        lbl_process.pack(side="left")
        
        lbl_master = ctk.CTkLabel(
            title_container,
            text=" Master",
            font=("Segoe UI", 18, "bold"),
            text_color="#B0050E"
        )
        lbl_master.pack(side="left")

        # Company Logo on the right
        try:
            logo_path = resource_path("settings/cherry_full_logo.png")
            pil_img = Image.open(logo_path)
            pil_img = pil_img.resize((144, 40), Image.Resampling.LANCZOS)
            pil_img = add_corners(pil_img, 10)
            self.logo_img2 = ctk.CTkImage(pil_img, size=(144, 40))
            logo_lbl = ctk.CTkLabel(header, text="", image=self.logo_img2)
            logo_lbl.pack(side="right", padx=(10, 20))
        except Exception as e:
            print(f"Error loading logo in ProcessMasterPage: {e}")

        # Form
        add_frame = ModernCardFrame(self)
        add_frame.pack(fill="x", padx=20, pady=(10, 6))

        LABEL_WIDTH = 100
        ENTRY_WIDTH = 260

        # Row 0: Code + Name
        ctk.CTkLabel(add_frame, text="Code:", width=LABEL_WIDTH, anchor="w", text_color="#202124", font=("Segoe UI", 11, "bold")).grid(
            row=0, column=0, sticky="w", pady=6, padx=(10, 0)
        )
        self.code_e = ctk.CTkEntry(add_frame, width=ENTRY_WIDTH, height=30, corner_radius=6, border_color="#007B43")
        self.code_e.grid(row=0, column=1, sticky="w", padx=(10, 30))

        ctk.CTkLabel(add_frame, text="Name:", width=LABEL_WIDTH, anchor="e", text_color="#202124", font=("Segoe UI", 11, "bold")).grid(
            row=0, column=2, sticky="e", padx=(0, 10), pady=6
        )
        self.name_e = ctk.CTkEntry(add_frame, width=ENTRY_WIDTH, height=30, corner_radius=6, border_color="#007B43")
        self.name_e.grid(row=0, column=3, sticky="w", padx=(10, 10))

        # Row 1: Description
        ctk.CTkLabel(add_frame, text="Description:", width=LABEL_WIDTH, anchor="w", text_color="#202124", font=("Segoe UI", 11, "bold")).grid(
            row=1, column=0, sticky="w", pady=6, padx=(10, 0)
        )
        self.desc_e = ctk.CTkEntry(add_frame, width=ENTRY_WIDTH * 2 + 130, height=30, corner_radius=6, border_color="#007B43")
        self.desc_e.grid(row=1, column=1, columnspan=3, sticky="w", padx=(10, 10))

        # Buttons row
        btn_frame = ctk.CTkFrame(add_frame, fg_color="transparent")
        btn_frame.grid(row=2, column=0, columnspan=4, sticky="w", pady=(6, 10), padx=(10, 0))

        self.action_btn = ModernButton(
            btn_frame,
            text="➕ Add Process",
            width=140,
            height=32,
            font=("Segoe UI", 11, "bold"),
            fg_color="#007B43",
            hover_color="#005C32",
            text_color="white",
            corner_radius=6,
            command=self._on_action_clicked
        )
        self.action_btn.pack(side="left", padx=(0, 10))

        del_btn = ModernButton(
            btn_frame,
            text="🗑 Delete Selected",
            width=150,
            height=32,
            font=("Segoe UI", 11, "bold"),
            fg_color="#B0050E",
            hover_color="#90040B",
            text_color="white",
            corner_radius=6,
            command=self._delete_selected
        )
        del_btn.pack(side="left")

        # Bind focus highlights
        for entry in (self.code_e, self.name_e, self.desc_e):
            entry.bind("<FocusIn>", lambda e, ent=entry: self.highlight_entry(ent, True))
            entry.bind("<FocusOut>", lambda e, ent=entry: self.highlight_entry(ent, False))

        # ====== TABLE ======
        table_card = ModernCardFrame(self)
        table_card.pack(fill="both", expand=True, padx=20, pady=(6, 12))
        table_card.grid_rowconfigure(0, weight=1)
        table_card.grid_columnconfigure(0, weight=1)

        # Try import tksheet
        try:
            from tksheet import Sheet
            SheetClass = Sheet
        except Exception:
            try:
                SheetClass = tksheet.Sheet
            except Exception:
                SheetClass = None

        cols = ["Code", "Name", "Description"]
        header_info = [
            ("Code", "🔑\ufe0e"),
            ("Name", "⚙\ufe0e"),
            ("Description", "📄\ufe0e")
        ]

        if SheetClass is not None:
            self.use_tksheet = True
            
            # --- Custom Horizontal Scroll-Synced Header Frame ---
            self.header_row_frame = ctk.CTkFrame(table_card, fg_color="white", height=42, corner_radius=0)
            self.header_row_frame.pack(fill="x", padx=10, pady=(10, 0))
            self.header_row_frame.pack_propagate(False)
            
            # Canvas inside header frame for scrolling
            self.header_canvas = tk.Canvas(self.header_row_frame, bg="white", highlightthickness=0, height=42)
            self.header_canvas.pack(side="left", fill="both", expand=True)
            
            # Frame inside canvas for custom header cells
            self.header_inner_frame = tk.Frame(self.header_canvas, bg="white")
            self.header_canvas.create_window(0, 0, window=self.header_inner_frame, anchor="nw")
            
            # Vertical scrollbar spacer on the right of header
            self.header_scrollbar_spacer = ctk.CTkFrame(
                self.header_row_frame,
                fg_color="white",
                corner_radius=0,
                width=16,
                height=42
            )
            self.header_scrollbar_spacer.pack(side="right", fill="y")
            
            # Row index spacer cell on the far left displaying green up-triangle sorting indicator
            row_index_cell = ctk.CTkFrame(
                self.header_inner_frame,
                fg_color="white",
                corner_radius=0,
                border_width=1,
                border_color="#E0E0E0",
                width=40,
                height=42
            )
            row_index_cell.pack(side="left", fill="y")
            row_index_cell.pack_propagate(False)
            
            lbl_tri = ctk.CTkLabel(
                row_index_cell,
                text="▲",
                font=("Segoe UI", 10, "bold"),
                text_color="#007B43",
                anchor="center"
            )
            lbl_tri.pack(expand=True)

            self.header_widgets = []
            for name, icon in header_info:
                cell = ctk.CTkFrame(
                    self.header_inner_frame,
                    fg_color="white",
                    corner_radius=0,
                    border_width=1,
                    border_color="#E0E0E0",
                    height=42
                )
                cell.pack(side="left", fill="y")
                cell.pack_propagate(False)
                
                content_frame = ctk.CTkFrame(cell, fg_color="transparent")
                content_frame.pack(expand=True)
                
                ctk.CTkLabel(
                    content_frame,
                    text=icon,
                    font=("Segoe UI", 12, "normal"),
                    text_color="#007B43",
                    anchor="center"
                ).pack(side="left", padx=(0, 4))
                
                lbl = ctk.CTkLabel(
                    content_frame,
                    text=name,
                    font=("Segoe UI", 11, "bold"),
                    text_color="#1A1A1A", # BLACK text color
                    anchor="center"
                )
                lbl.pack(side="left")
                self.header_widgets.append(cell)

            # Green border line under headers
            self.border_line = ctk.CTkFrame(table_card, fg_color="#007B43", height=2, corner_radius=0)
            self.border_line.pack(fill="x", padx=10, pady=(0, 0))

            # Table sheet frame
            self.table_sheet_frame = ctk.CTkFrame(table_card, fg_color="white", corner_radius=0)
            self.table_sheet_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))
            self.table_sheet_frame.grid_rowconfigure(0, weight=1)
            self.table_sheet_frame.grid_columnconfigure(0, weight=1)

            data = [
                [p.get("code", ""), p.get("name", ""), p.get("desc", "")]
                for p in self.processes
            ]

            self.sheet = SheetClass(
                self.table_sheet_frame,
                headers=cols,
                data=data,
                show_header=False,        # Hide native header
                show_row_index=True,       # Show row numbers index
                row_index_width=40,        # Width matches spacer (40px)
                show_x_scrollbar=True,
                show_y_scrollbar=True
            )
            self.sheet.grid(row=0, column=0, sticky="nsew")

            # Style tksheet cells to match
            try:
                self.sheet.set_options(
                    grid_fg="#E0E0E0",
                    table_bg="white",
                    table_fg="#202124",
                    frame_bg="white",
                    select_bg="#E8F5EE",
                    select_fg="#007B43",
                    font=("Segoe UI", 10, "normal")
                )
            except Exception:
                pass

            try:
                self.sheet.enable_bindings(("single_select", "row_select", "arrowkeys", "copy", "select_all","right_click_popup_menu"))
            except Exception:
                pass

            # Double click handler
            def _dbl(event):
                idx = self._get_selected_index()
                if idx is not None:
                    self._open_editor_dialog("edit", idx)

            try:
                self.sheet.bind("<Double-1>", _dbl)
            except Exception:
                pass

            # Sync horizontal scrolling to custom header
            try:
                orig_xscroll = self.sheet.MT.cget("xscrollcommand")
                def sync_scroll(first, last):
                    if orig_xscroll:
                        try:
                            self.sheet.tk.eval(f"{orig_xscroll} {first} {last}")
                        except Exception:
                            pass
                    try:
                        self.header_canvas.xview_moveto(first)
                    except Exception:
                        pass
                self.sheet.MT.configure(xscrollcommand=sync_scroll)
            except Exception as e:
                print("Failed to sync Process Master header scrollbar:", e)

            # Resize columns resizes both sheet and header cells
            def resize_sheet(event=None):
                try:
                    total_width = self.table_sheet_frame.winfo_width() or 1
                    if total_width < 120:
                        self.after(60, self.resize_sheet)
                        return
                    col_count = len(cols)
                    # Deduct spacing for the vertical scrollbar (16px) and row index column (40px)
                    available = max(500, total_width - 16 - 40)
                    col_width = max(80, int(available / col_count))
                    for c in range(col_count):
                        try:
                            self.sheet.column_width(column=c, width=col_width, only_set_if_too_small=False)
                        except TypeError:
                            self.sheet.column_width(c, col_width)
                        
                        # Apply width to header cells
                        if c < len(self.header_widgets):
                            self.header_widgets[c].configure(width=col_width)
                    try:
                        self.sheet.refresh()
                    except Exception:
                        pass
                    try:
                        self.header_inner_frame.update_idletasks()
                        self.header_canvas.configure(scrollregion=self.header_canvas.bbox("all"))
                    except Exception:
                        pass
                except Exception as e:
                    print("Error resizing Process Master sheet:", e)

            self.resize_sheet = resize_sheet
            self.table_sheet_frame.bind("<Configure>", self.resize_sheet)
            try:
                self.after(80, self.resize_sheet)
            except Exception:
                pass

        else:
            # TreeView fallback
            self.use_tksheet = False
            self.sheet = None

            self.tree = ttk.Treeview(
                table_card,
                columns=("code", "name", "desc"),
                show="headings",
                selectmode="browse",
            )
            self.tree.heading("code", text="Code")
            self.tree.column("code", width=260)
            self.tree.heading("name", text="Name")
            self.tree.column("name", width=320)
            self.tree.heading("desc", text="Description")
            self.tree.column("desc", width=500)

            vsb = ttk.Scrollbar(table_card, orient="vertical", command=self.tree.yview)
            hsb = ttk.Scrollbar(table_card, orient="horizontal", command=self.tree.xview)

            self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

            self.tree.grid(row=0, column=0, sticky="nsew")
            vsb.grid(row=0, column=1, sticky="ns")
            hsb.grid(row=1, column=0, sticky="ew")

            self.tree.bind(
                "<Double-1>",
                lambda e: self._open_editor_dialog("edit", self._get_selected_index()),
            )

        self.after(80, self.refresh_table)
        


    # ----------------------------------------------------------------
    # Utilities
    # ----------------------------------------------------------------
    def _get_selected_index(self):
        try:
            if self.use_tksheet and self.sheet:
                try:
                    sel = self.sheet.get_currently_selected()
                    if sel and isinstance(sel, tuple):
                        return sel[0]
                except Exception:
                    pass

                try:
                    rows = self.sheet.get_selected_rows()
                    if rows and rows[0] >= 0:
                        return rows[0]
                except Exception:
                    pass
                return None

            else:
                sel = self.tree.selection()
                if not sel:
                    return None
                vals = self.tree.item(sel[0], "values")
                code = str(vals[0])
                for i, p in enumerate(self.processes):
                    if str(p.get("code", "")) == code:
                        return i

        except Exception:
            return None

    def _clear_inputs(self):
        try:
            self.code_e.delete(0, "end")
            self.name_e.delete(0, "end")
            self.desc_e.delete(0, "end")
        except Exception:
            pass

    # ----------------------------------------------------------------
    # Add / Update / Delete
    # ----------------------------------------------------------------
    def _on_action_clicked(self):
        code = self.code_e.get().strip()
        name = self.name_e.get().strip()
        desc = self.desc_e.get().strip()

        if not code:
            messagebox.showwarning("Validation", "Code required.")
            return
        if not name:
            messagebox.showwarning("Validation", "Name required.")
            return

        # Add
        if getattr(self, "edit_index", None) is None:
            if any(str(p.get("code", "")) == code for p in self.processes):
                messagebox.showwarning("Validation", "Code already exists.")
                return

            try:
                conn = sqlite3.connect(resource_path("process_master.db"))
                cursor = conn.cursor()
                cursor.execute("INSERT INTO process_master (id, name, description) VALUES (?, ?, ?)",
                               (code, name, desc))
                conn.commit()
                conn.close()
            except Exception as e:
                messagebox.showerror("DB Error", f"Failed to add process: {e}")
                return

            self.load_process_data()
            self.refresh_table()

            idx = next((i for i, p in enumerate(self.processes) if str(p["code"]) == code), len(self.processes) - 1)

            try:
                self.after(120, lambda idx=idx: self._blink_new_row(idx))
                self.after(180, lambda idx=idx: self._fade_row(idx))
            except Exception:
                pass

            self._clear_inputs()
            try:
                self.code_e.focus_set()
            except Exception:
                pass

            try:
                self.app.status_label.configure(text=f"Process added: {code} ✅")
            except Exception:
                pass

        else:
            # Update
            idx = self.edit_index
            old_code = self.processes[idx].get("code")
            try:
                conn = sqlite3.connect(resource_path("process_master.db"))
                cursor = conn.cursor()
                cursor.execute("UPDATE process_master SET id=?, name=?, description=? WHERE id=?",
                               (code, name, desc, old_code))
                conn.commit()
                conn.close()
            except Exception as e:
                messagebox.showerror("Update Error", f"Failed to update entry: {e}")
                return

            self.load_process_data()
            self.refresh_table()

            def _restore():
                try:
                    if self.use_tksheet:
                        self.sheet.select_row(idx)
                except Exception:
                    pass

            self.after(80, _restore)

            self._clear_inputs()
            self.edit_index = None
            self.action_btn.configure(text="➕ Add Process")

            try:
                self.app.status_label.configure(text=f"Updated: {code} ✅")
            except Exception:
                pass

    def _delete_selected(self):
        idx = self._get_selected_index()
        if idx is None:
            messagebox.showwarning("No Selection", "Select a row.")
            return

        rec = self.processes[idx]

        if not messagebox.askyesno(
            "Confirm Delete",
            f"Delete process '{rec.get('name','')}' (Code: {rec.get('code','')})?",
        ):
            return

        try:
            conn = sqlite3.connect(resource_path("process_master.db"))
            cursor = conn.cursor()
            cursor.execute("DELETE FROM process_master WHERE id=?", (rec.get('code'),))
            conn.commit()
            conn.close()

            self.load_process_data()
            self.refresh_table()
        except Exception as e:
            messagebox.showerror("Delete Failed", f"Could not delete: {e}")
            return

        try:
            self.app.status_label.configure(text=f"Deleted {rec.get('code')} ✅")
        except Exception:
            pass

    # ----------------------------------------------------------------
    # Popup Editor
    # ----------------------------------------------------------------
    def _open_editor_dialog(self, mode="edit", index=None):
        if mode not in ("edit", "add"):
            return

        win = ctk.CTkToplevel(self)
        win.title("Edit Process" if mode == "edit" else "Add Process")
        win.geometry("580x320")
        win.grab_set()

        frm = ctk.CTkFrame(win)
        frm.pack(fill="both", expand=True, padx=12, pady=12)

        ctk.CTkLabel(frm, text="Code:", font=("Segoe UI", 11)).grid(
            row=0, column=0, sticky="w", padx=6, pady=6
        )
        code_e = ctk.CTkEntry(frm, width=300)
        code_e.grid(row=0, column=1, sticky="w", padx=6, pady=6)

        ctk.CTkLabel(frm, text="Name:", font=("Segoe UI", 11)).grid(
            row=1, column=0, sticky="w", padx=6, pady=6
        )
        name_e = ctk.CTkEntry(frm, width=420)
        name_e.grid(row=1, column=1, sticky="w", padx=6, pady=6)

        ctk.CTkLabel(frm, text="Description:", font=("Segoe UI", 11)).grid(
            row=2, column=0, sticky="w", padx=6, pady=6
        )
        desc_e = ctk.CTkEntry(frm, width=420)
        desc_e.grid(row=2, column=1, sticky="w", padx=6, pady=6)

        if mode == "edit" and index is not None:
            rec = self.processes[index]
            code_e.insert(0, str(rec.get("code", "")))
            name_e.insert(0, str(rec.get("name", "")))
            desc_e.insert(0, str(rec.get("desc", "")))

        btnf = ctk.CTkFrame(frm, fg_color="transparent")
        btnf.grid(row=3, column=0, columnspan=2, pady=(12, 0))

        def save_popup():
            code = code_e.get().strip()
            name = name_e.get().strip()
            desc = desc_e.get().strip()

            if not code or not name:
                messagebox.showwarning("Validation", "Code and Name required.")
                return

            if mode == "edit" and index is not None:
                if not messagebox.askyesno(
                    "Confirm Update", f"Apply changes to {code}?"
                ):
                    return
                old_code = self.processes[index].get("code")
                try:
                    conn = sqlite3.connect(resource_path("process_master.db"))
                    cursor = conn.cursor()
                    cursor.execute("UPDATE process_master SET id=?, name=?, description=? WHERE id=?",
                                   (code, name, desc, old_code))
                    conn.commit()
                    conn.close()
                except Exception as e:
                    messagebox.showerror("Update Error", f"Failed to update process: {e}")
                    return
            else:
                if any(str(p.get("code", "")) == code for p in self.processes):
                    messagebox.showwarning("Validation", "Code exists already.")
                    return
                try:
                    conn = sqlite3.connect(resource_path("process_master.db"))
                    cursor = conn.cursor()
                    cursor.execute("INSERT INTO process_master (id, name, description) VALUES (?, ?, ?)",
                                   (code, name, desc))
                    conn.commit()
                    conn.close()
                except Exception as e:
                    messagebox.showerror("DB Error", f"Failed to add process: {e}")
                    return

            self.load_process_data()
            self.refresh_table()
            win.destroy()

        ModernButton(btnf, text="Save", fg_color="#43A047", command=save_popup).pack(
            side="left", padx=8
        )
        ModernButton(btnf, text="Cancel", fg_color="#E53935", command=win.destroy).pack(
            side="left", padx=8
        )

    # ----------------------------------------------------------------
    # Animations
    # ----------------------------------------------------------------
    def _blink_new_row(self, row_index):
        if not (self.use_tksheet and self.sheet):
            return
        try:
            self.sheet.see(row_index, 0)
            self.sheet.highlight_rows(rows=[row_index], bg="#FFF176")
            self.after(500, lambda: self.sheet.highlight_rows(rows=[row_index], bg=None))
        except Exception:
            pass

    def _fade_row(self, row_index):
        if not (self.use_tksheet and self.sheet):
            return
        try:
            start_r, start_g, start_b = (255, 245, 120)
            end_r, end_g, end_b = (255, 255, 255)
            steps = 12
            duration = 15

            def interp(t):
                r = int(start_r + (end_r - start_r) * t)
                g = int(start_g + (end_g - start_g) * t)
                b = int(start_b + (end_b - start_b) * t)
                return f"#{r:02x}{g:02x}{b:02x}"

            def anim(i=0):
                try:
                    if i <= steps:
                        self.sheet.highlight_rows(rows=[row_index], bg=interp(i / steps))
                        self.after(duration, lambda: anim(i + 1))
                    else:
                        self.sheet.highlight_rows(rows=[row_index], bg=None)
                except Exception:
                    pass

            anim(0)
        except Exception:
            pass

    # ----------------------------------------------------------------
    # Refresh Table
    # ----------------------------------------------------------------
    def refresh_table(self):
        query = ""
        try:
            query = self.search_var.get().strip().lower()
        except Exception:
            pass

        def match(p):
            if not query:
                return True
            return (
                query in str(p.get("code", "")).lower()
                or query in str(p.get("name", "")).lower()
                or query in str(p.get("desc", "")).lower()
            )

        if self.use_tksheet and self.sheet:
            try:
                data = [
                    [p.get("code", ""), p.get("name", ""), p.get("desc", "")]
                    for p in self.processes
                    if match(p)
                ]
                self.sheet.set_sheet_data(data)

                # fixed widths
                for i, w in enumerate(self._col_widths):
                    self.sheet.column_width(i, w)

                try:
                    self.sheet.refresh()
                except Exception:
                    pass

            except Exception:
                pass

        else:
            try:
                # Clear
                for r in self.tree.get_children():
                    self.tree.delete(r)

                for p in self.processes:
                    if not match(p):
                        continue

                    self.tree.insert(
                        "", "end",
                        values=(p.get("code", ""), p.get("name", ""), p.get("desc", ""))
                    )
            except Exception:
                pass

    # ----------------------------------------------------------------
    # Reset
    # ----------------------------------------------------------------
    def reset_form(self):
        self._clear_inputs()
        self.edit_index = None
        try:
            self.action_btn.configure(text="➕ Add Process")
        except Exception:
            pass

    def go_back(self, event=None):
        try:
            self.app.load_settings_page()
        except:
            try: self.app.load_settings()
            except: pass

    def highlight_entry(self, entry, active):
        try:
            if active:
                entry.configure(border_width=2, border_color="#007B43")
                self.animate_glow(entry, 0)
            else:
                entry.configure(border_width=1, border_color="#007B43")
        except Exception:
            pass

    def animate_glow(self, entry, step):
        try:
            if self.focus_get() != entry:
                return
        except Exception:
            return
        wave_colors = ["#007B43", "#005C32", "#43A047", "#388E3C", "#005C32", "#007B43"]
        entry.configure(border_color=wave_colors[step % len(wave_colors)])
        self.after(250, lambda: self.animate_glow(entry, step + 1))

    def destroy(self):
        super().destroy()

# ==========================================================
# Operator Manager Page — Full Page (permanent inputs + table)
# ==========================================================
import os
import json
import tkinter as tk
from tkinter import ttk, messagebox

import customtkinter as ctk

# OperatorManagerPage: production-ready, responsive, tksheet preferred with Treeview fallback.
class OperatorManagerPage(ctk.CTkFrame):
    FNAME = "operators.json"

    def __init__(self, parent, app):
        """
        parent: container in which this page will live (should be packed/gridded with fill='both', expand=True)
        app: main application object (used for status_label updates, optional)
        """
        super().__init__(parent, fg_color=("white", "#1c1c1c"))
        self.app = app
        self.operators: list[dict] = []
        self._load_operators()

        # UI references
        self.edit_index = None
        self.use_tksheet = False
        self.sheet = None
        self.tree = None

        # Build UI
        self.build_ui()

        # Initial population and safe resize scheduling
        self.refresh_table()
        try:
            self.after(100, self._safe_resize)
        except Exception:
            pass

    # -----------------------
    # UI BUILD
    # -----------------------
    def build_ui(self):
        # Header
        header = ctk.CTkFrame(self, fg_color="transparent")
        header.pack(fill="x", side="top", pady=(15, 10), padx=20)

        # Back Button in Green
        ModernButton(
            header,
            text="← Back",
            width=80,
            height=32,
            font=("Segoe UI", 11, "bold"),
            fg_color="#007B43",
            hover_color="#005C32",
            text_color="white",
            corner_radius=6,
            command=self.go_back
        ).pack(side="left")
        
        # Branded Title: "Employee" (Green) + "Master" (Red) next to User silhouette
        title_container = ctk.CTkFrame(header, fg_color="transparent")
        title_container.pack(side="left", padx=20)
        
        lbl_worker = ctk.CTkLabel(
            title_container,
            text="👤\ufe0e",
            font=("Segoe UI", 18, "bold"),
            text_color="#007B43"
        )
        lbl_worker.pack(side="left")
        
        lbl_emp = ctk.CTkLabel(
            title_container,
            text=" Employee",
            font=("Segoe UI", 18, "bold"),
            text_color="#007B43"
        )
        lbl_emp.pack(side="left")
        
        lbl_master = ctk.CTkLabel(
            title_container,
            text=" Master",
            font=("Segoe UI", 18, "bold"),
            text_color="#B0050E"
        )
        lbl_master.pack(side="left")

        # Company Logo on the right
        try:
            logo_path = resource_path("settings/cherry_full_logo.png")
            pil_img = Image.open(logo_path)
            pil_img = pil_img.resize((144, 40), Image.Resampling.LANCZOS)
            pil_img = add_corners(pil_img, 10)
            self.logo_img = ctk.CTkImage(pil_img, size=(144, 40))
            logo_lbl = ctk.CTkLabel(header, text="", image=self.logo_img)
            logo_lbl.pack(side="right", padx=(10, 20))
        except Exception as e:
            print(f"Error loading logo in OperatorManagerPage: {e}")

        # Content - main area (this will expand)
        content = ctk.CTkFrame(self, fg_color="transparent")
        content.pack(fill="both", expand=True, padx=12, pady=10)

        # Top area: use grid so description can expand while left inputs are fixed-width
        top_frame = ModernCardFrame(content)
        top_frame.pack(fill="x", pady=(0, 12), anchor="n")

        # Configure top grid: left fixed inputs, right expands
        top_frame.grid_columnconfigure(0, weight=0)  # left column (inputs)
        top_frame.grid_columnconfigure(1, weight=1)  # right column (description expands)

        # Left inputs container
        left_inputs = ctk.CTkFrame(top_frame, fg_color="transparent")
        left_inputs.grid(row=0, column=0, sticky="nw")

        # Operator ID
        ctk.CTkLabel(left_inputs, text="Operator ID:", font=("Segoe UI", 11, "bold"), text_color="#202124").grid(row=0, column=0, sticky="w", padx=(10, 8), pady=(12, 6))
        self.id_entry = ctk.CTkEntry(left_inputs, width=220, height=30, corner_radius=6, border_color="#007B43")
        self.id_entry.grid(row=0, column=1, sticky="w", padx=4, pady=(12, 6))

        # Name
        ctk.CTkLabel(left_inputs, text="Name:", font=("Segoe UI", 11, "bold"), text_color="#202124").grid(row=1, column=0, sticky="w", padx=(10, 8), pady=(2, 6))
        self.name_entry = ctk.CTkEntry(left_inputs, width=220, height=30, corner_radius=6, border_color="#007B43")
        self.name_entry.grid(row=1, column=1, sticky="w", padx=4, pady=(2, 6))

        # Phone
        ctk.CTkLabel(left_inputs, text="Phone (optional):", font=("Segoe UI", 11, "bold"), text_color="#202124").grid(row=2, column=0, sticky="w", padx=(10, 8), pady=(2, 6))
        self.phone_entry = ctk.CTkEntry(left_inputs, width=220, height=30, corner_radius=6, border_color="#007B43")
        self.phone_entry.grid(row=2, column=1, sticky="w", padx=4, pady=(2, 6))

        # Add / Update button
        self.action_btn = ModernButton(
            left_inputs,
            text="+ Add Operator",
            width=160,
            height=32,
            font=("Segoe UI", 11, "bold"),
            fg_color="#007B43",
            hover_color="#005C32",
            text_color="white",
            corner_radius=6,
            command=self._on_action_clicked
        )
        self.action_btn.grid(row=6, column=0, sticky="w", padx=(10, 6), pady=(8, 12))

        # Delete button next to Add button
        self.del_btn = ModernButton(
            left_inputs,
            text="🗑 Delete",
            width=120,
            height=32,
            font=("Segoe UI", 11, "bold"),
            fg_color="#B0050E",
            hover_color="#90040B",
            text_color="white",
            corner_radius=6,
            command=self.delete_selected
        )
        self.del_btn.grid(row=6, column=1, sticky="w", padx=(0, 4), pady=(8, 12))


        left_inputs.grid_rowconfigure(7, weight=1)  # spacer

        # Right Frame (kept fixed width)
        right_frame = ctk.CTkFrame(top_frame, fg_color="transparent")
        right_frame.grid(row=0, column=1, sticky="nw", padx=(12, 0))

        # STOP it from expanding horizontally
        right_frame.grid_columnconfigure(0, weight=0)
        right_frame.grid_rowconfigure(1, weight=0)

        ctk.CTkLabel(right_frame, text="Description:", font=("Segoe UI", 11, "bold"), text_color="#202124").grid(
            row=0, column=0, sticky="nw", padx=4, pady=(12, 6)
        )

        desc_holder = ctk.CTkFrame(
            right_frame,
            border_width=1,
            border_color="#007B43",
            corner_radius=6
        )

        desc_holder.grid(row=1, column=0, sticky="nw")   # ← STOP expansion
        desc_holder.configure(width=300)                # ← Set fixed width

        self.desc_text = tk.Text(
            desc_holder,
            width=40,        # now width works
            height=3,
            wrap="word",
            relief="flat",
            bd=0,
            font=("Segoe UI", 11)
        )
        self.desc_text.pack(padx=6, pady=6)

        # Bind focus highlights
        for entry in (self.id_entry, self.name_entry, self.phone_entry):
            entry.bind("<FocusIn>", lambda e, ent=entry: self.highlight_entry(ent, True))
            entry.bind("<FocusOut>", lambda e, ent=entry: self.highlight_entry(ent, False))
        
        self.desc_text.bind("<FocusIn>", lambda e: self.highlight_entry(desc_holder, True))
        self.desc_text.bind("<FocusOut>", lambda e: self.highlight_entry(desc_holder, False))


        # Table area container: grid so scrollbars and table share space nicely
        table_container = ModernCardFrame(content)
        table_container.pack(fill="both", expand=True)
        # configure grid so widget inside can expand
        table_container.grid_rowconfigure(0, weight=1)
        table_container.grid_columnconfigure(0, weight=1)


        # Create table (tksheet preferred)
        self._create_table(table_container)



    # -----------------------
    # Create table (tksheet or treeview fallback)
    # -----------------------
    def _create_table(self, container):
        cols = ["Operator ID", "Name", "Description", "Phone"]
        header_info = [
            ("Operator ID", "👤\ufe0e"),
            ("Name", "📝\ufe0e"),
            ("Description", "📄\ufe0e"),
            ("Phone", "📞\ufe0e")
        ]
        data = [[op.get("id", ""), op.get("name", ""), op.get("description", ""), op.get("phone", "")] for op in self.operators]

        # Try tksheet
        try:
            import tksheet as _tksheet
            SheetClass = _tksheet.Sheet
        except Exception:
            try:
                import tksheet
                SheetClass = tksheet.Sheet
            except Exception:
                SheetClass = None

        try:
            if SheetClass is None:
                raise ImportError("No tksheet library found")

            self.use_tksheet = True

            # Configure container grid for header + border + sheet
            container.grid_rowconfigure(0, weight=0)
            container.grid_rowconfigure(1, weight=0)
            container.grid_rowconfigure(2, weight=1)
            container.grid_columnconfigure(0, weight=1)

            # --- Custom Horizontal Scroll-Synced Header Frame ---
            self.header_row_frame = ctk.CTkFrame(container, fg_color="white", height=42, corner_radius=0)
            self.header_row_frame.grid(row=0, column=0, sticky="ew")
            self.header_row_frame.pack_propagate(False)
            
            # Canvas inside header frame for scrolling
            self.header_canvas = tk.Canvas(self.header_row_frame, bg="white", highlightthickness=0, height=42)
            self.header_canvas.pack(side="left", fill="both", expand=True)
            
            # Frame inside canvas for custom header cells
            self.header_inner_frame = tk.Frame(self.header_canvas, bg="white")
            self.header_canvas.create_window(0, 0, window=self.header_inner_frame, anchor="nw")
            
            # Vertical scrollbar spacer on the right of header
            self.header_scrollbar_spacer = ctk.CTkFrame(
                self.header_row_frame,
                fg_color="white",
                corner_radius=0,
                width=16,
                height=42
            )
            self.header_scrollbar_spacer.pack(side="right", fill="y")
            
            # Row index spacer cell on the far left displaying green up-triangle sorting indicator
            row_index_cell = ctk.CTkFrame(
                self.header_inner_frame,
                fg_color="white",
                corner_radius=0,
                border_width=1,
                border_color="#E0E0E0",
                width=40,
                height=42
            )
            row_index_cell.pack(side="left", fill="y")
            row_index_cell.pack_propagate(False)
            
            lbl_tri = ctk.CTkLabel(
                row_index_cell,
                text="▲",
                font=("Segoe UI", 10, "bold"),
                text_color="#007B43",
                anchor="center"
            )
            lbl_tri.pack(expand=True)

            self.header_widgets = []
            for name, icon in header_info:
                cell = ctk.CTkFrame(
                    self.header_inner_frame,
                    fg_color="white",
                    corner_radius=0,
                    border_width=1,
                    border_color="#E0E0E0",
                    height=42
                )
                cell.pack(side="left", fill="y")
                cell.pack_propagate(False)
                
                content_frame = ctk.CTkFrame(cell, fg_color="transparent")
                content_frame.pack(expand=True)
                
                ctk.CTkLabel(
                    content_frame,
                    text=icon,
                    font=("Segoe UI", 11),
                    text_color="#007B43"
                ).pack(side="left", padx=(0, 4))
                
                lbl = ctk.CTkLabel(
                    content_frame,
                    text=name,
                    font=("Segoe UI", 11, "bold"),
                    text_color="#1A1A1A", # BLACK text color
                    anchor="center"
                )
                lbl.pack(side="left")
                self.header_widgets.append(cell)

            # Green border line under headers
            self.border_line = ctk.CTkFrame(container, fg_color="#007B43", height=2, corner_radius=0)
            self.border_line.grid(row=1, column=0, sticky="ew")

            # Table sheet frame
            self.table_sheet_frame = ctk.CTkFrame(container, fg_color="white", corner_radius=0)
            self.table_sheet_frame.grid(row=2, column=0, sticky="nsew")
            self.table_sheet_frame.grid_rowconfigure(0, weight=1)
            self.table_sheet_frame.grid_columnconfigure(0, weight=1)

            self.sheet = SheetClass(
                self.table_sheet_frame,
                headers=cols,
                data=data,
                show_header=False,         # Hide native header
                show_row_index=True,       # Show row numbers index
                row_index_width=40,        # Width matches spacer (40px)
                show_x_scrollbar=True,
                show_y_scrollbar=True
            )
            self.sheet.grid(row=0, column=0, sticky="nsew")

            # Style tksheet cells to match
            try:
                self.sheet.set_options(
                    grid_fg="#E0E0E0",
                    table_bg="white",
                    table_fg="#202124",
                    frame_bg="white",
                    select_bg="#E8F5EE",
                    select_fg="#007B43",
                    font=("Segoe UI", 10, "normal")
                )
            except Exception:
                pass

            try:
                self.sheet.enable_bindings(("single_select", "row_select", "arrowkeys", "copy", "select_all", "right_click_popup_menu"))
            except Exception:
                pass

            # double click
            try:
                self.sheet.bind("<Double-1>", lambda e: self._on_double_click_tksheet(e))
            except Exception:
                pass

            # Sync horizontal scrolling to custom header
            try:
                orig_xscroll = self.sheet.MT.cget("xscrollcommand")
                def sync_scroll(first, last):
                    if orig_xscroll:
                        try:
                            self.sheet.tk.eval(f"{orig_xscroll} {first} {last}")
                        except Exception:
                            pass
                    try:
                        self.header_canvas.xview_moveto(first)
                    except Exception:
                        pass
                self.sheet.MT.configure(xscrollcommand=sync_scroll)
            except Exception as e:
                print("Failed to sync Operator Master header scrollbar:", e)

            # resize handler: adapt columns proportionally with minimums
            def resize_sheet(event=None):
                try:
                    total_w = container.winfo_width() or 1
                    # Deduct spacing for the vertical scrollbar (16px) and row index column (40px)
                    avail = max(700, total_w - 16 - 40)

                    id_w = int(avail * 0.15)
                    name_w = int(avail * 0.25)
                    desc_w = int(avail * 0.45)
                    phone_w = max(80, avail - (id_w + name_w + desc_w))

                    # set widths
                    widths = [id_w, name_w, desc_w, phone_w]
                    for c in range(len(widths)):
                        try:
                            self.sheet.column_width(column=c, width=widths[c], only_set_if_too_small=False)
                        except TypeError:
                            self.sheet.column_width(c, widths[c])
                        
                        # Apply width to header cells
                        if c < len(self.header_widgets):
                            self.header_widgets[c].configure(width=widths[c])
                    
                    try:
                        self.sheet.refresh()
                    except Exception:
                        pass
                    try:
                        self.header_inner_frame.update_idletasks()
                        self.header_canvas.configure(scrollregion=self.header_canvas.bbox("all"))
                    except Exception:
                        pass
                except Exception as e:
                    print("Error resizing Operator Master sheet:", e)

            container.bind("<Configure>", resize_sheet)
            self.resize_table = resize_sheet

            # schedule an initial resize
            try:
                self.after(80, resize_sheet)
            except Exception:
                pass

            return
        except Exception as e:
            print("Failed to init tksheet for Operator Master, fallback to Treeview:", e)
            self.use_tksheet = False
            self.sheet = None

        # -----------------------
        # Treeview fallback
        # -----------------------
        cols_ids = ("id", "name", "description", "phone")
        self.tree = ttk.Treeview(container, columns=cols_ids, show="headings", height=18)
        self.tree.heading("id", text="Operator ID"); self.tree.column("id", width=120, anchor="center")
        self.tree.heading("name", text="Name"); self.tree.column("name", width=180, anchor="w")
        self.tree.heading("description", text="Description"); self.tree.column("description", width=420, anchor="w")
        self.tree.heading("phone", text="Phone"); self.tree.column("phone", width=140, anchor="center")

        # Grid layout so scrollbars don't eat width
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb = ttk.Scrollbar(container, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(container, orient="horizontal", command=self.tree.xview)
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.bind("<Double-1>", lambda e: self._on_tree_double_click(e))

        # resize handler for treeview
        def resize_tree(event=None):
            try:
                total_w = container.winfo_width() or 1
                min_total = 900
                avail = max(min_total, total_w - 20)

                id_w = int(avail * 0.12)
                name_w = int(avail * 0.22)
                desc_w = int(avail * 0.50)
                phone_w = max(100, avail - (id_w + name_w + desc_w))

                self.tree.column("id", width=id_w)
                self.tree.column("name", width=name_w)
                self.tree.column("description", width=desc_w)
                self.tree.column("phone", width=phone_w)
            except Exception:
                pass

        container.bind("<Configure>", resize_tree)
        self.resize_table = resize_tree
        try:
            self.after(80, resize_tree)
        except Exception:
            pass

    # -----------------------
    # Data load/save
    # -----------------------
    def _load_operators(self):
        try:
            conn = sqlite3.connect(resource_path("employee_master.db"))
            cursor = conn.cursor()
            # Ignore reserved IDs if they exist in the DB, as we handle them explicitly
            cursor.execute("SELECT id, name, description, phone FROM employee_master WHERE id NOT IN ('999', '000', '335', 999, 0, 335)")
            rows = cursor.fetchall()
            conn.close()
            self.operators = [
                {"id": "999", "name": "Guest", "description": "Default Guest User", "phone": ""}
            ] + [
                {"id": r[0], "name": r[1], "description": r[2] or "", "phone": r[3] or ""}
                for r in rows
            ]
        except Exception as e:
            messagebox.showerror("Load Error", f"Failed to load operators from DB: {e}")
            self.operators = [{"id": "999", "name": "Guest", "description": "Default Guest User", "phone": ""}]

    # -----------------------
    # Table helpers
    # -----------------------
    def refresh_table(self):
        """Populate the table from self.operators and trigger a resize."""
        if getattr(self, "use_tksheet", False) and getattr(self, "sheet", None):
            try:
                data = [[op.get("id", ""), op.get("name", ""), op.get("description", ""), op.get("phone", "")] for op in self.operators]
                # try to set sheet data with safe API name variants
                try:
                    self.sheet.set_sheet_data(data)
                except Exception:
                    try:
                        self.sheet.set_sheet_data(data, reset_index=True)
                    except Exception:
                        pass
            except Exception:
                pass

            # schedule resize and refresh
            def _delayed():
                try:
                    self.resize_table()
                    try:
                        self.sheet.highlight_rows(rows=[0], bg="#F5F5F5", fg="#9E9E9E")
                        self.sheet.readonly_rows(rows=[0])
                        self.sheet.refresh()
                    except Exception:
                        pass
                except Exception:
                    pass

            try:
                self.after_idle(_delayed)
            except Exception:
                self.after(100, _delayed)

        else:
            try:
                # repopulate treeview
                self.tree.tag_configure("guest", background="#F5F5F5", foreground="#9E9E9E")
                for r in self.tree.get_children():
                    self.tree.delete(r)
                for op in self.operators:
                    tags = ("guest",) if str(op.get("id")) == "999" else ()
                    self.tree.insert("", "end", values=(op.get("id", ""), op.get("name", ""), op.get("description", ""), op.get("phone", "")), tags=tags)
            except Exception:
                pass

            try:
                self.after(80, self.resize_table)
            except Exception:
                pass

    def _get_selected_index(self):
        """Return selected row index from tksheet or treeview."""
        try:
            if getattr(self, "use_tksheet", False) and getattr(self, "sheet", None):
                try:
                    cur = self.sheet.get_currently_selected()
                    if cur and isinstance(cur, tuple):
                        row = cur[0]
                        if isinstance(row, int) and row >= 0:
                            return row
                except Exception:
                    pass
                try:
                    rows = self.sheet.get_selected_rows()
                    if rows and isinstance(rows, list) and len(rows) > 0 and rows[0] >= 0:
                        return rows[0]
                except Exception:
                    pass
                return None
            else:
                sel = self.tree.selection()
                if not sel:
                    return None
                vals = self.tree.item(sel[0], "values")
                op_id = str(vals[0])
                for i, op in enumerate(self.operators):
                    if str(op.get("id", "")) == op_id:
                        return i
        except Exception:
            return None

    # -----------------------
    # Add / Update / Delete
    # -----------------------
    def _on_action_clicked(self):
        oid = self.id_entry.get().strip()
        name = self.name_entry.get().strip()
        desc = self.desc_text.get("1.0", "end").strip()
        phone = self.phone_entry.get().strip()

        if not oid or not name:
            messagebox.showwarning("Validation", "Operator ID and Name are required.")
            return

        if str(oid) in ["000", "0", "335", "999"]:
            messagebox.showinfo("Reserved ID", f"ID {oid} is reserved by Cherry.")
            return

        if self.edit_index is None:
            # Add
            if any(str(op.get("id", "")) == str(oid) for op in self.operators):
                messagebox.showwarning("Validation", "Operator ID already exists.")
                return
            try:
                conn = sqlite3.connect(resource_path("employee_master.db"))
                cursor = conn.cursor()
                cursor.execute("INSERT INTO employee_master (id, name, description, phone) VALUES (?, ?, ?, ?)",
                               (oid, name, desc, phone))
                conn.commit()
                conn.close()
            except Exception as e:
                messagebox.showerror("DB Error", f"Failed to add operator: {e}")
                return
            
            self._load_operators()
            new_idx = next((i for i, op in enumerate(self.operators) if str(op["id"]) == oid), len(self.operators)-1)
            self._post_mutation_actions(new_index=new_idx, action="added", id=oid)
        else:
            # Update
            idx = self.edit_index
            old_op = self.operators[idx]
            old_id = str(old_op.get("id", ""))
            
            if old_id == "999":
                messagebox.showinfo("Action Denied", "The Guest operator cannot be edited.")
                return
                
            try:
                conn = sqlite3.connect(resource_path("employee_master.db"))
                cursor = conn.cursor()
                cursor.execute("UPDATE employee_master SET id=?, name=?, description=?, phone=? WHERE id=?",
                               (oid, name, desc, phone, old_id))
                conn.commit()
                conn.close()
            except Exception as e:
                messagebox.showerror("Update Error", f"Failed to update DB entry: {e}")
                return
                
            self._load_operators()
            new_idx = next((i for i, op in enumerate(self.operators) if str(op["id"]) == oid), idx)
            self._post_mutation_actions(new_index=new_idx, action="updated", id=oid)
            
            # restore mode
            self.edit_index = None
            self.action_btn.configure(text="➕  Add Operator")

    def _post_mutation_actions(self, new_index, action: str, id: str):
        try:
            self.refresh_table()
            # blink/fade only if tksheet is present
            if getattr(self, "use_tksheet", False) and getattr(self, "sheet", None):
                try:
                    self._fade_row(new_index)
                except Exception:
                    pass
            # clear inputs and focus
            self._clear_inputs()
            try:
                self.id_entry.focus_set()
            except Exception:
                pass
            try:
                if self.app and hasattr(self.app, "status_label"):
                    self.app.status_label.configure(text=f"Operator {action}: {id} ✅")
            except Exception:
                pass
            try:
                self.after(80, self.resize_table)
            except Exception:
                pass
        except Exception:
            pass

    def delete_selected(self):
        idx = self._get_selected_index()
        if idx is None:
            messagebox.showinfo("Delete Operator", "Please select a row to delete.")
            return
        op = self.operators[idx]
        if str(op.get('id', '')) == "999":
            messagebox.showinfo("Action Denied", "The Guest operator cannot be deleted.")
            return
        if not messagebox.askyesno("Confirm Delete", f"Delete operator {op.get('id')} - {op.get('name')} ?"):
            return
        try:
            conn = sqlite3.connect(resource_path("employee_master.db"))
            cursor = conn.cursor()
            cursor.execute("DELETE FROM employee_master WHERE id=?", (op.get('id'),))
            conn.commit()
            conn.close()
            
            self._load_operators()
            self.refresh_table()
            try:
                if self.app and hasattr(self.app, "status_label"):
                    self.app.status_label.configure(text=f"Deleted operator {op.get('id')} ✅")
            except Exception:
                pass
            try:
                self.after(80, self.resize_table)
            except Exception:
                pass
        except Exception as e:
            messagebox.showerror("Delete Failed", f"Could not delete: {e}")

    # -----------------------
    # Double-click handlers
    # -----------------------
    def _on_double_click_tksheet(self, event):
        try:
            r = None
            try:
                cur = self.sheet.get_currently_selected()
                if cur and isinstance(cur, tuple):
                    r = cur[0]
            except Exception:
                pass
            if r is None:
                try:
                    rows = self.sheet.get_selected_rows()
                    if rows and len(rows) > 0:
                        r = rows[0]
                except Exception:
                    pass
            if r is None:
                return
            if str(self.operators[r].get("id")) == "999":
                messagebox.showinfo("Restricted", "The Guest operator cannot be edited.")
                return
            self._open_editor_dialog(mode="edit", index=r)
        except Exception:
            pass

    def _on_tree_double_click(self, event):
        try:
            sel = self.tree.selection()
            if not sel:
                return
            vals = self.tree.item(sel[0], "values")
            op_id = str(vals[0])
            for i, op in enumerate(self.operators):
                if str(op.get("id", "")) == op_id:
                    if str(op_id) == "999":
                        messagebox.showinfo("Restricted", "The Guest operator cannot be edited.")
                        return
                    self._open_editor_dialog(mode="edit", index=i)
                    break
        except Exception:
            pass

    # -----------------------
    # Visual helpers (tksheet-only animations)
    # -----------------------
    def _fade_row(self, row_index):
        if not (self.use_tksheet and self.sheet):
            return
        import math
        start_r, start_g, start_b = (255, 245, 120)
        end_r, end_g, end_b = (255, 255, 255)
        steps = 10
        duration = 20

        def interpolate(t):
            r = int(start_r + (end_r - start_r) * t)
            g = int(start_g + (end_g - start_g) * t)
            b = int(start_b + (end_b - start_b) * t)
            return f"#{r:02x}{g:02x}{b:02x}"

        def animate(frame=0):
            try:
                if frame <= steps:
                    color = interpolate(frame / steps)
                    self.sheet.highlight_rows(rows=[row_index], bg=color)
                    self.after(duration, lambda: animate(frame + 1))
                else:
                    self.sheet.highlight_rows(rows=[row_index], bg=None)
            except Exception:
                pass

        animate(0)
        try:
            self.sheet.see(row_index, 0)
        except Exception:
            pass

    # -----------------------
    # Editor dialog (modal)
    # -----------------------
    def _open_editor_dialog(self, mode="add", index=None):
        win = ctk.CTkToplevel(self)
        win.title("Add Operator" if mode == "add" else "Edit Operator")
        win.geometry("520x360")
        win.grab_set()
        win.focus_force()

        frame = ctk.CTkFrame(win)
        frame.pack(fill="both", expand=True, padx=12, pady=12)

        ctk.CTkLabel(frame, text="Operator ID:", font=("Segoe UI", 11)).grid(row=0, column=0, sticky="e", padx=6, pady=8)
        id_e = ctk.CTkEntry(frame, width=320); id_e.grid(row=0, column=1, sticky="w", padx=6, pady=8)

        ctk.CTkLabel(frame, text="Name:", font=("Segoe UI", 11)).grid(row=1, column=0, sticky="e", padx=6, pady=8)
        name_e = ctk.CTkEntry(frame, width=320); name_e.grid(row=1, column=1, sticky="w", padx=6, pady=8)

        ctk.CTkLabel(frame, text="Description:", font=("Segoe UI", 11)).grid(row=2, column=0, sticky="ne", padx=6, pady=8)
        desc_t = tk.Text(frame, width=36, height=6, bg="white", fg="black"); desc_t.grid(row=2, column=1, sticky="w", padx=6, pady=8)

        ctk.CTkLabel(frame, text="Phone (optional):", font=("Segoe UI", 11)).grid(row=3, column=0, sticky="e", padx=6, pady=8)
        phone_e = ctk.CTkEntry(frame, width=320); phone_e.grid(row=3, column=1, sticky="w", padx=6, pady=8)

        if mode == "edit" and index is not None:
            op = self.operators[index]
            id_e.insert(0, str(op.get("id", "")))
            name_e.insert(0, op.get("name", ""))
            desc_t.insert("1.0", op.get("description", ""))
            phone_e.insert(0, op.get("phone", ""))

        btnf = ctk.CTkFrame(frame, fg_color="transparent")
        btnf.grid(row=4, column=0, columnspan=2, pady=(8, 0))

        def on_save_clicked():
            oid = id_e.get().strip()
            name = name_e.get().strip()
            desc = desc_t.get("1.0", "end").strip()
            phone = phone_e.get().strip()

            if not oid or not name:
                messagebox.showwarning("Validation", "Operator ID and Name are required.")
                return

            if mode == "edit":
                if not messagebox.askyesno("Confirm Update", f"Apply changes to operator {oid}?"):
                    return
                old_id = self.operators[index].get("id")
                try:
                    conn = sqlite3.connect(resource_path("employee_master.db"))
                    cursor = conn.cursor()
                    cursor.execute("UPDATE employee_master SET id=?, name=?, description=?, phone=? WHERE id=?",
                                   (oid, name, desc, phone, old_id))
                    conn.commit()
                    conn.close()
                except Exception as e:
                    messagebox.showerror("Update Error", f"Failed to update DB entry: {e}")
                    return
                self.operators[index] = {"id": oid, "name": name, "description": desc, "phone": phone}
            else:
                if str(oid) in ["000", "0", "335", "999"]:
                    messagebox.showinfo("Reserved ID", f"ID {oid} is reserved by Cherry.")
                    return
                if any(str(op.get("id","")) == str(oid) for op in self.operators):
                    messagebox.showwarning("Validation", "Operator ID already exists.")
                    return
                try:
                    conn = sqlite3.connect(resource_path("employee_master.db"))
                    cursor = conn.cursor()
                    cursor.execute("INSERT INTO employee_master (id, name, description, phone) VALUES (?, ?, ?, ?)",
                                   (oid, name, desc, phone))
                    conn.commit()
                    conn.close()
                except Exception as e:
                    messagebox.showerror("DB Error", f"Failed to add operator: {e}")
                    return
                self.operators.append({"id": oid, "name": name, "description": desc, "phone": phone})

            self.refresh_table()
            # self.save_operators() # This was removed as DB operations handle persistence
            win.destroy()

        ModernButton(btnf, text="Save", fg_color="#43A047", command=on_save_clicked).pack(side="left", padx=8)
        ModernButton(btnf, text="Cancel", fg_color="#E53935", command=win.destroy).pack(side="left", padx=8)

    # -----------------------
    # Utilities
    # -----------------------
    def _clear_inputs(self):
        try:
            self.id_entry.delete(0, "end")
            self.name_entry.delete(0, "end")
            self.phone_entry.delete(0, "end")
            self.desc_text.delete("1.0", "end")
            self.edit_index = None
            self.action_btn.configure(text="➕  Add Operator")
        except Exception:
            pass

    def _safe_resize(self):
        try:
            if hasattr(self, "resize_table"):
                self.resize_table()
        except Exception:
            pass

    def go_back(self, event=None):
        try:
            self.app.load_settings_page()
        except:
            try: self.app.load_settings()
            except: pass

    def highlight_entry(self, entry, active):
        try:
            if active:
                entry.configure(border_width=2, border_color="#007B43")
                self.animate_glow(entry, 0)
            else:
                entry.configure(border_width=1, border_color="#007B43")
        except Exception:
            pass

    def animate_glow(self, entry, step):
        try:
            if self.focus_get() != entry:
                return
        except Exception:
            return
        wave_colors = ["#007B43", "#005C32", "#43A047", "#388E3C", "#005C32", "#007B43"]
        entry.configure(border_color=wave_colors[step % len(wave_colors)])
        self.after(250, lambda: self.animate_glow(entry, step + 1))

    def destroy(self):
        super().destroy()


# ==========================================================
# Customer Master Page — Full Page (permanent inputs + table)
# ==========================================================

class CustomerMasterPage(ctk.CTkFrame):
    """
    Customer Master — manages customers with:
     - Permanent Add form (Code, Customer Name, Description (single-line), Email, Phone)
     - Add / Update (single dynamic button)
     - Save All button
     - Delete Selected button
     - Search filter
     - Table implemented with tksheet (read-only) with Treeview fallback
     - Double-click row -> loads row into inputs (no large popup; small popup for edit available)
     - JSON persistence (customers.json)
    """
    FNAME = "customers.json"

    def __init__(self, parent, app):
        super().__init__(parent, fg_color="white")
        self.app = app

        # data model: list of dicts { "code","name","description","email","phone" }
        self.customers = []
        # UI state
        self.edit_index = None  # None => Add mode, int => Edit mode

        # sheet/tree placeholders
        self.use_tksheet = False
        self.sheet = None
        self.tree = None
        self.resize_sheet = lambda event=None: None

        # search variable
        self.search_var = tk.StringVar()

        # load existing data
        self._load_customers()

        # build UI
        self.build_ui()

        # initial populate
        self.refresh_table()
        try:
            self.after(60, self.resize_sheet)
        except Exception:
            pass

    # -------------------------
    # JSON helpers
    # -------------------------
    def _get_data_path(self):
        """Return path for customers.json located in cwd (consistent with other pages)."""
        return os.path.join(os.getcwd(), self.FNAME)

    def _load_customers(self):
        """Load customers list from SQLite DB."""
        try:
            conn = sqlite3.connect(resource_path("customers.db"))
            cursor = conn.cursor()
            cursor.execute("SELECT code, name, description, email, phone FROM customers")
            rows = cursor.fetchall()
            conn.close()
            self.customers = [
                {"code": r[0], "name": r[1], "description": r[2] or "", "email": r[3] or "", "phone": r[4] or ""}
                for r in rows
            ]
        except Exception as e:
            messagebox.showerror("Load Error", f"Failed to load customers from DB: {e}")
            self.customers = []

    # -------------------------
    # UI
    # -------------------------
    def build_ui(self):
        # ====== HEADER ======
        header = ctk.CTkFrame(self, fg_color="transparent")
        header.pack(fill="x", side="top", pady=(15, 10), padx=20)

        # Back Button in Green
        ModernButton(
            header,
            text="← Back",
            width=80,
            height=32,
            font=("Segoe UI", 11, "bold"),
            fg_color="#007B43",
            hover_color="#005C32",
            text_color="white",
            corner_radius=6,
            command=self.go_back
        ).pack(side="left")
        
        # Branded Title: "Customer" (Green) + "Master" (Red) next to User silhouette
        title_container = ctk.CTkFrame(header, fg_color="transparent")
        title_container.pack(side="left", padx=20)
        
        lbl_user = ctk.CTkLabel(
            title_container,
            text="👤\ufe0e",
            font=("Segoe UI", 18, "bold"),
            text_color="#007B43"
        )
        lbl_user.pack(side="left")
        
        lbl_customer = ctk.CTkLabel(
            title_container,
            text=" Customer",
            font=("Segoe UI", 18, "bold"),
            text_color="#007B43"
        )
        lbl_customer.pack(side="left")
        
        lbl_master = ctk.CTkLabel(
            title_container,
            text=" Master",
            font=("Segoe UI", 18, "bold"),
            text_color="#B0050E"
        )
        lbl_master.pack(side="left")

        # Company Logo on the right
        try:
            logo_path = resource_path("settings/cherry_full_logo.png")
            pil_img = Image.open(logo_path)
            pil_img = pil_img.resize((144, 40), Image.Resampling.LANCZOS)
            pil_img = add_corners(pil_img, 10)
            self.logo_img = ctk.CTkImage(pil_img, size=(144, 40))
            logo_lbl = ctk.CTkLabel(header, text="", image=self.logo_img)
            logo_lbl.pack(side="right", padx=(10, 20))
        except Exception as e:
            print(f"Error loading logo in CustomerMasterPage: {e}")

        # ====== FORM CONTAINER ======
        add_frame = ModernCardFrame(self)
        add_frame.pack(fill="x", padx=20, pady=(10, 6))

        LABEL_WIDTH = 90
        ENTRY_WIDTH = 220

        # ROW 0: Code + Email
        ctk.CTkLabel(add_frame, text="Code:", font=("Segoe UI", 11, "bold"),
                     width=LABEL_WIDTH, anchor="w", text_color="#202124").grid(row=0, column=0, sticky="w", pady=6, padx=(10, 0))
        self.code_e = ctk.CTkEntry(add_frame, width=ENTRY_WIDTH, height=30, corner_radius=6, border_color="#007B43")
        self.code_e.grid(row=0, column=1, sticky="w", padx=(10, 30))

        ctk.CTkLabel(add_frame, text="Email:", font=("Segoe UI", 11, "bold"),
                     width=LABEL_WIDTH, anchor="e", text_color="#202124").grid(row=0, column=2, sticky="e", padx=(0, 10), pady=6)
        self.email_e = ctk.CTkEntry(add_frame, width=ENTRY_WIDTH, height=30, corner_radius=6, border_color="#007B43")
        self.email_e.grid(row=0, column=3, sticky="w", padx=(10, 0))

        # ROW 1: Customer Name + Phone
        ctk.CTkLabel(add_frame, text="Customer Name:", font=("Segoe UI", 11, "bold"),
                     width=LABEL_WIDTH, anchor="w", text_color="#202124").grid(row=1, column=0, sticky="w", pady=6, padx=(10, 0))
        self.name_e = ctk.CTkEntry(add_frame, width=ENTRY_WIDTH, height=30, corner_radius=6, border_color="#007B43")
        self.name_e.grid(row=1, column=1, sticky="w", padx=(10, 30))

        ctk.CTkLabel(add_frame, text="Phone:", font=("Segoe UI", 11, "bold"),
                     width=LABEL_WIDTH, anchor="e", text_color="#202124").grid(row=1, column=2, sticky="e", padx=(0, 10), pady=6)
        self.phone_e = ctk.CTkEntry(add_frame, width=ENTRY_WIDTH, height=30, corner_radius=6, border_color="#007B43")
        self.phone_e.grid(row=1, column=3, sticky="w", padx=(10, 0))

        # ROW 2: Description (single-line long entry spanning right columns)
        ctk.CTkLabel(add_frame, text="Description:", font=("Segoe UI", 11, "bold"),
                     width=LABEL_WIDTH, anchor="w", text_color="#202124").grid(row=2, column=0, sticky="w", pady=(6, 6), padx=(10, 0))
        self.desc_e = ctk.CTkEntry(add_frame, width=(ENTRY_WIDTH * 2 + 255), height=30, corner_radius=6, border_color="#007B43")
        self.desc_e.grid(row=2, column=1, columnspan=3, sticky="w", padx=(10, 0), pady=(6, 12))

        # Bind entries highlighting
        for entry in [self.code_e, self.name_e, self.desc_e, self.email_e, self.phone_e]:
            entry.bind("<FocusIn>", lambda e, ent=entry: self.highlight_entry(ent, True))
            entry.bind("<FocusOut>", lambda e, ent=entry: self.highlight_entry(ent, False))

        # ROW 3: Buttons (left), Search (right)
        btn_frame = ctk.CTkFrame(add_frame, fg_color="transparent")
        btn_frame.grid(row=3, column=0, columnspan=2, sticky="w", pady=(0, 10), padx=(10, 0))

        # Add / Update dynamic action button
        self.action_btn = ModernButton(
            btn_frame,
            text="➕ Add Customer",
            fg_color="#007B43",
            hover_color="#005C32",
            height=32,
            width=140,
            corner_radius=6,
            font=("Segoe UI", 11, "bold"),
            command=self._on_action_clicked
        )
        self.action_btn.pack(side="left", padx=(0, 10))

        # Delete selected (left)
        del_btn = ModernButton(
            btn_frame,
            text="🗑 Delete Selected",
            fg_color="#B0050E",
            hover_color="#90040B",
            height=32,
            width=150,
            corner_radius=6,
            font=("Segoe UI", 11, "bold"),
            command=self._delete_selected
        )
        del_btn.pack(side="left")

        # Search area
        search_frame = ctk.CTkFrame(add_frame, fg_color="transparent")
        search_frame.grid(row=3, column=2, columnspan=2, sticky="e", padx=(0, 10), pady=(0, 10))
        ctk.CTkLabel(search_frame, text="Search:", font=("Segoe UI", 11, "bold"), text_color="#202124").pack(side="left", padx=(6, 4))
        search_e = ctk.CTkEntry(search_frame, textvariable=self.search_var, width=250, height=30, corner_radius=6, border_color="#007B43")
        search_e.pack(side="left")
        search_e.bind("<KeyRelease>", lambda e: self.refresh_table())
        search_e.bind("<FocusIn>", lambda e, ent=search_e: self.highlight_entry(ent, True))
        search_e.bind("<FocusOut>", lambda e, ent=search_e: self.highlight_entry(ent, False))

        # ====== TABLE ======
        table_card = ModernCardFrame(self)
        table_card.pack(fill="both", expand=True, padx=20, pady=(6, 12))
        table_card.grid_rowconfigure(0, weight=1)
        table_card.grid_columnconfigure(0, weight=1)

        # Try to create tksheet
        try:
            from tksheet import Sheet
            SheetClass = Sheet
        except Exception:
            try:
                SheetClass = tksheet.Sheet
            except Exception:
                SheetClass = None

        self._col_widths = [120, 300, 420, 260, 160]  # Code, Name, Description, Email, Phone
        cols = ["Code", "Customer Name", "Description", "Email", "Phone"]
        header_info = [
            ("Code", "🔑\ufe0e"),
            ("Customer Name", "👥\ufe0e"),
            ("Description", "📝\ufe0e"),
            ("Email", "✉\ufe0e"),
            ("Phone", "📞\ufe0e")
        ]

        if SheetClass is not None:
            self.use_tksheet = True
            
            # --- Custom Horizontal Scroll-Synced Header Frame ---
            self.header_row_frame = ctk.CTkFrame(table_card, fg_color="white", height=42, corner_radius=0)
            self.header_row_frame.pack(fill="x", padx=10, pady=(10, 0))
            self.header_row_frame.pack_propagate(False)
            
            # Canvas inside header frame for scrolling
            self.header_canvas = tk.Canvas(self.header_row_frame, bg="white", highlightthickness=0, height=42)
            self.header_canvas.pack(side="left", fill="both", expand=True)
            
            # Frame inside canvas for custom header cells
            self.header_inner_frame = tk.Frame(self.header_canvas, bg="white")
            self.header_canvas.create_window(0, 0, window=self.header_inner_frame, anchor="nw")
            
            # Vertical scrollbar spacer on the right of header
            self.header_scrollbar_spacer = ctk.CTkFrame(
                self.header_row_frame,
                fg_color="white",
                corner_radius=0,
                width=16,
                height=42
            )
            self.header_scrollbar_spacer.pack(side="right", fill="y")
            
            # Row index spacer cell on the far left displaying green up-triangle sorting indicator
            row_index_cell = ctk.CTkFrame(
                self.header_inner_frame,
                fg_color="white",
                corner_radius=0,
                border_width=1,
                border_color="#E0E0E0",
                width=40,
                height=42
            )
            row_index_cell.pack(side="left", fill="y")
            row_index_cell.pack_propagate(False)
            
            lbl_tri = ctk.CTkLabel(
                row_index_cell,
                text="▲",
                font=("Segoe UI", 10, "bold"),
                text_color="#007B43",
                anchor="center"
            )
            lbl_tri.pack(expand=True)

            self.header_widgets = []
            for name, icon in header_info:
                cell = ctk.CTkFrame(
                    self.header_inner_frame,
                    fg_color="white",
                    corner_radius=0,
                    border_width=1,
                    border_color="#E0E0E0",
                    height=42
                )
                cell.pack(side="left", fill="y")
                cell.pack_propagate(False)
                
                content_frame = ctk.CTkFrame(cell, fg_color="transparent")
                content_frame.pack(expand=True)
                
                ctk.CTkLabel(
                    content_frame,
                    text=icon,
                    font=("Segoe UI", 11),
                    text_color="#007B43"
                ).pack(side="left", padx=(0, 4))
                
                lbl = ctk.CTkLabel(
                    content_frame,
                    text=name,
                    font=("Segoe UI", 11, "bold"),
                    text_color="#1A1A1A", # BLACK text color
                    anchor="center"
                )
                lbl.pack(side="left")
                self.header_widgets.append(cell)

            # Green border line under headers
            self.border_line = ctk.CTkFrame(table_card, fg_color="#007B43", height=2, corner_radius=0)
            self.border_line.pack(fill="x", padx=10, pady=(0, 0))

            # Table sheet frame
            self.table_sheet_frame = ctk.CTkFrame(table_card, fg_color="white", corner_radius=0)
            self.table_sheet_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))
            self.table_sheet_frame.grid_rowconfigure(0, weight=1)
            self.table_sheet_frame.grid_columnconfigure(0, weight=1)

            data = [[c.get("code", ""), c.get("name", ""), c.get("description", ""), c.get("email", ""), c.get("phone", "")]
                    for c in self.customers]

            self.sheet = SheetClass(
                self.table_sheet_frame,
                headers=cols,
                data=data,
                show_header=False,        # Hide native header
                show_row_index=True,       # Show row numbers index
                row_index_width=40,        # Width matches spacer (40px)
                show_x_scrollbar=True,
                show_y_scrollbar=True
            )
            self.sheet.grid(row=0, column=0, sticky="nsew")

            # Style tksheet cells to match
            try:
                self.sheet.set_options(
                    grid_fg="#E0E0E0",
                    table_bg="white",
                    table_fg="#202124",
                    frame_bg="white",
                    select_bg="#E8F5EE",
                    select_fg="#007B43",
                    font=("Segoe UI", 10, "normal"),
                    auto_resize_columns=False,
                    column_width_resize=False,
                    headers_resizable=False
                )
            except Exception:
                pass

            try:
                self.sheet.enable_bindings(("single_select", "row_select", "arrowkeys", "copy", "select_all", "right_click_popup_menu"))
            except Exception:
                pass

            # set manual column widths
            try:
                for i, w in enumerate(self._col_widths):
                    self.sheet.column_width(i, w)
            except Exception:
                pass

            # Double-click loads editor
            def _on_sheet_double_click(event):
                idx = self._get_selected_index()
                if idx is None:
                    return
                try:
                    self._open_editor_dialog(mode="edit", index=idx)
                except Exception:
                    pass

            try:
                self.sheet.bind("<Double-1>", _on_sheet_double_click)
            except Exception:
                try:
                    self.sheet.bind("<Double-Button-1>", _on_sheet_double_click)
                except Exception:
                    pass

            # Sync horizontal scrolling to custom header
            try:
                orig_xscroll = self.sheet.MT.cget("xscrollcommand")
                def sync_scroll(first, last):
                    if orig_xscroll:
                        try:
                            self.sheet.tk.eval(f"{orig_xscroll} {first} {last}")
                        except Exception:
                            pass
                    try:
                        self.header_canvas.xview_moveto(first)
                    except Exception:
                        pass
                self.sheet.MT.configure(xscrollcommand=sync_scroll)
            except Exception as e:
                print("Failed to sync Customer Master header scrollbar:", e)

            # Resize columns resizes both sheet and header cells
            def resize_sheet(event=None):
                try:
                    total_width = self.table_sheet_frame.winfo_width() or 1
                    if total_width < 120:
                        self.after(60, self.resize_sheet)
                        return
                    col_count = len(cols)
                    # Deduct spacing for the vertical scrollbar (16px) and row index column (40px)
                    available = max(200, total_width - 16 - 40)
                    col_width = max(80, int(available / col_count))
                    for c in range(col_count):
                        try:
                            self.sheet.column_width(column=c, width=col_width, only_set_if_too_small=False)
                        except TypeError:
                            self.sheet.column_width(c, col_width)
                        
                        # Apply width to header cells
                        if c < len(self.header_widgets):
                            self.header_widgets[c].configure(width=col_width)
                    try:
                        self.sheet.refresh()
                    except Exception:
                        pass
                    try:
                        self.header_inner_frame.update_idletasks()
                        self.header_canvas.configure(scrollregion=self.header_canvas.bbox("all"))
                    except Exception:
                        pass
                except Exception:
                    pass

            self.resize_sheet = resize_sheet
            self.table_sheet_frame.bind("<Configure>", self.resize_sheet)

        else:
            # Treeview fallback
            self.use_tksheet = False
            cols = ("code", "name", "description", "email", "phone")
            
            style = ttk.Style()
            style.theme_use("default")
            style.configure(
                "Customer.Treeview",
                font=("Segoe UI", 11),
                rowheight=28,
                background="white",
                foreground="#212121",
                fieldbackground="white",
                borderwidth=0
            )
            style.configure(
                "Customer.Treeview.Heading",
                font=("Segoe UI", 11, "bold"),
                background="#205124",
                foreground="white",
                borderwidth=0,
                relief="flat"
            )
            style.map("Customer.Treeview.Heading", background=[("active", "#005C32")])
            style.layout("Customer.Treeview", [('Treeview.treearea', {'sticky': 'nswe'})])
            
            self.tree = ttk.Treeview(table_card, columns=cols, show="headings", height=16, style="Customer.Treeview")
            self.tree.heading("code", text="Code"); self.tree.column("code", width=120, anchor="center")
            self.tree.heading("name", text="Customer Name"); self.tree.column("name", width=300, anchor="w")
            self.tree.heading("description", text="Description"); self.tree.column("description", width=420, anchor="w")
            self.tree.heading("email", text="Email"); self.tree.column("email", width=260, anchor="w")
            self.tree.heading("phone", text="Phone"); self.tree.column("phone", width=140, anchor="center")

            vsb = ctk.CTkScrollbar(
                table_card,
                orientation="vertical",
                button_color="#007B43",
                button_hover_color="#005C32"
            )
            hsb = ctk.CTkScrollbar(
                table_card,
                orientation="horizontal",
                button_color="#007B43",
                button_hover_color="#005C32"
            )
            self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
            self.tree.pack(side="top", fill="both", expand=True)
            hsb.pack(side="bottom", fill="x")
            vsb.pack(side="right", fill="y")

            # bind double click
            self.tree.bind("<Double-1>", lambda e: self._on_tree_double_click(e))

            # resize logic for tree
            def resize_tree(event=None):
                try:
                    total_w = table_card.winfo_width() or 1
                    if total_w < 120:
                        self.after(60, resize_tree)
                        return
                    avail = max(200, total_w - 20)
                    code_w = int(avail * 0.12)
                    name_w = int(avail * 0.28)
                    desc_w = int(avail * 0.38)
                    email_w = int(avail * 0.16)
                    phone_w = max(80, avail - (code_w + name_w + desc_w + email_w))
                    self.tree.column("code", width=code_w)
                    self.tree.column("name", width=name_w)
                    self.tree.column("description", width=desc_w)
                    self.tree.column("email", width=email_w)
                    self.tree.column("phone", width=phone_w)
                except Exception:
                    pass

            self.resize_sheet = resize_tree
            table_card.bind("<Configure>", self.resize_sheet)

    # -------------------------
    # Utilities & animations
    # -------------------------
    def _get_selected_index(self):
        """Return selected row index from tksheet or treeview, else None."""
        try:
            if self.use_tksheet and self.sheet is not None:
                try:
                    cur = self.sheet.get_currently_selected()
                    if cur and isinstance(cur, tuple):
                        row = cur[0]
                        if isinstance(row, int) and row >= 0:
                            return row
                except Exception:
                    pass
                try:
                    rows = self.sheet.get_selected_rows()
                    if rows and isinstance(rows, list) and len(rows) > 0 and rows[0] >= 0:
                        return rows[0]
                except Exception:
                    pass
                return None
            else:
                sel = self.tree.selection()
                if not sel:
                    return None
                vals = self.tree.item(sel[0], "values")
                code = str(vals[0])
                for i, rec in enumerate(self.customers):
                    if str(rec.get("code", "")) == code:
                        return i
        except Exception:
            return None

    def _blink_new_row(self, row_index):
        """Blink newly added row once (tksheet only)."""
        if not (self.use_tksheet and self.sheet):
            return
        try:
            try:
                self.sheet.see(row_index, 0)
            except Exception:
                pass
            try:
                self.sheet.highlight_rows(rows=[row_index], bg="#FFF176")
            except Exception:
                pass

            def remove():
                try:
                    self.sheet.highlight_rows(rows=[row_index], bg=None)
                except Exception:
                    pass
            self.after(500, remove)
        except Exception:
            pass

    def _fade_row(self, row_index):
        """Fade highlight for a row (tksheet only)."""
        if not (self.use_tksheet and self.sheet):
            return
        try:
            start_r, start_g, start_b = (255, 245, 120)
            end_r, end_g, end_b = (255, 255, 255)
            steps = 14
            duration = 18

            def interpolate(t):
                r = int(start_r + (end_r - start_r) * t)
                g = int(start_g + (end_g - start_g) * t)
                b = int(start_b + (end_b - start_b) * t)
                return f"#{r:02x}{g:02x}{b:02x}"

            def animate(frame=0):
                try:
                    if frame <= steps:
                        color = interpolate(frame / steps)
                        self.sheet.highlight_rows(rows=[row_index], bg=color)
                        self.after(duration, lambda: animate(frame + 1))
                    else:
                        self.sheet.highlight_rows(rows=[row_index], bg=None)
                except Exception:
                    pass

            animate(0)
            try:
                self.sheet.see(row_index, 0)
            except Exception:
                pass
        except Exception:
            pass

    # -------------------------
    # Actions: Add / Update / Delete / Edit
    # -------------------------
    def _on_action_clicked(self):
        """Add or Update customer depending on edit_index."""
        code = self.code_e.get().strip()
        name = self.name_e.get().strip()
        desc = self.desc_e.get().strip()
        email = self.email_e.get().strip()
        phone = self.phone_e.get().strip()

        # validations
        if not code:
            messagebox.showwarning("Validation", "Code required.")
            return
        if not name:
            messagebox.showwarning("Validation", "Customer Name required.")
            return
        if email and ("@" not in email or "." not in email):
            messagebox.showwarning("Validation", "Invalid email.")
            return
        if phone and not phone.isdigit():
            messagebox.showwarning("Validation", "Phone must be digits.")
            return

        if self.edit_index is None:
            # Add mode: ensure unique code
            if any(str(c.get("code", "")) == code for c in self.customers):
                messagebox.showwarning("Validation", "Code already exists.")
                return
            try:
                conn = sqlite3.connect(resource_path("customers.db"))
                cursor = conn.cursor()
                cursor.execute("INSERT INTO customers (code, name, description, email, phone) VALUES (?, ?, ?, ?, ?)",
                               (code, name, desc, email, phone))
                conn.commit()
                conn.close()
            except Exception as e:
                messagebox.showerror("DB Error", f"Failed to add customer: {e}")
                return
                
            self._load_customers()
            self.refresh_table()
            new_index = next((i for i, c in enumerate(self.customers) if str(c["code"]) == code), len(self.customers) - 1)
            # animation
            self.after(120, lambda idx=new_index: self._fade_row(idx))
            self.after(150, lambda idx=new_index: self._blink_new_row(idx))
            # clear for next entry
            self._clear_inputs()
            try:
                self.code_e.focus_set()
            except Exception:
                pass
            try:
                self.app.status_label.configure(text=f"Customer added: {code} ✅")
            except Exception:
                pass
        else:
            # Update mode
            idx = self.edit_index
            old_code = self.customers[idx].get("code")
            try:
                conn = sqlite3.connect(resource_path("customers.db"))
                cursor = conn.cursor()
                cursor.execute("UPDATE customers SET code=?, name=?, description=?, email=?, phone=? WHERE code=?",
                               (code, name, desc, email, phone, old_code))
                conn.commit()
                conn.close()
            except Exception as e:
                messagebox.showerror("Update Error", f"Failed to update entry: {e}")
                return

            self._load_customers()
            self.refresh_table()

            # restore selection to updated row
            def _restore():
                try:
                    if self.use_tksheet and self.sheet is not None:
                        try:
                            self.sheet.select_row(idx)
                        except Exception:
                            try:
                                self.sheet.set_selected_rows([idx])
                            except Exception:
                                pass
                    else:
                        for iid in self.tree.get_children():
                            vals = self.tree.item(iid, "values")
                            if str(vals[0]) == str(self.customers[idx].get("code", "")):
                                self.tree.selection_set(iid)
                                self.tree.see(iid)
                                break
                except Exception:
                    pass

            try:
                self.after(80, _restore)
            except Exception:
                _restore()

            self._clear_inputs()
            self.edit_index = None
            self.action_btn.configure(text="➕ Add Customer")
            try:
                self.app.status_label.configure(text=f"Customer updated: {code} ✅")
            except Exception:
                pass

    def _delete_selected(self):
        """Delete selected row (single selection)."""
        idx = self._get_selected_index()
        if idx is None:
            messagebox.showwarning("No Selection", "Select a row to delete.")
            return
        rec = self.customers[idx]
        if not messagebox.askyesno("Confirm Delete", f"Delete customer '{rec.get('name','')}' (Code: {rec.get('code','')})?"):
            return
        try:
            conn = sqlite3.connect(resource_path("customers.db"))
            cursor = conn.cursor()
            cursor.execute("DELETE FROM customers WHERE code=?", (rec.get('code'),))
            conn.commit()
            conn.close()
            
            self._load_customers()
            self.refresh_table()
            try:
                self.app.status_label.configure(text=f"Deleted customer {rec.get('code')} ✅")
            except Exception:
                pass

            # restore selection to next available row
            def _restore_after_delete():
                try:
                    if self.use_tksheet and self.sheet is not None:
                        n = len(self.customers)
                        sel_idx = max(0, min(idx, n - 1))
                        if n > 0:
                            try:
                                self.sheet.select_row(sel_idx)
                            except Exception:
                                try:
                                    self.sheet.set_selected_rows([sel_idx])
                                except Exception:
                                    pass
                    else:
                        children = self.tree.get_children()
                        if children:
                            iid = children[min(idx, len(children) - 1)]
                            self.tree.selection_set(iid)
                            self.tree.see(iid)
                except Exception:
                    pass

            try:
                self.after(60, _restore_after_delete)
            except Exception:
                _restore_after_delete()

        except Exception as e:
            messagebox.showerror("Delete Failed", f"Could not delete: {e}")

    def _edit_selected(self):
        """Load selected row into small popup editor (optional) or into the form directly."""
        idx = self._get_selected_index()
        if idx is None:
            messagebox.showwarning("No Selection", "Please select a customer to edit.")
            return
        # Use popup editor for convenience
        try:
            self._open_editor_dialog(mode="edit", index=idx)
        except Exception:
            # fallback: load into permanent form
            self._load_into_form(idx)

    def _on_tree_double_click(self, event):
        idx = self._get_selected_index()
        if idx is None:
            return
        self._open_editor_dialog(mode="edit", index=idx)

    def _load_into_form(self, idx):
        """Fill the top form with data from customers[idx] for edit."""
        try:
            rec = self.customers[idx]
            self.code_e.delete(0, "end"); self.code_e.insert(0, str(rec.get("code","")))
            self.name_e.delete(0, "end"); self.name_e.insert(0, str(rec.get("name","")))
            self.desc_e.delete(0, "end"); self.desc_e.insert(0, str(rec.get("description","")))
            self.email_e.delete(0, "end"); self.email_e.insert(0, str(rec.get("email","")))
            self.phone_e.delete(0, "end"); self.phone_e.insert(0, str(rec.get("phone","")))
            self.edit_index = idx
            self.action_btn.configure(text="✏️ Update Customer")
        except Exception:
            pass

    def _open_editor_dialog(self, mode="edit", index=None):
        """Popup editor — mirrors OperatorManagerPage popup style but simplified for customers."""
        if mode not in ("edit", "add"):
            return

        win = ctk.CTkToplevel(self)
        win.title("Edit Customer" if mode == "edit" else "Add Customer")
        win.geometry("640x360")
        win.grab_set()

        frame = ctk.CTkFrame(win)
        frame.pack(fill="both", expand=True, padx=12, pady=12)

        # Code (row 0)
        ctk.CTkLabel(frame, text="Code:", font=("Segoe UI", 11)).grid(row=0, column=0, sticky="w", padx=6, pady=8)
        code_e = ctk.CTkEntry(frame, width=250)
        code_e.grid(row=0, column=1, sticky="w", padx=(2,0), pady=8)

        # customer Name (row 1)        
        ctk.CTkLabel(frame, text="Customer Name:", font=("Segoe UI", 11)).grid(row=0, column=2, sticky="w", padx=(0,6), pady=8)
        name_e = ctk.CTkEntry(frame, width=250)
        name_e.grid(row=0, column=3, sticky="w", padx=6, pady=8)

        # Description (row 2)
        ctk.CTkLabel(frame, text="Description:", font=("Segoe UI", 11)).grid(row=2, column=0, sticky="nw", padx=6, pady=8)
        desc_e = ctk.CTkEntry(frame, width=495)
        desc_e.grid(row=2, column=1, columnspan=3, sticky="w", padx=6, pady=8)
        
        # Email (row 3)
        ctk.CTkLabel(frame, text="Email:", font=("Segoe UI", 11)).grid(row=3, column=0, sticky="w", padx=6, pady=8)
        email_e = ctk.CTkEntry(frame, width=300)
        email_e.grid(row=3, column=1, sticky="w", padx=6, pady=8)

        # Phone (row 4)
        ctk.CTkLabel(frame, text="Phone:", font=("Segoe UI", 11)).grid(row=4, column=0, sticky="w", padx=6, pady=8)
        phone_e = ctk.CTkEntry(frame, width=200)
        phone_e.grid(row=4, column=1, sticky="w", padx=6, pady=8)
        


        # Prefill if edit
        if mode == "edit" and index is not None:
            rec = self.customers[index]
            code_e.insert(0, str(rec.get("code","")))
            name_e.insert(0, str(rec.get("name","")))
            desc_e.insert(0, str(rec.get("description","")))
            email_e.insert(0, str(rec.get("email","")))
            phone_e.insert(0, str(rec.get("phone","")))

        btnf = ctk.CTkFrame(frame, fg_color="transparent")
        btnf.grid(row=5, column=0, columnspan=4, pady=(12, 0))

        def on_save_clicked():
            code = code_e.get().strip()
            name = name_e.get().strip()
            desc = desc_e.get().strip()
            email = email_e.get().strip()
            phone = phone_e.get().strip()

            if not code or not name:
                messagebox.showwarning("Validation", "Code and Customer Name are required.")
                return
            if email and ("@" not in email or "." not in email):
                messagebox.showwarning("Validation", "Invalid email.")
                return
            if phone and not phone.isdigit():
                messagebox.showwarning("Validation", "Phone must be digits.")
                return

            if mode == "edit" and index is not None:
                # confirm update
                if not messagebox.askyesno("Confirm Update", f"Apply changes to customer {code}?"):
                    return
                old_code = self.customers[index].get("code")
                try:
                    conn = sqlite3.connect("customers.db")
                    cursor = conn.cursor()
                    cursor.execute("UPDATE customers SET code=?, name=?, description=?, email=?, phone=? WHERE code=?",
                                   (code, name, desc, email, phone, old_code))
                    conn.commit()
                    conn.close()
                except Exception as e:
                    messagebox.showerror("Update Error", f"Failed to update customer: {e}")
                    return
            else:
                # add mode
                if any(str(c.get("code","")) == code for c in self.customers):
                    messagebox.showwarning("Validation", "Code already exists.")
                    return
                try:
                    conn = sqlite3.connect("customers.db")
                    cursor = conn.cursor()
                    cursor.execute("INSERT INTO customers (code, name, description, email, phone) VALUES (?, ?, ?, ?, ?)",
                                   (code, name, desc, email, phone))
                    conn.commit()
                    conn.close()
                except Exception as e:
                    messagebox.showerror("DB Error", f"Failed to add customer: {e}")
                    return

            self._load_customers()
            self.refresh_table()
            win.destroy()

        ModernButton(btnf, text="Save", fg_color="#007B43", hover_color="#005C32", command=on_save_clicked).pack(side="left", padx=8)
        ModernButton(btnf, text="Cancel", fg_color="#757575", hover_color="#616161", command=win.destroy).pack(side="left", padx=8)

    def _clear_inputs(self):
        try:
            self.code_e.delete(0, "end")
            self.name_e.delete(0, "end")
            self.desc_e.delete(0, "end")
            self.email_e.delete(0, "end")
            self.phone_e.delete(0, "end")
        except Exception:
            pass

    # -------------------------
    # Table refresh / rendering
    # -------------------------
    def refresh_table(self):
        """Update table contents from self.customers applying search filter if present."""
        q = self.search_var.get().strip().lower()

        def matches(rec):
            if not q:
                return True
            return (q in str(rec.get("code","")).lower() or
                    q in str(rec.get("name","")).lower() or
                    q in str(rec.get("description","")).lower() or
                    q in str(rec.get("email","")).lower() or
                    q in str(rec.get("phone","")).lower())

        if getattr(self, "use_tksheet", False) and getattr(self, "sheet", None):
            try:
                data = [
                    [rec.get("code", ""), rec.get("name", ""), rec.get("description", ""), rec.get("email", ""), rec.get("phone", "")]
                    for rec in self.customers if matches(rec)
                ]
                self.sheet.set_sheet_data(data)
            except Exception:
                pass

            # schedule resize/redraw
            def _delayed():
                try:
                    self.resize_sheet()
                    try:
                        self.sheet.refresh()
                    except Exception:
                        pass
                except Exception:
                    pass
            try:
                self.after_idle(_delayed)
            except Exception:
                self.after(80, _delayed)
        else:
            try:
                for r in self.tree.get_children():
                    self.tree.delete(r)
                for rec in self.customers:
                    if not matches(rec):
                        continue
                    self.tree.insert("", "end", values=(rec.get("code",""), rec.get("name",""), rec.get("description",""), rec.get("email",""), rec.get("phone","")))
            except Exception:
                pass

            try:
                self.after(40, self.resize_sheet)
            except Exception:
                pass

    # -------------------------
    # Edit / helpers
    # -------------------------
    def reset_form(self):
        """Clear inputs and return to Add mode."""
        self._clear_inputs()
        self.edit_index = None
        self.action_btn.configure(text="➕ Add Customer")
        
    def go_back(self, event=None):
        try:
            self.app.load_settings_page()
        except:
            try: self.app.load_settings()
            except: pass

    def highlight_entry(self, entry, active):
        try:
            if active:
                entry.configure(border_width=2, border_color="#007B43")
                self.animate_glow(entry, 0)
            else:
                entry.configure(border_width=1, border_color="#007B43")
        except Exception:
            pass

    def animate_glow(self, entry, step):
        try:
            if self.focus_get() != entry:
                return
        except Exception:
            return
        wave_colors = ["#007B43", "#005C32", "#43A047", "#388E3C", "#005C32", "#007B43"]
        entry.configure(border_color=wave_colors[step % len(wave_colors)])
        self.after(150, lambda: self.animate_glow(entry, step + 1))

    def destroy(self):
        super().destroy()





# ==========================================================
# Run Chat Page — Recreated from Reference Logic
# ==========================================================
class RunChatPage(ctk.CTkFrame):
    STATE_FILE = "runchat_state.json"

    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self.chart_frames = []
        self.chart_sections = []
        self.chart_info = []
        self.saved_selections = []
        self.global_data = {}
        self.chart_device_map = {}      # ✅ fix for NameError (local mapping per chart)
        
        self.operator_map = {}
        self.load_operator_map()

        self.comp_map = self.load_component_specs()
        self._is_active = False   # becomes True when the page is shown

        # === Scrollable Canvas Setup ===
        self.canvas = tk.Canvas(self, bg="#fff", highlightthickness=0)
        self.canvas.pack(side="left", fill="both", expand=True, padx=10, pady=5)
        self.scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.scrollbar.pack(side="right", fill="y")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        self.scrollable_frame = tk.Frame(self.canvas, bg="#fff")
        self.window_id = self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")

        self.scrollable_frame.bind("<Configure>", lambda e: self.update_scrollregion())
        self.canvas.bind("<Configure>", lambda e: self.update_scrollregion())

        self.scrollable_frame.grid_columnconfigure(0, weight=1)
        self.scrollable_frame.grid_columnconfigure(1, weight=1)

        self.add_button = ModernButton(self.scrollable_frame, text="+", font=("Segoe UI", 18, "bold"),
                                    width=5, command=self.add_chart)
        self.add_button.grid(row=0, column=0, columnspan=2, pady=10, sticky="ew")

        # === Load previous layout ===
        initial_count, saved_state = self.load_runchat_state()
        if initial_count <= 0:
            initial_count = 1

        self.saved_selections = saved_state if saved_state else []

        for _ in range(initial_count):
            self.add_chart()

        self.fix_grid_layout()

        # === Subscribe to live updates ===
        # Initialise timing attrs BEFORE load_existing_buffer so update_with_parsed works
        self._last_draw_ts = 0.0
        self._min_draw_interval = 0.18  # seconds
        self._redraw_scheduled = False

        self.app.data_manager.subscribe(self)
        # === Load existing buffered data ===
        self.load_existing_buffer()

    # ------------------------------------------------------
    # Replay buffered data into charts
    # ------------------------------------------------------
    def load_existing_buffer(self):
        """Replay existing buffered data into charts so they appear filled when page opens."""
        try:
            buffer = getattr(self.app.data_manager, "buffer", [])
            if not buffer:
                print("RunChat preloaded 0 buffered data points.")
                return
            # Temporarily activate so update_with_parsed stores data
            was_active = self._is_active
            self._is_active = True
            for parsed in buffer[-200:]:
                self.update_with_parsed(parsed)
            self._is_active = was_active
            print(f"RunChat preloaded {len(buffer[-200:])} buffered data points.")
        except Exception as e:
            print("RunChat preload error:", e)

    # ------------------------------------------------------
    # Page-visibility hooks
    # ------------------------------------------------------
    def on_show(self):
        """Called by app when this page becomes visible."""
        self._is_active = True
        # optional: trigger an immediate redraw of latest data
        self.after(100, self._batch_redraw)

    def on_hide(self):
        """Called by app when navigating away."""
        self._is_active = False
        # ------------------------------------------------------
    # Basic Layout Helpers
    # ------------------------------------------------------
    # ------------------------------------------------------
    # Support Utilities
    # ------------------------------------------------------
    def load_operator_map(self):
        try:
            conn = sqlite3.connect(resource_path("employee_master.db"))
            cursor = conn.cursor()
            cursor.execute("SELECT id, name FROM employee_master")
            rows = cursor.fetchall()
            conn.close()
            self.operator_map = {str(r[0]): str(r[1]) for r in rows}
        except Exception:
            self.operator_map = {}

    def update_scrollregion(self):
        self.canvas.update_idletasks()
        bbox = self.canvas.bbox("all")
        if bbox:
            x1, y1, x2, y2 = bbox
            self.canvas.configure(scrollregion=(x1, y1, x2, y2 + 100))
            self.canvas.itemconfig(self.window_id, width=self.canvas.winfo_width())

    def fix_grid_layout(self):
        """Rebuild the RunChat grid ensuring equal box size and balanced columns."""
        n = len(self.chart_frames)
        for f in self.chart_frames:
            f.grid_forget()

        # Ensure all frames have consistent appearance
        for f in self.chart_frames:
            f.configure(fg_color="white", corner_radius=12, border_width=0)

        # Calculate number of rows needed
        cols = 2
        rows = (n + 1) // cols

        # Set uniform column & row weights for equal sizing
        for i in range(cols):
            self.scrollable_frame.grid_columnconfigure(i, weight=1, uniform="col")
        for r in range(rows):
            self.scrollable_frame.grid_rowconfigure(r, weight=1, uniform="row")

        # Re-grid all frames in order (2 per row)
        for i, f in enumerate(self.chart_frames):
            r, c = divmod(i, cols)
            f.grid(row=r, column=c, sticky="nsew", padx=8, pady=8)

        # Properly reposition Add button
        self.reposition_add_button()

        # Force geometry update
        self.scrollable_frame.update_idletasks()


    def reposition_add_button(self):
        n = len(self.chart_frames)
        self.add_button.grid_forget()
        row = n // 2
        if n % 2 == 0:
            self.add_button.grid(row=row, column=0, columnspan=2, pady=10, sticky="ew")
        else:
            if n == 1:
                self.add_button.grid(row=row + 1, column=0, columnspan=2, pady=10, sticky="ew")
            else:
                self.add_button.grid(row=row, column=1, pady=10, sticky="")
        self.scrollable_frame.grid_rowconfigure(row + 2, minsize=12)

    # ------------------------------------------------------
    # Add Chart Panel
    # ------------------------------------------------------
    def add_chart(self):
        idx = len(self.chart_frames)
        frame = ctk.CTkFrame(
            master=self.scrollable_frame,
            fg_color="white",
            corner_radius=12,
            border_width=0,
            height=280  # slightly taller
        )
        frame.pack_propagate(False)
        frame._chart_index = idx   # IMPORTANT: store index on frame


        # --- Add internal scrollable area (for chart + table) ---
        inner_canvas = tk.Canvas(frame, bg="#fff", highlightthickness=0)
        inner_scrollbar = ttk.Scrollbar(frame, orient="vertical", command=inner_canvas.yview)
        scrollable_inner = tk.Frame(inner_canvas, bg="#fff")

        inner_canvas.configure(yscrollcommand=inner_scrollbar.set)
        inner_canvas.pack(side="left", fill="both", expand=True)
        inner_scrollbar.pack(side="right", fill="y")

        inner_window = inner_canvas.create_window((0, 0), window=scrollable_inner, anchor="nw")

        # Update scroll region dynamically
        def update_inner_scroll(_=None):
            inner_canvas.configure(scrollregion=inner_canvas.bbox("all"))
            inner_canvas.itemconfig(inner_window, width=inner_canvas.winfo_width())
        # Enable mouse wheel scrolling
        def _on_mousewheel(event):
            inner_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        inner_canvas.bind("<Enter>", lambda e: inner_canvas.bind_all("<MouseWheel>", _on_mousewheel))
        inner_canvas.bind("<Leave>", lambda e: inner_canvas.unbind_all("<MouseWheel>"))


        scrollable_inner.bind("<Configure>", update_inner_scroll)
        inner_canvas.bind("<Configure>", update_inner_scroll)

        # Use scrollable_inner as the actual parent for chart content
        content_parent = scrollable_inner


        self.chart_frames.append(frame)

        # ensure storage
        self.global_data[idx] = {'x': [], 'y': [], 't': [], 'offset': []}
        while len(self.saved_selections) <= idx:
            self.saved_selections.append({'air_id': None, 'channel': None})

        chart_info = {'win_size': 10, 'start': 0, 'auto_follow': True}
        self.build_chart_panel(content_parent, idx, chart_info, frame)
        self.chart_info.append(chart_info)

        self.fix_grid_layout()
        self.save_runchat_state()

    # ------------------------------------------------------
    # Core chart builder
    # ------------------------------------------------------
    def build_chart_panel(self, parent, idx, chart_info, outer_frame):
        # Create clear vertical stack
        header_frame = tk.Frame(parent, bg="#fff")
        header_frame.pack(side="top", fill="x", pady=5)
        
        chart_frame = tk.Frame(parent, bg="#fff")
        chart_frame.pack(side="top", fill="both", expand=True)
        
        footer_frame = tk.Frame(parent, bg="#fff")
        footer_frame.pack(side="bottom", fill="x", pady=5)

        # === Top control row (selectors on left, buttons on right) ===
        control = tk.Frame(header_frame, bg="#f8f9fa", highlightbackground="#d0d0d0", highlightthickness=0)
        control.pack(fill="x", padx=5, ipady=3)

        expanded = False

        def close_chart():
            try:
                removed_index = getattr(outer_frame, "_chart_index", None)
                if outer_frame in self.chart_frames:
                    self.chart_frames.remove(outer_frame)
                outer_frame.destroy()

                if removed_index is not None:
                    if removed_index < len(self.chart_info):
                        del self.chart_info[removed_index]
                    if removed_index < len(self.chart_sections):
                        del self.chart_sections[removed_index]
                    if removed_index in self.global_data:
                        del self.global_data[removed_index]

                self.reindex_after_removal(removed_index)
            except Exception as e:
                print("Close chart error:", e)
        
        def toggle_expand():
            nonlocal expanded
            if not expanded:
                outer_frame.configure(height=600)
                for f in list(self.chart_frames):
                    if f is not outer_frame:
                        f.grid_remove()
                outer_frame.grid(row=0, column=0, columnspan=2, sticky="nsew", padx=8, pady=8)
                try:
                    self.add_button.grid_forget()
                except Exception:
                    pass
                expand_btn.config(text="🗗")
                
                # Show extended info (initially just show last known or Guest)
                # It will update on next data packet in update_with_parsed
                op_label.pack(side="left", padx=10)
                cpk_label.pack(side="left", padx=10)
                
                expanded = True
            else:
                outer_frame.configure(height=260)
                self.fix_grid_layout()
                expand_btn.config(text="🗖")
                
                # Hide extended info
                op_label.pack_forget()
                cpk_label.pack_forget()
                
                expanded = False
                try:
                    self.reposition_add_button()
                except Exception:
                    pass

        # --- AirGauge selector (left side)
        tk.Label(control, text="AirGauge:", bg="#f8f9fa", font=("Segoe UI", 11)).pack(side="left", padx=(8, 0))
        air_var = tk.StringVar()
        air_box = ttk.Combobox(control, textvariable=air_var,
                               values=sorted(list(self.comp_map.keys())), state="readonly", width=12)
        air_box.pack(side="left", padx=(5, 10))

        # --- Channel selector ---
        tk.Label(control, text="Channel:", bg="#f8f9fa", font=("Segoe UI", 11)).pack(side="left", padx=(5, 0))
        ch_var = tk.StringVar()
        ch_box = ttk.Combobox(control, textvariable=ch_var,
                              values=[], state="readonly", width=8)
        ch_box.pack(side="left", padx=(5, 10))


        # --- Latest value box ---
        val_frame = tk.Frame(control, bg="#ffffff", bd=1, relief="solid")
        val_frame.pack(side="left", padx=(15, 5))
        latest_var = tk.StringVar(value="--")
        latest_label = tk.Label(val_frame, textvariable=latest_var, bg="#fff",
                                fg="navy", font=("Segoe UI", 13, "bold"), width=10)
        latest_label.pack(padx=3, pady=2)

        # --- Spacer (to push buttons to right) ---
        tk.Label(control, text="", bg="#f8f9fa").pack(side="left", expand=True)

        # --- Buttons (rightmost side) ---
        expand_btn = tk.Button(control, text="🗖", font=("Segoe UI", 11, "bold"),
                               bg="#E0E0E0", width=3, relief="flat", command=toggle_expand)
        expand_btn.pack(side="right", padx=(5, 3))

        close_btn = tk.Button(control, text="✖", font=("Segoe UI", 11, "bold"),
                              bg="#FFCDD2", width=3, relief="flat", command=close_chart)
        close_btn.pack(side="right", padx=(0, 5))

        # --- Expanded View Labels (Operator & CpK) ---
        # Only visible when expanded
        op_var = tk.StringVar(value="Operator: --")
        op_label = tk.Label(control, textvariable=op_var, bg="#f8f9fa", fg="blue", font=("Segoe UI", 11, "bold"))
        
        cpk_var = tk.StringVar(value="CpK: --")
        cpk_label = tk.Label(control, textvariable=cpk_var, bg="#f8f9fa", fg="purple", font=("Segoe UI", 11, "bold"))


        # matplotlib chart
        fig, ax = plt.subplots(figsize=(5.2, 2))
        chart_info.update({'ax': ax, 'fig': fig})
        chart_info['canvas'] = FigureCanvasTkAgg(fig, master=chart_frame)
        chart_info['canvas'].get_tk_widget().pack(side="top", fill="both", expand=True)
        # === Move Left / Right Buttons (scroll window) ===
        nav_frame = tk.Frame(chart_frame, bg="#fff")
        nav_frame.pack(side="bottom", fill="x", pady=(0, 5))

        def move_left():
            data_len = len(self.global_data[idx]["x"])
            if data_len == 0:
                return
            chart_info["auto_follow"] = False
            chart_info["start"] = max(0, chart_info["start"] - 1)
            chart_info["update_visible"]()

        def move_right():
            data_len = len(self.global_data[idx]["x"])
            if data_len == 0:
                return
            max_start = max(0, data_len - chart_info["win_size"])
            chart_info["start"] = min(max_start, chart_info["start"] + 1)
            if chart_info["start"] >= max_start:
                chart_info["auto_follow"] = True
            chart_info["update_visible"]()

        ModernButton(nav_frame, text="◀", font=("Segoe UI", 11, "bold"), fg_color="#E0E0E0", text_color="black", hover_color="#BDBDBD", width=30, command=move_left).pack(side="left", padx=5)
        ModernButton(nav_frame, text="▶", font=("Segoe UI", 11, "bold"), fg_color="#E0E0E0", text_color="black", hover_color="#BDBDBD", width=30, command=move_right).pack(side="right", padx=5)

        # style setup
        def setup_chart():
            ax.clear()
            try:
                ag, ch = air_var.get(), ch_var.get()
                spec = self.comp_map.get(ag, {}).get(ch, {})
                
                # Robust guard: Ensure basic tolerance keys exist
                if spec and "drawing_value" in spec and "low_tolerance" in spec and "high_tolerance" in spec:
                    c = float(spec["drawing_value"])
                    lo = float(spec["low_tolerance"])
                    hi = float(spec["high_tolerance"])
                    diff = hi - lo
                    m = diff * 0.5 if diff >= 1e-9 else 1.0

                    ax.axhline(c, color="black", linewidth=1, alpha=0.5)
                    
                    # --- Determine Colors based on Type ---
                    comp_type = spec.get("type", "Shaft")
                    if comp_type == "Hole":
                        # Hole: Low=Orange (Rework), High=Red (Reject)
                        low_color, high_color = "#D4700A", "#C0392B"
                        low_bg, high_bg = "#F5C98A", "#F4A8A8"
                    else:
                        # Shaft: Low=Red (Reject), High=Orange (Rework)
                        low_color, high_color = "#C0392B", "#D4700A"
                        low_bg, high_bg = "#F4A8A8", "#F5C98A"

                    ax.axhspan(lo - m, lo, color=low_bg)
                    ax.axhspan(lo, hi, color="#AADDAA") # Moderate green for safe zone
                    ax.axhspan(hi, hi + m, color=high_bg)

                    ax.axhline(lo, color=low_color, ls="--", linewidth=1.8)
                    ax.axhline(hi, color=high_color, ls="--", linewidth=1.8)
                    
                    # Set explicit limits based on tolerances
                    ax.set_ylim(lo - m, hi + m)
                    ax.set_yticks([lo, c, hi])
                    ax.set_yticklabels([f"{lo:.3f}", f"{c:.3f}", f"{hi:.3f}"])
                    ax.tick_params(axis='y', labelsize=7)
                else:
                    # Clear fallback when no spec
                    ax.set_ylim(0, 1)
                    ax.set_yticks([0, 0.5, 1])
                    ax.set_yticklabels(["0.000", "No Spec", "1.000"])
                    ax.text(5, 0.5, "Please Select Valid Configuration", ha='center', va='center', alpha=0.5)

            except Exception as e:
                print(f"Setup chart error: {e}")
                ax.set_ylim(0, 100)

            ax.grid(True, linestyle="--", alpha=0.5)
            chart_info["line"], = ax.plot([], [], marker="o", color="black", markersize=4, linewidth=1.5)
            chart_info["texts"] = [] 



        setup_chart()
        
        for spine in ["top", "right", "left", "bottom"]:
            ax.spines[spine].set_visible(False)
            
        try:
            fig.subplots_adjust(left=0.15, right=0.95, top=0.90, bottom=0.25)
        except Exception:
            pass

        # === Enhanced Data Table ===
        # === Bordered Data Table using Canvas ===
        # Colors
        ROW_COLORS  = ["#F0F4FF", "#FFFFFF"]   # alternate: light blue-white, pure white
        HEADER_BG   = "#1565C0"                 # same blue as USB page header
        HEADER_FG   = "white"
        BORDER_CLR  = "#BDBDBD"
        CELL_FG     = "#212121"
        LABEL_W     = 72    # px width of the row-label column
        MIN_COL_W   = 75    # minimum px per value column
        ROW_H       = 22    # px per row
        WIN         = chart_info["win_size"]   # number of value columns = 10

        # outer frame — gives us the border around the whole table
        table_frame = tk.Frame(footer_frame, bg=BORDER_CLR, bd=1, relief="flat")
        table_frame.pack(fill="x", padx=5)

        # canvas + horizontal scrollbar (in case very narrow screen)
        tbl_canvas  = tk.Canvas(table_frame, bg="white", highlightthickness=0,
                                height=(ROW_H * 6 + 1))  # 1 header + 5 data rows
        tbl_canvas.pack(side="top", fill="x", expand=True)

        # We'll draw everything on the canvas as text + rectangles.
        # Store column label widget references so update_visible can refresh them.
        row_labels = ["S.No.", "Reading", "Time", "Date", "Offset"]

        # cell_vars[row][col] → StringVar holding the cell text
        # col 0 = label column, cols 1..WIN = value columns
        cell_vars = []
        for r_idx, lbl in enumerate(row_labels):
            row_vars = [tk.StringVar(value=lbl)]   # col-0: always the label
            for _ in range(WIN):
                row_vars.append(tk.StringVar(value=""))
            cell_vars.append(row_vars)

        # Keep canvas item ids so we can update text in-place
        cell_items = {}   # (row, col) → canvas text item id
        rect_items = {}   # (row, col) → canvas rectangle item id

        def _draw_table(total_w):
            """(Re-)draw the whole table to fit total_w pixels."""
            tbl_canvas.delete("all")
            col_w = max(MIN_COL_W, (total_w - LABEL_W - 2) // max(1, WIN))

            for r_idx in range(len(row_labels)):
                y0 = r_idx * ROW_H + 1   # +1 for top border
                y1 = y0 + ROW_H

                for c_idx in range(WIN + 1):   # 0 = label col
                    x0 = (LABEL_W if c_idx == 0 else LABEL_W + (c_idx - 1) * col_w) + 1
                    x1 = (LABEL_W if c_idx == 0 else LABEL_W + c_idx * col_w)
                    if c_idx == 0:
                        x0 = 1
                        x1 = LABEL_W

                    bg  = ROW_COLORS[r_idx % 2]
                    fg  = CELL_FG
                    anc = "w" if c_idx == 0 else "center"
                    tx  = x0 + 4 if c_idx == 0 else (x0 + x1) // 2

                    rid = tbl_canvas.create_rectangle(
                        x0, y0, x1, y1,
                        fill=bg, outline=BORDER_CLR, width=1
                    )
                    tid = tbl_canvas.create_text(
                        tx, (y0 + y1) // 2,
                        text=cell_vars[r_idx][c_idx].get(),
                        fill=fg, font=("Segoe UI", 9),
                        anchor=anc
                    )
                    rect_items[(r_idx, c_idx)] = rid
                    cell_items[(r_idx, c_idx)] = tid

            # set canvas scroll-region width
            total_cols_w = LABEL_W + WIN * col_w + 2
            tbl_canvas.configure(scrollregion=(0, 0, total_cols_w, ROW_H * len(row_labels) + 2))
            chart_info['_tbl_col_w'] = col_w

        def _refresh_cells():
            """Update only the text values without redrawing rectangles."""
            col_w = chart_info.get('_tbl_col_w', MIN_COL_W)
            for r_idx in range(len(row_labels)):
                for c_idx in range(WIN + 1):
                    tid = cell_items.get((r_idx, c_idx))
                    if tid:
                        tbl_canvas.itemconfig(tid, text=cell_vars[r_idx][c_idx].get())

        def _on_tbl_resize(event):
            _draw_table(event.width)

        tbl_canvas.bind("<Configure>", _on_tbl_resize)
        # initial draw after geometry settles
        self.after(80, lambda: _draw_table(tbl_canvas.winfo_width() or 700))

        chart_info['table_frame']         = table_frame
        chart_info['tbl_canvas']          = tbl_canvas
        chart_info['cell_vars']           = cell_vars
        chart_info['_refresh_cells']      = _refresh_cells
        chart_info['_draw_table']         = _draw_table
        chart_info['row_labels']          = row_labels
        chart_info['resize_table_columns'] = lambda: _draw_table(
            tbl_canvas.winfo_width() or 700
        )

        # legacy 'rows' shim so update_visible can still reference rows dict
        rows = {lbl: i for i, lbl in enumerate(row_labels)}
        chart_info['rows'] = rows
        chart_info['table'] = None   # no ttk.Treeview


        # update visible window
        def update_visible():
            """Redraw chart and update data table for this chart window."""
            data = self.global_data[idx]
            n = len(data["x"])
            win = chart_info["win_size"]

            # ---- Handle window start logic ----
            if chart_info["auto_follow"]:
                s = max(0, n - win)
                chart_info["start"] = s
            else:
                s = chart_info["start"]

            e = min(s + win, n)
            xw, yw = data["x"][s:e], data["y"][s:e]
            tw, ow = data["t"][s:e], data["offset"][s:e]
            try:
                resize_fn = chart_info.get('resize_table_columns')
                if callable(resize_fn):
                    resize_fn()
            except Exception:
                pass

            # --- Date Extraction for table (safe) ---
            try:
                all_buf = self.app.data_manager.buffer
                recent_buf = all_buf[max(0, len(all_buf) - n + s):max(0, len(all_buf) - n + e)]
                dates = [b[3] for b in recent_buf] if recent_buf else ["" for _ in range(len(yw))]
            except Exception:
                dates = ["" for _ in range(len(yw))]

            # ---- Update Chart ----
            # Clear old text labels
            for txt in chart_info.get("texts", []):
                txt.remove()
            chart_info["texts"] = []

            if xw:
                ax.set_xlim(min(xw) - 0.1, max(xw) + 0.1)
                chart_info["line"].set_data(xw, yw)
                
                # Add text labels for points
                # Re-calculate color logic for text to match point status
                try:
                    spec = self.comp_map.get(air_var.get(), {}).get(ch_var.get(), {})
                    lo = float(spec.get("low_tolerance", 0))
                    hi = float(spec.get("high_tolerance", 0))
                    c_type = spec.get("type", "Shaft")
                except:
                    lo, hi = 0, 0
                    c_type = "Shaft"

                for xi, yi in zip(xw, yw):
                    # Determine color
                    color = "green"
                    if yi < lo:
                        color = "red" if c_type == "Shaft" else "orange"
                    elif yi > hi:
                        color = "orange" if c_type == "Shaft" else "red"
                    
                    # Add offset to y for label position
                    offset_y = (hi - lo) * 0.05 if (hi - lo) > 0 else 0.001
                    
                    t_obj = ax.text(xi, yi + offset_y, f"{yi:.3f}", 
                                    color=color, fontsize=8, ha='center', va='bottom', fontweight='bold')
                    chart_info["texts"].append(t_obj)

                chart_info["canvas"].draw()
            else:
                ax.set_xlim(0, 10)
                chart_info["canvas"].draw()

            # ---- Update Table (canvas-based) ----
            WIN = chart_info["win_size"]
            cv  = chart_info.get("cell_vars")
            if cv:
                padded = {
                    "S.No."  : list(range(s + 1, e + 1)),
                    "Reading": [f"{v:.4f}" for v in yw],
                    "Time"   : list(tw),
                    "Date"   : list(dates),
                    "Offset" : list(ow),
                }
                for r_idx, lbl in enumerate(chart_info["row_labels"]):
                    vals = padded.get(lbl, [])
                    pad  = vals + [""] * (WIN - len(vals))
                    cv[r_idx][0].set(lbl)          # col-0: label
                    for c_idx, v in enumerate(pad[:WIN], start=1):
                        cv[r_idx][c_idx].set(str(v))
                rfn = chart_info.get("_refresh_cells")
                if callable(rfn):
                    rfn()

            # ---- Update latest numeric value + color ----
            if yw:
                latest_val = yw[-1]
                latest_var.set(f"{latest_val:.4f}")
                latest_var.set(f"{latest_val:.4f}")
                self.update_val_color(latest_label, air_var.get(), ch_var.get(), latest_val)

            # ---- Calculate CpK (only after 50 readings) ----
            try:
                all_y     = data["y"]
                total_n   = len(all_y)
                last_50_y = all_y[-50:]

                if total_n < 50:
                    # Show a waiting progress indicator
                    cpk_var.set(f"CpK: waiting ({total_n}/50)")
                elif len(last_50_y) >= 50:
                    import statistics
                    mean_val  = statistics.mean(last_50_y)
                    stdev_val = statistics.stdev(last_50_y)

                    spec = self.comp_map.get(air_var.get(), {}).get(ch_var.get(), {})
                    usl  = float(spec.get("high_tolerance", 0))
                    lsl  = float(spec.get("low_tolerance",  0))

                    if stdev_val > 1e-9:
                        cpu     = (usl - mean_val) / (3 * stdev_val)
                        cpl     = (mean_val - lsl)  / (3 * stdev_val)
                        cpk_val = min(cpu, cpl)
                        cpk_var.set(f"CpK: {cpk_val:.2f} (n=50)")
                    else:
                        cpk_var.set("CpK: ∞ (n=50)")
                else:
                    cpk_var.set("CpK: --")
            except Exception:
                cpk_var.set("CpK: Err")


        chart_info["update_visible"] = update_visible

        def on_spec_change(*args):
            # Refresh map from server/storage if possible before redrawing
            self.comp_map = self.load_component_specs()
            setup_chart()
            update_visible()
            self.saved_selections[idx] = {"air_id": air_var.get(), "channel": ch_var.get()}
            self.save_runchat_state()

        def update_channels(*_):
            ag = air_var.get()
            if ag in self.comp_map:
                chans = sorted(list(self.comp_map[ag].keys()))
                ch_box["values"] = chans
                if ch_var.get() not in chans:
                    ch_var.set(chans[0] if chans else "")
            else:
                ch_box["values"] = []
                ch_var.set("")
            on_spec_change()

        air_var.trace_add("write", update_channels)
        ch_var.trace_add("write", on_spec_change)


        air_box.bind("<<ComboboxSelected>>", on_spec_change)
        ch_box.bind("<<ComboboxSelected>>", on_spec_change)

        # defaults
        sel = self.saved_selections[idx]
        
        # Fetch IDs from DB
        try:
            conn = sqlite3.connect("airgauge_master.db")
            ids = [str(r[0]) for r in conn.execute("SELECT airgauge_id FROM airgauge_master").fetchall()]
            conn.close()
        except:
            ids = []

        if sel.get("air_id"):
            air_var.set(sel["air_id"])
        elif ids:
            air_var.set(ids[idx % len(ids)])
        ch_var.set(sel.get("channel", f"CH{(idx % 4) + 1}"))
        on_spec_change()

        chart_info.update({
            "air_var": air_var, "ch_var": ch_var,
            "latest_var": latest_var, "latest_label": latest_label,
            "op_var": op_var, "cpk_var": cpk_var,
        })
    def reindex_after_removal(self, removed_index):
        """
        Maintain contiguous indices after a chart is removed.
        - Rebuild global_data, saved_selections, chart_device_map to be 0..n-1
        - Update each frame._chart_index attribute
        """
        # Rebuild global_data based on order of self.chart_frames
        rebuilt_global = {}
        new_saved = []

        for new_idx, frm in enumerate(self.chart_frames):
            old_idx = getattr(frm, "_chart_index", None)
            if old_idx is None:
                old_idx = new_idx
            rebuilt_global[new_idx] = self.global_data.get(old_idx, {'x': [], 'y': [], 't': [], 'offset': []})
            # keep saved selection if present
            if old_idx is not None and old_idx < len(self.saved_selections):
                new_saved.append(self.saved_selections[old_idx])
            else:
                new_saved.append({'air_id': None, 'channel': None})

        # replace with rebuilt structures
        self.global_data.clear()
        self.global_data.update(rebuilt_global)
        self.saved_selections = new_saved
        # ✅ Compact chart_device_map (instance attribute, not global)
        new_map = {}
        for new_idx, frm in enumerate(self.chart_frames):
            old_idx = getattr(frm, "_chart_index", None)
            if old_idx is None:
                old_idx = new_idx
            new_map[new_idx] = self.chart_device_map.get(
                old_idx, {"air_id": None, "channel": None}
            )
        self.chart_device_map = new_map
        # update _chart_index on frames
        for i, frm in enumerate(self.chart_frames):
            try:
                frm._chart_index = i
            except Exception:
                pass

        # relayout & save state
        self.fix_grid_layout()
        self.save_runchat_state()

    # ------------------------------------------------------
    # Support Utilities
    # ------------------------------------------------------
    def load_component_specs(self):
        try:
            conn = sqlite3.connect(resource_path("component_setup.db"))
            cursor = conn.cursor()
            cursor.execute("SELECT airgauge_id, channel, properties FROM component_setup")
            rows = cursor.fetchall()
            conn.close()
            comp_map = {}
            for ag_id, ch, props_str in rows:
                if ag_id not in comp_map:
                    comp_map[ag_id] = {}
                try:
                    comp_map[ag_id][ch] = json.loads(props_str)
                except:
                    comp_map[ag_id][ch] = {}
            return comp_map
        except Exception:
            return {}

    def load_runchat_state(self):
        try:
            with open(self.STATE_FILE, "r") as f:
                s = json.load(f)
            return int(s.get("chart_count", 1)), s.get("saved_selections", [])
        except Exception:
            return 1, []

    def refresh_airgauge_ids(self):
        """Refresh AirGauge ID dropdowns in all charts after config change."""
        try:
            # Reload updated component map
            self.comp_map = self.load_component_specs()

            # Get latest IDs from airgauge_master.db
            try:
                conn = sqlite3.connect("airgauge_master.db")
                ids = [str(r[0]) for r in conn.execute("SELECT airgauge_id FROM airgauge_master").fetchall()]
                conn.close()
            except:
                ids = []

            # Update all chart comboboxes
            for info in self.chart_info:
                air_box = info.get("air_var")
                if air_box:
                    # reset dropdown list if supported
                    try:
                        air_box.set("")  # clear selection if outdated
                    except Exception:
                        pass
                ch_box = info.get("ch_var")
                # Optionally update values if using ttk.Combobox
                # Example:
                # info["air_box"]["values"] = ids
            print("✅ RunChatPage AirGauge IDs refreshed.")
        except Exception as e:
            print("RunChat refresh error:", e)


    def save_runchat_state(self):
        """Save how many charts exist and their airgauge/channel selections."""
        try:
            selections = []
            for info in self.chart_info:
                air = info["air_var"].get()
                ch = info["ch_var"].get()
                selections.append({"air_id": air, "channel": ch})

            state = {
                "chart_count": len(self.chart_frames),
                "saved_selections": selections
            }
            with open(self.STATE_FILE, "w") as f:
                json.dump(state, f, indent=2)
        except Exception as e:
            print("Save RunChat state error:", e)


    def update_val_color(self, label, air, ch, val):
        try:
            spec = self.comp_map[air][ch]
            lo, hi = float(spec["low_tolerance"]), float(spec["high_tolerance"])
            comp_type = spec.get("type", "Shaft")
            
            # Default Hole: < Lo (Orange), > Hi (Red)
            # Shaft: < Lo (Red), > Hi (Orange)
            
            if val < lo:
                label.config(fg="red" if comp_type == "Shaft" else "orange")
            elif val > hi:
                label.config(fg="orange" if comp_type == "Shaft" else "red")
            else:
                label.config(fg="green")
        except Exception:
            label.config(fg="navy")

    def update_with_parsed(self, parsed):
        """Append data to matching charts and schedule a redraw."""
        try:
            (hh, mm, ss, datestr, Reading, offset, status,
             air_id, ch,
             drawing_val, userID,
             compID, item_name, item_code, cncID) = parsed

            # Normalize incoming air_id and channel to plain strings for matching
            air_id_str = str(air_id).strip().upper().lstrip("AG").lstrip("0") or "0"
            ch_str     = str(ch).strip().upper().lstrip("CH").lstrip("0") or "0"

            matched_any = False
            for i, info in enumerate(self.chart_info):
                # Normalize selected values the same way
                sel_air = str(info["air_var"].get()).strip().upper().lstrip("AG").lstrip("0") or "0"
                sel_ch  = str(info["ch_var"].get()).strip().upper().lstrip("CH").lstrip("0") or "0"

                if sel_air != air_id_str or sel_ch != ch_str:
                    continue

                matched_any = True

                # Update Operator Name
                if info.get("op_var"):
                    op_name = self.operator_map.get(str(userID), "Guest")
                    info["op_var"].set(f"Operator: {op_name}")

                # Parse the reading value
                try:
                    val = float(Reading)
                except Exception:
                    try:
                        val = float(drawing_val)
                    except Exception:
                        continue

                self.global_data[i]["x"].append(len(self.global_data[i]["x"]) + 1)
                self.global_data[i]["y"].append(val)
                self.global_data[i]["t"].append(f"{hh}:{mm}:{ss}")
                self.global_data[i]["offset"].append(offset)

            # Always schedule a redraw if page is active (even if no match —
            # so the chart stays current and doesn't appear frozen)
            if getattr(self, "_is_active", False):
                now = time.time()
                if now - getattr(self, "_last_draw_ts", 0) >= getattr(self, "_min_draw_interval", 0.18):
                    self._last_draw_ts = now
                    self.after(1, self._batch_redraw)
                elif not getattr(self, "_redraw_scheduled", False):
                    self._redraw_scheduled = True
                    delay = int((self._min_draw_interval - (now - self._last_draw_ts)) * 1000)
                    self.after(max(delay, 50), self._batch_redraw)

        except Exception as e:
            print("RunChat update error:", e)


    def _batch_redraw(self):
        self._redraw_scheduled = False
        if not getattr(self, "_is_active", False):
            return
        for info in self.chart_info:
            try:
                info["update_visible"]()
            except Exception:
                pass

# ==========================================================
# USB Data Assignment Popup
# ==========================================================
class UsbAssignmentPopup(ctk.CTkToplevel):
    def __init__(self, parent, unique_combos, customers_list, items_list, on_confirm):
        super().__init__(parent)
        self.title("USB Data Assignment")
        self.geometry("750x450")
        self.attributes('-topmost', True)
        self.grab_set()

        self.unique_combos = unique_combos
        self.on_confirm = on_confirm
        
        self.selections = {} # (ag, ch, dwg) -> {"customer": var, "item": var}
        
        # UI Setup
        ctk.CTkLabel(self, text="Assign Customer & Item to New USB Data", font=("Segoe UI", 18, "bold")).pack(pady=10)
        
        scroll_frame = ctk.CTkScrollableFrame(self)
        scroll_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        # Header Row
        headers = ctk.CTkFrame(scroll_frame, fg_color="transparent")
        headers.pack(fill="x", pady=5)
        ctk.CTkLabel(headers, text="Drawing Val", width=120, font=("Segoe UI", 11, "bold")).pack(side="left", padx=5)
        ctk.CTkLabel(headers, text="AG ID", width=60, font=("Segoe UI", 11, "bold")).pack(side="left", padx=5)
        ctk.CTkLabel(headers, text="CH", width=40, font=("Segoe UI", 11, "bold")).pack(side="left", padx=5)
        ctk.CTkLabel(headers, text="Customer", width=200, font=("Segoe UI", 11, "bold")).pack(side="left", padx=5)
        ctk.CTkLabel(headers, text="Item", width=200, font=("Segoe UI", 11, "bold")).pack(side="left", padx=5)

        customer_options = ["None"] + [c.get("name", "") for c in customers_list if c.get("name")]
        item_options = ["None"] + [i.get("name", "") for i in items_list if i.get("name")]

        for combo in self.unique_combos:
            ag_id, ch_no, dwg = combo
            row = ctk.CTkFrame(scroll_frame)
            row.pack(fill="x", pady=2)
            
            ctk.CTkLabel(row, text=str(dwg), width=120).pack(side="left", padx=5)
            ctk.CTkLabel(row, text=str(ag_id), width=60).pack(side="left", padx=5)
            ctk.CTkLabel(row, text=str(ch_no), width=40).pack(side="left", padx=5)
            
            cust_var = ctk.StringVar(value="None")
            item_var = ctk.StringVar(value="None")
            self.selections[combo] = {"customer": cust_var, "item": item_var}
            
            # Using dynamic_resizing=False ensures wide options don't break the UI layout
            ctk.CTkOptionMenu(row, variable=cust_var, values=customer_options, width=200, dynamic_resizing=False).pack(side="left", fill="x", expand=True, padx=5)
            ctk.CTkOptionMenu(row, variable=item_var, values=item_options, width=200, dynamic_resizing=False).pack(side="left", fill="x", expand=True, padx=5)

        # Buttons
        btn_frame = ctk.CTkFrame(self, fg_color="transparent")
        btn_frame.pack(pady=10)
        
        self.confirm_btn = ModernButton(btn_frame, text="Confirm Assignments", fg_color="#4CAF50", hover_color="#388E3C", command=self.confirm)
        self.confirm_btn.pack(side="left", padx=10)
        ModernButton(btn_frame, text="Cancel", fg_color="#F44336", hover_color="#D32F2F", command=self.destroy).pack(side="left", padx=10)

    def confirm(self):
        # Disable button immediately to prevent double-clicks
        self.confirm_btn.configure(state="disabled")
        
        assignments = {}
        for combo, vars_dict in self.selections.items():
            cust = vars_dict["customer"].get()
            item = vars_dict["item"].get()
            assignments[combo] = {
                "customer": "" if cust == "None" else cust,
                "item": "" if item == "None" else item
            }
        
        # Withdraw window immediately so it feels responsive
        self.withdraw()
        
        # Execute the callback which will insert data
        self.on_confirm(assignments)
        
        # Finally destroy
        self.destroy()

# ==========================================================
# Live Data Page (Optimized Version from Reference)
# ==========================================================

class UsbDataPage(ctk.CTkFrame):
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        
        self.comp_json = getattr(self.app, 'comp_json', {})

        # Load ID Maps
        self.employee_map = self._load_json_map("operators.json", "id", "name")
        self.machine_map = self._load_json_map("machines.json", "code", "name")
        self.customers_list = self._load_json_list("customers.json")
        self.items_list = self._load_json_list("items.json")
        
        # Track rows to prevent duplicate saves
        self.loaded_usb_rows = []
        self.saved_usb_indices = set()

        self.build_ui()

    def _load_json_map(self, filename, key_field, val_field):
        table = None
        if "operators" in filename: table = "employee_master"
        elif "machines" in filename: table = "machine_master"
        if table:
            try:
                conn = sqlite3.connect(resource_path("employee_master.db"))
                cursor = conn.cursor()
                k = "id" if table == "employee_master" else "code"
                cursor.execute(f"SELECT {k}, name FROM {table}")
                rows = cursor.fetchall()
                conn.close()
                return {str(r[0]): str(r[1]) for r in rows}
            except Exception: pass
        return {}

    def _load_json_list(self, filename):
        table = None
        if "customers" in filename: table = "customers"
        elif "items" in filename: table = "items"
        if table:
            try:
                conn = sqlite3.connect(f"{table}.db")
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                cursor.execute(f"SELECT * FROM {table}")
                rows = cursor.fetchall()
                conn.close()
                return [dict(r) for r in rows]
            except Exception: pass
        return []

    def build_ui(self):
        # Helper to dynamically generate a clean white-on-transparent USB trident logo
        def create_usb_icon():
            from PIL import Image, ImageDraw
            img = Image.new("RGBA", (80, 80), (0, 0, 0, 0))
            draw = ImageDraw.Draw(img)
            
            # White circular outline
            draw.ellipse([12, 12, 68, 68], outline="white", width=4)
            
            # USB Trident Bottom Circle
            draw.ellipse([34, 50, 46, 62], fill="white")
            
            # Central stem
            draw.line([(40, 50), (40, 24)], fill="white", width=4)
            # Top arrow
            draw.polygon([(40, 15), (33, 24), (47, 24)], fill="white")
            
            # Left branch (ending in a square)
            draw.line([(40, 42), (28, 30)], fill="white", width=4)
            draw.line([(28, 30), (28, 24)], fill="white", width=4)
            draw.rectangle([23, 19, 33, 29], fill="white")
            
            # Right branch (ending in a circle)
            draw.line([(40, 42), (52, 30)], fill="white", width=4)
            draw.line([(52, 30), (52, 24)], fill="white", width=4)
            draw.ellipse([47, 19, 57, 29], fill="white")
            
            return ctk.CTkImage(img, size=(28, 28))

        # === Modern Header Card ===
        header_card = ctk.CTkFrame(
            self,
            fg_color="white",
            corner_radius=15,
            border_width=1,
            border_color="#E0E0E0"
        )
        header_card.pack(fill="x", padx=20, pady=(10, 5))
        
        # Green vertical accent bar on the left edge
        green_bar = ctk.CTkFrame(header_card, fg_color="#007B43", width=6, corner_radius=0)
        green_bar.place(x=0, y=0, relheight=1)
        
        # Header content
        header_content = ctk.CTkFrame(header_card, fg_color="transparent")
        header_content.pack(fill="x", padx=(25, 20), pady=10)
        
        # Title with icon
        title_frame = ctk.CTkFrame(header_content, fg_color="transparent")
        title_frame.pack(side="left")
        
        # Rounded green square for logo
        logo_frame = ctk.CTkFrame(title_frame, fg_color="#007B43", width=40, height=40, corner_radius=8)
        logo_frame.pack(side="left", padx=(0, 12))
        logo_frame.pack_propagate(False)
        
        logo_lbl = ctk.CTkLabel(logo_frame, image=create_usb_icon(), text="", fg_color="transparent")
        logo_lbl.pack(expand=True)
        
        # Text label for title
        ctk.CTkLabel(
            title_frame,
            text="USB Data Upload",
            font=("Segoe UI", 20, "bold"),
            text_color="#007B43"
        ).pack(side="left")
        
        # Buttons Frame (Right)
        btn_frame = ctk.CTkFrame(header_content, fg_color="transparent")
        btn_frame.pack(side="right")

        # Save to DB Button
        self.save_btn = ModernButton(
            btn_frame,
            text="💾 Save to DB",
            font=("Segoe UI", 12, "bold"),
            fg_color="#007B43", hover_color="#005C32",
            text_color="white",
            height=36,
            corner_radius=6,
            command=self.save_usb_data_to_db,
            state="disabled"
        )
        self.save_btn.pack(side="left", padx=8)

        # Edit Assignments Button
        edit_btn = ModernButton(
            btn_frame,
            text="✎ Edit Assignments",
            font=("Segoe UI", 12, "bold"),
            fg_color="white",
            border_color="#007B43",
            border_width=1,
            hover_color="#F1F8E9",
            text_color="#007B43",
            height=36,
            corner_radius=6,
            command=self.edit_usb_assignments
        )
        edit_btn.pack(side="left", padx=8)

        # Refresh Button
        refresh_btn = ModernButton(
            btn_frame,
            text="↻ Refresh",
            font=("Segoe UI", 12, "bold"),
            fg_color="#FF9800", hover_color="#E65100",
            text_color="white",
            height=36,
            corner_radius=6,
            command=self.refresh_usb_data
        )
        refresh_btn.pack(side="left", padx=8)

        # Clear Button
        clear_btn = ModernButton(
            btn_frame,
            text="🗑 Clear Data",
            font=("Segoe UI", 12, "bold"),
            fg_color="#D32F2F", hover_color="#C62828",
            text_color="white",
            height=36,
            corner_radius=6,
            command=self.clear_usb_data
        )
        clear_btn.pack(side="left", padx=8)

        # Upload Button
        upload_btn = ModernButton(
            btn_frame,
            text="📤 Upload USB Files",
            font=("Segoe UI", 12, "bold"),
            fg_color="#007B43", hover_color="#005C32",
            text_color="white",
            height=36,
            corner_radius=6,
            command=self.upload_usb_file
        )
        upload_btn.pack(side="left", padx=8)

        # Custom header frame
        self.header_row_frame = ctk.CTkFrame(self, fg_color="white", height=42, corner_radius=0)
        self.header_row_frame.pack(fill="x", padx=20, pady=(0, 0))
        self.header_row_frame.pack_propagate(False)
        
        # Canvas inside the frame for horizontal scrolling
        self.header_canvas = tk.Canvas(self.header_row_frame, bg="white", highlightthickness=0, height=42)
        self.header_canvas.pack(side="left", fill="both", expand=True)
        
        # Frame inside the canvas to hold header cells
        self.header_inner_frame = tk.Frame(self.header_canvas, bg="white")
        self.header_canvas.create_window(0, 0, window=self.header_inner_frame, anchor="nw")
        
        # Right-side spacer to account for the vertical scrollbar width
        self.header_scrollbar_spacer = ctk.CTkFrame(
            self.header_row_frame,
            fg_color="white",
            corner_radius=0,
            width=16,
            height=42
        )
        self.header_scrollbar_spacer.pack(side="right", fill="y")
        
        # Green border line under headers
        self.border_line = ctk.CTkFrame(self, fg_color="#007B43", height=2, corner_radius=0)
        self.border_line.pack(fill="x", padx=20, pady=(0, 5))

        # === Modern Table Container ===
        self.table_frame = ctk.CTkFrame(
            self,
            corner_radius=15,
            border_width=1,
            border_color="#E0E0E0",
            height=250
        )
        self.table_frame.pack(fill="both", expand=True, padx=20, pady=(0, 10))

        # === Column Configuration (Identical to LiveDataPage) ===
        cols = [
            "Date",
            "Time",
            "Reading",
            "Offset",
            "Status",
            "AirGauge ID",
            "Channel",
            "Drawing",
            "User ID",
            "Component ID",  
            "Item",
            "Machine ID",
            "Customer",
            "UTL",
            "LTL"
        ]

        self.create_table(cols)

    def create_table(self, cols):
        """Create empty table identical to Live Data."""
        # Holder grid
        for w in self.table_frame.winfo_children():
            try: w.destroy()
            except: pass

        # Emojis and column display names
        header_info = [
            ("Date", "📅"),
            ("Time", "🕐"),
            ("Reading", "∿"),
            ("Offset", "⌖"),
            ("Status", "ℹ️"),
            ("AirGauge ID", "🏷️"),
            ("Channel", "📊"),
            ("Drawing", "📄"),
            ("User ID", "👤"),
            ("Component ID", "📦"),
            ("Item", "📋"),
            ("Machine ID", "⚙️"),
            ("Customer", "👥"),
            ("UTL", "🔗"),
            ("LTL", "🔗")
        ]

        # Clear custom header inner frame children
        for child in self.header_inner_frame.winfo_children():
            try: child.destroy()
            except: pass

        self.header_widgets = []
        
        # Create custom header cells packed horizontally inside self.header_inner_frame
        for name, emoji in header_info:
            cell = ctk.CTkFrame(
                self.header_inner_frame,
                fg_color="white",
                corner_radius=0,
                height=42,
                border_width=1,
                border_color="#E0E0E0"
            )
            cell.pack(side="left", fill="y")
            cell.pack_propagate(False)
            
            content_frame = ctk.CTkFrame(cell, fg_color="transparent")
            content_frame.pack(expand=True)
            
            icon_lbl = ctk.CTkLabel(
                content_frame,
                text=emoji,
                font=("Segoe UI", 12, "normal"),
                text_color="#007B43",
                anchor="center"
            )
            icon_lbl.pack(side="left", padx=(0, 4))
            
            lbl = ctk.CTkLabel(
                content_frame,
                text=name,
                font=("Segoe UI", 11, "bold"),
                text_color="#1A1A1A",
                anchor="center"
            )
            lbl.pack(side="left")
            self.header_widgets.append(cell)

        try:
            self.sheet = tksheet.Sheet(
                self.table_frame,
                headers=cols,
                data=[],
                show_header=False, # HIDE DEFAULT HEADERS
                show_x_scrollbar=True,
                show_y_scrollbar=True,
                show_row_index=False,
                show_top_left=False
            )
            self.sheet.grid(row=0, column=0, sticky="nsew")
            self.table_frame.grid_rowconfigure(0, weight=1)
            self.table_frame.grid_columnconfigure(0, weight=1)

            # Basic configuration
            self.sheet.enable_bindings()
            self.sheet.disable_bindings("edit_cell", "rc_insert_column", "rc_delete_column",
                                         "rc_insert_row", "rc_delete_row")
            try:
                self.sheet.set_options(
                    theme="light",
                )
                self.sheet.hide_row_index()
            except Exception:
                pass

            # Apply initial column widths after widget is fully drawn
            self.sheet.after(100, self._apply_column_widths)

            # Bind resize to the table frame (like LiveDataPage) for reliable width detection
            self.table_frame.bind("<Configure>", self._on_table_resize)
            
            # Bind scrollbar syncing after sheet is initialized
            try:
                orig_xscroll = self.sheet.MT.cget("xscrollcommand")
                def sync_scroll(first, last):
                    if orig_xscroll:
                        try:
                            # Safely evaluate as a Tcl script to split command name and arguments correctly
                            self.sheet.tk.eval(f"{orig_xscroll} {first} {last}")
                        except Exception:
                            pass
                    try:
                        self.header_canvas.xview_moveto(first)
                    except Exception:
                        pass
                self.sheet.MT.configure(xscrollcommand=sync_scroll)
            except Exception as e:
                print("Failed to sync scrollbar:", e)
            
            try:
                self.sheet.redraw()
            except Exception:
                pass

        except Exception as e:
            print("=== USB create_table failed ===")
            import traceback
            traceback.print_exc()
            err_lbl = ctk.CTkLabel(self.table_frame, text=f"Error loading table: {e}", text_color="red")
            err_lbl.grid(row=0, column=0, pady=20)

    # ── Minimum pixel widths per column (Date, Time, Reading, Offset, Status,
    #    AirGauge ID, Channel, Drawing, User ID, Component ID, Item,
    #    Machine ID, Customer, UTL, LTL)
    _COL_MIN_WIDTHS = [90, 80, 90, 65, 105, 90, 70, 90, 105, 110, 110, 105, 110, 80, 80]
    # Proportional weights that control how remaining space is shared
    _COL_WEIGHTS    = [9,  8,  9,  6,  10,  9,  7,  9,  10,  11,  11,  10,  11,  8,  8]

    def _apply_column_widths(self):
        """Calculate and apply column widths so columns fill 100% of the table — zero trailing whitespace."""
        if not hasattr(self, "sheet") or not self.sheet:
            return
        try:
            # Get the actual rendered width of the sheet canvas area
            w = self.sheet.winfo_width()
            if w < 200:
                w = self.table_frame.winfo_width()
            if w < 200:
                return

            scrollbar_pad = 18                       # vertical scrollbar
            usable = w - scrollbar_pad

            n = len(self._COL_MIN_WIDTHS)
            total_min = sum(self._COL_MIN_WIDTHS)
            total_weight = sum(self._COL_WEIGHTS)
            extra = max(0, usable - total_min)

            widths = [
                self._COL_MIN_WIDTHS[i] + int(extra * self._COL_WEIGHTS[i] / total_weight)
                for i in range(n)
            ]

            # Absorb any leftover rounding pixels into the last column
            leftover = usable - sum(widths)
            if leftover > 0:
                widths[-1] += leftover

            self.sheet.set_column_widths(widths)
            self.sheet.redraw()

            # Sync header cells widths
            for idx, col_w in enumerate(widths):
                if idx < len(self.header_widgets):
                    self.header_widgets[idx].configure(width=col_w)
            
            # Update canvas scroll region
            self.header_canvas.configure(scrollregion=(0, 0, sum(widths), 42))
        except Exception:
            pass

    def _on_table_resize(self, event):
        """Dynamically resize columns to fill the exact screen width without empty whitespace."""
        if not hasattr(self, "sheet") or not self.sheet:
            return
        # Debounce: cancel any pending resize call
        if hasattr(self, "_resize_after_id") and self._resize_after_id:
            try:
                self.sheet.after_cancel(self._resize_after_id)
            except Exception:
                pass
        self._resize_after_id = self.sheet.after(80, self._apply_column_widths)

    def edit_usb_assignments(self):
        """Re-open popup for currently loaded rows to amend Customer/Item."""
        if not hasattr(self, "loaded_usb_rows") or not self.loaded_usb_rows:
            messagebox.showinfo("Edit", "No USB data is currently loaded to edit.")
            return

        unique_combos = set()
        for row_data in self.loaded_usb_rows:
            values = row_data[0]
            ag_id = values[5]
            ch_no = values[6]
            dwg = values[7]
            unique_combos.add((ag_id, ch_no, dwg))

        unique_combinations = list(unique_combos)
        UsbAssignmentPopup(
            self.winfo_toplevel(), 
            unique_combinations, 
            self.customers_list, 
            self.items_list, 
            self._apply_edited_assignments
        )

    def _apply_edited_assignments(self, assignments):
        """Apply the edited mapping to existing rows in memory and UI."""
        updated_count = 0
        for i, row_data in enumerate(self.loaded_usb_rows):
            values, status = row_data
            ag_id = values[5]
            ch_no = values[6]
            dwg = values[7]
            combo = (ag_id, ch_no, dwg)
            
            if combo in assignments:
                new_cust = assignments[combo]["customer"]
                new_item = assignments[combo]["item"]

                if values[10] != new_item or values[12] != new_cust:
                    self.loaded_usb_rows[i][0][10] = new_item
                    self.loaded_usb_rows[i][0][12] = new_cust
                    if getattr(self, "sheet", None):
                        self.sheet.set_cell_data(i, 10, new_item, redraw=False)
                        self.sheet.set_cell_data(i, 12, new_cust, redraw=False)
                    updated_count += 1
                        
        if getattr(self, "sheet", None):
            self.sheet.redraw()

        if updated_count > 0:
            messagebox.showinfo("Edit Complete", f"Successfully updated {updated_count} rows with new assignments.")

    def save_usb_data_to_db(self):
        """Permanently save un-saved loaded rows to production_data.db measurements table."""
        if not hasattr(self, "loaded_usb_rows") or not self.loaded_usb_rows:
            messagebox.showinfo("Save", "No USB data loaded to save.")
            return
            
        unsaved_indices = [i for i in range(len(self.loaded_usb_rows)) if i not in self.saved_usb_indices]
        if not unsaved_indices:
            messagebox.showinfo("Save", "All currently loaded data has already been saved.")
            return

        if not messagebox.askyesno("Confirm Save", f"Are you sure you want to permanently save {len(unsaved_indices)} new rows to the database?"):
            return

        try:
            import sqlite3
            conn = sqlite3.connect(resource_path("production_data.db"))
            cursor = conn.cursor()
            
            # Create table if it doesn't exist
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS measurements (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    source TEXT,
                    date TEXT,
                    time TEXT,
                    reading TEXT,
                    offset TEXT,
                    status TEXT,
                    airgauge_id TEXT,
                    channel TEXT,
                    drawing TEXT,
                    user_id TEXT,
                    component_id TEXT,
                    item TEXT,
                    machine_id TEXT,
                    customer TEXT,
                    utl TEXT,
                    ltl TEXT,
                    upload_timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            # Add columns if they do not exist
            try:
                cursor.execute("SELECT utl, ltl FROM measurements LIMIT 1")
            except sqlite3.OperationalError:
                try:
                    cursor.execute("ALTER TABLE measurements ADD COLUMN utl TEXT")
                    cursor.execute("ALTER TABLE measurements ADD COLUMN ltl TEXT")
                except Exception:
                    pass
            
            insert_stmt = '''
                INSERT INTO measurements (
                    source, date, time, reading, offset, status, airgauge_id, channel, 
                    drawing, user_id, component_id, item, machine_id, customer, utl, ltl
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            '''
            
            # Extract just the values array from (values, status) tuples for UN-SAVED rows
            data_to_insert = [
                ["USB"] + self.loaded_usb_rows[idx][0] for idx in unsaved_indices
            ]
            
            cursor.executemany(insert_stmt, data_to_insert)
            conn.commit()
            conn.close()
            
            # Mark all as saved
            self.saved_usb_indices.update(unsaved_indices)
            self.save_btn.configure(state="disabled")
            
            messagebox.showinfo("Save Success", f"Successfully saved {len(data_to_insert)} new records to the database!")
            
        except Exception as e:
            messagebox.showerror("Database Error", f"Failed to save USB data: {e}")

    def clear_usb_data(self):
        """Clear all loaded USB data from memory and UI."""
        if messagebox.askyesno("Confirm Clear", "Are you sure you want to clear all loaded USB data?"):
            self.loaded_usb_rows = []
            if getattr(self, "sheet", None):
                self.sheet.set_sheet_data([])
                messagebox.showinfo("Clear Data", "All USB Data cleared successfully.")

    def refresh_usb_data(self):
        """Re-read Master settings and update Item/Customer for all loaded rows."""
        self.comp_json = getattr(self.app, 'comp_json', {}) # Refresh from memory (which is refreshed elsewhere)
        
        # If no data is loaded, do nothing
        if not hasattr(self, "loaded_usb_rows") or not self.loaded_usb_rows:
            messagebox.showinfo("Refresh", "No USB data is currently loaded to refresh.")
            return

        updated_count = 0
        for i, row_data in enumerate(self.loaded_usb_rows):
            values, status = row_data
            airgauge_id = values[5]
            channel_no = values[6]
            
            # Recalculate based on current comp_json
            new_item_name = ""
            new_customer = ""
            try:
                ag_key = str(airgauge_id).upper()
                if not ag_key.startswith("AG"):
                    ag_key = f"AG{airgauge_id}"
                    
                ch_key = str(channel_no).upper()
                if not ch_key.startswith("CH"):
                    ch_key = f"CH{channel_no}"
                    
                if ag_key in self.comp_json and ch_key in self.comp_json[ag_key]:
                    new_item_name = self.comp_json[ag_key][ch_key].get("item_name", "")
                    new_customer = self.comp_json[ag_key][ch_key].get("customer_name", "")
            except Exception:
                pass

            # Update if changed (Note: UTL/LTL are preserved from original USB string)
            if values[10] != new_item_name or values[12] != new_customer:
                self.loaded_usb_rows[i][0][10] = new_item_name # Item
                self.loaded_usb_rows[i][0][12] = new_customer  # Customer
                
                # Update UI immediately
                if getattr(self, "sheet", None):
                    self.sheet.set_cell_data(i, 10, new_item_name, redraw=False)
                    self.sheet.set_cell_data(i, 12, new_customer, redraw=False)
                updated_count += 1
                
        if getattr(self, "sheet", None):
            self.sheet.redraw()

        messagebox.showinfo("Refresh Complete", f"Refreshed USB Data. {updated_count} rows were updated with new Master settings.")


    def upload_usb_file(self):
        """Parse raw USB files, aggregate unique combos, and trigger assignment popup."""
        filenames = filedialog.askopenfilenames(
            title="Select USB Data File(s)",
            filetypes=(("Text Files", "*.txt"), ("CSV Files", "*.csv"), ("All Files", "*.*"))
        )
        
        if not filenames:
            return

        if not hasattr(self, "loaded_usb_rows"):
            self.loaded_usb_rows = []

        temp_parsed_rows = []
        unique_combos = set()
        error_files = []

        for filename in filenames:
            try:
                basename = os.path.basename(filename).split('.')[0]
                if len(basename) < 8:
                    error_files.append((basename, "Filename too short to parse properties."))
                    continue

                # Example: 210618B2 (DD=21, MM=06, YY=18, B=MachineID, 2=ProgramNo)
                date_str = f"{basename[0:2]}/{basename[2:4]}/20{basename[4:6]}"
                
                ag_char = basename[6].upper()
                airgauge_id = str(ord(ag_char) - 64) if ag_char.isalpha() else str(ag_char)
                channel_no = int(basename[7])

                with open(filename, 'r', encoding='utf-8', errors='replace') as f:
                    for line in f:
                        line = line.strip()
                        if not (line.startswith("*") and line.endswith("#") and len(line) >= 54):
                            continue
                        
                        # Time
                        hours = line[1:3]
                        minutes = line[3:5]
                        seconds = line[5:7]
                        time_str = f"{hours}:{minutes}:{seconds}"

                        offset_value = "0"
                        
                        # DWG DIA
                        draw_raw = line[22:29]
                        if not draw_raw.isdigit():
                            draw_raw = "".join(c for c in draw_raw if c.isdigit()).ljust(7, "0")
                        drawing_val = f"{draw_raw[:3]}.{draw_raw[3:]}" if len(draw_raw) >= 6 else draw_raw
                        
                        # Calc Value (Reading)
                        calc_raw = line[31:38]
                        if not calc_raw.isdigit():
                            calc_raw = "".join(c for c in calc_raw if c.isdigit()).ljust(7, "0")
                        Reading = f"{calc_raw[:3]}.{calc_raw[3:]}" if len(calc_raw) >= 6 else calc_raw

                        # Status
                        status_raw = line[38].upper() if len(line) > 38 else "U"
                        if status_raw == "A":
                            status = "ACCEPTED"
                        elif status_raw == "B":
                            status = "REJECTED"
                        elif status_raw == "C":
                            status = "REWORK"
                        else:
                            status = "UNKNOWN"
                        
                        compID = line[39:49].strip() if len(line) >= 49 else ""
                        
                        # User ID Resolution
                        raw_userID = line[49:52].strip() if len(line) >= 52 else ""
                        if raw_userID and raw_userID in self.employee_map:
                            userID = f"{self.employee_map[raw_userID]} ({raw_userID})"
                        else:
                            userID = raw_userID

                        # Machine ID Resolution
                        raw_cncID = line[52:55].strip() if len(line) >= 55 else ""
                        if raw_cncID and raw_cncID in self.machine_map:
                            cncID = f"{self.machine_map[raw_cncID]} ({raw_cncID})"
                        else:
                            cncID = raw_cncID

                        # Note: Item and Customer will be assigned via Popup
                        item_name = ""
                        customer_name = ""
                        
                        # High Tolerance (8:15)
                        utl_raw = line[8:15]
                        if not utl_raw.isdigit():
                            utl_raw = "".join(c for c in utl_raw if c.isdigit()).ljust(7, "0")
                        utl = f"{utl_raw[:3]}.{utl_raw[3:]}" if len(utl_raw) >= 6 else utl_raw

                        # Low Tolerance (15:22)
                        ltl_raw = line[15:22]
                        if not ltl_raw.isdigit():
                            ltl_raw = "".join(c for c in ltl_raw if c.isdigit()).ljust(7, "0")
                        ltl = f"{ltl_raw[:3]}.{ltl_raw[3:]}" if len(ltl_raw) >= 6 else ltl_raw
                        
                        values = [
                            date_str,      # cols[0]: Date
                            time_str,      # cols[1]: Time
                            Reading,       # cols[2]: Reading
                            offset_value,  # cols[3]: Offset
                            status,        # cols[4]: Status
                            airgauge_id,   # cols[5]: AirGauge ID
                            channel_no,    # cols[6]: Channel
                            drawing_val,   # cols[7]: Drawing
                            userID,        # cols[8]: User ID
                            compID,        # cols[9]: Component ID
                            item_name,     # cols[10]: Item (Pending)
                            cncID,         # cols[11]: Machine ID
                            customer_name, # cols[12]: Customer (Pending)
                            utl,           # cols[13]: UTL
                            ltl            # cols[14]: LTL
                        ]
                        
                        temp_parsed_rows.append((values, status))
                        unique_combos.add((airgauge_id, channel_no, drawing_val))

            except Exception as e:
                error_files.append((os.path.basename(filename), str(e)))

        if not temp_parsed_rows:
            if error_files:
                msg = "Errors encountered:\n" + "\n".join(f"- {f}: {e}" for f, e in error_files)
                messagebox.showerror("Upload Error", msg)
            else:
                messagebox.showinfo("USB Upload", "No valid data lines found in the selected files.")
            return

        # Show Assignment Popup
        unique_combinations = list(unique_combos)
        UsbAssignmentPopup(
            self.winfo_toplevel(), 
            unique_combinations, 
            self.customers_list, 
            self.items_list, 
            lambda assignments: self._finish_usb_upload(temp_parsed_rows, assignments, len(filenames), error_files)
        )

    def _finish_usb_upload(self, temp_parsed_rows, assignments, total_files, error_files):
        """Callback from UsbAssignmentPopup to apply assignments and insert into table."""
        start_row = 0
        if getattr(self, "sheet", None):
            start_row = self.sheet.get_total_rows()

        total_new_rows = 0

        # Apply assignments
        for row_data in temp_parsed_rows:
            values, status = row_data
            ag_id = values[5]
            ch_no = values[6]
            dwg = values[7]
            
            combo = (ag_id, ch_no, dwg)
            if combo in assignments:
                values[10] = assignments[combo]["item"]     # Item
                values[12] = assignments[combo]["customer"] # Customer
            
            self.loaded_usb_rows.append((values, status))
            total_new_rows += 1

            # Insert into Table
            if getattr(self, "sheet", None):
                try:
                    self.sheet.insert_row(data=values, index="end")
                except TypeError:
                    self.sheet.insert_row(values)
                
                row_idx = self.sheet.get_total_rows() - 1
                if status == "REJECTED":
                    self.sheet.highlight_rows([row_idx], bg="#FFCDD2")

        # Scroll to the first newly added row
        if getattr(self, "sheet", None) and start_row < self.sheet.get_total_rows():
            self.sheet.see(start_row, 0)
            
        # Enable save button if we added new rows
        if total_new_rows > 0:
            self.save_btn.configure(state="normal")

        # Build summary statement
        summary = f"Successfully appended {total_new_rows} entries from {total_files - len(error_files)} file(s)!"
        if error_files:
            summary += f"\n\nErrors encountered in {len(error_files)} file(s):"
            for ef, msg in error_files:
                summary += f"\n- {ef}: {msg}"
                
        messagebox.showinfo("USB Upload Summary", summary)


class LiveDataPage(ctk.CTkFrame):
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        
        # --- Performance: Batching ---
        self._data_buffer = []
        self._batch_interval_ms = 300  # flush every 300ms (was 200ms)
        
        self.serial_no = 0
        self.rows = []

        # Load component setup JSON once (cache)
        self.comp_json = self.load_component_json()

        self.build_ui()
        self.load_existing_data()
        self.app.data_manager.subscribe(self)
        
        # Start UI flush loop
        self.after(self._batch_interval_ms, self._ui_update_loop)

        # Background DB writer: keeps SQLite completely off the main/UI thread
        self._db_write_queue = queue.Queue(maxsize=2000)
        self._db_writer_thread = threading.Thread(target=self._db_writer_loop, daemon=True)
        self._db_writer_thread.start()

    def _ui_update_loop(self):
        """Periodically flush buffer to UI — max 5 rows per tick to stay responsive."""
        try:
            if getattr(self.app, '_closing', False) or not self.winfo_exists():
                return
        except Exception:
            return
        if self._data_buffer:
            # Process at most 5 rows per tick; remaining stay buffered for next tick
            chunk = self._data_buffer[:5]
            del self._data_buffer[:5]
            for parsed in chunk:
                self._process_row_visuals(parsed)
        self.after(self._batch_interval_ms, self._ui_update_loop)

    def _make_header_icon(self, kind, size=24, color="#009A55"):
        """Create simple green line icons for the Live Data header."""
        img = Image.new("RGBA", (size, size), (0, 0, 0, 0))
        draw = ImageDraw.Draw(img)
        w = size
        c = color
        lw = max(2, size // 12)

        if kind == "date":
            draw.rectangle((4, 6, w - 4, w - 4), outline=c, width=lw)
            draw.line((4, 11, w - 4, 11), fill=c, width=lw)
            draw.line((8, 3, 8, 8), fill=c, width=lw)
            draw.line((w - 8, 3, w - 8, 8), fill=c, width=lw)
        elif kind == "time":
            draw.ellipse((4, 4, w - 4, w - 4), outline=c, width=lw)
            draw.line((w // 2, w // 2, w // 2, 8), fill=c, width=lw)
            draw.line((w // 2, w // 2, w - 8, w // 2 + 4), fill=c, width=lw)
        elif kind == "reading":
            draw.arc((4, 5, w - 4, w + 10), 200, 340, fill=c, width=lw)
            draw.line((w // 2, w // 2 + 5, w - 7, 8), fill=c, width=lw)
            for x, y in ((8, 15), (11, 10), (17, 8), (20, 14)):
                draw.ellipse((x - 1, y - 1, x + 1, y + 1), fill=c)
        elif kind == "offset":
            draw.ellipse((5, 5, w - 5, w - 5), outline=c, width=lw)
            draw.ellipse((w // 2 - 3, w // 2 - 3, w // 2 + 3, w // 2 + 3), fill=c)
            draw.line((w // 2, 1, w // 2, 7), fill=c, width=lw)
            draw.line((w // 2, w - 7, w // 2, w - 1), fill=c, width=lw)
            draw.line((1, w // 2, 7, w // 2), fill=c, width=lw)
            draw.line((w - 7, w // 2, w - 1, w // 2), fill=c, width=lw)
        elif kind == "status":
            draw.ellipse((5, 5, w - 5, w - 5), outline=c, width=lw)
            draw.line((w // 2, 11, w // 2, w - 8), fill=c, width=lw)
            draw.ellipse((w // 2 - 1, 7, w // 2 + 1, 9), fill=c)
        elif kind == "tag":
            draw.line((5, 5, 14, 5, w - 4, 15, 15, w - 4, 5, 14, 5, 5), fill=c, width=lw)
            draw.ellipse((8, 8, 11, 11), fill=c)
        elif kind == "channel":
            for i, h in enumerate((7, 11, 15, 20)):
                x = 5 + i * 5
                draw.line((x, w - 4, x, w - h), fill=c, width=lw)
        elif kind == "drawing":
            draw.rectangle((6, 4, w - 6, w - 4), outline=c, width=lw)
            draw.line((w - 10, 4, w - 6, 8, w - 10, 8), fill=c, width=lw)
        elif kind == "user":
            draw.ellipse((w // 2 - 4, 4, w // 2 + 4, 12), outline=c, width=lw)
            draw.arc((5, 11, w - 5, w + 6), 205, 335, fill=c, width=lw)
        elif kind == "cube":
            draw.line((w // 2, 4, w - 5, 9, w - 5, w - 7, w // 2, w - 3, 5, w - 7, 5, 9, w // 2, 4), fill=c, width=lw)
            draw.line((5, 9, w // 2, 14, w - 5, 9), fill=c, width=lw)
            draw.line((w // 2, 14, w // 2, w - 3), fill=c, width=lw)
        elif kind == "live_dot":
            draw.ellipse((4, 4, w - 4, w - 4), fill=c)
        else:
            draw.ellipse((5, 5, w - 5, w - 5), outline=c, width=lw)

        return ctk.CTkImage(img, size=(size, size))
    

    def build_ui(self):
        # === Modern Header Card ===
        header_card = ctk.CTkFrame(
            self,
            fg_color="white",
            corner_radius=15,
            border_width=1,
            border_color="#E0E0E0"
        )
        header_card.pack(fill="x", padx=20, pady=(10, 5))
        
        # Green vertical accent bar on the left edge
        green_bar = ctk.CTkFrame(header_card, fg_color="#007B43", width=6, corner_radius=0)
        green_bar.place(x=0, y=0, relheight=1)
        
        # Header content
        header_content = ctk.CTkFrame(header_card, fg_color="transparent")
        header_content.pack(fill="x", padx=(25, 20), pady=10)
        
        # Title with icon
        title_frame = ctk.CTkFrame(header_content, fg_color="transparent")
        title_frame.pack(side="left")
        
        # Rounded green square for logo
        logo_frame = ctk.CTkFrame(title_frame, fg_color="#007B43", width=40, height=40, corner_radius=8)
        logo_frame.pack(side="left", padx=(0, 12))
        logo_frame.pack_propagate(False)
        
        # White cherry emoji inside logo
        logo_lbl = ctk.CTkLabel(logo_frame, text="🍒", font=("Segoe UI", 18), text_color="white", fg_color="transparent")
        logo_lbl.pack(expand=True)
        
        logo_lbl.configure(text="T", font=("Segoe UI", 24, "bold"))
        
        ctk.CTkLabel(
            title_frame,
            text="Live Data Monitoring",
            font=("Segoe UI", 20, "bold"),
            text_color="#007B43"
        ).pack(side="left")
        
        # Status badge
        status_badge = ctk.CTkFrame(
            header_content,
            fg_color="#E8F5E9",
            corner_radius=18
        )
        status_badge.pack(side="right", padx=10)
        
        ctk.CTkLabel(
            status_badge,
            text="● LIVE",
            font=("Segoe UI", 11, "bold"),
            text_color="#43A047"
        ).pack(padx=15, pady=6)
        for child in status_badge.winfo_children():
            child.destroy()
        self.live_dot_img = self._make_header_icon("live_dot", size=18, color="#00A95C")
        ctk.CTkLabel(status_badge, image=self.live_dot_img, text="", fg_color="transparent").pack(side="left", padx=(14, 6), pady=8)
        ctk.CTkLabel(status_badge, text="LIVE", font=("Segoe UI", 12, "bold"), text_color="#007B43").pack(side="left", padx=(0, 14), pady=8)

        # Custom header frame, matching USB Data page
        self.header_row_frame = ctk.CTkFrame(self, fg_color="white", height=42, corner_radius=0)
        self.header_row_frame.pack(fill="x", padx=20, pady=(0, 0))
        self.header_row_frame.pack_propagate(False)
        
        self.header_canvas = tk.Canvas(self.header_row_frame, bg="white", highlightthickness=0, height=42)
        self.header_canvas.pack(side="left", fill="both", expand=True)
        
        self.header_inner_frame = tk.Frame(self.header_canvas, bg="white")
        self.header_canvas.create_window(0, 0, window=self.header_inner_frame, anchor="nw")
        
        self.header_scrollbar_spacer = ctk.CTkFrame(
            self.header_row_frame,
            fg_color="white",
            corner_radius=0,
            width=16,
            height=42
        )
        self.header_scrollbar_spacer.pack(side="right", fill="y")
        
        # Green border line under headers
        self.border_line = ctk.CTkFrame(self, fg_color="#007B43", height=2, corner_radius=0)
        self.border_line.pack(fill="x", padx=20, pady=(0, 5))

        # === Modern Table Container ===
        self.table_frame = ctk.CTkFrame(
            self,
            corner_radius=15,
            border_width=1,
            border_color="#E0E0E0",
            height=250
        )
        self.table_frame.pack(fill="both", expand=True, padx=20, pady=(0, 10))

        # === Column Configuration ===
        cols = [
            "Date",
            "Time",
            "Reading",
            "Offset",
            "Status",
            "AirGauge ID",
            "Channel",
            "Drawing",
            "User ID",
            "Component ID"
        ]

        self.create_table(cols)
        
    def load_component_json(self):
        try:
            conn = sqlite3.connect("component_setup.db")
            cursor = conn.cursor()
            cursor.execute("SELECT airgauge_id, channel, properties FROM component_setup")
            rows = cursor.fetchall()
            conn.close()
            comp_map = {}
            for ag_id, ch, props_str in rows:
                if ag_id not in comp_map:
                    comp_map[ag_id] = {}
                try:
                    comp_map[ag_id][ch] = json.loads(props_str)
                except:
                    comp_map[ag_id][ch] = {}
            return comp_map
        except Exception:
            return {}

    def get_customer_name(self, airgauge_id, ch):
        """
        Resolve customer name from component JSON using AirGauge ID + Channel
        """
        try:
            ag_key = str(airgauge_id).upper()
            if not ag_key.startswith("AG"):
                ag_key = f"AG{airgauge_id}"
                
            ch_key = str(ch).upper()
            if not ch_key.startswith("CH"):
                ch_key = f"CH{ch}"

            return (
                self.comp_json
                .get(ag_key, {})
                .get(ch_key, {})
                .get("customer_name", "")
            )
        except Exception:
            return ""

    def create_table(self, cols):
        """
        Create a live data table using tksheet if available, otherwise fallback to ttk.Treeview.
        Keeps tags/colors compatible with previous logic (evenrow/oddrow/mismatch).
        """
        self.display_col_count = len(cols)
        # Holder grid
        for w in self.table_frame.winfo_children():
            try: w.destroy()
            except: pass

        # Emojis and column display names
        header_info = [
            ("Date", "📅"),
            ("Time", "🕐"),
            ("Reading", "📊"),
            ("Offset", "🎯"),
            ("Status", "ℹ️"),
            ("AirGauge ID", "🏷️"),
            ("Channel", "📶"),
            ("Drawing", "📄"),
            ("User ID", "👤"),
            ("Component ID", "📦"),
            ("Item", "📋"),
            ("CNC ID", "💻"),
            ("Customer", "🏢")
        ]

        header_info = [header for header in header_info if header[0] in cols]

        # Clear custom header inner frame children
        for child in self.header_inner_frame.winfo_children():
            try: child.destroy()
            except: pass

        self.header_widgets = []
        
        # Create custom header cells packed horizontally inside self.header_inner_frame
        for name, emoji in header_info:
            cell = ctk.CTkFrame(
                self.header_inner_frame,
                fg_color="white",
                corner_radius=0,
                height=42,
                border_width=1,
                border_color="#E0E0E0"
            )
            cell.pack(side="left", fill="y")
            cell.pack_propagate(False)
            
            content_frame = ctk.CTkFrame(cell, fg_color="transparent")
            content_frame.pack(expand=True)
            
            icon_lbl = ctk.CTkLabel(
                content_frame,
                text=emoji,
                font=("Segoe UI", 12, "normal"),
                text_color="#007B43",
                anchor="center"
            )
            icon_lbl.pack(side="left", padx=(0, 4))

            lbl = ctk.CTkLabel(
                content_frame,
                text=name,
                font=("Segoe UI", 11, "bold"),
                text_color="#1A1A1A",
                anchor="center"
            )
            lbl.pack(side="left")
            self.header_widgets.append(cell)

        self.use_tksheet = False
        try:
            # Try creating tksheet
            self.sheet = tksheet.Sheet(
                self.table_frame,
                headers=cols,
                data=[],
                show_header=False,
                show_row_index=False,
                show_x_scrollbar=True,
                show_y_scrollbar=True,
                font=("Segoe UI", 11, "normal"),
                row_height=28
            )
            self.sheet.grid(row=0, column=0, sticky="nsew")
            # Grid config
            self.table_frame.grid_rowconfigure(0, weight=1)
            self.table_frame.grid_columnconfigure(0, weight=1)

            # Column width ratio-based resizer
            def resize_sheet(ev=None):
                try:
                    total_w = self.table_frame.winfo_width()
                    if total_w < 100:
                        total_w = 800
                    
                    # Deduct spacing for the vertical scrollbar
                    usable_w = total_w - 18
                    
                    ratios = [0.09, 0.10, 0.10, 0.09, 0.10, 0.12, 0.10, 0.10, 0.09, 0.11]
                    widths = [int(r * usable_w) for r in ratios]
                    # Adjust last column to match exactly
                    widths[-1] += usable_w - sum(widths)
                    
                    # Apply widths to tksheet
                    for idx, w in enumerate(widths):
                        try:
                            self.sheet.column_width(column=idx, width=w, only_set_if_too_small=False)
                        except TypeError:
                            self.sheet.column_width(idx, w)
                    
                    try:
                        self.sheet.refresh()
                    except Exception:
                        pass

                    # Sync header cells widths
                    for idx, w in enumerate(widths):
                        if idx < len(self.header_widgets):
                            self.header_widgets[idx].configure(width=w)
                    try:
                        self.header_inner_frame.update_idletasks()
                        self.header_canvas.configure(scrollregion=self.header_canvas.bbox("all"))
                    except Exception:
                        pass
                except Exception as e:
                    print("Resize sheet error:", e)

            self.resize_sheet = resize_sheet
            self.table_frame.bind("<Configure>", self.resize_sheet)
            
            # Enable bindings
            try:
                self.sheet.enable_bindings(("single_select", "row_select", "rc_select"))
            except Exception:
                pass

            try:
                orig_xscroll = self.sheet.MT.cget("xscrollcommand")
                def sync_scroll(first, last):
                    if orig_xscroll:
                        try:
                            self.sheet.tk.eval(f"{orig_xscroll} {first} {last}")
                        except Exception:
                            pass
                    try:
                        self.header_canvas.xview_moveto(first)
                    except Exception:
                        pass
                self.sheet.MT.configure(xscrollcommand=sync_scroll)
            except Exception as e:
                print("Failed to sync Live Data header scrollbar:", e)

            self.use_tksheet = True
            self._sheet_row_count = 0

        except Exception as e:
            # fallback to Treeview (keeps previous behavior but hides heading)
            self.use_tksheet = False
            style = ttk.Style()
            style.configure("LiveData.Treeview", font=("Segoe UI", 11), rowheight=28)
            # Hide the headings by configuring minimal size
            style.configure("LiveData.Treeview.Heading", font=("Segoe UI", 1), rowheight=1)
            style.map("LiveData.Treeview", background=[("selected", "#BBDEFB")])

            self.table = ttk.Treeview(
                self.table_frame, 
                columns=cols, 
                show="headings", 
                height=12,
                style="LiveData.Treeview"
            )
            for c in cols:
                self.table.heading(c, text="") # Empty header text
                self.table.column(c, width=100, anchor="center")

            y_scroll = ttk.Scrollbar(self.table_frame, orient="vertical", command=self.table.yview)
            x_scroll = ttk.Scrollbar(self.table_frame, orient="horizontal", command=self.table.xview)
            self.table.configure(yscroll=y_scroll.set, xscroll=x_scroll.set)
            self.table.grid(row=1, column=0, sticky="nsew")
            y_scroll.grid(row=1, column=1, sticky="ns")
            x_scroll.grid(row=2, column=0, sticky="ew")

            self.table_frame.grid_rowconfigure(1, weight=1)
            self.table_frame.grid_columnconfigure(0, weight=1)

            # Column width ratio-based resizer for fallback Treeview
            def resize_tree(ev=None):
                try:
                    total_w = self.table_frame.winfo_width()
                    if total_w < 100:
                        total_w = 800
                    
                    usable_w = total_w - 18
                    ratios = [0.09, 0.10, 0.10, 0.09, 0.10, 0.12, 0.10, 0.10, 0.09, 0.11]
                    widths = [int(r * usable_w) for r in ratios]
                    widths[-1] += usable_w - sum(widths)
                    
                    for idx, w in enumerate(widths):
                        col_name = cols[idx]
                        self.table.column(col_name, width=w, anchor="center")
                    
                    # Sync header cells widths
                    for idx, w in enumerate(widths):
                        if idx < len(self.header_widgets):
                            self.header_widgets[idx].configure(width=w)
                    try:
                        self.header_inner_frame.update_idletasks()
                        self.header_canvas.configure(scrollregion=self.header_canvas.bbox("all"))
                    except Exception:
                        pass
                except Exception as e:
                    print("Resize tree error:", e)

            self.resize_sheet = resize_tree
            self.table_frame.bind("<Configure>", self.resize_sheet)

            # preserve tags for compatibility
            self.table.tag_configure("evenrow", background="#f9fff4")
            self.table.tag_configure("oddrow", background="#f4f6f8")
            self.table.tag_configure("mismatch", background="#FFCDD2")   # light red

            self.sheet = None
    def update_with_parsed(self, parsed):
        try:
            if parsed and len(parsed) >= 13:
                self.add_data_row(parsed)
        except Exception as e:
            print("LiveData update error:", e)

    def load_existing_data(self):
        """Restore previous session rows into the sheet or treeview."""
        if not os.path.exists(LIVE_DATA_FILE):
            return
        try:
            with open(LIVE_DATA_FILE, "r") as f:
                data = json.load(f)
            # ensure counter starts correctly
            self.serial_no = 0
            if self.use_tksheet and getattr(self, "sheet", None):
                # tksheet expects a list of lists
                rows = []
                for row in data:
                    self.serial_no += 1
                    rows.append(row[:getattr(self, "display_col_count", len(row))])
                try:
                    self.sheet.set_sheet_data(rows)
                except Exception:
                    # fallback: insert rows one by one
                    for r in rows:
                        try:
                            self.sheet.insert_row(data=r, index="end")
                        except Exception:
                            pass
                self._sheet_row_count = len(rows)
            else:
                # Treeview fallback
                for row in data:
                    self.serial_no += 1
                    tag = "evenrow" if self.serial_no % 2 == 0 else "oddrow"
                    try:
                        self.table.insert("", "end", values=row[:getattr(self, "display_col_count", len(row))], tags=(tag,))
                    except Exception:
                        pass
            self.rows = [row[:getattr(self, "display_col_count", len(row))] for row in data]
        except Exception as e:
            print("Load error:", e)
    def save_to_file(self):
        try:
            with open(LIVE_DATA_FILE, "w") as f:
                json.dump(self.rows, f, indent=2)
        except Exception as e:
            print("Save error:", e)

    def append_to_all_data(self, parsed_tuple, customer_name):
        """Queue a DB write to the background thread — never blocks the UI."""
        try:
            self._db_write_queue.put_nowait((parsed_tuple, customer_name))
        except Exception:
            pass  # Queue full: silently drop (rare; only if disk is extremely slow)

    def _db_writer_loop(self):
        """Background thread: persist live data rows to production_data.db.
        Uses a persistent connection + WAL mode for maximum write throughput.
        """
        import sqlite3 as _sqlite3
        db_path = resource_path("production_data.db")
        CREATE_SQL = '''
            CREATE TABLE IF NOT EXISTS measurements (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                source TEXT, date TEXT, time TEXT, reading TEXT, offset TEXT,
                status TEXT, airgauge_id TEXT, channel TEXT, drawing TEXT,
                user_id TEXT, component_id TEXT, item TEXT, machine_id TEXT,
                customer TEXT, utl TEXT, ltl TEXT,
                upload_timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
            )'''
        INSERT_SQL = '''
            INSERT INTO measurements (
                source, date, time, reading, offset, status, airgauge_id, channel,
                drawing, user_id, component_id, item, machine_id, customer, utl, ltl
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)'''

        conn = None
        while True:
            try:
                item = self._db_write_queue.get()  # blocks until item available
                if item is None:  # shutdown signal
                    break

                parsed_tuple, customer_name = item

                # --- Prepare insert data ---
                data = list(parsed_tuple)
                date_old = str(data[3])
                parts = date_old.split("/")
                if len(parts) == 3 and len(parts[2]) == 2:
                    date_val = f"{parts[0]}/{parts[1]}/20{parts[2]}"
                else:
                    date_val = date_old
                time_str = f"{data[0]}:{data[1]}:{data[2]}"
                item_display = f"{data[13]} - {data[12]}".strip(" -")
                utl = ltl = "0"
                try:
                    ag_key = str(data[7]).upper()
                    if not ag_key.startswith("AG"):
                        ag_key = f"AG{data[7]}"
                    ch_key = str(data[8]).upper()
                    if not ch_key.startswith("CH"):
                        ch_key = f"CH{data[8]}"
                    if hasattr(self, 'app') and hasattr(self.app, 'comp_json'):
                        comp = self.app.comp_json.get(ag_key, {}).get(ch_key, {})
                        utl = comp.get("high_tolerance", "0")
                        ltl = comp.get("low_tolerance", "0")
                except Exception:
                    pass
                insert_data = (
                    "LIVE", date_val, time_str, str(data[4]), str(data[5]),
                    str(data[6]), str(data[7]), str(data[8]), str(data[9]),
                    str(data[10]), str(data[11]), item_display, str(data[14]),
                    customer_name, utl, ltl
                )

                # --- Open/reuse persistent connection ---
                if conn is None:
                    try:
                        conn = _sqlite3.connect(db_path, timeout=10, check_same_thread=False)
                        conn.execute("PRAGMA journal_mode=WAL")  # WAL = much faster writes
                        conn.execute("PRAGMA synchronous=NORMAL")  # safe but faster than FULL
                        conn.execute(CREATE_SQL)
                        conn.commit()
                        # Add tolerance columns if missing (migration guard)
                        try:
                            conn.execute("SELECT utl, ltl FROM measurements LIMIT 1")
                        except _sqlite3.OperationalError:
                            try:
                                conn.execute("ALTER TABLE measurements ADD COLUMN utl TEXT")
                                conn.execute("ALTER TABLE measurements ADD COLUMN ltl TEXT")
                                conn.commit()
                            except Exception:
                                pass
                    except Exception as e:
                        print("DB connect error:", e)
                        conn = None
                        continue

                # --- Insert row ---
                try:
                    conn.execute(INSERT_SQL, insert_data)
                    conn.commit()
                except Exception as e:
                    print("DB insert error:", e)
                    try:
                        conn.close()
                    except Exception:
                        pass
                    conn = None  # reconnect next iteration

            except Exception as e:
                print("DB writer loop error:", e)

        # Cleanup on shutdown
        if conn:
            try:
                conn.close()
            except Exception:
                pass

    def add_data_row(self, parsed_tuple):
        """Buffer incoming data. Called by DataManager thread."""
        self._data_buffer.append(parsed_tuple)

    def _process_row_visuals(self, parsed_tuple):
        try:
            # 🔒 HARD GUARD: page already destroyed
            if not self.winfo_exists():
                return

            # 🔒 HARD GUARD: tksheet already destroyed
            if self.use_tksheet:
                if not getattr(self, "sheet", None):
                    return
                try:
                    if not self.sheet.winfo_exists():
                        return
                except Exception:
                    return

            # 🔒 UI readiness check
            if not (
                (self.use_tksheet and getattr(self, "sheet", None)) or
                (getattr(self, "table", None) and self.table.winfo_exists())
            ):
                self.rows.append(parsed_tuple)
                return

            # -------- SAFE TO CONTINUE --------

            (hh, mm, ss, date, Reading, offset, status,
             airgauge_id, ch,
             drawing_val, userID,
             compID, item_name, item_code, cncID) = parsed_tuple

            try:
                ag_key = str(airgauge_id).upper()
                if not ag_key.startswith("AG"):
                    ag_key = f"AG{airgauge_id}"
                    
                ch_key = str(ch).upper()
                if not ch_key.startswith("CH"):
                    ch_key = f"CH{ch}"
                    
                fresh_item_name = self.comp_json.get(ag_key, {}).get(ch_key, {}).get("item_name", "")
                if fresh_item_name:
                    item_name = fresh_item_name
            except Exception:
                pass

            customer_name = self.get_customer_name(airgauge_id, ch)

            values = [
                date,
                f"{hh}:{mm}:{ss}",
                Reading,
                offset,
                status,
                airgauge_id,
                ch,
                drawing_val,
                userID,
                compID,
                item_name,
                cncID,
                customer_name
            ]
            display_values = values[:getattr(self, "display_col_count", len(values))]

            # Uses cached self.comp_json (performance optimization)

            # Mismatch check
            try:
                ag_key = str(airgauge_id)
                ch_key = f"CH{ch}"
                saved_draw = (
                    self.comp_json.get(ag_key, {})
                    .get(ch_key, {})
                    .get("drawing_value", "")
                )
                mismatch = saved_draw and str(drawing_val).strip() != str(saved_draw).strip()
            except Exception:
                mismatch = False

            # ================ INSERT INTO TABLE =================
            if self.use_tksheet and self.sheet:
                try:
                    self.sheet.insert_row(data=display_values, index="end")
                except TypeError:
                    self.sheet.insert_row(display_values)

                row_idx = self.sheet.get_total_rows() - 1

                if mismatch:
                    self.sheet.highlight_rows([row_idx], bg="#FFCDD2")
                else:
                    bg = "#f9fff4" if row_idx % 2 == 0 else "#f4f6f8"
                    self.sheet.highlight_rows([row_idx], bg=bg)

                self.sheet.see(row_idx, 0)

            else:
                tag = "mismatch" if mismatch else ("evenrow" if self.serial_no % 2 == 0 else "oddrow")
                self.table.insert("", "end", values=display_values, tags=(tag,))
                self.table.yview_moveto(1)

            # bookkeeping
            self.rows.append(display_values)
            self.serial_no += 1

            # Save JSON every 100 rows in background thread (was every 10 rows, on main thread)
            if self.serial_no % 100 == 0:
                threading.Thread(target=self.save_to_file, daemon=True).start()

            # Queue DB write — non-blocking, handled by _db_writer_loop background thread
            self.append_to_all_data(parsed_tuple, customer_name)

        except Exception as e:
            print("Add error:", e)


    def destroy(self):
        try:
            self.app.data_manager.unsubscribe(self)
        except Exception:
            pass
        # Shutdown background DB writer thread gracefully
        try:
            self._db_write_queue.put(None)  # None = shutdown signal
        except Exception:
            pass
        super().destroy()

#===========================================================
#---------------------REPORT PAGE---------------------------
#===========================================================
import json
import datetime
import threading
import tkinter as tk
from tkinter import ttk, messagebox

import customtkinter as ctk

# tksheet and tkcalendar as module-level imports to avoid local shadowing issues.
try:
    from tksheet import Sheet
except Exception:
    Sheet = None

try:
    from tkcalendar import Calendar  # module-level variable used by open_calendar_popup()
except Exception:
    Calendar = None


# Paste-replace your existing ReportPage class with this (uses airgauge_component_config.json)
class ReportPage(ctk.CTkFrame):
    """
    Full ReportPage with:
      - background load of JSONL (non blocking)
      - filters (date/time, item, operator, machine, airgauge)
      - fast filter worker thread + debounce
      - calendar popup (tkcalendar) and inline time spinner
      - tksheet or treeview fallback
      - Refresh and Delete buttons (delete requires password every time)
      - exact-file-line deletion by stored file_index
    """
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app

        # Configure premium styling for ttk.Combobox drop-downs
        style = ttk.Style()
        try:
            style.configure('TCombobox',
                            fieldbackground='white',
                            background='white',
                            foreground='#333333',
                            bordercolor='#E0E0E0',
                            arrowcolor='#2E7D32',
                            arrowsize=12,
                            font=("Segoe UI", 11))
            style.map('TCombobox',
                      fieldbackground=[('readonly', 'white')],
                      bordercolor=[('focus', '#2E7D32'), ('hover', '#43A047')])
        except Exception as e:
            print("Error styling combobox:", e)

        # Config
        self.AIR_CONFIG_FILE = "airgauge_component_config.json"
        self.DELETE_PASSWORD = "delete@cherry"

        # Filter data sources
        self.items_list = []
        self.operators_list = []
        self.machines_list = []

        # UI vars (defaults = wide range for USB data)
        today = datetime.date.today().strftime("%d/%m/%Y")
        past = "01/01/2020"
        self.from_date_var = tk.StringVar(value=past)
        self.from_time_var = tk.StringVar(value="00:00:00")
        self.to_date_var = tk.StringVar(value=today)
        self.to_time_var = tk.StringVar(value="23:59:59")

        self.item_var = tk.StringVar(value="All")
        self.operator_var = tk.StringVar(value="All")
        self.machine_var = tk.StringVar(value="All")
        self.airgauge_var = tk.StringVar(value="All")
        self.channel_var = tk.StringVar(value="All") # Default to All
        self.drawing_var = tk.StringVar(value="All")
        self.customer_var = tk.StringVar(value="All")
        # Internal state
        # all_data_rows: list of tuples -> (row_list, rec_dt_or_None, metadata_dict)
        # metadata_dict must include 'file_index' (int) and other filter fields
        self.all_data_rows = []
        # visible mapping after last filter: list of tuples (row_list, file_index)
        self._current_visible = []
        self._visible_file_indices = []

        self._suspend_auto_filter = False
        self._load_thread = None
        self._filter_thread = None

        # Table/backing widgets
        self.sheet = None
        self.use_tksheet = Sheet is not None
        self.tree = None

        # Debounce/filter tokens
        self._filter_job = None
        self._filter_debounce_ms = 200
        self._filter_token = 0

        # Loading UI
        self.loading_label = None
        self.loading_overlay = None

        # Build UI
        self._load_filter_data_files()
        self._build_header()           # includes Export + Delete + Refresh
        self._build_filter_grid()
        self._build_table_area()
        self._wire_traces_and_bindings()
        self._create_loading_overlay()

        self.data_loaded = False
        self._load_token = 0
        self._load_thread = None

        # --- RESTORE PREVIOUS STATE IF AVAILABLE ---
        self._restore_filter_state()

        # --- INITIAL LOAD ---
        self.after(500, lambda: self._schedule_filter(immediate=True))



    def on_window_resized(self):
        try:
            if self.use_tksheet and self.sheet:
                self.sheet.refresh()
        except:
            pass

    # ---------------------------
    # Utilities to load config files
    # ---------------------------
    def _load_filter_data_files(self):
        try:
            conn1 = sqlite3.connect(resource_path("items.db"))
            conn1.row_factory = sqlite3.Row
            c1 = conn1.cursor()
            c1.execute("SELECT code, name FROM items")
            self.items_list = [dict(r) for r in c1.fetchall()]
            conn1.close()
        except Exception: self.items_list = []

        try:
            conn2 = sqlite3.connect(resource_path("employee_master.db"))
            conn2.row_factory = sqlite3.Row
            c2 = conn2.cursor()
            c2.execute("SELECT id, name, description, phone FROM employee_master")
            self.operators_list = [dict(r) for r in c2.fetchall()]
            conn2.close()
        except Exception: self.operators_list = []

        try:
            conn3 = sqlite3.connect(resource_path("machine_master.db"))
            conn3.row_factory = sqlite3.Row
            c3 = conn3.cursor()
            c3.execute("SELECT code, name, description FROM machine_master")
            self.machines_list = [dict(r) for r in c3.fetchall()]
            conn3.close()
        except Exception: self.machines_list = []

    def _load_airgauge_ids(self):
        try:
            conn = sqlite3.connect(resource_path("component_setup.db"))
            cursor = conn.cursor()
            cursor.execute("SELECT DISTINCT airgauge_id FROM component_setup")
            ids = [f"AG{r[0]}" for r in cursor.fetchall()]
            conn.close()
            # 🆕 RESTORE 'All' option
            return ["All"] + sorted(list(set(ids)))
        except Exception:
            return ["All"]
    def _load_channels_for_airgauge(self, airgauge_id):
        """
        Return channel list for selected airgauge from SQLite.
        REMOVED "All" - returns specific channels only.
        """
        out = []

        if airgauge_id in ("", "All"):
            return ["CH1"] 

        try:
            conn = sqlite3.connect(resource_path("component_setup.db"))
            cursor = conn.cursor()
            cursor.execute("SELECT DISTINCT channel FROM component_setup WHERE airgauge_id=?", (str(airgauge_id).strip(),))
            out = [str(r[0]).strip() for r in cursor.fetchall()]
            conn.close()
        except Exception:
            pass

        # unique + ordered
        seen = set()
        final = []
        for v in out:
            if v not in seen:
                final.append(v)
                seen.add(v)
        
        # Sort naturally: CH1, CH2, ... CH10
        def natural_keys(text):
            # extract int if possible
            import re
            c = re.findall(r'\d+', text)
            return int(c[0]) if c else 0

        final.sort(key=natural_keys)
        
        if not final:
            final = ["CH1"]

        return final

    # ---------------------------
    # Header with Delete + Refresh
    # ---------------------------
    # --- PATCH 1: Modify _build_header() to add the PRINT button ---
    def _build_header_legacy(self):
        # === Modern Header Card with Glassmorphism ===
        header = ModernCardFrame(self)
        header.pack(fill="x", padx=20, pady=(15, 10))
        
        header_inner = ctk.CTkFrame(header, fg_color="transparent")
        header_inner.pack(fill="x", padx=20, pady=15)

        # ---------------------------------------------------------
        # LEFT SECTION → Report Title + Export
        # ---------------------------------------------------------
        left_frame = ctk.CTkFrame(header_inner, fg_color="transparent")
        left_frame.pack(side="left")

        ctk.CTkLabel(
            left_frame,
            text="📊  Report",
            font=("Segoe UI", 18, "bold"),
            text_color="#1976D2"
        ).pack(side="left", padx=(0, 15))

        # Export Excel Button (Deep Green theme)
        self.excel_btn = ModernButton(
            left_frame,
            text="📊 Export Excel",
            text_color="white",
            width=120,
            height=38,
            font=("Segoe UI", 13, "bold"),
            fg_color="#1B5E20",
            hover_color="#154d1a",
            corner_radius=8,
            border_width=0,
            command=lambda: self.export_data("excel")
        )
        self.excel_btn.pack(side="left", padx=6)

        # Export PDF Button (Deep Green theme)
        self.pdf_btn = ModernButton(
            left_frame,
            text="📄 Export PDF",
            text_color="white",
            width=120,
            height=38,
            font=("Segoe UI", 13, "bold"),
            fg_color="#1B5E20",
            hover_color="#154d1a",
            corner_radius=8,
            border_width=0,
            command=lambda: self.export_data("pdf")
        )
        self.pdf_btn.pack(side="left", padx=6)

        # ---------------------------------------------------------
        # CENTER SECTION → Analyze Button (Dark Blue theme)
        # ---------------------------------------------------------
        center_frame = ctk.CTkFrame(header_inner, fg_color="transparent")
        center_frame.pack(side="left", expand=True)

        self.analyze_btn = ModernButton(
            center_frame,
            text="📈 Analyze",
            width=130,
            height=38,
            font=("Segoe UI", 13, "bold"),
            fg_color="#0D47A1",
            hover_color="#0a3c85",
            corner_radius=8,
            border_width=0,
            command=self.open_analyze_page
        )
        self.analyze_btn.pack()

        # ---------------------------------------------------------
        # RIGHT SECTION → Delete + Refresh
        # ---------------------------------------------------------
        right_frame = ctk.CTkFrame(header_inner, fg_color="transparent")
        right_frame.pack(side="right")

        del_btn = ModernButton(
            right_frame,
            text="🗑 Delete",
            width=100,
            height=38,
            font=("Segoe UI", 13, "bold"),
            fg_color="#C62828",
            hover_color="#a82222",
            corner_radius=8,
            border_width=0,
            command=self._on_delete_clicked
        )
        del_btn.pack(side="left", padx=6)

        ref_btn = ModernButton(
            right_frame,
            text="🔄 Refresh",
            width=100,
            height=38,
            font=("Segoe UI", 13, "bold"),
            fg_color="#00695C",
            hover_color="#00574b",
            corner_radius=8,
            border_width=0,
            command=self.refresh_table_data
        )
        ref_btn.pack(side="left", padx=6)

    def _build_header(self):
        """Compact Report toolbar matching the USB-style action bar."""
        header = ctk.CTkFrame(
            self,
            fg_color="white",
            corner_radius=8,
            border_width=1,
            border_color="#E0E0E0",
            height=72
        )
        header.pack(fill="x", padx=20, pady=(10, 5))
        header.pack_propagate(False)

        green_bar = ctk.CTkFrame(header, fg_color="#007B43", width=6, corner_radius=0)
        green_bar.place(x=0, y=0, relheight=1)

        header_inner = ctk.CTkFrame(header, fg_color="transparent")
        header_inner.pack(fill="both", expand=True, padx=(25, 20), pady=10)

        left_frame = ctk.CTkFrame(header_inner, fg_color="transparent")
        left_frame.pack(side="left", fill="y")

        logo_frame = ctk.CTkFrame(left_frame, fg_color="#007B43", width=44, height=44, corner_radius=8)
        logo_frame.pack(side="left", padx=(0, 12))
        logo_frame.pack_propagate(False)

        ctk.CTkLabel(
            logo_frame,
            text="\U0001F4CA",
            font=("Segoe UI", 20, "bold"),
            text_color="white",
            fg_color="transparent"
        ).pack(expand=True)

        ctk.CTkLabel(
            left_frame,
            text="Report",
            font=("Segoe UI", 22, "bold"),
            text_color="#007B43"
        ).pack(side="left", padx=(0, 15))

        self.excel_btn = ModernButton(
            left_frame,
            text="\U0001F4CA Export Excel",
            text_color="white",
            width=120,
            height=38,
            font=("Segoe UI", 13, "bold"),
            fg_color="#007B43",
            hover_color="#005C32",
            corner_radius=6,
            border_width=0,
            command=lambda: self.export_data("excel")
        )
        self.excel_btn.pack(side="left", padx=6)

        self.pdf_btn = ModernButton(
            left_frame,
            text="\U0001F4C4 Export PDF",
            text_color="white",
            width=120,
            height=38,
            font=("Segoe UI", 13, "bold"),
            fg_color="#007B43",
            hover_color="#005C32",
            corner_radius=6,
            border_width=0,
            command=lambda: self.export_data("pdf")
        )
        self.pdf_btn.pack(side="left", padx=6)

        center_frame = ctk.CTkFrame(header_inner, fg_color="transparent")
        center_frame.pack(side="left", fill="y", expand=True)

        self.analyze_btn = ModernButton(
            center_frame,
            text="\U0001F4C8 Analyze",
            width=120,
            height=38,
            font=("Segoe UI", 13, "bold"),
            fg_color="#0D47A1",
            hover_color="#0a3c85",
            corner_radius=6,
            border_width=0,
            command=self.open_analyze_page
        )
        self.analyze_btn.pack()

        right_frame = ctk.CTkFrame(header_inner, fg_color="transparent")
        right_frame.pack(side="right", fill="y")

        del_btn = ModernButton(
            right_frame,
            text="\U0001F5D1 Delete",
            width=100,
            height=38,
            font=("Segoe UI", 13, "bold"),
            fg_color="#E53935",
            hover_color="#C62828",
            corner_radius=6,
            border_width=0,
            command=self._on_delete_clicked
        )
        del_btn.pack(side="left", padx=6)

        ref_btn = ModernButton(
            right_frame,
            text="\U0001F504 Refresh",
            width=100,
            height=38,
            font=("Segoe UI", 13, "bold"),
            fg_color="#00695C",
            hover_color="#00574b",
            corner_radius=6,
            border_width=0,
            command=self.refresh_table_data
        )
        ref_btn.pack(side="left", padx=6)

    def _load_logo_base64(self):
        """Load cherry.png safely (supports EXE and any PC)."""
        import os, base64, sys
        try:
            # For EXE
            if getattr(sys, 'frozen', False):
                base_dir = sys._MEIPASS
            else:
                base_dir = os.path.dirname(os.path.abspath(__file__))

            logo_path = os.path.join(base_dir, "images", "cherry.png")  # << UPDATED PATH

            with open(logo_path, "rb") as f:
                return base64.b64encode(f.read()).decode("utf-8")

        except Exception as e:
            print("Logo load error:", e)
            return ""

    def export_data(self, fmt=None):
        """
        Generate strict professional PDF or Excel exports.
        fmt: "pdf", "excel", or None (both)
        Exports only visible/filtered data.
        Prompts user to select save path via dialog; auto-opens file after save.
        """
        try:
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

            # ── Ask where to save ──────────────────────────────────────
            from tkinter import filedialog
            if fmt == "pdf":
                pdf_path = filedialog.asksaveasfilename(
                    title="Save PDF Report As",
                    initialfile=f"SPC_Report_{timestamp}.pdf",
                    defaultextension=".pdf",
                    filetypes=[("PDF File", "*.pdf")]
                )
                if not pdf_path:
                    return  # user cancelled
                xlsx_path = None
            elif fmt == "excel":
                xlsx_path = filedialog.asksaveasfilename(
                    title="Save Excel Report As",
                    initialfile=f"SPC_Report_{timestamp}.xlsx",
                    defaultextension=".xlsx",
                    filetypes=[("Excel File", "*.xlsx")]
                )
                if not xlsx_path:
                    return  # user cancelled
                pdf_path = None
            else:
                # Both — ask separately
                pdf_path = filedialog.asksaveasfilename(
                    title="Save PDF Report As",
                    initialfile=f"SPC_Report_{timestamp}.pdf",
                    defaultextension=".pdf",
                    filetypes=[("PDF File", "*.pdf")]
                )
                if not pdf_path:
                    return
                xlsx_path = filedialog.asksaveasfilename(
                    title="Save Excel Report As",
                    initialfile=f"SPC_Report_{timestamp}.xlsx",
                    defaultextension=".xlsx",
                    filetypes=[("Excel File", "*.xlsx")]
                )
                if not xlsx_path:
                    return


            # --- GATHER DATA ---
            header_cols = [
                "S.No", "Time", "Date", "Reading", "Offset", "Status",
                "AirGauge ID", "Channel", "Drawing", "User", 
                "CompID", "Item", "CNC ID", "Customer"
            ]
            
            rows = []
            try:
                # Prefer source of truth (all filtered) or visible
                source = getattr(self, "_current_visible", [])
                
                for i, (r, fi) in enumerate(source):
                    new_r = [
                        i + 1,          # S.No
                        r[1],           # Time
                        r[2],           # Date
                        r[3],           # Reading
                        r[4],           # Offset
                        r[5],           # Status
                        r[6],           # AirGauge ID
                        r[7],           # Channel
                        r[8],           # Drawing
                        r[9],           # User
                        r[10],          # CompID
                        r[11],          # Item
                        r[12],          # CNC ID
                        r[13] if len(r) > 13 else "" # Customer
                    ]
                    rows.append(new_r)
            except Exception as e:
                print("Data gather error:", e)

            if not rows:
                messagebox.showwarning("Export", "No data available to export.")
                self._set_loading_state(False)
                return
            
            self._set_loading_state(True)
            
            # --- PREPARE FILTER SUMMARY TEXT ---
            f_from = f"{self.from_date_var.get()} {self.from_time_var.get()}"
            f_to   = f"{self.to_date_var.get()} {self.to_time_var.get()}"
            f_ag   = self.airgauge_var.get()
            f_ch   = self.channel_var.get()
            f_item = self.item_var.get()
            f_draw = self.drawing_var.get()
            f_op   = self.operator_var.get()
            f_IP  = self.machine_var.get()
            f_cust = self.customer_var.get()

            # --- GENERATE EXCEL ---
            if (fmt == "excel" or fmt is None) and openpyxl:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "SPC Report"
                
                # Styles
                bold_font = Font(bold=True)
                header_font = Font(bold=True, color="FFFFFF")
                fill_header = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
                center_align = Alignment(horizontal="center", vertical="center")
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                # Title
                ws.merge_cells('A1:N1')
                ws['A1'] = "SPC INSPECTION REPORT"
                ws['A1'].font = Font(bold=True, size=14, color="0D47A1")
                ws['A1'].alignment = center_align

                # Filter Grid (Rows 2-3) using merged cells to simulate layout
                # Layout:
                # Row 2: Date From | Date To | AirGauge | Channel | Item
                # Row 3: Operator | Machine | Customer | Drawing | (Empty)
                
                # We'll just list them in pairs for simplicity or spanning
                # Let's do a strict grid 2 rows.
                # A2: "Filters:"
                ws['A2'] = "Date Range:"
                ws['B2'] = f"{f_from} to {f_to}"
                ws['D2'] = "AirGauge:"
                ws['E2'] = f"{f_ag}"
                ws['F2'] = "Channel:"
                ws['G2'] = f"{f_ch}"
                ws['H2'] = "Item:"
                ws['I2'] = f"{f_item}"
                
                ws['A3'] = "Operator:"
                ws['B3'] = f"{f_op}"
                ws['D3'] = "Machine:"
                ws['E3'] = f"{f_IP}"
                ws['F3'] = "Customer:"
                ws['G3'] = f"{f_cust}"
                ws['H3'] = "Drawing:"
                ws['I3'] = f"{f_draw}"

                for cell in ['A2','D2','F2','H2', 'A3','D3','F3','H3']:
                    ws[cell].font = bold_font

                # Table Header (Row 5)
                ws.append([]) # spacer
                ws.append(header_cols) # Row 5
                
                for col_num, val in enumerate(header_cols, 1):
                    cell = ws.cell(row=5, column=col_num)
                    cell.font = header_font
                    cell.fill = fill_header
                    cell.alignment = center_align
                    cell.border = thin_border
                
                # Data Rows
                for r_data in rows:
                    ws.append(r_data)
                    for col_num in range(1, len(r_data) + 1):
                        cell = ws.cell(row=ws.max_row, column=col_num)
                        cell.border = thin_border
                        cell.alignment = center_align

                # Adjust widths
                from openpyxl.utils import get_column_letter
                for i, col_cells in enumerate(ws.columns, 1):
                    max_length = 0
                    col_letter = get_column_letter(i)
                    for cell in col_cells:
                        try:
                            # Skip merged cells logic safety
                            if hasattr(cell, 'value') and cell.value:
                                s_val = str(cell.value)
                                if len(s_val) > max_length:
                                    max_length = len(s_val)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[col_letter].width = adjusted_width

                wb.save(xlsx_path)

            # --- GENERATE PDF ---
            if (fmt == "pdf" or fmt is None) and 'SimpleDocTemplate' in globals():
                doc = SimpleDocTemplate(pdf_path, pagesize=landscape(A4),
                                        rightMargin=10*mm, leftMargin=10*mm,
                                        topMargin=18*mm, bottomMargin=10*mm)
                elements = []
                styles = getSampleStyleSheet()

                # ── Per-page logo callback ─────────────────────────────
                _logo_path = resource_path(os.path.join("settings", "cherry_full_logo.png"))

                def _draw_page_header(canvas_obj, doc_obj):
                    """Draw the Cherry logo in the top-left on every page."""
                    canvas_obj.saveState()
                    try:
                        from reportlab.lib.utils import ImageReader
                        logo_img = ImageReader(_logo_path)
                        ph = landscape(A4)[1]
                        canvas_obj.drawImage(logo_img, 10*mm, ph - 16*mm,
                                             width=44*mm, height=12*mm,
                                             preserveAspectRatio=True, mask='auto')
                    except Exception:
                        pass
                    canvas_obj.restoreState()

                # Title
                title_style = ParagraphStyle('Title', parent=styles['Heading1'],
                                             alignment=1,
                                             textColor=colors.HexColor("#0D47A1"))
                elements.append(Paragraph("SPC Inspection Report", title_style))
                elements.append(Spacer(1, 4*mm))

                # ── Filter table: 8 columns (label | value | label | value …) ──
                ag_ch = f"{f_ag}  /  CH: {f_ch}"

                lbl_style = ParagraphStyle('lbl', fontName='Helvetica-Bold',
                                           fontSize=8, textColor=colors.white)
                val_style = ParagraphStyle('val', fontName='Helvetica', fontSize=8)

                def L(text): return Paragraph(text, lbl_style)
                def V(text): return Paragraph(str(text), val_style)

                filter_data = [
                    [L("Date Time (from):"), V(f_from),  L("Item:"),     V(f_item),
                     L("AirGauge ID:"),      V(ag_ch),   L("Drawing:"),  V(f_draw)],
                    [L("Date Time (to):"),   V(f_to),    L("Operator:"), V(f_op),
                     L("Machine ID:"),       V(f_IP),    L("Customer:"), V(f_cust)],
                ]

                page_w = landscape(A4)[0] - 20*mm
                lbl_w = page_w * 0.10
                val_w = page_w * 0.15
                col_widths = [lbl_w, val_w, lbl_w, val_w, lbl_w, val_w, lbl_w, val_w]

                f_table = Table(filter_data, colWidths=col_widths)
                f_table.setStyle(TableStyle([
                    ('GRID',          (0, 0), (-1, -1), 0.4, colors.lightgrey),
                    ('BACKGROUND',    (0, 0), (-1, -1), colors.whitesmoke),
                    ('BACKGROUND',    (0, 0), (0, -1),  colors.HexColor("#0D47A1")),
                    ('BACKGROUND',    (2, 0), (2, -1),  colors.HexColor("#0D47A1")),
                    ('BACKGROUND',    (4, 0), (4, -1),  colors.HexColor("#0D47A1")),
                    ('BACKGROUND',    (6, 0), (6, -1),  colors.HexColor("#0D47A1")),
                    ('FONTSIZE',      (0, 0), (-1, -1), 8),
                    ('TOPPADDING',    (0, 0), (-1, -1), 5),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                    ('LEFTPADDING',   (0, 0), (-1, -1), 6),
                    ('RIGHTPADDING',  (0, 0), (-1, -1), 4),
                    ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
                ]))
                elements.append(f_table)
                elements.append(Spacer(1, 5*mm))

                # Main Data Table
                table_data = [header_cols] + rows

                # 14 columns
                t_col_w = page_w / 14.0

                t = Table(table_data, colWidths=[t_col_w]*14, repeatRows=1)
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#0D47A1")),

                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('FONTSIZE', (0, 1), (-1, -1), 7),
                    ('valign', (0, 0), (-1, -1), 'MIDDLE'),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.whitesmoke])
                ]))
                elements.append(t)
                
                doc.build(elements,
                          onFirstPage=_draw_page_header,
                          onLaterPages=_draw_page_header)

            self._set_loading_state(False)

            # ── Auto-open the saved file(s) ───────────────────────────
            opened = []
            try:
                if xlsx_path and os.path.exists(xlsx_path):
                    os.startfile(xlsx_path)
                    opened.append(os.path.basename(xlsx_path))
            except Exception:
                pass
            try:
                if pdf_path and os.path.exists(pdf_path):
                    os.startfile(pdf_path)
                    opened.append(os.path.basename(pdf_path))
            except Exception:
                pass

            if opened:
                messagebox.showinfo("Export Successful", f"Saved & opened:\n" + "\n".join(opened))
            else:
                messagebox.showinfo("Export Successful", "Files exported successfully.")

        except Exception as e:
            self._set_loading_state(False)
            messagebox.showerror("Export Error", f"Failed to export: {str(e)}")




    # ---------------------------
    # Filter grid (2x4)
    # ---------------------------
    def _build_filter_grid_legacy(self):
        # === Modern Filter Panel Container (Transparent since cards themselves have borders) ===
        filter_container = ctk.CTkFrame(
            self,
            fg_color="transparent"
        )
        filter_container.pack(fill="x", padx=20, pady=(0, 10))
        
        filter_frame = ctk.CTkFrame(filter_container, fg_color="transparent")
        filter_frame.pack(fill="x", padx=0, pady=0)

        # 4 columns for clean grid:
        # Col 0: Date Time (From/To) [Wide]
        # Col 1: Item / Operator
        # Col 2: AirGauge ID / Machine ID
        # Col 3: Drawing / Customer
        filter_frame.grid_columnconfigure(0, weight=3)
        filter_frame.grid_columnconfigure(1, weight=2)
        filter_frame.grid_columnconfigure(2, weight=2)
        filter_frame.grid_columnconfigure(3, weight=2)

        self.filter_cells = []
        ROWS, COLS = 2, 4
        for r in range(ROWS):
            row_list = []
            for c in range(COLS):
                card = ctk.CTkFrame(
                    filter_frame,
                    fg_color="white",
                    corner_radius=12,
                    border_width=1,
                    border_color="#E0E0E0",
                    height=85
                )
                card.grid(row=r, column=c, sticky="nsew", padx=6, pady=6)
                card.grid_propagate(False)
                row_list.append(card)
            self.filter_cells.append(row_list)

        # ---------------------------------------------------------
        # Card 0,0: Date Time (From)
        # ---------------------------------------------------------
        from_card = self.filter_cells[0][0]
        ctk.CTkLabel(
            from_card, 
            text="Date Time (From)", 
            font=("Segoe UI", 11, "bold"),
            text_color="#757575"
        ).pack(anchor="w", padx=12, pady=(8, 2))
        
        from_dt_container = ctk.CTkFrame(from_card, fg_color="transparent")
        from_dt_container.pack(fill="x", padx=10, pady=(2, 8))
        from_dt_container.grid_columnconfigure(0, weight=3)
        from_dt_container.grid_columnconfigure(1, weight=2)

        # Date Entry Container (to place the 📅 calendar icon inside)
        from_date_container = ctk.CTkFrame(from_dt_container, fg_color="transparent")
        from_date_container.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        
        self.from_date_entry = ctk.CTkEntry(
            from_date_container, 
            textvariable=self.from_date_var,
            placeholder_text="dd/mm/yyyy", 
            height=30, 
            corner_radius=6, 
            border_color="#E0E0E0",
            fg_color="white",
            text_color="#333333",
            font=("Segoe UI", 12)
        )
        self.from_date_entry.pack(side="left", fill="x", expand=True)
        
        from_cal_btn = ctk.CTkLabel(
            from_date_container, 
            text="📅", 
            font=("Segoe UI", 13), 
            text_color="#2E7D32", 
            cursor="hand2"
        )
        from_cal_btn.pack(side="right", padx=(4, 0))
        if Calendar is not None:
            self.from_date_entry.bind("<Button-1>", lambda e: self.open_calendar_popup(self.from_date_entry, self.from_date_var))
            from_cal_btn.bind("<Button-1>", lambda e: self.open_calendar_popup(self.from_date_entry, self.from_date_var))

        # Time spinner (From)
        self.from_time_spinner = self._create_time_spinner(from_dt_container, self.from_time_var)
        self.from_time_spinner.grid(row=0, column=1, sticky="nsew")

        # ---------------------------------------------------------
        # Card 1,0: Date Time (To)
        # ---------------------------------------------------------
        to_card = self.filter_cells[1][0]
        ctk.CTkLabel(
            to_card, 
            text="Date Time (To)", 
            font=("Segoe UI", 11, "bold"),
            text_color="#757575"
        ).pack(anchor="w", padx=12, pady=(8, 2))
        
        to_dt_container = ctk.CTkFrame(to_card, fg_color="transparent")
        to_dt_container.pack(fill="x", padx=10, pady=(2, 8))
        to_dt_container.grid_columnconfigure(0, weight=3)
        to_dt_container.grid_columnconfigure(1, weight=2)

        # Date Entry Container (To)
        to_date_container = ctk.CTkFrame(to_dt_container, fg_color="transparent")
        to_date_container.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        
        self.to_date_entry = ctk.CTkEntry(
            to_date_container, 
            textvariable=self.to_date_var,
            placeholder_text="dd/mm/yyyy", 
            height=30, 
            corner_radius=6, 
            border_color="#E0E0E0",
            fg_color="white",
            text_color="#333333",
            font=("Segoe UI", 12)
        )
        self.to_date_entry.pack(side="left", fill="x", expand=True)
        
        to_cal_btn = ctk.CTkLabel(
            to_date_container, 
            text="📅", 
            font=("Segoe UI", 13), 
            text_color="#2E7D32", 
            cursor="hand2"
        )
        to_cal_btn.pack(side="right", padx=(4, 0))
        if Calendar is not None:
            self.to_date_entry.bind("<Button-1>", lambda e: self.open_calendar_popup(self.to_date_entry, self.to_date_var))
            to_cal_btn.bind("<Button-1>", lambda e: self.open_calendar_popup(self.to_date_entry, self.to_date_var))

        # Time spinner (To)
        self.to_time_spinner = self._create_time_spinner(to_dt_container, self.to_time_var)
        self.to_time_spinner.grid(row=0, column=1, sticky="nsew")

        # ---------------------------------------------------------
        # Card 0,1: Item
        # ---------------------------------------------------------
        item_card = self.filter_cells[0][1]
        ctk.CTkLabel(
            item_card, 
            text="Item", 
            font=("Segoe UI", 11, "bold"),
            text_color="#757575"
        ).pack(anchor="w", padx=12, pady=(8, 2))
        self.item_combo = self._create_searchable_combobox(item_card, self.item_var, self._items_display_list())

        # ---------------------------------------------------------
        # Card 0,2: AirGauge ID
        # ---------------------------------------------------------
        ag_card = self.filter_cells[0][2]
        ctk.CTkLabel(
            ag_card, 
            text="AirGauge ID", 
            font=("Segoe UI", 11, "bold"),
            text_color="#757575"
        ).pack(anchor="w", padx=12, pady=(8, 2))
        
        ag_container = ctk.CTkFrame(ag_card, fg_color="transparent")
        ag_container.pack(fill="x", padx=10, pady=(2, 8))
        ag_container.grid_columnconfigure(0, weight=1)
        ag_container.grid_columnconfigure(1, weight=1)

        # AirGauge ID combo
        self.airgauge_combo = ttk.Combobox(
            ag_container,
            values=self._load_airgauge_ids(),
            textvariable=self.airgauge_var,
            width=8
        )
        self.airgauge_combo.grid(row=0, column=0, sticky="nsew", padx=(0, 4))
        self.airgauge_combo._base_values = list(self.airgauge_combo["values"])

        # Channel combo
        self.channel_combo = ttk.Combobox(
            ag_container,
            values=["All"],
            textvariable=self.channel_var,
            width=8
        )
        self.channel_combo.grid(row=0, column=1, sticky="nsew")
        self.channel_combo._base_values = ["All"]

        # ---------------------------------------------------------
        # Card 0,3: Drawing
        # ---------------------------------------------------------
        drawing_card = self.filter_cells[0][3]
        ctk.CTkLabel(
            drawing_card, 
            text="Drawing", 
            font=("Segoe UI", 11, "bold"),
            text_color="#757575"
        ).pack(anchor="w", padx=12, pady=(8, 2))
        self.drawing_combo = self._create_searchable_combobox(drawing_card, self.drawing_var, ["All"])

        # ---------------------------------------------------------
        # Card 1,1: Operator
        # ---------------------------------------------------------
        operator_card = self.filter_cells[1][1]
        ctk.CTkLabel(
            operator_card, 
            text="Operator", 
            font=("Segoe UI", 11, "bold"),
            text_color="#757575"
        ).pack(anchor="w", padx=12, pady=(8, 2))
        self.operator_combo = self._create_searchable_combobox(operator_card, self.operator_var, self._operators_display_list())

        # ---------------------------------------------------------
        # Card 1,2: Machine ID
        # ---------------------------------------------------------
        machine_card = self.filter_cells[1][2]
        ctk.CTkLabel(
            machine_card, 
            text="Machine ID", 
            font=("Segoe UI", 11, "bold"),
            text_color="#757575"
        ).pack(anchor="w", padx=12, pady=(8, 2))
        self.machine_combo = self._create_searchable_combobox(machine_card, self.machine_var, self._machines_display_list())

        # ---------------------------------------------------------
        # Card 1,3: Customer
        # ---------------------------------------------------------
        customer_card = self.filter_cells[1][3]
        ctk.CTkLabel(
            customer_card, 
            text="Customer", 
            font=("Segoe UI", 11, "bold"),
            text_color="#757575"
        ).pack(anchor="w", padx=12, pady=(8, 2))
        self.customer_combo = self._create_searchable_combobox(customer_card, self.customer_var, ["All"])

    def _build_filter_grid(self):
        """Compact two-row Report filter grid matching the requested layout."""
        filter_container = ctk.CTkFrame(self, fg_color="transparent")
        filter_container.pack(fill="x", padx=0, pady=(0, 8))

        filter_frame = ctk.CTkFrame(filter_container, fg_color="transparent")
        filter_frame.pack(fill="x", padx=0, pady=0)

        for col, weight in enumerate((2, 1, 1, 1)):
            filter_frame.grid_columnconfigure(col, weight=weight, uniform="report_filter")
        filter_frame.grid_rowconfigure(0, weight=1)
        filter_frame.grid_rowconfigure(1, weight=1)

        def make_card(row, col, title):
            card = ctk.CTkFrame(
                filter_frame,
                fg_color="white",
                corner_radius=8,
                border_width=1,
                border_color="#E6E9ED",
                height=78
            )
            card.grid(row=row, column=col, sticky="nsew", padx=3, pady=5)
            card.grid_propagate(False)
            ctk.CTkLabel(
                card,
                text=title,
                font=("Segoe UI", 11, "bold"),
                text_color="#202124",
                anchor="w"
            ).pack(anchor="w", padx=12, pady=(8, 3))
            return card

        def add_combo(card, tk_var, values):
            combo = self._create_searchable_combobox(card, tk_var, values)
            try:
                combo.configure(font=("Segoe UI", 11))
            except Exception:
                pass
            return combo

        def add_date_time(card, date_var, time_var):
            body = ctk.CTkFrame(card, fg_color="transparent")
            body.pack(fill="x", padx=10, pady=(0, 8))
            body.grid_columnconfigure(0, weight=3)
            body.grid_columnconfigure(1, weight=2)

            date_wrap = ctk.CTkFrame(body, fg_color="transparent")
            date_wrap.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
            entry = ctk.CTkEntry(
                date_wrap,
                textvariable=date_var,
                placeholder_text="dd/mm/yyyy",
                height=30,
                corner_radius=6,
                border_width=1,
                border_color="#D8E3DC",
                fg_color="white",
                text_color="#333333",
                font=("Segoe UI", 11),
                state="readonly"
            )
            entry.pack(side="left", fill="x", expand=True)

            cal_btn = ctk.CTkLabel(
                date_wrap,
                text="\U0001F4C5",
                font=("Segoe UI", 13),
                text_color="#008A4D",
                cursor="hand2",
                width=18
            )
            cal_btn.pack(side="right", padx=(4, 0))
            if Calendar is not None:
                entry.bind("<Button-1>", lambda e: self.open_calendar_popup(entry, date_var))
                cal_btn.bind("<Button-1>", lambda e: self.open_calendar_popup(entry, date_var))

            spinner = self._create_time_spinner(body, time_var)
            spinner.grid(row=0, column=1, sticky="nsew")
            return entry, spinner

        from_card = make_card(0, 0, "Date Time (From)")
        self.from_date_entry, self.from_time_spinner = add_date_time(from_card, self.from_date_var, self.from_time_var)

        item_card = make_card(0, 1, "Item")
        self.item_combo = add_combo(item_card, self.item_var, self._items_display_list())

        ag_card = make_card(0, 2, "AirGauge ID")
        ag_body = ctk.CTkFrame(ag_card, fg_color="transparent")
        ag_body.pack(fill="x", padx=10, pady=(0, 8))
        ag_body.grid_columnconfigure(0, weight=1)
        ag_body.grid_columnconfigure(1, weight=1)
        self.airgauge_combo = ttk.Combobox(
            ag_body,
            values=self._load_airgauge_ids(),
            textvariable=self.airgauge_var,
            font=("Segoe UI", 11)
        )
        self.airgauge_combo.grid(row=0, column=0, sticky="nsew", padx=(0, 6))
        self.airgauge_combo._base_values = list(self.airgauge_combo["values"])
        self.channel_combo = ttk.Combobox(
            ag_body,
            values=["All"],
            textvariable=self.channel_var,
            font=("Segoe UI", 11)
        )
        self.channel_combo.grid(row=0, column=1, sticky="nsew")
        self.channel_combo._base_values = ["All"]

        drawing_card = make_card(0, 3, "Drawing")
        self.drawing_combo = add_combo(drawing_card, self.drawing_var, ["All"])

        to_card = make_card(1, 0, "Date Time (To)")
        self.to_date_entry, self.to_time_spinner = add_date_time(to_card, self.to_date_var, self.to_time_var)

        operator_card = make_card(1, 1, "Operator")
        self.operator_combo = add_combo(operator_card, self.operator_var, self._operators_display_list())

        machine_card = make_card(1, 2, "Machine ID")
        self.machine_combo = add_combo(machine_card, self.machine_var, self._machines_display_list())

        customer_card = make_card(1, 3, "Customer")
        self.customer_combo = add_combo(customer_card, self.customer_var, ["All"])

    # ---------------------------
    # Searchable combobox helper
    # ---------------------------
    def _create_searchable_combobox(self, parent_cell, tk_var, options_list):
        frm = ctk.CTkFrame(parent_cell, fg_color="transparent")
        frm.pack(fill="x", padx=10, pady=(2, 8))
        combo = ttk.Combobox(frm, values=options_list, textvariable=tk_var)
        combo.pack(fill="x", expand=True)
        combo.config(state="normal")
        combo._base_values = list(options_list)
        def on_keyrelease(e):
            typed = combo.get().strip().lower()
            if typed:
                filtered = [v for v in combo._base_values if typed in v.lower()]
            else:
                filtered = combo._base_values
            if "All" not in filtered:
                filtered = ["All"] + filtered
            try:
                combo['values'] = filtered
            except Exception:
                pass
        def on_select(event=None):
            val = combo.get().strip() or "All"
            tk_var.set(val)
            self._schedule_filter()
        combo.bind("<KeyRelease>", on_keyrelease)
        combo.bind("<<ComboboxSelected>>", on_select)
        combo.bind("<FocusOut>", lambda e: on_select())
        return combo

    def _items_display_list(self):
        out = ["All"]
        for it in self.items_list:
            code = str(it.get("code", "")).strip()
            name = str(it.get("name", "")).strip()
            display = f"{code} - {name}" if code else name
            out.append(display)
        return out
    
    def _on_airgauge_changed(self, *_):
        ag = self.airgauge_var.get().strip()

        channels = self._load_channels_for_airgauge(ag)

        self.channel_combo._base_values = channels
        self.channel_combo["values"] = channels

        if self.channel_var.get() not in channels:
            self.channel_var.set("All")

        self._schedule_filter()

    def _operators_display_list(self):
        out = ["All"]
        for op in self.operators_list:
            name = str(op.get("name", "")).strip()
            display = f"{op.get('id','')} - {name}" if op.get("id") else name
            out.append(display)
        return out

    def _machines_display_list(self):
        out = ["All"]
        for m in self.machines_list:
            code = str(m.get("code", "")).strip()
            name = str(m.get("name", "")).strip()
            display = f"{code} - {name}" if code else name
            out.append(display)
        return out

    def _update_all_dynamic_filters(self, data_list):
        """
        Dynamically repopulate all dropdowns (Item, Operator, Machine, AirGauge, Channel, etc.)
        based on what is actually present in the provided data_list.
        """
        try:
            items = set()
            operators = set()
            machines = set()
            airgauges = set()
            drawings = set()
            customers = set()

            for row, _, _ in data_list:
                # row structure:
                # 6: AirGauge, 7: Channel, 8: Drawing, 9: User/Op, 11: Item, 12: Machine, 13: Customer
                if len(row) > 11 and row[11]: items.add(str(row[11]).strip())
                if len(row) > 9 and row[9]: operators.add(str(row[9]).strip())
                if len(row) > 12 and row[12]: machines.add(str(row[12]).strip())
                if len(row) > 6 and row[6]: 
                    ag = str(row[6]).strip()
                    if not ag.startswith("AG"): ag = f"AG{ag}"
                    airgauges.add(ag)
                if len(row) > 8 and row[8]: drawings.add(str(row[8]).strip())
                if len(row) > 13 and row[13]: customers.add(str(row[13]).strip())

            def update_combo(combo, var, new_vals):
                current = var.get()
                sorted_vals = ["All"] + sorted(list(new_vals))
                combo._base_values = sorted_vals
                combo["values"] = sorted_vals
                if current not in sorted_vals:
                    var.set("All")

            update_combo(self.item_combo, self.item_var, items)
            update_combo(self.operator_combo, self.operator_var, operators)
            update_combo(self.machine_combo, self.machine_var, machines)
            update_combo(self.airgauge_combo, self.airgauge_var, airgauges)
            update_combo(self.drawing_combo, self.drawing_var, drawings)
            update_combo(self.customer_combo, self.customer_var, customers)

        except Exception as e:
            print(f"Filter update error: {e}")

    # ---------------------------
    # Time spinner (inline)
    # ---------------------------
    def _create_time_spinner(self, parent, time_var):
        outer = ctk.CTkFrame(parent, corner_radius=4, border_width=1, border_color="#D8E3DC", fg_color="white")
        outer.grid_columnconfigure(0, weight=1)
        outer.grid_columnconfigure(1, weight=0)
        display = ctk.CTkFrame(outer, corner_radius=0, border_width=0, fg_color="white")
        display.grid(row=0, column=0, sticky="nsew", padx=(0, 3))
        
        # Configure columns to be uniform and equal width
        display.grid_columnconfigure(0, weight=1, uniform="time_col")
        display.grid_columnconfigure(1, weight=0)
        display.grid_columnconfigure(2, weight=1, uniform="time_col")
        display.grid_columnconfigure(3, weight=0)
        display.grid_columnconfigure(4, weight=1, uniform="time_col")
        
        def parse_time(s):
            try:
                p = s.split(":")
                hh = int(p[0]) if p and p[0] else 0
                mm = int(p[1]) if len(p) > 1 and p[1] else 0
                ss = int(p[2]) if len(p) > 2 and p[2] else 0
            except Exception:
                hh, mm, ss = 0, 0, 0
            hh %= 24; mm %= 60; ss %= 60
            return hh, mm, ss
            
        hh_val, mm_val, ss_val = parse_time(time_var.get())
        NORMAL_BG = "white"; SELECT_BG = "#008A4D"; NORMAL_TEXT = "#222"; SELECT_TEXT = "white"
        LABEL_FONT = ("Segoe UI", 11, "bold"); COLON_FONT = ("Segoe UI", 11)
        
        # Perfectly uniform labels with width=35 and uniform padding
        hh_lbl = ctk.CTkLabel(display, text=f"{hh_val:02d}", font=LABEL_FONT, text_color=NORMAL_TEXT, fg_color=NORMAL_BG, corner_radius=4, anchor="center", width=35)
        hh_lbl.grid(row=0, column=0, sticky="nsew", padx=2, pady=3)
        
        colon1 = ctk.CTkLabel(display, text=":", font=COLON_FONT, text_color=NORMAL_TEXT)
        colon1.grid(row=0, column=1, sticky="nsew", padx=(1, 1))
        
        mm_lbl = ctk.CTkLabel(display, text=f"{mm_val:02d}", font=LABEL_FONT, text_color=NORMAL_TEXT, fg_color=NORMAL_BG, corner_radius=4, anchor="center", width=35)
        mm_lbl.grid(row=0, column=2, sticky="nsew", padx=2, pady=3)
        
        colon2 = ctk.CTkLabel(display, text=":", font=COLON_FONT, text_color=NORMAL_TEXT)
        colon2.grid(row=0, column=3, sticky="nsew", padx=(1, 1))
        
        ss_lbl = ctk.CTkLabel(display, text=f"{ss_val:02d}", font=LABEL_FONT, text_color=NORMAL_TEXT, fg_color=NORMAL_BG, corner_radius=4, anchor="center", width=35)
        ss_lbl.grid(row=0, column=4, sticky="nsew", padx=2, pady=3)
        
        arrow_frame = ctk.CTkFrame(outer, corner_radius=0, border_width=0, fg_color="white")
        arrow_frame.grid(row=0, column=1, sticky="ns")
        
        # Up and Down buttons configured with hover=False, border_width=0, and hover_color matching the normal background to prevent blue color
        up_btn = ModernButton(
            arrow_frame,
            text="▲",
            width=22,
            height=13,
            fg_color="#F5FBF8",
            text_color="#008A4D",
            hover_color="#F5FBF8",
            hover=False,
            corner_radius=4,
            command=lambda: change_value(1),
            takefocus=False
        )
        up_btn.pack(side="top", padx=0, pady=(2, 4))
        
        down_btn = ModernButton(
            arrow_frame,
            text="▼",
            width=22,
            height=13,
            fg_color="#F5FBF8",
            text_color="#008A4D",
            hover_color="#F5FBF8",
            hover=False,
            corner_radius=4,
            command=lambda: change_value(-1),
            takefocus=False
        )
        down_btn.pack(side="top", padx=0, pady=(0, 2))
        
        selected = {"field": "hour"}
        typed_buffer = []
        
        def update_visual():
            # Reset all label colors
            hh_lbl.configure(fg_color=NORMAL_BG, text_color=NORMAL_TEXT)
            mm_lbl.configure(fg_color=NORMAL_BG, text_color=NORMAL_TEXT)
            ss_lbl.configure(fg_color=NORMAL_BG, text_color=NORMAL_TEXT)
            
            if selected["field"] == "hour":
                hh_lbl.configure(fg_color=SELECT_BG, text_color=SELECT_TEXT)
            elif selected["field"] == "minute":
                mm_lbl.configure(fg_color=SELECT_BG, text_color=SELECT_TEXT)
            else:
                ss_lbl.configure(fg_color=SELECT_BG, text_color=SELECT_TEXT)

        def select_field(f):
            selected["field"] = f
            typed_buffer.clear()
            update_visual()
            try:
                root = outer.winfo_toplevel()
                root.focus_set()
            except:
                pass
            
        hh_lbl.bind("<Button-1>", lambda e: select_field("hour"))
        mm_lbl.bind("<Button-1>", lambda e: select_field("minute"))
        ss_lbl.bind("<Button-1>", lambda e: select_field("second"))
        
        def write_time(h, m, s):
            hh_lbl.configure(text=f"{h:02d}")
            mm_lbl.configure(text=f"{m:02d}")
            ss_lbl.configure(text=f"{s:02d}")
            time_var.set(f"{h:02d}:{m:02d}:{s:02d}")
            self._schedule_filter()

        def change_value(delta):
            nonlocal hh_val, mm_val, ss_val
            typed_buffer.clear()
            if selected["field"] == "hour":
                hh_val = (hh_val + delta) % 24
            elif selected["field"] == "minute":
                mm_val = (mm_val + delta) % 60
            else:
                ss_val = (ss_val + delta) % 60
            write_time(hh_val, mm_val, ss_val)

        def on_wheel(e):
            d = 1 if (e.delta > 0) else -1
            change_value(d)
        
        # Bind scroll to everything
        for w in [outer, hh_lbl, mm_lbl, ss_lbl, up_btn, down_btn, arrow_frame, colon1, colon2]:
            try:
                w.bind("<MouseWheel>", on_wheel)
            except:
                pass

        # Helper to check if user is typing inside standard entry fields elsewhere in the app
        def is_typing_elsewhere():
            try:
                root = outer.winfo_toplevel()
                focused = root.focus_get()
                if focused:
                    cls = focused.winfo_class()
                    if cls in ("Entry", "Text", "TEntry", "TCombobox", "Listbox"):
                        return True
            except:
                pass
            return False

        # Keypress Handlers for manual time entry
        def handle_digit(digit):
            if is_typing_elsewhere():
                return
            nonlocal hh_val, mm_val, ss_val
            val = int(digit)
            
            if len(typed_buffer) >= 2:
                typed_buffer.clear()
            typed_buffer.append(digit)
            
            if len(typed_buffer) == 1:
                # If first digit determines completion (e.g. >=3 for hour, >=6 for minute/second)
                limit = 3 if selected["field"] == "hour" else 6
                if val >= limit:
                    if selected["field"] == "hour":
                        hh_val = val
                        write_time(hh_val, mm_val, ss_val)
                        select_field("minute")
                    elif selected["field"] == "minute":
                        mm_val = val
                        write_time(hh_val, mm_val, ss_val)
                        select_field("second")
                    else:
                        ss_val = val
                        write_time(hh_val, mm_val, ss_val)
                    typed_buffer.clear()
                else:
                    if selected["field"] == "hour":
                        hh_val = val
                    elif selected["field"] == "minute":
                        mm_val = val
                    else:
                        ss_val = val
                    write_time(hh_val, mm_val, ss_val)
            else:
                first = typed_buffer[0]
                combined = int(first + digit)
                
                if selected["field"] == "hour":
                    if combined > 23:
                        combined = 23
                    hh_val = combined
                    write_time(hh_val, mm_val, ss_val)
                    select_field("minute")
                elif selected["field"] == "minute":
                    if combined > 59:
                        combined = 59
                    mm_val = combined
                    write_time(hh_val, mm_val, ss_val)
                    select_field("second")
                else:
                    if combined > 59:
                        combined = 59
                    ss_val = combined
                    write_time(hh_val, mm_val, ss_val)
                typed_buffer.clear()

        def handle_backspace():
            if is_typing_elsewhere():
                return
            nonlocal hh_val, mm_val, ss_val
            if selected["field"] == "hour":
                hh_val = 0
            elif selected["field"] == "minute":
                mm_val = 0
            else:
                ss_val = 0
            typed_buffer.clear()
            write_time(hh_val, mm_val, ss_val)

        def navigate_fields(direction):
            if is_typing_elsewhere():
                return
            if direction == -1: # Left
                if selected["field"] == "second":
                    select_field("minute")
                elif selected["field"] == "minute":
                    select_field("hour")
            else: # Right
                if selected["field"] == "hour":
                    select_field("minute")
                elif selected["field"] == "minute":
                    select_field("second")

        def change_value_key(direction):
            if is_typing_elsewhere():
                return
            change_value(direction)

        # Keyboard support (hover to enable)
        def on_enter(e):
            try:
                root = outer.winfo_toplevel()
                root.bind("<Up>", lambda ev: change_value_key(1))
                root.bind("<Down>", lambda ev: change_value_key(-1))
                root.bind("<Left>", lambda ev: navigate_fields(-1))
                root.bind("<Right>", lambda ev: navigate_fields(1))
                root.bind("<BackSpace>", lambda ev: handle_backspace())
                for digit in "0123456789":
                    root.bind(digit, lambda ev, d=digit: handle_digit(d))
            except Exception as ex:
                pass

        def on_leave(e):
            try:
                root = outer.winfo_toplevel()
                root.unbind("<Up>")
                root.unbind("<Down>")
                root.unbind("<Left>")
                root.unbind("<Right>")
                root.unbind("<BackSpace>")
                for digit in "0123456789":
                    root.unbind(digit)
            except Exception as ex:
                pass

        outer.bind("<Enter>", on_enter)
        outer.bind("<Leave>", on_leave)
        
        select_field("hour")
        write_time(hh_val, mm_val, ss_val)
        return outer

    # ---------------------------
    # Calendar popup (click outside to close)
    # ---------------------------
    def open_calendar_popup(self, entry_widget, target_var):
        if Calendar is None:
            return
        try:
            if hasattr(self, "_cal_popup") and self._cal_popup.winfo_exists():
                self._cal_popup.destroy()
            if hasattr(self, "_cal_overlay") and self._cal_overlay.winfo_exists():
                self._cal_overlay.destroy()
        except:
            pass
        root = self.winfo_toplevel()
        overlay = tk.Toplevel(root)
        overlay.overrideredirect(True)
        overlay.geometry(f"{root.winfo_screenwidth()}x{root.winfo_screenheight()}+0+0")
        overlay.attributes("-alpha", 0.01)
        overlay.lift()
        self._cal_overlay = overlay
        def close_all(event=None):
            try: popup.destroy()
            except: pass
            try: overlay.destroy()
            except: pass
        overlay.bind("<Button-1>", close_all)
        popup = tk.Toplevel(root); self._cal_popup = popup
        popup.overrideredirect(True); popup.attributes("-topmost", True)
        x = entry_widget.winfo_rootx(); y = entry_widget.winfo_rooty() + entry_widget.winfo_height()
        popup.geometry(f"+{x}+{y}"); popup.lift(overlay)
        try:
            d, m, y = target_var.get().split("/"); d=int(d); m=int(m); y=int(y) if int(y)>50 else 2000+int(y)
        except:
            t = datetime.date.today(); d, m, y = t.day, t.month, t.year
        cal = Calendar(popup, selectmode="day", year=y, month=m, day=d, date_pattern="dd/mm/yyyy")
        cal.pack(padx=5, pady=5)
        def on_select(event=None):
            target_var.set(cal.get_date()); close_all()
        cal.bind("<<CalendarSelected>>", on_select)
        popup.bind("<Button-1>", lambda e: "break"); cal.bind("<Button-1>", lambda e: "break")
        popup.bind("<Escape>", lambda e: close_all())

    # ---------------------------
    # Table area build
    # ---------------------------
    def _build_table_area_legacy(self):
        table_frame = ModernCardFrame(self)
        table_frame.pack(fill="both", expand=True, padx=12, pady=(6, 12))
        table_frame.grid_rowconfigure(0, weight=1); table_frame.grid_columnconfigure(0, weight=1)

        cols = (
            "S.No", "Time", "Date","Reading","Offset", "Status","AirGauge ID", "Channel",
             "Drawing", "User", "CompID", "Item", "CNC ID", "Customer"
        )
        
        cols_display = {
            "S.No": "# S.No",
            "Time": "🕒 Time",
            "Date": "📅 Date",
            "Reading": "📈 Reading",
            "Offset": "⚙️ Offset",
            "Status": "🟢 Status",
            "AirGauge ID": "🏷️ AirGauge ID",
            "Channel": "📊 Channel",
            "Drawing": "📄 Drawing",
            "User": "👤 User",
            "CompID": "📦 CompID",
            "Item": "📋 Item",
            "CNC ID": "⚙️ CNC ID",
            "Customer": "👥 Customer"
        }
        cols_display_list = [cols_display[h] for h in cols]

        if self.use_tksheet:
            self.sheet = Sheet(table_frame, headers=cols_display_list, show_x_scrollbar=True, show_y_scrollbar=True, height=500)
            try:
                self.sheet.set_options(auto_resize_columns=False)
                self.sheet.set_options(column_width_resize=False)
            except:
                pass
            self.sheet.grid(row=0, column=0, sticky="nsew", padx=8, pady=8)
            make_sheet_auto_resize(self.sheet, table_frame, cols)
            try:
                self.sheet.enable_bindings(("single_select", "row_select", "arrowkeys", "copy", "select_all", "right_click_popup_menu"))
            except:
                pass
        else:
            self.tree = ttk.Treeview(table_frame, columns=cols, show="headings", selectmode="browse")
            for h in cols:
                self.tree.heading(h, text=cols_display[h])
                self.tree.column(h, width=120, anchor="center")
            vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
            hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
            self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
            self.tree.grid(row=0, column=0, sticky="nsew"); vsb.grid(row=0, column=1, sticky="ns"); hsb.grid(row=1, column=0, sticky="ew")

            self.tree.bind("<Double-1>", lambda e: None)
            
        # Create empty state overlay frame (white to blend with card bg)
        self.empty_state_frame = ctk.CTkFrame(table_frame, fg_color="white", corner_radius=12)
        
        # Center container
        center_content = ctk.CTkFrame(self.empty_state_frame, fg_color="transparent")
        center_content.place(relx=0.5, rely=0.5, anchor="center")
        
        # Empty state graphic, title and helper text
        icon_lbl = ctk.CTkLabel(
            center_content, 
            text="📥", 
            font=("Segoe UI", 48), 
            text_color="#BDBDBD"
        )
        icon_lbl.pack(pady=(0, 10))
        
        msg_lbl = ctk.CTkLabel(
            center_content, 
            text="No data available", 
            font=("Segoe UI", 15, "bold"), 
            text_color="#333333"
        )
        msg_lbl.pack(pady=(0, 4))
        
        sub_lbl = ctk.CTkLabel(
            center_content, 
            text="Please select filters and click Analyze to view data.", 
            font=("Segoe UI", 11), 
            text_color="#757575"
        )
        sub_lbl.pack()

    def _build_table_area(self):
        """Compact Report table with a custom green icon header strip."""
        table_frame = ctk.CTkFrame(
            self,
            fg_color="white",
            corner_radius=8,
            border_width=1,
            border_color="#E6E9ED"
        )
        table_frame.pack(fill="both", expand=True, padx=0, pady=(4, 12))
        table_frame.grid_rowconfigure(2, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        cols = (
            "S.No", "Time", "Date", "Reading", "Offset", "Status",
            "AirGauge ID", "Channel", "Drawing", "User", "CompID",
            "Item", "CNC ID", "Customer"
        )

        self.report_columns = cols
        header_info = [
            ("S.No", "#"),
            ("Time", "\u25F4"),
            ("Date", "\U0001F4C5"),
            ("Reading", "~"),
            ("Offset", "\u2192"),
            ("Status", "\u2139"),
            ("AirGauge ID", "\u25C7"),
            ("Channel", "\U0001F4CA"),
            ("Drawing", "\U0001F4C4"),
            ("User", "\U0001F464"),
            ("CompID", "\U0001F4E6"),
            ("Item", "\u2630"),
            ("CNC ID", "\u2699"),
            ("Customer", "\U0001F465"),
        ]

        self.report_header_frame = ctk.CTkFrame(table_frame, fg_color="white", height=42, corner_radius=0)
        self.report_header_frame.grid(row=0, column=0, sticky="ew", padx=0, pady=(0, 0))
        self.report_header_frame.grid_propagate(False)

        self.report_header_canvas = tk.Canvas(self.report_header_frame, bg="white", highlightthickness=0, height=42)
        self.report_header_canvas.pack(side="left", fill="both", expand=True)
        self.report_header_inner = tk.Frame(self.report_header_canvas, bg="white")
        self.report_header_canvas.create_window(0, 0, window=self.report_header_inner, anchor="nw")

        ctk.CTkFrame(
            self.report_header_frame,
            fg_color="white",
            corner_radius=0,
            width=16,
            height=42
        ).pack(side="right", fill="y")

        self.report_header_widgets = []
        for name, icon in header_info:
            cell = ctk.CTkFrame(
                self.report_header_inner,
                fg_color="white",
                corner_radius=0,
                border_width=1,
                border_color="#E6E9ED",
                height=42
            )
            cell.pack(side="left", fill="y")
            cell.pack_propagate(False)

            content_frame = ctk.CTkFrame(cell, fg_color="transparent")
            content_frame.pack(expand=True)

            ctk.CTkLabel(
                content_frame,
                text=icon,
                font=("Segoe UI", 12, "normal"),
                text_color="#007B43",
                anchor="center"
            ).pack(side="left", padx=(0, 4))

            ctk.CTkLabel(
                content_frame,
                text=name,
                font=("Segoe UI", 11, "bold"),
                text_color="#1A1A1A",
                anchor="center"
            ).pack(side="left")
            self.report_header_widgets.append(cell)

        ctk.CTkFrame(table_frame, fg_color="#E6E9ED", height=1, corner_radius=0).grid(row=1, column=0, sticky="ew")

        sheet_holder = ctk.CTkFrame(table_frame, fg_color="white", corner_radius=0)
        sheet_holder.grid(row=2, column=0, sticky="nsew", padx=0, pady=0)
        sheet_holder.grid_rowconfigure(0, weight=1)
        sheet_holder.grid_columnconfigure(0, weight=1)

        self._report_col_min_widths = [56, 70, 76, 86, 70, 70, 102, 78, 82, 70, 82, 78, 82, 92]
        self._report_col_weights = [6, 7, 8, 9, 7, 7, 11, 8, 8, 7, 8, 8, 8, 10]

        def apply_report_widths(event=None):
            try:
                total_w = sheet_holder.winfo_width()
                if total_w < 200:
                    total_w = table_frame.winfo_width()
                if total_w < 200:
                    return

                scrollbar_pad = 18
                usable = max(100, total_w - scrollbar_pad)
                total_min = sum(self._report_col_min_widths)
                total_weight = sum(self._report_col_weights)
                extra = max(0, usable - total_min)
                widths = [
                    self._report_col_min_widths[i] + int(extra * self._report_col_weights[i] / total_weight)
                    for i in range(len(cols))
                ]
                widths[-1] += usable - sum(widths)

                if self.use_tksheet and self.sheet:
                    for idx, width in enumerate(widths):
                        try:
                            self.sheet.column_width(column=idx, width=width, only_set_if_too_small=False)
                        except TypeError:
                            self.sheet.column_width(idx, width)
                    try:
                        self.sheet.refresh()
                    except Exception:
                        pass
                elif self.tree:
                    for idx, width in enumerate(widths):
                        self.tree.column(cols[idx], width=width, anchor="center")

                for idx, width in enumerate(widths):
                    if idx < len(self.report_header_widgets):
                        self.report_header_widgets[idx].configure(width=width)
                self.report_header_inner.update_idletasks()
                self.report_header_canvas.configure(scrollregion=self.report_header_canvas.bbox("all"))
            except Exception as e:
                print("Report table resize error:", e)

        self._apply_report_widths = apply_report_widths

        if self.use_tksheet:
            self.sheet = Sheet(
                sheet_holder,
                headers=list(cols),
                show_header=False,
                show_row_index=False,
                show_x_scrollbar=True,
                show_y_scrollbar=True,
                height=500
            )
            self.sheet.grid(row=0, column=0, sticky="nsew")
            try:
                self.sheet.set_options(auto_resize_columns=False)
                self.sheet.set_options(column_width_resize=False)
            except Exception:
                pass
            try:
                self.sheet.enable_bindings(("single_select", "row_select", "arrowkeys", "copy", "select_all", "right_click_popup_menu"))
            except Exception:
                pass
            try:
                orig_xscroll = self.sheet.MT.cget("xscrollcommand")
                def sync_scroll(first, last):
                    if orig_xscroll:
                        try:
                            self.sheet.tk.eval(f"{orig_xscroll} {first} {last}")
                        except Exception:
                            pass
                    try:
                        self.report_header_canvas.xview_moveto(first)
                    except Exception:
                        pass
                self.sheet.MT.configure(xscrollcommand=sync_scroll)
            except Exception as e:
                print("Failed to sync Report header scrollbar:", e)
        else:
            self.tree = ttk.Treeview(sheet_holder, columns=cols, show="headings", selectmode="browse")
            for h in cols:
                self.tree.heading(h, text="")
                self.tree.column(h, width=90, anchor="center")
            vsb = ttk.Scrollbar(sheet_holder, orient="vertical", command=self.tree.yview)
            hsb = ttk.Scrollbar(sheet_holder, orient="horizontal", command=self.tree.xview)
            self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
            self.tree.grid(row=0, column=0, sticky="nsew")
            vsb.grid(row=0, column=1, sticky="ns")
            hsb.grid(row=1, column=0, sticky="ew")
            self.tree.bind("<Double-1>", lambda e: None)

        sheet_holder.bind("<Configure>", self._apply_report_widths)
        self.after(100, self._apply_report_widths)

        self.empty_state_frame = ctk.CTkFrame(sheet_holder, fg_color="white", corner_radius=8)
        center_content = ctk.CTkFrame(self.empty_state_frame, fg_color="transparent")
        center_content.place(relx=0.5, rely=0.5, anchor="center")
        ctk.CTkLabel(
            center_content,
            text="No data available",
            font=("Segoe UI", 13, "bold"),
            text_color="#333333"
        ).pack(pady=(0, 4))
        ctk.CTkLabel(
            center_content,
            text="Please select filters and click Analyze to view data.",
            font=("Segoe UI", 10),
            text_color="#757575"
        ).pack()

    # ---------------------------
    # Loading overlay
    # ---------------------------
    def _create_loading_overlay(self):
        self._overlay = ctk.CTkFrame(self, fg_color="#000000", corner_radius=0)
        self._overlay.place_forget()
        self._overlay_label = tk.Label(self._overlay, text="Loading data…", font=("Segoe UI", 13, "bold"), bg="#000000", fg="white")
        self._overlay_label.pack(pady=(20,10))
        self._overlay_pb = ttk.Progressbar(self._overlay, mode="indeterminate", length=240)
        self._overlay_pb.pack(pady=(0,20))

    def _show_loading(self, text="Loading data…"):
        try:
            w = self.winfo_width() or self.winfo_toplevel().winfo_width()
            h = self.winfo_height() or self.winfo_toplevel().winfo_height()
            self._overlay.place(x=0, y=0, width=w, height=h)
            self._overlay_label.config(text=text)
            self._overlay.lift()
            self._overlay_pb.start(10)
            self.update_idletasks()
        except Exception:
            pass

    def _hide_loading(self):
        try:
            self._overlay_pb.stop()
            self._overlay.place_forget()
        except Exception:
            pass

    # ---------------------------
    # Traces & bindings
    # ---------------------------
    def _wire_traces_and_bindings(self):
        self.from_date_var.trace_add("write", lambda *a: self._schedule_filter())
        self.to_date_var.trace_add("write", lambda *a: self._schedule_filter())
        self.from_time_var.trace_add("write", lambda *a: self._schedule_filter())
        self.to_time_var.trace_add("write", lambda *a: self._schedule_filter())
        self.item_var.trace_add("write", lambda *a: self._schedule_filter())
        self.operator_var.trace_add("write", lambda *a: self._schedule_filter())
        self.machine_var.trace_add("write", lambda *a: self._schedule_filter())
        #self.airgauge_var.trace_add("write", lambda *a: self._schedule_filter())
        self.airgauge_var.trace_add("write", self._on_airgauge_changed)
        self.channel_var.trace_add("write", lambda *a: self._schedule_filter())
        self.drawing_var.trace_add("write", lambda *a: self._schedule_filter())
        self.customer_var.trace_add("write", lambda *a: self._schedule_filter())


    # ---------------------------
    # Background JSONL loader
    # ---------------------------
    def _load_json_background(self, fname):
        """
        Returns list of tuples: (row_list, rec_dt_or_None, metadata)
        metadata includes 'file_index' and 'db_id' filtering keys.
        """
        result = []
        ops = self.operators_list
        machs = self.machines_list
        try:
            import sqlite3
            conn = sqlite3.connect(resource_path("production_data.db"))
            cursor = conn.cursor()
            
            # ensure table exists to prevent errors on strict first load
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS measurements (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    source TEXT,
                    date TEXT,
                    time TEXT,
                    reading TEXT,
                    offset TEXT,
                    status TEXT,
                    airgauge_id TEXT,
                    channel TEXT,
                    drawing TEXT,
                    user_id TEXT,
                    component_id TEXT,
                    item TEXT,
                    machine_id TEXT,
                    customer TEXT,
                    upload_timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            # Use chronological retrieval
            cursor.execute("SELECT id, source, date, time, reading, offset, status, airgauge_id, channel, drawing, user_id, component_id, item, machine_id, customer, utl, ltl FROM measurements ORDER BY id DESC")
            rows = cursor.fetchall()
            conn.close()
            
            for file_idx, row_record in enumerate(rows):
                db_id = row_record[0]
                source = row_record[1]
                date_str = str(row_record[2] or "")
                time_ = str(row_record[3] or "")
                reading = str(row_record[4] or "")
                offset = str(row_record[5] or "")
                status = str(row_record[6] or "")
                airgauge = str(row_record[7] or "")
                channel = str(row_record[8] or "")
                drawing = str(row_record[9] or "")
                userid = str(row_record[10] or "")
                compid = str(row_record[11] or "")
                item_display = str(row_record[12] or "") 
                cncid = str(row_record[13] or "")
                customer = str(row_record[14] or "")
                utl_val = str(row_record[15]) if len(row_record) > 15 and row_record[15] is not None else ""
                ltl_val = str(row_record[16]) if len(row_record) > 16 and row_record[16] is not None else ""
                
                rec_dt = self.parse_record_datetime(date_str, time_)
                
                # operator lookup
                op_name = ""
                for op in ops:
                    try:
                        if str(op.get("id","")).strip() == userid.strip():
                            op_name = str(op.get("name","")).strip(); break
                    except:
                        continue
                user_display = userid.strip()
                if op_name:
                    user_display = f"{user_display} - {op_name}"
                
                # machine lookup
                cnc_display = cncid.strip()
                for m in machs:
                    try:
                        if str(m.get("code","")).strip() == cncid.strip():
                            name = str(m.get("name","")).strip()
                            if name:
                                cnc_display = f"{cncid} - {name}"
                            break
                    except:
                        continue
                
                # final row order per requirement
                row = [
                    None,               # 0  S.No
                    time_,              # 1  Time
                    date_str,           # 2  Date
                    reading,            # 3  Reading
                    offset,             # 4  Offset
                    status,             # 5  Status
                    airgauge,           # 6  AirGauge ID
                    channel,            # 7  Channel
                    drawing,            # 8  Drawing
                    user_display,       # 9  User
                    compid,             # 10 CompID
                    item_display,       # 11 Item
                    cnc_display,        # 12 CNC ID
                    customer            # 13 Customer
                ]

                # split item safely for strict metadata filters
                item_parts = item_display.split(" - ")
                meta_code = item_parts[0] if item_parts else ""
                meta_name = item_parts[1] if len(item_parts)>1 else ""

                metadata = {
                    "file_index": int(file_idx),
                    "db_id": int(db_id),
                    "airgauge": airgauge.strip(),
                    "item_name": meta_name.strip().lower(),
                    "item_code": meta_code.strip().lower(),
                    "operator": userid.strip(),
                    "machine_code": cncid.strip(),
                    "parsed_dt": rec_dt,
                    "utl": utl_val,
                    "ltl": ltl_val,
                }
                result.append((row, rec_dt, metadata))
                
        except Exception as e:
            print("SQL Load error:", e)
            return []
        
        return result

    def _start_loading_thread(self):
        if self._load_thread and self._load_thread.is_alive():
            return
        self._show_loading("Loading data…")
        self._suspend_auto_filter = True
        def worker():
            parsed = self._load_json_background(None)
            self.after(10, lambda: self._finish_loading(parsed))
        self._load_thread = threading.Thread(target=worker, daemon=True)
        self._load_thread.start()

    def _finish_loading(self, parsed_data):
        # keep master dataset
        self.all_data_rows = parsed_data

        # build visible rows
        table_rows = []
        vis = []
        for i, (row, rec_dt, meta) in enumerate(parsed_data):
            display_row = row.copy()
            display_row[0] = i + 1
            table_rows.append(display_row)
            vis.append((display_row, meta["file_index"]))

        # store visible mapping
        self._current_visible = vis
        self._visible_file_indices = [fi for (_, fi) in vis]

        # populate UI table
        try:
            if self.use_tksheet and self.sheet:
                self.sheet.set_sheet_data(table_rows)
                self.sheet.refresh()
            else:
                for r in self.tree.get_children():
                    self.tree.delete(r)
                for rrow in table_rows:
                    self.tree.insert("", "end", values=rrow)
        except Exception:
            pass

        # ✅ NOW table is final → update all dynamic filters at once
        self._update_all_dynamic_filters(parsed_data)

        self._update_empty_state(len(table_rows))

        self._suspend_auto_filter = False
        self._set_loading_state(False)
        
        try:
            self.app.status_label.configure(text="Data Loaded Successfully")
            # Clear status after 3 seconds
            self.after(3000, lambda: self.app.status_label.configure(text="Ready"))
        except:
            pass

    def _update_empty_state(self, row_count):
        try:
            if row_count == 0:
                self.empty_state_frame.place(relx=0, rely=0, relwidth=1, relheight=1)
                self.empty_state_frame.lift()
            else:
                self.empty_state_frame.place_forget()
        except Exception as e:
            print("Error updating empty state:", e)



    # ---------------------------
    # Debounced filter scheduler (on-demand file scanning)
    # ---------------------------
    def _schedule_filter(self, immediate=False):
        if self._suspend_auto_filter:
            return

        # PERSIST STATE
        self._persist_filter_state()
        self._set_loading_state(True)

        try:
            if self._filter_job:
                try: self.after_cancel(self._filter_job)
                except: pass
                self._filter_job = None
        except Exception:
            pass

        if immediate:
            # We must use 'after' slightly to ensure UI updates for loading state
            self.after(10, self._run_file_filter_in_background)
        else:
            self._filter_job = self.after(self._filter_debounce_ms, self._run_file_filter_in_background)

    def _restore_filter_state(self):
        """Restore all variables from global state if present."""
        global REPORT_FILTER_STATE
        if not REPORT_FILTER_STATE:
            return
            
        try:
            # Restore if key exists
            if "from_date" in REPORT_FILTER_STATE: self.from_date_var.set(REPORT_FILTER_STATE["from_date"])
            if "to_date" in REPORT_FILTER_STATE: self.to_date_var.set(REPORT_FILTER_STATE["to_date"])
            if "from_time" in REPORT_FILTER_STATE: self.from_time_var.set(REPORT_FILTER_STATE["from_time"])
            if "to_time" in REPORT_FILTER_STATE: self.to_time_var.set(REPORT_FILTER_STATE["to_time"])
            
            if "item" in REPORT_FILTER_STATE: self.item_var.set(REPORT_FILTER_STATE["item"])
            if "operator" in REPORT_FILTER_STATE: self.operator_var.set(REPORT_FILTER_STATE["operator"])
            if "machine" in REPORT_FILTER_STATE: self.machine_var.set(REPORT_FILTER_STATE["machine"])
            if "airgauge" in REPORT_FILTER_STATE: self.airgauge_var.set(REPORT_FILTER_STATE["airgauge"])
            if "channel" in REPORT_FILTER_STATE: self.channel_var.set(REPORT_FILTER_STATE["channel"])
            if "customer" in REPORT_FILTER_STATE: self.customer_var.set(REPORT_FILTER_STATE["customer"])
            if "drawing" in REPORT_FILTER_STATE: self.drawing_var.set(REPORT_FILTER_STATE["drawing"])
            
        except Exception as e:
            print("Error restoring filter state:", e)

    def _persist_filter_state(self):
        """Save current variables to global state."""
        global REPORT_FILTER_STATE
        try:
            REPORT_FILTER_STATE["from_date"] = self.from_date_var.get()
            REPORT_FILTER_STATE["to_date"] = self.to_date_var.get()
            REPORT_FILTER_STATE["from_time"] = self.from_time_var.get()
            REPORT_FILTER_STATE["to_time"] = self.to_time_var.get()
            
            REPORT_FILTER_STATE["item"] = self.item_var.get()
            REPORT_FILTER_STATE["operator"] = self.operator_var.get()
            REPORT_FILTER_STATE["machine"] = self.machine_var.get()
            REPORT_FILTER_STATE["airgauge"] = self.airgauge_var.get()
            REPORT_FILTER_STATE["channel"] = self.channel_var.get()
            REPORT_FILTER_STATE["customer"] = self.customer_var.get()
            REPORT_FILTER_STATE["drawing"] = self.drawing_var.get()
        except Exception:
            pass

    def _set_loading_state(self, is_loading):
        """
        Show/Hide spinner label over table. 
        Disable/Enable Analyze and Export buttons.
        """
        try:
            # 1. Overlay
            if is_loading:
                self._show_loading("Processing...")
            else:
                self._hide_loading()
                
            # 2. Buttons & Inputs
            state = "disabled" if is_loading else "normal"
            if hasattr(self, "export_btn") and self.export_btn:
                self.export_btn.configure(state=state)
            if hasattr(self, "analyze_btn") and self.analyze_btn:
                self.analyze_btn.configure(state=state)
            
            # 3. Disable table interaction (optional but safe)
            # (Requires tksheet disable logic or just blocking overlay)
            
        except Exception as e:
            print("Loading state error:", e)

    def _run_file_filter_in_background(self):
        from_dt = self._parse_filter_datetime(self.from_date_var.get(), self.from_time_var.get())
        to_dt   = self._parse_filter_datetime(self.to_date_var.get(), self.to_time_var.get())

        if from_dt is None:
            from_dt = datetime.datetime.min
        if to_dt is None:
            to_dt = datetime.datetime.max

        self._load_token += 1
        my_token = self._load_token

        self._suspend_auto_filter = True
        self._set_loading_state(True)

        def worker():
            try:
                # 1. Load ALL data from SQLite (it already parses metadata)
                all_data = self._load_json_background(None)
                
                out = []
                for row, rec_dt, meta in all_data:
                    if my_token != self._load_token:
                        return

                    # --- Date/Time Filter ---
                    if rec_dt is None or rec_dt < from_dt or rec_dt > to_dt:
                        continue
                    
                    # --- Operator Filter ---
                    selected_operator = self.operator_var.get().strip()
                    if selected_operator != "All":
                        sel_op_id = selected_operator.split(" - ")[0].strip()
                        if meta["operator"] != sel_op_id:
                            continue

                    # --- Item Filter ---
                    selected_item = self.item_var.get().strip()
                    if selected_item != "All":
                        sel_item_code = selected_item.split(" - ")[0].strip()
                        if meta["item_code"] != sel_item_code.lower():
                            continue

                    # --- Machine Filter ---
                    selected_machine = self.machine_var.get().strip()
                    if selected_machine != "All":
                        sel_machine_code = selected_machine.split(" - ")[0].strip()
                        if meta["machine_code"] != sel_machine_code:
                            continue

                    # --- AirGauge Filter ---
                    selected_airgauge = self.airgauge_var.get().strip()
                    if selected_airgauge != "All":
                        # Flexible matching: "1" == "AG1"
                        db_ag = meta["airgauge"]
                        sel_ag = selected_airgauge
                        
                        # Strip non-digits for comparison if direct match fails
                        if db_ag != sel_ag:
                            db_clean = "".join(filter(str.isdigit, db_ag))
                            sel_clean = "".join(filter(str.isdigit, sel_ag))
                            if not (db_clean and sel_clean and db_clean == sel_clean):
                                continue
                        
                    # --- Channel Filter ---
                    selected_channel = self.channel_var.get().strip()
                    if selected_channel != "All":
                        sel_ch = selected_channel.replace("CH", "").strip()
                        if str(row[7]).strip() != sel_ch: # row[7] is Channel
                            continue

                    # --- Drawing Filter ---
                    selected_drawing = self.drawing_var.get().strip()
                    if selected_drawing != "All":
                        if str(row[8]).strip() != selected_drawing: # row[8] is Drawing
                            continue

                    # --- Customer Filter ---
                    selected_customer = self.customer_var.get().strip()
                    if selected_customer != "All":
                        if str(row[13]).strip().lower() != selected_customer.lower(): # row[13] is Customer
                            continue

                    out.append((row, rec_dt, meta))

                if my_token != self._load_token:
                    return

                self.after(10, lambda: self._finish_loading(out))

            except Exception as e:
                self.after(10, lambda: (
                    self._hide_loading(),
                    setattr(self, "_suspend_auto_filter", False),
                    self.app.status_label.configure(text=f"Filter failed: {e}")
                ))

        self._load_thread = threading.Thread(target=worker, daemon=True)
        self._load_thread.start()

    # ---------------------------
    # Refresh (reset + trigger reload)
    # ---------------------------
    def refresh_table_data(self):
        try:
            self._load_token += 1  # cancel existing loads
            # REMOVED: self._suspend_auto_filter = True (This was preventing the filter from running)

            if self.use_tksheet and getattr(self, "sheet", None):
                self.sheet.set_sheet_data([])
                self.sheet.refresh()
            elif getattr(self, "tree", None):
                for r in self.tree.get_children():
                    self.tree.delete(r)

            self.app.status_label.configure(text="Refreshing…")
            self.data_loaded = False

            # Explicitly set to False before calling to ensure it runs
            self._suspend_auto_filter = False
            self._schedule_filter(immediate=True)
        except Exception as e:
            messagebox.showerror("Refresh failed", str(e))


    # ---------------------------
    # Date parsing helpers
    # ---------------------------
    def parse_record_datetime(self, date_str, time_str):
        try:
            parts = date_str.strip().split("/")
            if len(parts) == 3:
                d, m, y = parts
                if d == "" or m == "" or y == "": return None
                
                # Robust Year parsing
                yi = int(y)
                if yi < 100:
                    year = 2000 + yi
                else:
                    year = yi
                    
                month = int(m); day = int(d)
            else:
                return None
            th = tm = ts = 0
            try:
                th, tm, ts = [int(x) for x in time_str.split(":")]
            except:
                pass
            return datetime.datetime(year, month, day, th, tm, ts)
        except:
            return None

    def _parse_filter_datetime(self, date_str, time_str):
        try:
            parts = date_str.strip().split("/")
            if len(parts) == 3:
                d, m, y = parts
                
                # Robust Year parsing
                yi = int(y)
                if yi < 100:
                    year = 2000 + yi
                else:
                    year = yi
                    
                month = int(m); day = int(d)
            else:
                return None
            th = tm = ts = 0
            try:
                th, tm, ts = [int(x) for x in time_str.split(":")]
            except:
                pass
            return datetime.datetime(year, month, day, th, tm, ts)
        except:
            return None


    def _apply_filtered_rows(self, filtered_tuples):
        """
        filtered_tuples: list of (display_row, file_index)
        Update displayed table and internal visible mapping.
        """
        try:
            # assign S.No sequentially
            vis = []
            table_rows = []
            for i, (r, file_idx) in enumerate(filtered_tuples):
                display = r.copy()
                display[0] = i + 1
                table_rows.append(display)
                vis.append((display, file_idx))
            self._current_visible = vis
            self._visible_file_indices = [fi for (_, fi) in vis]

            if self.use_tksheet and self.sheet:
                self.sheet.set_sheet_data(table_rows)
                try: self.sheet.refresh()
                except: pass
            else:
                for iid in self.tree.get_children(): self.tree.delete(iid)
                for rrow in table_rows: self.tree.insert("", "end", values=rrow)
            self._update_empty_state(len(table_rows))
        except Exception:
            pass
        try:
            self.app.status_label.configure(text=f"Filtered {len(self._current_visible)} records")
        except Exception:
            pass


    # ---------------------------
    # Get currently selected visible row index and file_index
    # Returns (visible_index, file_index) or (None, None)
    # ---------------------------
    def _get_selected_visible(self):
        try:
            if self.use_tksheet and self.sheet:
                # try sheet API
                try:
                    sel = self.sheet.get_selected_rows()
                    if sel and isinstance(sel, list) and len(sel) > 0:
                        vis_idx = sel[0]
                    else:
                        cur = self.sheet.get_currently_selected()
                        if cur and isinstance(cur, tuple):
                            vis_idx = cur[0]
                        else:
                            return (None, None)
                except Exception:
                    return (None, None)
            else:
                sel = self.tree.selection()
                if not sel:
                    return (None, None)
                # find index of selected item in display order
                items = self.tree.get_children()
                try:
                    vis_idx = items.index(sel[0])
                except Exception:
                    # fallback: linear search by comparing values
                    vis_idx = None
                    vals = self.tree.item(sel[0], "values")
                    for i, iid in enumerate(items):
                        if str(self.tree.item(iid, "values")) == str(vals):
                            vis_idx = i; break
                    if vis_idx is None:
                        return (None, None)
            # map to file_index
            if vis_idx is None or vis_idx < 0 or vis_idx >= len(self._current_visible):
                return (None, None)
            file_idx = self._current_visible[vis_idx][1]
            return (vis_idx, file_idx)
        except Exception:
            return (None, None)

    # ---------------------------
    # Delete flow: prompt for password then delete the exact file line
    # ---------------------------
    def _on_delete_clicked(self):
        # get selection
        vis_idx, file_idx = self._get_selected_visible()
        if vis_idx is None:
            messagebox.showwarning("No selection", "Select a row to delete.")
            return
        # ask password modal
        pw = self._ask_password_modal()
        if pw is None:
            return  # user cancelled
        if pw != self.DELETE_PASSWORD:
            messagebox.showerror("Wrong password", "Password incorrect.")
            return
        # Confirm final
        if not messagebox.askyesno("Confirm Delete", "Delete selected record permanently from file?"):
            return
        try:
            self._perform_delete_by_file_index(file_idx, vis_idx)
            messagebox.showinfo("Deleted", "Record deleted successfully.")
        except Exception as e:
            messagebox.showerror("Delete failed", str(e))

    def _ask_password_modal(self):
            """
            Modern, professional modal password prompt returning entered string or None if cancelled.
            """
            # Create a Toplevel window with customtkinter for consistent theming
            dlg = ctk.CTkToplevel(self)
            dlg.title("Authentication Required")
            
            # Geometry and centering
            w, h = 380, 240
            # Center the modal relative to the parent window
            x_pos = self.winfo_x() + (self.winfo_width() // 2) - (w // 2)
            y_pos = self.winfo_y() + (self.winfo_height() // 2) - (h // 2)
            dlg.geometry(f"{w}x{h}+{x_pos}+{y_pos}")
            dlg.resizable(False, False)
            
            # Make the dialog modal
            dlg.transient(self.winfo_toplevel())
            dlg.grab_set()
            
            # Main container with padding for a clean look
            main_frame = ctk.CTkFrame(dlg, corner_radius=0, fg_color="transparent")
            main_frame.pack(fill="both", expand=True, padx=24, pady=24)
            # Header Section
            header_label = ctk.CTkLabel(
                main_frame, 
                text="Enter Admin Password", 
                font=("Segoe UI", 18, "bold"),
                text_color=("gray10", "gray90")
            )
            header_label.pack(anchor="w", pady=(0, 8))
            
            sub_label = ctk.CTkLabel(
                main_frame,
                text="Please verify your identity to continue.",
                font=("Segoe UI", 11),
                text_color=("gray40", "gray70")
            )
            sub_label.pack(anchor="w", pady=(0, 20))
            # Password Entry
            pw_var = tk.StringVar()
            entry = ctk.CTkEntry(
                main_frame, 
                textvariable=pw_var, 
                show="●", 
                width=300, 
                height=36,
                font=("Segoe UI", 13),
                placeholder_text="Password"
            )
            entry.pack(fill="x", pady=(0, 24))
            entry.focus_set()
            # Action Buttons
            button_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
            button_frame.pack(fill="x", side="bottom")
            result = {"val": None}
            def on_confirm(event=None):
                result["val"] = pw_var.get()
                dlg.destroy()
            def on_cancel(event=None):
                dlg.destroy()
            # Confirm Button (Primary)
            btn_confirm = ModernButton(
                button_frame, 
                text="Unlock", 
                width=100, 
                height=36,
                font=("Segoe UI", 13, "bold"),
                command=on_confirm
            )
            btn_confirm.pack(side="right", padx=(12, 0))
            # Cancel Button (Secondary - Ghost Style)
            btn_cancel = ModernButton(
                button_frame, 
                text="Cancel", 
                width=100, 
                height=36,
                font=("Segoe UI", 13),
                fg_color="transparent",
                border_width=1,
                border_color=("gray70", "gray40"),
                text_color=("gray10", "gray90"),
                hover_color=("gray90", "gray25"),
                command=on_cancel
            )
            btn_cancel.pack(side="right")
            # Bind keys for keyboard accessibility
            dlg.bind("<Return>", on_confirm)
            dlg.bind("<Escape>", on_cancel)
            self.wait_window(dlg)
            return result["val"]
            # Header Section

    def _perform_delete_by_file_index(self, file_index_to_delete, visible_index):
        """
        Delete exact line using db_id from SQLite production_data.db.
        """
        target_db_id = None
        for (row, rec_dt, meta) in self.all_data_rows:
            if meta.get("file_index") == file_index_to_delete:
                target_db_id = meta.get("db_id")
                break
                
        if target_db_id is None:
            raise Exception("Target line not found in memory map.")
            
        # 1) Remove line from DB
        try:
            import sqlite3
            conn = sqlite3.connect(resource_path("production_data.db"))
            cursor = conn.cursor()
            cursor.execute("DELETE FROM measurements WHERE id = ?", (target_db_id,))
            conn.commit()
            conn.close()
        except Exception as e:
            raise Exception(f"Database deletion failed: {e}")

        # 2) Update in-memory data structures
        new_all = []
        for (row, rec_dt, meta) in self.all_data_rows:
            if meta.get("file_index") == file_index_to_delete:
                continue
            new_all.append((row, rec_dt, meta))
        # replace master list
        self.all_data_rows = new_all

        # 3) Decrement file_index for entries that were after deleted line
        for i, (row, rec_dt, meta) in enumerate(self.all_data_rows):
            fi = meta.get("file_index")
            if fi is not None and fi > file_index_to_delete:
                meta["file_index"] = fi - 1

        # 4) Update current visible mapping and UI (remove only visible row)
        # remove that visible entry at visible_index
        try:
            if 0 <= visible_index < len(self._current_visible):
                del self._current_visible[visible_index]
        except Exception:
            pass
        # rebuild table rows from current visible mapping with new S.No
        table_rows = []
        new_vis = []
        for i, (r, fi) in enumerate(self._current_visible):
            display = r.copy()
            display[0] = i + 1
            table_rows.append(display)
            new_vis.append((display, fi))
        self._current_visible = new_vis
        self._visible_file_indices = [fi for (_, fi) in new_vis]

        # Update UI table
        try:
            if self.use_tksheet and self.sheet:
                self.sheet.set_sheet_data(table_rows)
                try: self.sheet.refresh()
                except: pass
            else:
                for iid in self.tree.get_children(): self.tree.delete(iid)
                for rrow in table_rows: self.tree.insert("", "end", values=rrow)
            self._update_empty_state(len(table_rows))
        except Exception:
            pass

        try:
            self.app.status_label.configure(text="Deleted record and updated view.")
        except Exception:
            pass

    # ---------------------------
    # Reload filter files (items/operators/machines/airgauge)
    # ---------------------------
    def reload_filter_files(self):
        self._load_filter_data_files()
        try:
            self.item_combo._base_values = self._items_display_list(); self.item_combo['values'] = self.item_combo._base_values
            self.operator_combo._base_values = self._operators_display_list(); self.operator_combo['values'] = self.operator_combo._base_values
            self.machine_combo._base_values = self._machines_display_list(); self.machine_combo['values'] = self.machine_combo._base_values
            vals = self._load_airgauge_ids()
            self.airgauge_combo._base_values = vals; self.airgauge_combo['values'] = vals
        except Exception:
            pass


        
    def open_analyze_page(self):
        """Collect readings from the currently filtered table and open AnalysisPage."""
            # Save ReportPage filter values for AnalysisPage
        self.app.last_report_filters = {
            "from_date": self.from_date_var.get(),
            "from_time": self.from_time_var.get(),
            "to_date":   self.to_date_var.get(),
            "to_time":   self.to_time_var.get(),
            "item":      self.item_var.get(),
            "operator":  self.operator_var.get(),
            "machine":   self.machine_var.get(),
            "airgauge":  self.airgauge_var.get(),
            "channel":   self.channel_var.get(),
            "drawing":   self.drawing_var.get(),
            "customer":  self.customer_var.get(),
        }
        try:
            # Show loading immediately
            try:
                self.app.status_label.configure(text="Preparing Analysis...")
                self.update_idletasks()
            except: pass

            # Prefer using stored filtered rows from source of truth
            readings = []
            
            # Use _current_visible which contains (row_tuple, file_index)
            # index 3 is Reading in the internal data structure
            source = getattr(self, "_current_visible", [])
            # Also gather UTL/LTL from the full all_data_rows metadata
            all_rows_meta = {i: m for i, (_, _, m) in enumerate(getattr(self, "all_data_rows", []))}
            ltl_vals = []
            utl_vals = []
            for row_item in source:
                r = row_item[0]
                fi = row_item[1]  # file_index == position in all_data_rows
                try:
                    val = float(r[3])
                    readings.append(val)
                    meta = all_rows_meta.get(fi, {})
                    ltl_vals.append(meta.get("ltl", ""))
                    utl_vals.append(meta.get("utl", ""))
                except Exception:
                    pass

            # Only show LTL/UTL if ALL rows in the filtered data agree on the same value.
            # If mixed values exist → leave blank (show nothing on Analysis page).
            valid_ltls = [v for v in ltl_vals if v and v not in ("", "0", "0.0")]
            valid_utls = [v for v in utl_vals if v and v not in ("", "0", "0.0")]
            row_ltl = valid_ltls[0] if valid_ltls and len(set(valid_ltls)) == 1 else ""
            row_utl = valid_utls[0] if valid_utls and len(set(valid_utls)) == 1 else ""
            
            if not readings:
                self._show_no_readings_alert()
                return

            # Open AnalysisPage inside content_frame with loading overlay
            # Set manual_dismiss=True because AnalysisPage uses threads to build charts
            self.app._switch_page_with_overlay(
                lambda: self._do_open_analysis(readings, row_ltl, row_utl),
                manual_dismiss=True
            )

        except Exception as e:
            print("Error opening analysis:", e)
            messagebox.showerror("Error", f"Could not open analysis: {e}")

    def _do_open_analysis(self, readings, row_ltl, row_utl):
        """Inner function: actually builds and packs AnalysisPage (called inside overlay)."""
        for w in self.app.content_frame.winfo_children():
            if isinstance(w, AnalysisPage):
                w.destroy()
            else:
                w.pack_forget()

        ap = AnalysisPage(self.app.content_frame, self.app, readings=readings,
                          title="📊 SPC Analysis",
                          row_ltl=row_ltl, row_utl=row_utl)
        ap.pack(fill="both", expand=True)

        try:
            self.app.status_label.configure(text=f"Opened Analysis ({len(readings)} readings)")
        except Exception:
            pass

    def _show_no_readings_alert(self):
        try:
            win = ctk.CTkToplevel(self)
            win.title("Analysis")
            win.geometry("400x180")
            win.resizable(False, False)
            win.transient(self)
            win.grab_set()
            
            # Center on screen
            try:
                win.update_idletasks()
                x = (self.winfo_screenwidth() - 400) // 2
                y = (self.winfo_screenheight() - 180) // 2
                win.geometry(f"+{x}+{y}")
            except: pass

            bg = ctk.CTkFrame(win, fg_color="white", corner_radius=0)
            bg.pack(fill="both", expand=True)

            # Warning Icon/Text
            ctk.CTkLabel(bg, text="⚠️", font=("Segoe UI", 24, "bold"), text_color="#FFB300").pack(pady=(20, 5))
            ctk.CTkLabel(bg, text="No valid numerical readings found in current view.", 
                         font=("Segoe UI", 11, "bold"), text_color="#333333").pack(pady=5)

            ModernButton(bg, text="OK", width=90, fg_color="#1976D2", hover_color="#1565C0",
                          command=win.destroy).pack(pady=(15, 20))
            
            win.focus_force()
        except:
            pass

    def go_back(self, event=None):
        if not self.winfo_viewable():
            return
        # ReportPage usually has filters, check focus
        focused = self.focus_get()
        if isinstance(focused, (ctk.CTkEntry, tk.Entry)):
            return
        # If open_calendar_popup is active, maybe don't go back?
        # But simple focus check should cover generic inputs only.
        try:
            self.app.load_settings()
        except:
            pass

    def destroy(self):
        super().destroy()


import statistics
import tempfile
import json
import os
import math
import tkinter as tk
from tkinter import ttk
import customtkinter as ctk
import tksheet
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

class AnalysisPage(ctk.CTkFrame):
    """
    AnalysisPage: builds SPC main 10x table and histogram calculation table
    using a list of reading values (floats). Layout and visuals derived from
    your reference SPC page.
    """
    def __init__(self, parent, app, readings=None, title="📊 SPC Analysis", row_ltl="", row_utl=""):
        super().__init__(parent)

        self.app = app
        self.row_ltl = row_ltl  # LTL directly from the filtered dataset rows
        self.row_utl = row_utl  # UTL directly from the filtered dataset rows
        try:
            if hasattr(self.app, "sidebar") and self.app.sidebar.winfo_ismapped():
                self.app.sidebar.pack_forget()

            if hasattr(self.app, "content_frame"):
                self.app.content_frame.pack_forget()
                self.app.content_frame.pack(side="left", fill="both", expand=True)

        except Exception as e:
            print("Sidebar hide failed:", e)
        self.app = app
        self.readings = readings or []       # list of floats
        self.title_text = title
        # columns for SPC main table (S.No + C1..C10)
        self.columns = ["S.No"] + [f"C{i}" for i in range(1, 11)]
        # columns for histogram calc table (C1..C13 in ref)
        self.calc_columns = [f"C{i}" for i in range(1, 14)]
        # placeholder for frequency table
        self.freq_table = None
        
            
        self._build_ui()
        
        # Defer calculation to allow UI to render first (Performance Fix)
        if self.readings:
            self.after(200, lambda: self.load_from_readings(self.readings))
            # Show initial status
            try: self.status_label.configure(text="Computing SPC metrics...")
            except: pass
        # Set focus on the AnalysisPage
        try:
            self.focus_set()
        except Exception:
            pass

    def destroy(self):
        try:
            self.unbind_all("<MouseWheel>")
            self.unbind_all("<Shift-MouseWheel>")
        except Exception:
            pass
        super().destroy()

    # ------------------------------
    # UI building (based on your reference)
    # ------------------------------
    def _build_ui(self):
        # container with canvas for auto-fade scrollbars
        self.container = ctk.CTkFrame(self, corner_radius=0)
        # DEFER PACKING til end of method to avoid piece-by-piece rendering
        # self.container.pack(fill="both", expand=True)

        # Smooth scroll: yscrollincrement=2 allows pixel-based scrolling (2px precision)
        self.canvas = tk.Canvas(self.container, bg="white", highlightthickness=0, bd=0, relief="flat", yscrollincrement=2)
        self.canvas.grid(row=0, column=0, sticky="nsew")

        # scrollbar style
        scroll_style = ttk.Style()
        try:
            scroll_style.theme_use("default")
        except Exception:
            pass
        scroll_style.configure("Modern.Vertical.TScrollbar",
                               troughcolor="#F0F0F0", background="#C0C0C0",
                               bordercolor="#E0E0E0", arrowcolor="#7A7A7A", relief="flat")
        scroll_style.map("Modern.Vertical.TScrollbar",
                         background=[("active", "#A8A8A8"), ("pressed", "#909090")])
        scroll_style.configure("Modern.Horizontal.TScrollbar",
                               troughcolor="#F0F0F0", background="#C0C0C0",
                               bordercolor="#E0E0E0", arrowcolor="#7A7A7A", relief="flat")
        scroll_style.map("Modern.Horizontal.TScrollbar",
                         background=[("active", "#A8A8A8"), ("pressed", "#909090")])

        vbar_frame = ctk.CTkFrame(self.container, fg_color="#F5F5F5", width=14)
        vbar_frame.grid(row=0, column=1, sticky="ns")
        hbar_frame = ctk.CTkFrame(self.container, fg_color="#F5F5F5", height=14)
        hbar_frame.grid(row=1, column=0, sticky="ew")

        v_scroll = ttk.Scrollbar(vbar_frame, orient="vertical", command=self.canvas.yview, style="Modern.Vertical.TScrollbar")
        v_scroll.pack(fill="y", expand=True, pady=(2,2))
        h_scroll = ttk.Scrollbar(hbar_frame, orient="horizontal", command=self.canvas.xview, style="Modern.Horizontal.TScrollbar")
        h_scroll.pack(fill="x", expand=True, padx=(2,2))

        self.canvas.configure(xscrollcommand=h_scroll.set, yscrollcommand=v_scroll.set)

        self.scroll_frame = tk.Frame(self.canvas, bg="white")
        scroll_frame = self.scroll_frame
        self.canvas.create_window((0,0), window=scroll_frame, anchor="nw")

        self.container.grid_rowconfigure(0, weight=1)
        self.container.grid_columnconfigure(0, weight=1)

        def on_frame_configure(event=None):
            self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        scroll_frame.bind("<Configure>", on_frame_configure)

        # Mousewheel handling similar to reference
        # Mousewheel handling
        def is_cursor_inside(widget):
            try:
                if not widget.winfo_exists(): return False
                x, y = widget.winfo_pointerx(), widget.winfo_pointery()
                gx, gy = widget.winfo_rootx(), widget.winfo_rooty()
                w, h = widget.winfo_width(), widget.winfo_height()
                return gx <= x <= gx + w and gy <= y <= gy + h
            except Exception:
                return False

        def on_mousewheel(event):
            try:
                if not self.winfo_exists(): return
                inside_spc = hasattr(self, "table") and is_cursor_inside(self.table)
                inside_hist = hasattr(self, "calc_table") and is_cursor_inside(self.calc_table)
                
                if inside_spc or inside_hist:
                    # Mouse is over a table -> Let tksheet handle it
                    return 
                
                # Mouse is over background -> Smooth Scroll Main Canvas
                # Since yscrollincrement=2, we scroll by roughly 25*2 = 50px per notch (delta=120)
                # Adjust '25' to change speed.
                scroll_units = int(-1 * (event.delta / 120) * 25)
                canvas.yview_scroll(scroll_units, "units")
                return "break"
            except Exception:
                pass

        def _bind_mouse(_):
            self.canvas.bind_all("<MouseWheel>", on_mousewheel)
            self.canvas.bind_all("<Shift-MouseWheel>", lambda e: self.canvas.xview_scroll(-1 * (e.delta // 120), "units"))
            
            # Keypad/Arrow Scrolling
            try:
                top = self.winfo_toplevel()
                top.bind("<Up>", lambda e: self.canvas.yview_scroll(-5, "units"))
                top.bind("<Down>", lambda e: self.canvas.yview_scroll(5, "units"))
                top.bind("<KP_Up>", lambda e: self.canvas.yview_scroll(-5, "units"))
                top.bind("<KP_Down>", lambda e: self.canvas.yview_scroll(5, "units")) # Keypad 8/2 (NumLock Off/On behavior varies but KP_Up/Down usually handles arrows)
                # Windows might map Keypad 8 with NumLock OFF as Up, but if ON it sends a number. 
                # Tkinter <KP_Up> is usually the arrow on keypad. 
                # Just in case, we can also bind the numbers if simpler, but <KP_Up> is standard.
            except: pass
        
        def _unbind_mouse(_):
            self.canvas.unbind_all("<MouseWheel>")
            self.canvas.unbind_all("<Shift-MouseWheel>")
            try:
                top = self.winfo_toplevel()
                top.unbind("<Up>")
                top.unbind("<Down>")
                top.unbind("<KP_Up>")
                top.unbind("<KP_Down>")
            except: pass

        self.canvas.bind("<Enter>", _bind_mouse)
        self.canvas.bind("<Leave>", _unbind_mouse)

        # Header
        header = ctk.CTkFrame(scroll_frame, fg_color="#E3F2FD", corner_radius=8, height=58)
        header.pack(fill="x", padx=12, pady=(10,6))
        ctk.CTkLabel(header, text=self.title_text, font=("Segoe UI", 18, "bold"), text_color="#0D47A1").pack(side="left", padx=18, pady=8)
        

        # small right area with a back button and optional refresh (kept minimal)
        right_btn_frame = ctk.CTkFrame(header, fg_color="transparent")
        right_btn_frame.pack(side="right", padx=10)
        back_btn = ModernButton(right_btn_frame, text="← Back", width=110, fg_color="#1976D2", hover_color="#1565C0",
                                command=self.go_back)
        back_btn.pack(side="left", padx=6)
        print_btn = ModernButton(
            right_btn_frame,
            text="🖨 Print",
            width=110,
            fg_color="#2E7D32",
            hover_color="#1B5E20",
            command=self.print_spc_report
        )
        print_btn.pack(side="left", padx=6)

        # -------------------------------------------------------------------
        #  FILTER SUMMARY GRID (Excel-style 2 rows × 10 columns)
        # -------------------------------------------------------------------
        filter_grid = ctk.CTkFrame(
            scroll_frame,
            border_color="#BDBDBD",
            border_width=1,
            corner_radius=0
        )
        filter_grid.pack(padx=18, pady=(4, 8), fill="x")

        # Each cell is wrapped inside a frame so borders become visible
        def cell(frame, text, bg, row, col):
            cell_frame = tk.Frame(frame, bg="white", highlightbackground="#BDBDBD",
                                  highlightthickness=1)
            cell_frame.grid(row=row, column=col, sticky="nsew")
            lbl = ctk.CTkLabel(
                cell_frame,
                text=text,
                fg_color=bg,
                corner_radius=0,
                anchor="w",
                padx=6
            )
            lbl.pack(fill="both", expand=True)

        # Make 8 equal columns
        for c in range(10):
            filter_grid.grid_columnconfigure(c, weight=1)

        # Read ReportPage filter values
        rp = getattr(self.app, "last_report_filters", {})

        fd   = rp.get("from_date",  "")
        ft   = rp.get("from_time",  "")
        td   = rp.get("to_date",    "")
        tt   = rp.get("to_time",    "")
        item = rp.get("item",       "All")
        opr  = rp.get("operator",   "All")
        mach = rp.get("machine",    "All")
        air  = rp.get("airgauge",   "All")
        ch = rp.get("channel", "All")
        drawing  = rp.get("drawing",  "All")
        customer = rp.get("customer", "All")


        # ---------------- ROW 1 ----------------
        row1 = [
            ("From Date", fd),
            ("From Time", ft),
            ("Item", item),
            ("AirGauge ID", f"{air} - {ch}" if ch != "All" else air),
            ("Drawing", drawing),     # ✅ NEW
        ]

        col = 0
        for title, val in row1:
            cell(filter_grid, title, "white", 0, col)
            cell(filter_grid, val,   "#EEECE1", 0, col+1)
            col += 2

        # ---------------- ROW 2 ----------------
        row2 = [
            ("To Date", td),
            ("To Time", tt),
            ("Operator", opr),
            ("Machine ID", mach),
            ("Customer", customer),   # ✅ NEW
        ]


        col = 0
        for title, val in row2:
            cell(filter_grid, title, "white", 1, col)
            cell(filter_grid, val,   "#EEECE1", 1, col+1)
            col += 2

        data_row = tk.Frame(scroll_frame, bg="white")
        data_row.pack(padx=20, pady=(4, 6), fill="x")

        # Column configuration for centering
        data_row.grid_columnconfigure(0, weight=0)  # label
        data_row.grid_columnconfigure(1, weight=1)  # left spacer
        data_row.grid_columnconfigure(2, weight=0)  # LTL/UTL table (center)
        data_row.grid_columnconfigure(3, weight=1)  # right spacer
        data_row.grid_columnconfigure(4, weight=0)  # status


        # -------------------------------------------------
        # LEFT : Data Collection label
        # -------------------------------------------------
        ctk.CTkLabel(
            data_row,
            text="💾 Data Collection (10×N)",
            font=("Segoe UI", 13, "bold"),
            text_color="#2E7D32"
        ).grid(row=0, column=0, sticky="w")

        tol_frame = tk.Frame(
            data_row,
            bg="white",
            highlightbackground="#BDBDBD",
            highlightthickness=1
        )
        tol_frame.grid(row=0, column=1, sticky="n", padx=(12, 0))


        def tol_cell(col, text, bg):
            cell = tk.Frame(
                tol_frame,
                bg=bg,
                height=28,
                bd=0
            )
            cell.grid(row=0, column=col, sticky="nsew")
            cell.grid_propagate(False)

            lbl = ctk.CTkLabel(
                cell,
                text=text,
                font=("Segoe UI", 11, "bold" if bg == "#E3F2FD" else "normal")
            )
            lbl.pack(expand=True, fill="both")

            return lbl   # 🔥 THIS WAS MISSING


        WIDTHS = [55, 85, 55, 85]  # LTL | value | UTL | value
        for i, w in enumerate(WIDTHS):
            tol_frame.grid_columnconfigure(i, minsize=w)

        LABEL_BG = "#E3F2FD"
        VALUE_BG = "White"

        self.lbl_ltl = tol_cell(0, "LTL", LABEL_BG)
        self.val_ltl = tol_cell(1, "–",  VALUE_BG)
        self.lbl_utl = tol_cell(2, "UTL", LABEL_BG)
        self.val_utl = tol_cell(3, "–",  VALUE_BG)

        self.cpk_status_label = ctk.CTkLabel(
            data_row,
            text="",
            font=("Segoe UI", 18, "bold"),
            anchor="e"
        )
        self.cpk_status_label.grid(row=0, column=3, sticky="e")



        # SPC card
        # -------- SPC + Right Summary Wrapper --------
        spc_row = tk.Frame(scroll_frame, bg="white")
        spc_row.pack(anchor="w", padx=18, pady=(0, 8))

        # SPC card (LEFT)
        spc_card = ModernCardFrame(spc_row)
        spc_card.pack(side="left", padx=(0, 12), anchor="n")
        right_stack = tk.Frame(spc_row, bg="white")
        right_stack.pack(side="left", padx=12, anchor="n")


        # -------- Cp / Cpk Summary Card (BELOW UTL/LTL) --------
        self.capability_card = ModernCardFrame(spc_row, width=400, height=340)
        self.capability_card.pack(in_=right_stack,pady=0)
        self.capability_card.pack_propagate(False)

        FONT_L = ("Segoe UI", 11, "bold")
        FONT_V = ("Segoe UI", 11, "bold")
        FONT_H = ("Segoe UI", 13, "bold")

        cap = tk.Frame(self.capability_card, bg="white")
        cap.pack(fill="x", padx=10, pady=2)

        # ---------- CONTROL LIMITS ----------
        ctk.CTkLabel(
            cap, text="CONTROL LIMITS", font=FONT_H, text_color="#37474F"
        ).grid(row=0, column=0, columnspan=5, sticky="w", pady=(0, 2))

        self.cap_ucl_x = self._cap_row(cap, 1, "UCL (X̄)")
        self.cap_lcl_x = self._cap_row(cap, 2, "LCL (X̄)")
        self.cap_ucl_r = self._cap_row(cap, 3, "UCL (R)")
        self.cap_lcl_r = self._cap_row(cap, 4, "LCL (R)")

        tk.Frame(cap, bg="#E0E0E0", height=1).grid(
            row=5, column=0, columnspan=5, sticky="ew", pady=2
        )

        # ---------- GRID ----------
        for c in range(5):
            cap.grid_columnconfigure(c, weight=1)

        ROW = 6

        ctk.CTkLabel(
            cap, text="PROCESS CAPABILITY", font=FONT_H, text_color="#37474F"
        ).grid(row=ROW, column=0, columnspan=2, sticky="w", pady=(2, 1))

        ctk.CTkLabel(
            cap, text="PROCESS PERFORMANCE", font=FONT_H, text_color="#37474F"
        ).grid(row=ROW, column=3, columnspan=2, sticky="w", pady=(2, 1))

        ROW += 1

        CAP_COLOR = "#C62828"    # 🔴 red for Process Capability
        PERF_COLOR = "#EF6C00"   # ��� orange for Process Performance

        def kpi(row, cl, cv, label, color):
            ctk.CTkLabel(
                cap,
                text=label,
                font=FONT_L,
                text_color=color
            ).grid(row=row, column=cl, sticky="w", pady=0)

            val = ctk.CTkLabel(
                cap,
                text="–",
                font=FONT_V,
                text_color=color
            )
            val.grid(row=row, column=cv, sticky="e", pady=0)

            return val


        self.cap_sigma  = kpi(ROW, 0, 1, "σ",    CAP_COLOR)
        self.pp_stddev = kpi(ROW, 3, 4, "Std Dev", PERF_COLOR)
        ROW += 1

        self.cap_cpk_l  = kpi(ROW, 0, 1, "CPKₗ", CAP_COLOR)
        self.pp_ppk_l  = kpi(ROW, 3, 4, "Ppkₗ", PERF_COLOR)
        ROW += 1

        self.cap_cpk_u  = kpi(ROW, 0, 1, "CPKᵤ", CAP_COLOR)
        self.pp_ppk_u  = kpi(ROW, 3, 4, "Ppkᵤ", PERF_COLOR)
        ROW += 1

        self.cap_cpk    = kpi(ROW, 0, 1, "CPK",  CAP_COLOR)
        self.pp_ppk    = kpi(ROW, 3, 4, "Ppk",  PERF_COLOR)
        ROW += 1

        self.cap_cp     = kpi(ROW, 0, 1, "CP",   CAP_COLOR)
        self.pp_p      = kpi(ROW, 3, 4, "Pp",   PERF_COLOR)    

        # ---------- FORCE COMPACT ROW HEIGHT ----------
        for r in range(20):
            cap.grid_rowconfigure(r, minsize=18)


        # NEW: horizontal container
        spc_holder = tk.Frame(spc_card, bg="#FFFFFF")
        spc_holder.pack(padx=10, pady=10, fill="both", expand=True)

        # FIXED: use spc_holder
        self.table = tksheet.Sheet(
            spc_holder,   # ✅ correct parent
            headers=self.columns,
            height=320,
            width=1075,
            show_x_scrollbar=False,
            show_y_scrollbar=True,
            theme="light green"
        )
        self.table.pack(fill="both", expand=True)
        make_sheet_auto_resize(self.table, spc_holder, self.columns)

        # hide index robustly
        try:
            if hasattr(self.table, "hide_index"):
                self.table.hide_index()
            elif hasattr(self.table, "show_index"):
                self.table.show_index(False)
            elif hasattr(self.table, "set_index_visible"):
                self.table.set_index_visible(False)
            elif hasattr(self.table, "show_row_index"):
                self.table.show_row_index(False)
            else:
                self.table.set_index_width(0)
        except Exception:
            pass

        self.table.enable_bindings((
            "single_select", "row_select", "column_select", "arrowkeys", "copy", "right_click_popup_menu"
        ))
        self.table.pack(fill="both", expand=True)

        # visual tuning
        try:
            self.table.set_options(
                font=("Segoe UI", 11, "normal"),
                table_bg="#FFFFFF",
                grid_color="#E0E0E0",
                row_height=30,
                align="center",
                empty_vertical=True,
                empty_horizontal=True,
                header_bg="#E3F2FD",
                header_fg="#0D47A1",
                outline_color="#CFD8DC",
                outline_thickness=1,
            )

            total_cols = len(self.columns)
            for c in range(total_cols):
                self.table.highlight_cells(row=-1, column=c, bg="#E3F2FD", fg="#0D47A1")

            total_rows = self.table.get_total_rows()
            for r in range(total_rows):
                bg = "#FAFAFA" if r % 2 == 0 else "#FFFFFF"
                self.table.highlight_rows(rows=[r], bg=bg)
                self.table.highlight_cells(row=r, column=0, bg="#E8EAF6", fg="#1A237E")

            self.table.refresh()
        except Exception:
            pass

        # column widths
        try:
            self.table.column_width(0, width=55)
            for i in range(1, len(self.columns)):
                self.table.column_width(i, width=102)
        except Exception:
            pass

        # click handler to select rows by S.No column (same robust approach)
        def _select_row_on_sno_click_improved(event):
            def resolve():
                try:
                    tbl = self.table
                    row = None; col = None
                    try:
                        if hasattr(tbl, "get_row_clicked") and hasattr(tbl, "get_column_clicked"):
                            row = tbl.get_row_clicked(event); col = tbl.get_column_clicked(event)
                    except Exception:
                        row = col = None
                    if row is None:
                        try:
                            if hasattr(tbl, "get_row_col_from_xy"):
                                r, c = tbl.get_row_col_from_xy(event.x, event.y); row, col = r, c
                            elif hasattr(tbl, "get_cell_from_xy"):
                                r, c = tbl.get_cell_from_xy(event.x, event.y); row, col = r, c
                        except Exception:
                            row = col = None
                    if row is None:
                        try:
                            sel = tbl.get_selected_cells()
                            if sel:
                                first = sel[0]
                                if isinstance(first, (list, tuple)) and len(first) >= 2:
                                    row, col = int(first[0]), int(first[1])
                                else:
                                    row = int(first); col = 0
                        except Exception:
                            row = col = None
                    if row is not None and col is not None and int(col) == 0:
                        try:
                            if hasattr(tbl, "select_row"):
                                tbl.select_row(int(row), redraw=True)
                            elif hasattr(tbl, "select_rows"):
                                tbl.select_rows([int(row)])
                            else:
                                if hasattr(tbl, "set_selected_rows"):
                                    tbl.set_selected_rows([int(row)])
                        except Exception:
                            pass
                except Exception:
                    pass
            try:
                self.after(10, resolve)
            except Exception:
                resolve()

        try:
            self.table.unbind("<ButtonRelease-1>")
        except Exception:
            pass
        try:
            self.table.bind("<ButtonRelease-1>", _select_row_on_sno_click_improved, add="+")
            self.table.bind("<Double-Button-1>", _select_row_on_sno_click_improved, add="+")
        except Exception:
            pass

        # ------------------- Histogram area (no S.No) -------------------
        ctk.CTkLabel(scroll_frame, text="📈 Calculations for Histogram",
                     font=("Segoe UI", 13, "bold"), text_color="#2E7D32").pack(anchor="w", padx=20, pady=(3,4))

        hist_card = ctk.CTkFrame(scroll_frame, fg_color="#FFFFFF",
                                 border_color="#CFD8DC", border_width=1, corner_radius=8)
        hist_card.pack(anchor="w", padx=18, pady=(0,14))

        calc_outer = tk.Frame(hist_card, bg="#FFFFFF")
        calc_outer.pack(padx=8, pady=(12,0), anchor="w")

        initial_hist_width = 1320
        calc_outer.configure(width=initial_hist_width)
        hist_card.configure(width=initial_hist_width)

        # ← Make sure self.calc_columns DO NOT include "S.No"
        # Example: C1..C12 (no S.No)
        self.calc_columns = [f"C{i}" for i in range(1, 14)]

        self.calc_table = tksheet.Sheet(
            calc_outer,
            headers=self.calc_columns,
            height=100,
            width=initial_hist_width,
            show_x_scrollbar=False,
            show_y_scrollbar=False,
            theme="light green",
            show_header=False,     # you already hide header for visual parity
            show_index=False       # ask tksheet to hide the index column
        )
        self.calc_table.enable_bindings(("single_select", "column_select", "arrowkeys", "copy"))
        self.calc_table.pack(fill="both", expand=True)
        make_sheet_auto_resize(self.calc_table, calc_outer, self.calc_columns)
        # -------- Second histogram table (6-column metrics) --------
        lower_hist_holder = tk.Frame(hist_card, bg="#FFFFFF")
        lower_hist_holder.pack(anchor="w", padx=8, pady=(0,12), fill="x", expand=True)

        lower_left = tk.Frame(lower_hist_holder, bg="#FFFFFF")
        lower_left.pack(side="left", fill="both", expand=True)


        # ===============================================================
        # STEP 4️⃣  6-column Histogram Table (Cp / Cpk / PPM)
        # ===============================================================

        self.calc_table_small = tksheet.Sheet(
            lower_left,
            headers=[f"C{i}" for i in range(1, 7)],   # 6 columns only
            height=80,
            width=13 * 101.5,  
            show_x_scrollbar=False,
            show_y_scrollbar=False,
            show_header=False,
            show_index=False,
            theme="light green"
        ) 

        self.calc_table_small.enable_bindings((
            "single_select", "row_select", "arrowkeys", "copy"
        ))

        self.calc_table_small.pack(fill="both", expand=True)
        make_sheet_auto_resize(self.calc_table_small, lower_left, [f"C{i}" for i in range(1, 7)])


        # ===============================================================
        # STEP 6️⃣  Load data into 6-column histogram table
        # ===============================================================

        small_data = [
            ["Process width (P)",  "–", "Specification Width (S)", "–", "Total Range", "–"],
            ["Design center (D)", "–", "Interval", "–", "Selecting No. of Classes", "–"],
            ["-", "–", "No. of Readings", "–", "Shift of X bar from 'D'", "–"],
        ]

        self.calc_table_small.set_sheet_data(small_data)
        self.after_idle(self._lock_small_hist_widths)
        self._force_hide_index(self.calc_table_small)

        # ---- Populate Histogram / Frequency Distribution table (NEW) ---
        # Note: this uses vals, Xmin, Xmax, interval variables computed later
        # so actual populate call is inserted in load_from_readings after they are computed.

        # ---- Visual tuning to match upper histogram table ----
        try:
            self.calc_table_small.set_options(
                font=("Segoe UI", 11, "normal"),
                table_bg="#FFFFFF",
                grid_color="#E0E0E0",
                row_height=28,
                align="center",
                empty_vertical=True,
                empty_horizontal=True,
                outline_thickness=1,
                outline_color="#CFD8DC"
            )

            # ---- Column width mapping to match upper 13-column table ----
            try:
                COL = 102  # single column width of main table

                width_map = [
                    2 * COL,  # col-1
                    2 * COL,  # col-2
                    2 * COL,  # col-3
                    2 * COL,  # col-4
                    3 * COL,  # col-5 (spans 3 columns)
                    2 * COL,  # col-6
                ]

                for i, w in enumerate(width_map):
                    self.calc_table_small.column_width(
                        column=i,
                        width=w,
                        only_set_if_too_small=False
                    )

                self.calc_table_small.refresh()
            except Exception:
                pass


            # Highlight label column (Cp / Cpk / PPM)
            for r in range(3):
                self.calc_table_small.highlight_cells(
                    row=r, column=0, bg="#E3F2FD", fg="#0D47A1"
                )
                self.calc_table_small.highlight_cells(
                    row=r, column=2, bg="#E3F2FD", fg="#0D47A1"
                )

                self.calc_table_small.highlight_cells(
                    row=r, column=4, bg="#E3F2FD", fg="#0D47A1"
                )

            self.calc_table_small.refresh()
        except Exception:
            pass


        # Robustly force-hide any index column regardless of tksheet version
        self._force_hide_index(self.calc_table)

        # Visual options (keep as you had)
        try:
            self.calc_table.set_options(
                font=("Segoe UI", 11, "normal"),
                table_bg="#FFFFFF",
                grid_color="#E0E0E0",
                row_height=28,
                align="center",
                empty_vertical=True,
                empty_horizontal=True,
                outline_thickness=1,
                outline_color="#CFD8DC"
            )
        except Exception:
            pass


        # blank histogram default rows (4 summary rows)
        # ------------------- Histogram default rows (NO S.No, KEEP LAST 2 COLUMNS) -------------------
        blank_data = [
            ["Xlarge"] + ["–"] * 10 + ["X Max=", "–"],
            ["Xsmall"] + ["–"] * 10 + ["X Min=", "–"],
            ["Range"]  + ["–"] * 10 + ["R-Bar=", "–"],
            ["Avg"]    + ["–"] * 10 + ["X-Bar=", "–"],
            
        ]

        # Ensure row size matches expected column count
        expected_cols = len(self.calc_columns)

        for i in range(len(blank_data)):
            row = blank_data[i]
            if len(row) > expected_cols:
                blank_data[i] = row[:expected_cols]
            elif len(row) < expected_cols:
                blank_data[i] = row + [""] * (expected_cols - len(row))

        # Load into histogram table
        self.calc_table.set_sheet_data(blank_data)
        self.calc_table.refresh()

        # Column widths — NO S.No column anymore
        try:
            for ci in range(expected_cols):
                self.calc_table.column_width(ci, width=102, only_set_if_too_small=False)
        except:
            pass

        # Highlight ONLY the last 2 columns
        try:
            last1 = expected_cols - 2
            last2 = expected_cols - 1
            for r in range(len(blank_data)):
                self.calc_table.highlight_cells(row=r, column=0, bg="#C8E6C9", fg="#0B3D0B")
                for col in range(1, 11):   # 1 → 10
                    self.calc_table.highlight_cells(row=r,column=col,bg="#FFF3E0")
                self.calc_table.highlight_cells(row=r, column=last1, bg="#C8E6C9", fg="#0B3D0B")
                self.calc_table.highlight_cells(row=r, column=last2, bg="#FFF3E0", fg="#E65100")
        except:
            pass

        # Status label (unchanged)
        self.status_label = ctk.CTkLabel(
            scroll_frame, text="", font=("Segoe UI", 11), text_color="#1565C0"
        )
        self.status_label.pack(pady=(6, 10))


        # lock histogram widths after idle
        def _lock_calc_widths():
            try:
                column_widths = [102] * (len(self.calc_columns) - 2) + [97, 97]
                for i, w in enumerate(column_widths):
                    try:
                        self.calc_table.column_width(column=i, width=w, only_set_if_too_small=False)
                    except Exception:
                        pass
                self.calc_table.refresh()
            except Exception:
                pass
        self.after_idle(_lock_calc_widths)
        self.after(50, self._update_capability_ui)
        
    def _force_hide_index(self, table):
        try:
            if hasattr(table, "hide_index"):
                table.hide_index()

            if hasattr(table, "show_index"):
                try:
                    table.show_index(False)
                except Exception:
                    pass

            if hasattr(table, "set_index_visible"):
                table.set_index_visible(False)

            if hasattr(table, "show_row_index"):
                try:
                    table.show_row_index(False)
                except Exception:
                    pass

            # Final hard fallback
            if hasattr(table, "set_index_width"):
                try:
                    table.set_index_width(0)
                except Exception:
                    pass
        except Exception:
            pass

        # -------------------------------------------------------------
        # FINAL LAYOUT REVEAL (Fix for piece-by-piece rendering)
        # -------------------------------------------------------------
        # Force layout calculation on the hidden container
        # Force layout calculation on the hidden container
        try:
            self.container.update_idletasks()
        except:
            pass
        # Now show the fully constructed page at once
        self.container.pack(fill="both", expand=True)


        
    def _lock_small_hist_widths(self):
        try:
            COL = 102
            width_map = [
                2 * COL,
                2 * COL,
                2 * COL,
                2 * COL,
                3 * COL,
                2 * COL,
            ]
            for i, w in enumerate(width_map):
                self.calc_table_small.column_width(
                    column=i, width=w, only_set_if_too_small=False
                )
            self.calc_table_small.refresh()
        except Exception:
            pass

    # ---------------------------
    # NEW: Histogram Frequency Table UI + Calculation methods
    # These are class-level methods (not nested) so they can be reused.
    # ---------------------------
    def _build_histogram_frequency_table(self):
        # 1. Create shared horizontal container for Charts (Left) and Table (Right)
        if not hasattr(self, "charts_and_table_row"):
            self.charts_and_table_row = tk.Frame(self.scroll_frame, bg="white")
            self.charts_and_table_row.pack(fill="x", padx=18, pady=(10, 20))
            # Configure grid columns: fixed width for charts, remaining for table
            self.charts_and_table_row.grid_columnconfigure(0, weight=0, minsize=540)  # Chart column (520px chart + 20px padding)
            self.charts_and_table_row.grid_columnconfigure(1, weight=1, minsize=535)  # Table column (flexible)

        # 2. Create the Frequency Holder in grid column 1 (right side)
        self.freq_holder = tk.Frame(self.charts_and_table_row, bg="white")
        self.freq_holder.grid(row=0, column=1, sticky="nw", padx=(16, 0))

        headers = ["INTERVAL FROM", "INTERVAL TO", "CUM FREQ", "FREQ"]
        TOTAL_WIDTH = 490 

        try:
            self.freq_table = tksheet.Sheet(
                self.freq_holder,
                headers=headers,
                height=340, # Match chart height
                width=TOTAL_WIDTH,
                show_x_scrollbar=False,
                show_y_scrollbar=True,
                theme="light green",
                show_index=True
            )
            self.freq_table.pack()
            COL_WIDTHS = [120, 120, 120, 120]
            for i, w in enumerate(COL_WIDTHS):
                self.freq_table.column_width(i, width=w, only_set_if_too_small=False)

            self.freq_table.set_index_width(0)
            self.freq_table.refresh()

            
            # Apply visual styles (omitted for brevity, keep your existing set_options)
            self.freq_table.set_index_width(55)
            self._force_hide_index(self.freq_table) # Use your existing helper
            
        except Exception as e:
            print(f"Freq Table Error: {e}")


    def _compute_histogram_frequency(self, vals, Xmin_val, Xmax_val, interval):
        """
        Compute chained intervals and Excel-style cumulative frequencies (FREQUENCY behavior).

        Strict interval generation rules:
          - Row1 Lower = Xmin - interval
          - Row1 Upper = Lower + interval
          - Row2 Lower = previous Upper
          - Row i Upper = Lower + interval
          - Stop when Upper >= Xmax

        CUM FREQ(i): count of values <= Upper(i)  (Excel FREQUENCY-like cumulative behavior)
        FREQ(i): difference between successive cumulative frequencies (first row equals cum1)

        Returns: list of rows [interval_from (str), interval_to (str), cum_freq (int), freq (int)]
                 - numeric interval bounds formatted to 4 decimal places for display
        """
        rows_out = []
        try:
            # Basic validation
            if vals is None or interval is None:
                return rows_out

            # Build sorted numeric list for efficient cumulative counts
            numeric_vals = []
            for v in vals:
                try:
                    numeric_vals.append(float(v))
                except Exception:
                    continue

            if not numeric_vals:
                return rows_out

            numeric_vals.sort()

            import bisect

            # Start Lower as Xmin - interval (per spec)
            lower = float(Xmin_val) - float(interval)
            prev_cum = 0

            def fmt_num(x):
                try:
                    return f"{float(x):.5f}"
                except Exception:
                    return str(x)

            while True:
                upper = lower + float(interval)
                # CUM FREQ: count of values <= upper (bisect_right)
                cum = bisect.bisect_right(numeric_vals, upper)
                # FREQ: difference from previous cumulative
                freq = int(cum - prev_cum) if cum >= prev_cum else 0

                rows_out.append([fmt_num(lower), fmt_num(upper), int(cum), int(freq)])

                # Stop when upper >= Xmax (inclusive)
                try:
                    if upper >= float(Xmax_val):
                        break
                except Exception:
                    break

                # advance
                prev_cum = cum
                lower = upper

        except Exception as e:
            # On any error, return what we've built so far (or empty)
            print("Histogram frequency compute error:", e)

        return rows_out


    def _populate_histogram_frequency_table(self, freq_rows):
        """
        Ensure the frequency table exists (build if needed) and populate it with freq_rows.
        freq_rows should be a list of lists as returned by _compute_histogram_frequency.
        """
        # Lazy-build the UI if not already present
        if not hasattr(self, "freq_table") or self.freq_table is None:
            try:
                self._build_histogram_frequency_table()
            except Exception:
                pass

        if not getattr(self, "freq_table", None):
            return

        try:
            # replace sheet data and refresh view
            self.freq_table.set_sheet_data(freq_rows)
            try:
                self.freq_table.refresh()
            except Exception:
                pass
        except Exception as e:
            print("Failed to populate histogram frequency table:", e)

    # ------------------------------
    # Charts generation
    # ------------------------------
    # ------------------------------
    # Threaded Analysis Logic
    # ------------------------------
    def load_from_readings(self, readings):
        """
        readings: list of floats (or strings convertible to floats)
        Refactored to run heavy analysis in a background thread.
        1. Setup initial table immediately (fast).
        2. Show loading overlay.
        3. Start thread.
        """
        try:
            # 1. Normalize data (Fast)
            vals = []
            for v in readings:
                try:
                    vals.append(float(v))
                except Exception:
                    pass

            total = len(vals)
            usable = total - (total % 10)
            
            # Basic validation
            if usable <= 0:
                self._reset_to_blank_state()
                return

            vals = vals[:usable]

            # 2. Populate Main Table (Fast enough to do on main thread usually, 
            # or could be deferred. We do it here to show "something" immediately)
            table_data = []
            row_no = 0
            for i in range(0, len(vals), 10):
                row_no += 1
                group = vals[i:i+10]
                row = [row_no] + [f"{x:.5f}" for x in group]
                table_data.append(row)

            self.table.set_sheet_data(table_data)
            self.after_idle(self._lock_main_table_widths)

            # 3. Show Loading
            self._show_loading_over_charts()

            # 4. Prepare data for thread
            # Need tolerance inputs for calculations
            air = getattr(self.app, "last_report_filters", {}).get("airgauge", "All")
            ch  = getattr(self.app, "last_report_filters", {}).get("channel", "All")
            
            # Run in thread
            import threading
            t = threading.Thread(target=self._run_analysis_thread, 
                                 args=(vals, air, ch), 
                                 daemon=True)
            t.start()

        except Exception as e:
            print("Load from readings failed:", e)
            self._hide_loading_over_charts()

    def _reset_to_blank_state(self):
        try:
            self.table.set_sheet_data([])
            blank_data = [
                ["Xlarge"] + ["–"] * 10 + ["X Max=", "–"],
                ["Xsmall"] + ["–"] * 10 + ["X Min=", "–"],
                ["Range"]  + ["–"] * 10 + ["R-Bar=", "–"],
                ["Avg"]    + ["–"] * 10 + ["X-Bar=", "–"],
            ]
            for i in range(len(blank_data)):
                blank_data[i] = blank_data[i][:len(self.calc_columns)]
            self.calc_table.set_sheet_data(blank_data)
            if hasattr(self, "status_label"):
                self.status_label.configure(text="No complete 10-value rows found.")
        except: pass

    def _lock_main_table_widths(self):
        try:
            SNO_WIDTH = 55
            self.table.column_width(0, width=SNO_WIDTH, only_set_if_too_small=False)
            for i in range(1, len(self.columns)):
                self.table.column_width(i, width=102, only_set_if_too_small=False)
            self.table.refresh()
            if hasattr(self, "canvas"):
                self.canvas.yview_moveto(0.0)
        except: pass

    # ------------------------------
    # Background Worker
    # ------------------------------
    def _run_analysis_thread(self, vals, air, ch):
        """
        Executed in a separate thread. 
        Performs all calculations and generates chart images.
        """
        try:
            # A. Get Tolerances (File I/O)
            low_tol, high_tol = self._get_tolerance_for_airgauge(air, ch)
            
            # B. Core Calculations
            # 1. Table columnar (Xmax, Xmin, etc.) - needed for Calc Table
            rows = [vals[i:i + 10] for i in range(0, len(vals), 10)]
            numeric_rows = [ [float(x) for x in r] for r in rows]
            cols = list(zip(*numeric_rows))
            
            stats_res = {
                "xlarge": [max(c) if c else 0.0 for c in cols],
                "xsmall": [min(c) if c else 0.0 for c in cols],
                "rng":    [(max(c) - min(c)) if c else 0.0 for c in cols],
                "avg":    [statistics.mean(c) if c else 0.0 for c in cols],
                "Xmax":   max(max(c) for c in cols),
                "Xmin":   min(min(c) for c in cols),
                "Rbar":   statistics.mean([max(c) - min(c) for c in cols]),
                "Xbar":   statistics.mean([statistics.mean(c) for c in cols])
            }
            
            # 2. Histogram Calc (Sturges etc)
            N = len(vals)
            process_width = stats_res["Xmax"] - stats_res["Xmin"]
            total_range = process_width
            k = round(1 + 3.322 * math.log10(N)) if N > 1 else 1
            interval = (total_range / k) if k > 0 else 0
            
            hist_params = {
                "N": N,
                "process_width": process_width,
                "total_range": total_range,
                "k": k,
                "interval": interval,
                "LTL": low_tol,
                "UTL": high_tol,
                "Xbar": stats_res["Xbar"]
            }

            # 3. SPC Metrics (Cp, Cpk)
            spc_metrics = self._compute_spc_metrics(vals)
            
            # 4. PP Metrics
            pp_metrics = {}
            tol_valid = False
            try:
                ltl_f = float(low_tol)
                utl_f = float(high_tol)
                if utl_f > ltl_f:
                    tol_valid = True
                    pp_metrics = self._compute_pp_metrics(vals, ltl_f, utl_f)
            except: pass

            # 5. GENERATE CHARTS (Thread-Safe Way)
            # We must use Figure/FigureCanvas, NOT plt.plot()
            chart_paths = self._generate_charts_thread_safe(
                vals, 
                stats_res["avg"],   # xbars
                stats_res["rng"],   # ranges
                stats_res["Xbar"],  # xbarbar
                stats_res["Rbar"],  # rbar
                spc_metrics,
                interval
            )
            
            # 6. Frequency Table Data
            freq_rows = self._compute_histogram_frequency(vals, stats_res["Xmin"], stats_res["Xmax"], interval)
            
            # 7. Histogram Interval Image (Optional)
            hist_interval_path = None
            if hasattr(self, "_generate_histogram_from_frequency"):
                 hist_interval_path = self._generate_histogram_from_frequency(freq_rows)

            # Pack results
            result = {
                "stats_res": stats_res,
                "hist_params": hist_params,
                "spc_metrics": spc_metrics,
                "pp_metrics": pp_metrics,
                "chart_paths": chart_paths,
                "freq_rows": freq_rows,
                "hist_interval_path": hist_interval_path,
                "tol_valid": tol_valid
            }

            # Schedule UI Update on Main Thread
            self.after(10, lambda: self._on_analysis_complete(result))

        except Exception as e:
            print("Analysis Thread Error:", e)
            import traceback
            traceback.print_exc()
            self.after(10, self._hide_loading_over_charts)

    def _generate_charts_thread_safe(self, vals, xbars, ranges, xbarbar, rbar, metrics, interval):
        """
        Generates plots using OO interface for thread safety.
        Returns (xbar_path, rbar_path, hist_path)
        """
        from matplotlib.figure import Figure
        from matplotlib.backends.backend_agg import FigureCanvasAgg
        
        tmp_dir = tempfile.gettempdir()
        x_indices = list(range(1, len(xbars) + 1))
        
        # --- Helper: Save Figure ---
        def save_fig(fig, path):
            canvas = FigureCanvasAgg(fig)
            fig.subplots_adjust(right=0.82) # Margin for legend
            canvas.print_figure(path, dpi=100)
        
        # 1. X-Bar Chart
        fig1 = Figure(figsize=(6, 2.8), dpi=100)
        ax1 = fig1.add_subplot(111)
        
        # Convert strings to floats if needed
        xbars_f = [float(x) for x in xbars]
        
        ax1.plot(x_indices, xbars_f, marker="o", label="AVG")
        ax1.axhline(float(xbarbar), linestyle="--", label="X-BAR")
        
        if metrics.get("ucl_x") is not None:
            ax1.axhline(metrics["ucl_x"], linestyle="--", color="red", label="UCL")
        if metrics.get("lcl_x") is not None:
            ax1.axhline(metrics["lcl_x"], linestyle="--", color="orange", label="LCL")
            
        ax1.set_title("X-Bar Chart")
        ax1.set_xlabel("Sample")
        ax1.set_ylabel("Average")
        ax1.legend(bbox_to_anchor=(1.02, 1), loc='upper left', borderaxespad=0.)
        
        xbar_path = os.path.join(tmp_dir, "xbar_chart.png")
        save_fig(fig1, xbar_path)
        
        # 2. R-Bar Chart
        fig2 = Figure(figsize=(6, 2.8), dpi=100)
        ax2 = fig2.add_subplot(111)
        
        ranges_f = [float(x) for x in ranges]
        
        ax2.plot(x_indices, ranges_f, marker="o", label="RANGE")
        ax2.axhline(float(rbar), linestyle="--", label="R-BAR")
        
        if metrics.get("ucl_r") is not None:
            ax2.axhline(metrics["ucl_r"], linestyle="--", color="red", label="UCL")
        if metrics.get("lcl_r") is not None:
            ax2.axhline(metrics["lcl_r"], linestyle="--", color="orange", label="LCL")

        ax2.set_title("R-Bar Chart")
        ax2.set_xlabel("Sample")
        ax2.set_ylabel("Range")
        ax2.legend(bbox_to_anchor=(1.02, 1), loc='upper left', borderaxespad=0.)
        
        rbar_path = os.path.join(tmp_dir, "rbar_chart.png")
        save_fig(fig2, rbar_path)
        
        # 3. Histogram
        # For histogram, we only need vals
        # To match previous bins logic: round(1 + 3.322 * math.log10(len(vals)))
        try:
            bins_count = round(1 + 3.322 * math.log10(len(vals)))
        except: 
            bins_count = 10

        fig3 = Figure(figsize=(6, 2.8), dpi=100)
        ax3 = fig3.add_subplot(111)
        ax3.hist(vals, bins=bins_count, edgecolor="black")
        ax3.set_title("Histogram")
        ax3.set_xlabel("Measurement")
        ax3.set_ylabel("Frequency")
        # Reuse save_fig but maybe reset margin if we want tight layout? 
        # The previous code used tight_layout. 
        # Let's manual adjust for consistency or just use tight_layout logic
        fig3.tight_layout()
        
        hist_path = os.path.join(tmp_dir, "histogram.png")
        canvas3 = FigureCanvasAgg(fig3)
        canvas3.print_figure(hist_path, dpi=100)
        
        return xbar_path, rbar_path, hist_path

    # ------------------------------
    # Completion Handler (Main Thread)
    # ------------------------------
    def _on_analysis_complete(self, res):
        """
        Called when background thread finishes.
        Updates ALL UI elements first, forces a full repaint, then — and
        only then — dismisses the loading overlay so the page appears
        complete in a single instant (no one-by-one content rendering).
        """
        try:
            # ── Unpack results from thread ────────────────────────────
            stats = res["stats_res"]
            hp    = res["hist_params"]
            spc   = res["spc_metrics"]
            pp    = res["pp_metrics"]
            charts       = res["chart_paths"]
            freq_rows    = res["freq_rows"]
            hist_int_path = res["hist_interval_path"]
            tol_valid    = res["tol_valid"]

            # ── 1. Fill Calc Table ────────────────────────────────────
            disp_xlarge = [f"{v:.7f}" for v in stats["xlarge"]]
            disp_xsmall = [f"{v:.7f}" for v in stats["xsmall"]]
            disp_rng    = [f"{v:.7f}" for v in stats["rng"]]
            disp_avg    = [f"{v:.7f}" for v in stats["avg"]]

            calc_data = [
                ["Xlarge"] + disp_xlarge + ["X Max=", f"{stats['Xmax']:.7f}"],
                ["Xsmall"] + disp_xsmall + ["X Min=", f"{stats['Xmin']:.7f}"],
                ["Range"]  + disp_rng    + ["R-Bar=", f"{stats['Rbar']:.7f}"],
                ["Avg"]    + disp_avg    + ["X-Bar=", f"{stats['Xbar']:.7f}"],
            ]
            for i in range(len(calc_data)):
                calc_data[i] = calc_data[i][:len(self.calc_columns)]

            self.calc_table.set_sheet_data(calc_data)
            self._lock_calc_widths_helper()

            # ── 2. Fill Small Histogram Table ─────────────────────────
            def f(v): return "–" if v is None else f"{v:.7f}"

            design_center = spec_width = shift_xbar = None
            if tol_valid:
                ltl = float(hp["LTL"])
                utl = float(hp["UTL"])
                design_center = (ltl + utl) / 2
                spec_width    = utl - ltl
                shift_xbar    = hp["Xbar"] - design_center

            small_data = [
                ["Process width (P)",  f(hp["process_width"]),
                 "Specification Width (S)", f(spec_width),
                 "Total Range (R)", f(hp["total_range"])],

                ["Design center (D)", f(design_center),
                 "Interval", f(hp["interval"]),
                 "Selecting No. of Classes", str(hp["k"])],

                ["-", "–",
                 "No. of Readings", str(hp["N"]),
                 "Shift of X\u0304 from 'D'", f(shift_xbar)],
            ]
            self.calc_table_small.set_sheet_data(small_data)
            self.after_idle(self._lock_small_hist_widths)

            # ── 3. Update Tolerance Labels ────────────────────────────
            self.val_ltl.configure(text=str(hp["LTL"]))
            self.val_utl.configure(text=str(hp["UTL"]))

            # ── 4. Update SPC Metrics (Cp, Cpk, UCL/LCL …) ──────────
            self._last_spc_metrics = spc
            try:
                self.cap_ucl_x.configure(text=f(spc.get("ucl_x")))
                self.cap_lcl_x.configure(text=f(spc.get("lcl_x")))
                self.cap_ucl_r.configure(text=f(spc.get("ucl_r")))
                self.cap_lcl_r.configure(text=f(spc.get("lcl_r")))
                self.cap_sigma.configure(text=f(spc.get("sigma")))
                self.cap_cp.configure(text=f(spc.get("cp")))
                self.cap_cpk.configure(text=f(spc.get("cpk")))
                self.cap_cpk_l.configure(text=f(spc.get("cpk_l")))
                self.cap_cpk_u.configure(text=f(spc.get("cpk_u")))
                self._update_cpk_status_text(spc.get("cpk"))
            except Exception:
                pass

            # ── 5. Update PP Metrics ──────────────────────────────────
            if tol_valid and pp:
                self.pp_stddev.configure(text=f(pp.get("std")))
                self.pp_p.configure(text=f(pp.get("pp")))
                self.pp_ppk.configure(text=f(pp.get("ppk")))
                self.pp_ppk_l.configure(text=f(pp.get("ppk_l")))
                self.pp_ppk_u.configure(text=f(pp.get("ppk_u")))

            # ── 6. Populate Frequency Table ───────────────────────────
            self._populate_histogram_frequency_table(freq_rows)

            # ── 7. Build Chart UI (loads PNG images from disk) ────────
            self._xbar_img_path, self._rbar_img_path, self._hist_img_path = charts
            if hist_int_path:
                self._hist_interval_img_path = hist_int_path
            self._build_spc_chart_ui()

            # ── 8. Clear status text ──────────────────────────────────
            if hasattr(self, "status_label"):
                self.status_label.configure(text="")

            # ── 9. Force Tkinter to finish ALL pending redraws ────────
            # update_idletasks() flushes every pending configure/pack/grid
            # call so every widget is fully painted BEFORE the overlay
            # fades away.  This is the key step that prevents the
            # "content appearing one-by-one" effect.
            self.update_idletasks()

            # ── 10. NOW dismiss overlays — page is 100% ready ─────────
            # Both the inner chart overlay and the global app overlay are
            # removed only after all data is visible.
            self._hide_loading_over_charts()

            if hasattr(self.app, "_dismiss_overlay_callback") and self.app._dismiss_overlay_callback:
                self.app._dismiss_overlay_callback()
                self.app._dismiss_overlay_callback = None

        except Exception as e:
            print("UI Update Failed:", e)
            import traceback
            traceback.print_exc()
            # Safety: always dismiss overlays even on error
            try:
                self._hide_loading_over_charts()
            except Exception:
                pass
            try:
                if hasattr(self.app, "_dismiss_overlay_callback") and self.app._dismiss_overlay_callback:
                    self.app._dismiss_overlay_callback()
                    self.app._dismiss_overlay_callback = None
            except Exception:
                pass


    def _lock_calc_widths_helper(self):
         try:
            column_widths = [102] * (len(self.calc_columns) - 2) + [97, 97]
            for i, w in enumerate(column_widths):
                self.calc_table.column_width(column=i, width=w, only_set_if_too_small=False)
            self.calc_table.refresh()
         except: pass

    # ------------------------------
    # Loading Overlay Helpers
    # ------------------------------
    def _show_loading_over_charts(self):
        try:
            if hasattr(self, "loading_overlay") and self.loading_overlay.winfo_exists():
                return
            
            # Create an overlay frame on top of the charts area
            # We'll place it relative to the 'spc_row' or scroll_frame 
            # ideally covering the chart area.
            # Simplified: Just create a centered toplevel-like frame or pack at top
            
            self.loading_overlay = ctk.CTkFrame(self, fg_color="rgba(255,255,255,0.7)", corner_radius=8)
            self.loading_overlay.place(relx=0.5, rely=0.5, anchor="center")
            
            lbl = ctk.CTkLabel(self.loading_overlay, text="⟳ Calculating & Plotting...", 
                               font=("Segoe UI", 18, "bold"), text_color="#1565C0")
            lbl.pack(padx=30, pady=20)
            
            # Disable interactions if needed
            
        except Exception: 
            pass

    def _hide_loading_over_charts(self):
        try:
            if hasattr(self, "loading_overlay"):
                self.loading_overlay.destroy()
                del self.loading_overlay
        except Exception:
            pass


    def _build_spc_chart_ui(self):
        # 1. Ensure shared row container exists (created by _build_histogram_frequency_table if needed)
        if not hasattr(self, "charts_and_table_row"):
            self.charts_and_table_row = tk.Frame(self.scroll_frame, bg="white")
            self.charts_and_table_row.pack(fill="x", padx=18, pady=(10, 20))
            # Configure grid columns: fixed width for charts, remaining for table
            self.charts_and_table_row.grid_columnconfigure(0, weight=0, minsize=540)  # Chart column (520px chart + 20px padding)
            self.charts_and_table_row.grid_columnconfigure(1, weight=1, minsize=535)  # Table column (flexible)

        from PIL import Image
        import customtkinter as ctk

        # ================= LEFT : Charts Container (Grid Column 0) =================
        # Create left container for charts (X-Bar and R-Bar) with fixed width
        # Create charts container only once
        if not hasattr(self, "charts_left"):
            self.charts_left = tk.Frame(self.charts_and_table_row, bg="white", width=540)
            self.charts_left.grid(row=0, column=0, sticky="nw")
        left = self.charts_left
        for w in left.winfo_children():
            w.destroy()

        def load_img(path, w=720, h=230):
            pil_img = Image.open(path)
            return ctk.CTkImage(
                light_image=pil_img,
                dark_image=pil_img,
                size=(w, h)
            )

        # X-Bar Chart
        xbar_img = load_img(self._xbar_img_path)
        lbl_x = ctk.CTkLabel(left, image=xbar_img, text="")
        lbl_x.image = xbar_img   # 🔥 still REQUIRED
        lbl_x.pack(pady=(0, 10), anchor="nw")

        # R-Bar Chart
        rbar_img = load_img(self._rbar_img_path)
        lbl_r = ctk.CTkLabel(left, image=rbar_img, text="")
        lbl_r.image = rbar_img
        lbl_r.pack(anchor="nw")

        # ================= INTERVAL HISTOGRAM =================
        if hasattr(self, "_hist_interval_img_path") and self._hist_interval_img_path:
            hist_img = load_img(self._hist_interval_img_path, w=760, h=260)
            lbl_h = ctk.CTkLabel(left, image=hist_img, text="")
            lbl_h.image = hist_img
            lbl_h.pack(pady=(12, 0), anchor="nw")


        # Note: Frequency table (freq_holder) is already placed in grid column 1
        # in _build_histogram_frequency_table, so no need to repack it here
        



            
    def _update_capability_ui(self):
        metrics = getattr(self, "_last_spc_metrics", None)
        if not metrics:
            return

        def fmt(v):
            return "–" if v is None else f"{v:.7f}"

        try:
            self.cap_ucl_x.configure(text=fmt(metrics.get("ucl_x")))
            self.cap_lcl_x.configure(text=fmt(metrics.get("lcl_x")))
            self.cap_ucl_r.configure(text=fmt(metrics.get("ucl_r")))
            self.cap_lcl_r.configure(text=fmt(metrics.get("lcl_r")))

            self.cap_sigma.configure(text=fmt(metrics.get("sigma")))
            self.cap_cp.configure(text=fmt(metrics.get("cp")))
            self.cap_cpk.configure(text=fmt(metrics.get("cpk")))
            
        except Exception as e:
            print("Capability UI update failed:", e)

    def _get_tolerance_for_airgauge(self, airgauge, channel):
        """
        Returns (low_tol, high_tol) for given airgauge + channel.
        If not found, returns ("–", "–")
        If "All" is selected for AirGauge or Channel, checks if a uniform tolerance exists across all matching components.
        """
        try:
            ag = str(airgauge).strip()
            ch = str(channel).strip()
            
            # PRIORITY: If we came from Report page, row_ltl/row_utl were explicitly set.
            # Use them ALWAYS — if empty it means mixed data → show "–", never fall through to comp_json.
            if hasattr(self, "row_ltl") and hasattr(self, "row_utl"):
                ltl = getattr(self, "row_ltl", "")
                utl = getattr(self, "row_utl", "")
                if ltl and ltl not in ("", "0", "0.0") and utl and utl not in ("", "0", "0.0"):
                    return ltl, utl
                else:
                    return "–", "–"

            # Format channel properly if not 'All'
            if ch != "All" and not ch.upper().startswith("CH"):
                ch = f"CH{ch}"

            # Fallback: Access comp_json from the main app instance
            if hasattr(self, "app") and hasattr(self.app, "comp_json") and self.app.comp_json:
                comp_data = self.app.comp_json
            else:
                # Last resort: load directly from DB
                try:
                    conn = sqlite3.connect(resource_path("component_setup.db"))
                    cursor = conn.cursor()
                    cursor.execute("SELECT airgauge_id, channel, properties FROM component_setup")
                    rows = cursor.fetchall()
                    conn.close()
                    comp_data = {}
                    for ag_id, ch_db, props_str in rows:
                        if ag_id not in comp_data:
                            comp_data[ag_id] = {}
                        try:
                            comp_data[ag_id][ch_db] = json.loads(props_str)
                        except:
                            comp_data[ag_id][ch_db] = {}
                except Exception:
                    return "–", "–"

            # CASE 1: Specific AirGauge & Specific Channel
            if ag != "All" and ch != "All":
                if ag in comp_data and ch in comp_data[ag]:
                    low  = comp_data[ag][ch].get("low_tolerance",  "–")
                    high = comp_data[ag][ch].get("high_tolerance", "–")
                    return low, high
                return "–", "–"

            # CASE 2: "All" AirGauges and/or "All" Channels
            # Scan all matching combinations to see if they share uniform limits.
            valid_lows = set()
            valid_highs = set()
            
            for d_ag, channels in comp_data.items():
                if ag != "All" and d_ag != ag: continue
                
                for d_ch, props in channels.items():
                    if ch != "All" and d_ch != ch: continue
                    
                    l = props.get("low_tolerance", "")
                    h = props.get("high_tolerance", "")
                    try:
                        if l != "" and l != "–": valid_lows.add(float(l))
                    except: pass
                    try:
                        if h != "" and h != "–": valid_highs.add(float(h))
                    except: pass

            # If strictly 1 unified limit exists across all tracked data, display it.
            if len(valid_lows) == 1 and len(valid_highs) == 1:
                return str(valid_lows.pop()), str(valid_highs.pop())

            return "–", "–"

        except Exception:
            return "–", "–"

    def _get_spc_constants(self, n):
        """
        Returns SPC constants (d2, A2, D3, D4) for given subgroup size n.
        Source: Standard SPC control chart constants (n = 1 to 25)

        If n is out of range, returns None values.
        """

        SPC_CONSTANTS = {
            1:  {"d2": 1.123, "A2": 2.560, "D3": 0.000, "D4": 3.270},
            2:  {"d2": 1.128, "A2": 1.880, "D3": 0.000, "D4": 3.267},
            3:  {"d2": 1.693, "A2": 1.023, "D3": 0.000, "D4": 2.574},
            4:  {"d2": 2.059, "A2": 0.729, "D3": 0.000, "D4": 2.282},
            5:  {"d2": 2.326, "A2": 0.577, "D3": 0.000, "D4": 2.114},
            6:  {"d2": 2.534, "A2": 0.483, "D3": 0.000, "D4": 2.004},
            7:  {"d2": 2.704, "A2": 0.419, "D3": 0.076, "D4": 1.924},
            8:  {"d2": 2.847, "A2": 0.373, "D3": 0.136, "D4": 1.864},
            9:  {"d2": 2.970, "A2": 0.337, "D3": 0.184, "D4": 1.816},
            10: {"d2": 3.078, "A2": 0.308, "D3": 0.000, "D4": 1.777},
            11: {"d2": 3.173, "A2": 0.285, "D3": 0.256, "D4": 1.744},
            12: {"d2": 3.258, "A2": 0.266, "D3": 0.283, "D4": 1.717},
            13: {"d2": 3.336, "A2": 0.249, "D3": 0.307, "D4": 1.693},
            14: {"d2": 3.407, "A2": 0.235, "D3": 0.328, "D4": 1.672},
            15: {"d2": 3.472, "A2": 0.223, "D3": 0.347, "D4": 1.653},
            16: {"d2": 3.532, "A2": 0.212, "D3": 0.363, "D4": 1.637},
            17: {"d2": 3.588, "A2": 0.203, "D3": 0.378, "D4": 1.622},
            18: {"d2": 3.640, "A2": 0.194, "D3": 0.391, "D4": 1.608},
            19: {"d2": 3.689, "A2": 0.187, "D3": 0.403, "D4": 1.597},
            20: {"d2": 3.735, "A2": 0.180, "D3": 0.415, "D4": 1.585},
            21: {"d2": 3.778, "A2": 0.173, "D3": 0.425, "D4": 1.575},
            22: {"d2": 3.819, "A2": 0.167, "D3": 0.434, "D4": 1.566},
            23: {"d2": 3.858, "A2": 0.162, "D3": 0.443, "D4": 1.557},
            24: {"d2": 3.895, "A2": 0.157, "D3": 0.451, "D4": 1.548},
            25: {"d2": 3.931, "A2": 0.153, "D3": 0.459, "D4": 1.541},
        }

        return SPC_CONSTANTS.get(int(n), {
            "d2": None,
            "A2": None,
            "D3": None,
            "D4": None
        })

    def _compute_spc_metrics(self, values):
        """
        Computes SPC metrics using true SPC formulas.
        Returns a dict with all final values needed for GUI/PDF.

        values: flat list of floats (already validated)
        """

        import statistics
        import datetime

        result = {
            "Xbarbar": None,
            "Rbar": None,
            "sigma": None,
            "ucl_x": None,
            "lcl_x": None,
            "ucl_r": None,
            "lcl_r": None,
            "cp": None,
            "cpk": None,
            "date": datetime.date.today().strftime("%d-%m-%Y")
        }

        if not values or len(values) < 2:
            return result

        # --------------------------------------------------
        # Subgrouping (10 per group)
        # --------------------------------------------------
        subgroup_size = 10
        rows = [values[i:i + subgroup_size]
                for i in range(0, len(values), subgroup_size)
                if len(values[i:i + subgroup_size]) == subgroup_size]

        if not rows:
            return result

        # --------------------------------------------------
        # Column-based Rbar AND Xbarbar (matches display table)
        # Transpose rows to get columns C1..C10, then
        # compute range and mean per column across all subgroups
        # --------------------------------------------------
        cols = list(zip(*rows))
        col_ranges = [max(c) - min(c) for c in cols]
        col_means  = [statistics.mean(c) for c in cols]
        Rbar    = statistics.mean(col_ranges)
        Xbarbar = round(statistics.mean(col_means), 4)

        result["Xbarbar"] = Xbarbar
        result["Rbar"] = Rbar

        # --------------------------------------------------
        # SPC constants lookup
        # --------------------------------------------------
        c = self._get_spc_constants(subgroup_size)
        d2, A2, D3, D4 = c["d2"], c["A2"], c["D3"], c["D4"]

        if not d2:
            return result

        # --------------------------------------------------
        # True SPC sigma
        # Round to 7 decimals so displayed value matches
        # the value actually used in CPK formulas
        # --------------------------------------------------
        sigma = round(Rbar / d2, 7)
        result["sigma"] = sigma

        # --------------------------------------------------
        # Control limits (use rounded values)
        # --------------------------------------------------
        result["ucl_x"] = Xbarbar + A2 * Rbar
        result["lcl_x"] = Xbarbar - A2 * Rbar
        result["ucl_r"] = D4 * Rbar
        result["lcl_r"] = D3 * Rbar

        # --------------------------------------------------
        # Tolerance (from config)
        # --------------------------------------------------
        rp = getattr(self.app, "last_report_filters", {})
        air = rp.get("airgauge", "All")
        ch  = rp.get("channel", "All")

        low_tol, high_tol = self._get_tolerance_for_airgauge(air, ch)

        try:
            LSL = float(low_tol)
            USL = float(high_tol)
        except Exception:
            return result

        # --------------------------------------------------
        # Cp & Cpk (true SPC) — uses rounded sigma
        # --------------------------------------------------
        try:
            cp = (USL - LSL) / (6 * sigma)

            cpk_l = (Xbarbar - LSL) / (3 * sigma)
            cpk_u = (USL - Xbarbar) / (3 * sigma)
            cpk   = min(cpk_l, cpk_u)

            result["cp"]    = cp
            result["cpk"]   = cpk
            result["cpk_l"] = cpk_l
            result["cpk_u"] = cpk_u

        except Exception:
            pass

        return result

    def _compute_pp_metrics(self, values, LSL, USL):
        import statistics

        if not values or LSL is None or USL is None:
            return {}

        xbar = round(statistics.mean(values), 4)
        # Round to 7 decimals so displayed value matches
        # the value actually used in PPK formulas
        s = round(statistics.stdev(values), 7)  # overall std dev (n-1)

        ppk_l = (xbar - LSL) / (3 * s)
        ppk_u = (USL - xbar) / (3 * s)

        return {
            "std": s,
            "pp": (USL - LSL) / (6 * s),
            "ppk": min(ppk_l, ppk_u),
            "ppk_l": ppk_l,
            "ppk_u": ppk_u
        }
    def _generate_histogram_from_frequency(self, freq_rows):
        import os, tempfile, math
        from matplotlib.figure import Figure
        from matplotlib.backends.backend_agg import FigureCanvasAgg

        labels = []
        freqs = []

        # -----------------------------
        # Build X-axis labels & Y values
        # -----------------------------
        for row in freq_rows:
            interval_from = row[0]
            interval_to   = row[1]
            freq          = row[3]   # ONLY FREQ

            label = f"{interval_from}\n|\n{interval_to}"
            labels.append(label)
            freqs.append(freq)

        if not freqs:
            return None

        # -----------------------------
        # Y-axis ticks (4 values only)
        # -----------------------------
        max_freq = max(freqs)
        step = math.ceil(max_freq / 3) if max_freq > 0 else 1
        y_ticks = [0, step, step * 2, step * 3]

        # -----------------------------
        # Plot histogram (Thread-safe OO approach)
        # -----------------------------
        tmp_dir = tempfile.gettempdir()
        path = os.path.join(tmp_dir, "histogram_interval_freq.png")

        fig = Figure(figsize=(10.5, 3.5), dpi=150)
        ax = fig.add_subplot(111)
        ax.bar(range(len(freqs)), freqs, color="#42A5F5")

        ax.set_xticks(range(len(labels)))
        ax.set_xticklabels(labels, fontsize=8)
        ax.set_yticks(y_ticks)
        ax.set_xlabel("Interval (From | To)")
        ax.set_ylabel("Frequency")
        ax.set_title("Histogram \u2013 Interval Frequency Distribution")

        try:
            fig.tight_layout()
        except Exception:
            pass
        canvas = FigureCanvasAgg(fig)
        canvas.print_figure(path, dpi=150)

        return path

    # ------------------------------
    # Navigation
    # ------------------------------
    def go_back(self, event=None):
        try:
            # Restore sidebar if user had it visible
            if hasattr(self.app, "sidebar") and getattr(self.app, "sidebar_user_visible", False) and not self.app.sidebar.winfo_ismapped():
                self.app.sidebar.pack(side="left", fill="y", before=self.app.content_frame)

            # Reset content frame position
            if hasattr(self.app, "content_frame"):
                self.app.content_frame.pack_forget()
                self.app.content_frame.pack(side="left", fill="both", expand=True)

            # Clear current page
            for w in self.app.content_frame.winfo_children():
                w.destroy()

            # Load ReportPage
            if hasattr(self.app, "load_report_page"):
                self.app.load_report_page()
            else:
                from report import ReportPage  # if exists; fallback suppressed
                rp = ReportPage(self.app.content_frame, self.app)
                rp.pack(fill="both", expand=True)

        except Exception as e:
            print("Failed to go back:", e)

    def print_spc_report(self):
        """
        Generates a COMPLETE Professional SPC PDF Report.
        Replicates ALL content from AnalysisPage.
        """
        try:
            if 'SimpleDocTemplate' not in globals():
                messagebox.showerror("Error", "ReportLab library not available.")
                return

            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"SPC_Analysis_{timestamp}.pdf"
            report_dir = os.path.join(os.getcwd(), "Reports")
            os.makedirs(report_dir, exist_ok=True)
            filepath = os.path.join(report_dir, filename)

            # INCREASE TOP MARGIN TO 20mm (Approx 2cm) as requested
            doc = SimpleDocTemplate(filepath, pagesize=A4, rightMargin=8*mm, leftMargin=8*mm, topMargin=20*mm, bottomMargin=15*mm)
            elements = []
            styles = getSampleStyleSheet()

            # --- HEADER ---
            title_style = ParagraphStyle('Title', parent=styles['Heading1'], alignment=1, textColor=colors.HexColor("#0D47A1"), fontSize=14)
            elements.append(Paragraph("SPC Analysis Report", title_style))
            elements.append(Spacer(1, 3*mm))

            # --- FILTER GRID (Metadata) ---
            # Load Operator Map for correct name display
            op_map = {}
            try:
                with open("operators.json", "r") as f:
                    op_data = json.load(f)
                    op_map = {str(item.get("id", "")): item.get("name", "") for item in op_data}
            except: pass

            rp = getattr(self.app, "last_report_filters", {})
            f_from = rp.get("from_date", ""); f_to = rp.get("to_date", "")
            f_ag = rp.get("airgauge", ""); f_ch = rp.get("channel", "")
            f_item = rp.get("item", ""); f_drw = rp.get("drawing", "")
            f_op_raw = rp.get("operator", "")
            f_mac = rp.get("machine", "")
            
            # Map ID to Name if possible
            f_op = op_map.get(str(f_op_raw), f_op_raw)
            if not f_op and f_op_raw: f_op = f_op_raw 
            
            ag_ch = f"{f_ag}  /  CH: {f_ch}"

            lbl_sty = ParagraphStyle('lbl2', fontName='Helvetica-Bold',
                                     fontSize=8, textColor=colors.white)
            val_sty = ParagraphStyle('val2', fontName='Helvetica', fontSize=8)

            def La(t): return Paragraph(t, lbl_sty)
            def Va(t): return Paragraph(str(t), val_sty)

            filter_data = [
                [La("Item:"),       Va(f_item), La("AirGauge ID:"), Va(ag_ch), La("Drawing:"),  Va(f_drw)],
                [La("Operator:"),   Va(f_op),   La("Machine ID:"),  Va(f_mac), La("Customer:"), Va("")],
            ]

            page_w = A4[0] - 16*mm
            lbl_w = page_w * 0.12
            val_w = page_w * 0.21
            col_widths = [lbl_w, val_w, lbl_w, val_w, lbl_w, val_w]

            f_table = Table(filter_data, colWidths=col_widths)
            f_table.setStyle(TableStyle([
                ('GRID',          (0, 0), (-1, -1), 0.4, colors.lightgrey),
                ('BACKGROUND',    (0, 0), (-1, -1), colors.whitesmoke),
                ('BACKGROUND',    (0, 0), (0, -1),  colors.HexColor("#0D47A1")),
                ('BACKGROUND',    (2, 0), (2, -1),  colors.HexColor("#0D47A1")),
                ('BACKGROUND',    (4, 0), (4, -1),  colors.HexColor("#0D47A1")),
                ('FONTSIZE',      (0, 0), (-1, -1), 8),
                ('TOPPADDING',    (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ('LEFTPADDING',   (0, 0), (-1, -1), 6),
                ('RIGHTPADDING',  (0, 0), (-1, -1), 4),
                ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            elements.append(f_table)
            elements.append(Spacer(1, 4*mm))


            # Helper for Tables (Grouped)
            def add_section_table(title, header, data, widths, heading_style=styles["Heading3"], 
                                  parent_list=elements, is_header_row=True, keep_together=True, custom_style=None):
                if not data: return
                group = []
                group.append(Paragraph(f"<b>{title}</b>", heading_style))
                t_data = [header] + data if header else data
                safe_data = []
                for row in t_data:
                    safe_row = [str(cell) if cell is not None else "" for cell in row]
                    safe_data.append(safe_row)
                
                t = Table(safe_data, colWidths=widths, repeatRows=1 if is_header_row else 0)
                
                style_cmds = [
                    ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
                    ('FONTSIZE', (0,0), (-1,-1), 7),
                    ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                    ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                    ('PADDING', (0,0), (-1,-1), 3),
                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ]
                
                if is_header_row:
                    style_cmds.extend([
                        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#0D47A1")),
                        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.whitesmoke]),
                    ])
                else:
                    style_cmds.extend([
                        ('ROWBACKGROUNDS', (0,0), (-1,-1), [colors.white, colors.whitesmoke]),
                    ])
                
                if custom_style:
                    style_cmds.extend(custom_style)

                t.setStyle(TableStyle(style_cmds))
                group.append(t)
                
                if keep_together:
                    parent_list.append(KeepTogether(group))
                else:
                    parent_list.extend(group)
                
                parent_list.append(Spacer(1, 1*mm)) # Reduce spacing (was 3mm)

            # --- PROCESS CAPABILITY RESULTS (Computed Early for Layout) ---
            # 1. Flatten all data from the TABLE
            all_values = []
            if hasattr(self, "table") and self.table:
                try:
                    raw_data = self.table.get_sheet_data()
                    for row in raw_data:
                        if len(row) > 1:
                            for cell in row[1:]:
                                try:
                                    val = float(str(cell))
                                    all_values.append(val)
                                except: pass
                except: pass

            # 2. Compute Metrics
            spc = self._compute_spc_metrics(all_values)
            
            rp = getattr(self.app, "last_report_filters", {})
            f_ag = rp.get("airgauge", "All")
            f_ch = rp.get("channel", "All")
            f_from = rp.get("from_date", "Unknown")
            f_to = rp.get("to_date", "Unknown")

            l_tol, h_tol = self._get_tolerance_for_airgauge(f_ag, f_ch)
            try: LSL, USL = float(l_tol), float(h_tol)
            except: LSL, USL = None, None

            pp = self._compute_pp_metrics(all_values, LSL, USL)

            def fmt(v, p=7): return f"{v:.{p}f}" if isinstance(v, (int, float)) else str(v)

            # --- DATE HEADER ---
            # "DATA COLLECTION FROM {start} TO {end}"
            date_header_style = ParagraphStyle('DateHeader', parent=styles['Heading3'], alignment=1, spaceBefore=2*mm, spaceAfter=2*mm)
            elements.append(Paragraph(f"<b>DATA COLLECTION FROM {f_from} TO {f_to}</b>", date_header_style))

            # --- 10-COLUMN GRID (5 PAIRS) ---
            # Row 1: UCL(X), LCL(X), UCL(R), LCL(R), (Empty)
            r1 = ["UCL (X)", fmt(spc.get('ucl_x')), "LCL (X)", fmt(spc.get('lcl_x')), 
                  "UCL (R)", fmt(spc.get('ucl_r')), "LCL (R)", fmt(spc.get('lcl_r')), "", ""]
            
            # Row 2: Sigma, CPK_l, CPK_u, CP, CPK (Swapped as requested)
            r2 = ["σ", fmt(spc.get('sigma')), "CPK_l", fmt(spc.get('cpk_l')), 
                  "CPK_u", fmt(spc.get('cpk_u')), "CP", fmt(spc.get('cp')), "CPK", fmt(spc.get('cpk'))]
            
            # Row 3: Std Dev, PPK_l, PPK_u, PP, PPK (Swapped as requested)
            r3 = ["Std Dev", fmt(pp.get('std')), "PPK_l", fmt(pp.get('ppk_l')),
                  "PPK_u", fmt(pp.get('ppk_u')), "PP", fmt(pp.get('pp')), "PPK", fmt(pp.get('ppk'))]


            res_data = [r1, r2, r3]
            col_w = page_w / 10.0
            
            # Add Result Table (10 cols)
            # Bold labels: cols 0, 2, 4, 6, 8
            res_style = [
                ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
                ('FONTNAME', (2,0), (2,-1), 'Helvetica-Bold'),
                ('FONTNAME', (4,0), (4,-1), 'Helvetica-Bold'),
                ('FONTNAME', (6,0), (6,-1), 'Helvetica-Bold'),
                ('FONTNAME', (8,0), (8,-1), 'Helvetica-Bold'),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('FONTSIZE', (0,0), (-1,-1), 8), # Slightly smaller font to fit 10 cols
            ]
            add_section_table("", [], res_data, [col_w]*10, is_header_row=False, keep_together=True, custom_style=res_style)

            # --- RESULT TEXT ---
            cpk_val = spc.get('cpk')
            process_status = "UNKNOWN"
            status_color = colors.black
            
            if cpk_val is not None:
                if cpk_val > 1.67:
                    process_status = "PROCESS IS EXCELLENT"
                    status_color = colors.green
                elif cpk_val > 1.33:
                    process_status = "PROCESS IS STABLE"
                    status_color = colors.blue
                elif cpk_val > 1.0:
                    process_status = "PROCESS IS CAPABLE"
                    status_color = colors.orange
                else:
                    process_status = "PROCESS IS POOR, ACTION REQUIRED"
                    status_color = colors.red

            # Reduced spacing for result text
            p_style = ParagraphStyle('Result', parent=styles['Heading2'], alignment=0, textColor=status_color, spaceBefore=2*mm, spaceAfter=4*mm)
            elements.append(Paragraph(f"<b>RESULT: {process_status}</b>", p_style))


            # 1. OBSERVATIONS / READINGS
            full_obs_rows = []
            if hasattr(self, "table") and self.table:
                try:
                    raw_data = self.table.get_sheet_data()
                    for row in raw_data: # [SNo, v1..v10]
                        try:
                            s_no = row[0]
                            disp_vals = [str(v) for v in row[1:]]
                            while len(disp_vals) < 10: disp_vals.append("")
                            # Only SNo + 10 Values (No Avg/Range)
                            full_obs_rows.append([str(s_no)] + disp_vals[:10])
                        except: pass
                except: pass

            if full_obs_rows:
                # 11 cols: SNo, C1-C10
                obs_header = ["S.No"] + [f"C{i}" for i in range(1, 11)]
                w_sno = 15*mm
                w_val = (page_w - w_sno) / 10.0
                add_section_table("Data Collection (10xN)", obs_header, full_obs_rows, [w_sno] + [w_val]*10, is_header_row=True, keep_together=False)

            # 2. COLUMN STATISTICS
            calc_rows = []
            if hasattr(self, "calc_table") and self.calc_table:
                try: calc_rows = self.calc_table.get_sheet_data()
                except: pass
            
            if calc_rows:
                n_cols = len(calc_rows[0]) if calc_rows else 10
                w_calc = page_w / float(n_cols)
                
                # Custom Styling
                # User request: "bg is blue text is white" for 1st and 12th column
                stats_style = [
                    ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),   # First Col
                    ('FONTNAME', (-2,0), (-2,-1), 'Helvetica-Bold'), # 2nd to last Col (Col 11 if 13 cols)
                    
                    ('BACKGROUND', (0,0), (0,-1), colors.HexColor("#0D47A1")),   # Blue BG First Col
                    ('TEXTCOLOR', (0,0), (0,-1), colors.white),                  # White Text First Col
                    
                    ('BACKGROUND', (-2,0), (-2,-1), colors.HexColor("#0D47A1")), # Blue BG 12th Col
                    ('TEXTCOLOR', (-2,0), (-2,-1), colors.white),                # White Text 12th Col
                    
                    ('FONTNAME', (-1,0), (-1,-1), 'Helvetica-Bold'),             # Bold Last Col (Values)
                ]
                
                add_section_table("Column Statistics", [], calc_rows, [w_calc]*n_cols, is_header_row=False, keep_together=True, custom_style=stats_style)

            # 3. FREQUENCY DISTRIBUTION
            freq_rows = []
            if hasattr(self, "freq_table") and self.freq_table:
               try: freq_rows = self.freq_table.get_sheet_data()
               except: pass
            
            if freq_rows:
                freq_header = ["Interval From", "Interval To", "Cumulative", "Frequency"]
                w_freq = page_w / 4.0
                add_section_table("Frequency Distribution", freq_header, freq_rows, [w_freq]*4, keep_together=True, is_header_row=True)

            elements.append(Spacer(1, 4*mm))

            # --- CHARTS ---
            # Regenerate high res if needed, or use cached
            flat_vals = []
            if hasattr(self, "table") and self.table:
                 try:
                     raw_rows = self.table.get_sheet_data()
                     for r in raw_rows:
                         for c in r[1:]: 
                             try: flat_vals.append(float(c))
                             except: pass
                 except: pass

            if flat_vals:
                if hasattr(self, "_xbar_img_path"):
                    xbar_img = self._xbar_img_path
                    rbar_img = self._rbar_img_path
                    hist_img = self._hist_img_path
                else:
                    xbar_img, rbar_img, hist_img = self._generate_spc_charts(flat_vals)

                # Images (Vertical Layout)
                img_w = 170*mm
                img_h = 60*mm 

                elements.append(KeepTogether([
                    Paragraph("X-Bar Chart", styles["Heading3"]),
                    RLImage(xbar_img, width=img_w, height=img_h),
                    Spacer(1, 2*mm)
                ]))

                elements.append(KeepTogether([
                    Paragraph("R-Bar Chart", styles["Heading3"]),
                    RLImage(rbar_img, width=img_w, height=img_h),
                    Spacer(1, 2*mm)
                ]))
                
                elements.append(KeepTogether([
                    Paragraph("Histogram", styles["Heading3"]),
                    RLImage(hist_img, width=img_w, height=img_h)
                ]))

            # --- HEADER & FOOTER ON EACH PAGE ---
            # --- HEADER & FOOTER ON EACH PAGE ---
            def _header_footer(canvas, doc):
                canvas.saveState()
                width, height = A4
                
                # 1. LOGO (Top-Left)
                # Robust Path Finding
                script_dir = os.path.dirname(os.path.abspath(__file__))
                logo_path = os.path.join(script_dir, "settings", "cherry_full_logo.png")
                
                if not os.path.exists(logo_path):
                     logo_path = os.path.join("settings", "cherry_full_logo.png") # Retry relative

                if os.path.exists(logo_path):
                    try:
                        # Draw logo top-left
                        # Position: 8mm from left, 12mm from top (y ~ 285mm)
                        # A4 height is 297mm.
                        canvas.drawImage(logo_path, 8*mm, height - 15*mm, width=35*mm, height=12*mm, preserveAspectRatio=True, mask='auto', anchor='nw')
                    except Exception as e: 
                        print(f"Logo draw error: {e}")

                # 2. HEADER TEXT (Customer Name)
                right_margin = width - 20*mm
                left_margin = 20*mm
                top_y = height - 10*mm
                
                canvas.setFont('Helvetica', 10)
                canvas.setFillColor(colors.gray)
                c_name = rp.get("customer", "")
                header_text = f"Customer name: {c_name}"
                canvas.drawRightString(right_margin, top_y, header_text)
                
                # Dark Line Below Header - Adjusted down (was 3mm)
                line_y = top_y - 10*mm
                canvas.setLineWidth(1.5)
                canvas.setStrokeColor(colors.black)
                canvas.line(left_margin, line_y, right_margin, line_y)
                
                # 3. FOOTER (Page Number & Company Name)
                footer_y = 10*mm 
                
                # Separator Line (Light Grey)
                canvas.setLineWidth(1)
                canvas.setStrokeColor(colors.lightgrey)
                canvas.line(left_margin, footer_y + 4*mm, right_margin, footer_y + 4*mm)

                page_num = doc.page
                canvas.setFont('Helvetica', 9)
                canvas.setFillColor(colors.black)
                
                # Page Number (Right)
                canvas.drawRightString(right_margin, footer_y, f"Page {page_num}")
                
                # Company Name (Left)
                try:
                    company_name = SetupDatabase.get_company_name()
                    canvas.drawString(left_margin, footer_y, company_name)
                except:
                    pass
                
                canvas.restoreState()

            doc.build(elements, onFirstPage=_header_footer, onLaterPages=_header_footer)
            try: os.startfile(filepath)
            except: pass
            messagebox.showinfo("Export Success", f"Analysis Report generated:\n{filepath}")

        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to generate SPC PDF: {str(e)}")                # 11 cols: SNo, C1-C10

    def _update_cpk_status_text(self, cpk):
        if cpk is None:
            self.cpk_status_label.configure(text="")
            return

        try:
            cpk = float(cpk)
        except Exception:
            self.cpk_status_label.configure(text="")
            return

        if cpk < 1:
            text = "PROCESS IS VERY POOR, IMMEDIATE ACTION REQUIRED"
            color = "#C62828"   # dark red
        elif cpk < 1.33:
            text = "PROCESS NEEDS CORRECTION"
            color = "#EF6C00"   # orange
        elif cpk < 2:
            text = "PROCESS IS GOOD BUT STILL IMPROVEMENT IS REQUIRED"
            color = "#F9A825"   # amber
        else:  # cpk >= 2
            text = "PROCESS IS EXCELLENT"
            color = "#2E7D32"   # green

        self.cpk_status_label.configure(text=text, text_color=color)
    def _cap_row(self, parent, row, label):
        ctk.CTkLabel(
            parent,
            text=label,
            font=("Segoe UI", 11),
            anchor="w",
            text_color="#37474F"
        ).grid(row=row, column=0, columnspan=2, sticky="w", pady=0)

        val = ctk.CTkLabel(
            parent,
            text="–",
            font=("Segoe UI", 11, "bold"),
            anchor="e",
            text_color="#37474F"
        )
        val.grid(row=row, column=3, columnspan=2, sticky="e", pady=0)
        return val

    
    def go_back(self):
        """
        Return to ReportPage.
        Since we used pack_forget() on ReportPage, calling load_report_page()
        should restore the existing instance.
        Also need to RESTORE the sidebar if it was hidden.
        """
        # Restore Sidebar & Layout if user had it visible
        try:
            if hasattr(self.app, "sidebar") and getattr(self.app, "sidebar_user_visible", False) and not self.app.sidebar.winfo_ismapped():
                self.app.sidebar.pack(side="left", fill="y", padx=0, pady=0, before=self.app.content_frame)
            
            if hasattr(self.app, "content_frame"):
                # Reset content frame to original position (right of sidebar)
                self.app.content_frame.pack_forget()
                self.app.content_frame.pack(side="right", fill="both", expand=True)
        except Exception as e:
            print("Error restoring layout:", e)

        # Destroy self (AnalysisPage) to free resources
        self.destroy()
        
        # Restore Report Page
        if hasattr(self.app, "load_report_page"):
            self.app.load_report_page()


# ─────────────────────────────────────────────────────────────────────────────
#  Home Page — Displays welcome_bg.png filling the entire content area
# ─────────────────────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────
#  Logout Page — Displays confirmation modal card
# ─────────────────────────────────────────────────────────────────────────────
class LogoutPage(ctk.CTkFrame):
    """Clean centered logout confirmation card."""
    def __init__(self, parent, app):
        super().__init__(parent, fg_color="#F5F5F5", corner_radius=0, border_width=0)
        self.app = app

        import tkinter as _tk

        # ── Centered white card ───────────────────────────────────────────────
        card = ctk.CTkFrame(
            self,
            fg_color="white",
            corner_radius=16,
            border_width=1,
            border_color="#E8E8E8",
            width=360,
            height=268,
        )
        card.place(relx=0.5, rely=0.5, anchor="center")
        card.pack_propagate(False)

        # ── Exit icon in soft green circle ────────────────────────────────────
        icon_bg = ctk.CTkFrame(card, width=56, height=56,
                               corner_radius=28, fg_color="#EEF7EE")
        icon_bg.place(relx=0.5, rely=0.0, anchor="n", y=18)
        icon_bg.pack_propagate(False)

        icon_cv = _tk.Canvas(icon_bg, width=56, height=56,
                             bg="#EEF7EE", highlightthickness=0)
        icon_cv.place(x=0, y=0)
        # Door body (dark green)
        icon_cv.create_rectangle(12, 8, 34, 46, outline="#1B5E20", width=2, fill="")
        # Door inner indent
        icon_cv.create_rectangle(15, 12, 31, 42, outline="#2E7D32", width=1, fill="")
        # Door knob
        icon_cv.create_oval(28, 25, 33, 30, fill="#1B5E20", outline="")
        # Red exit arrow
        icon_cv.create_line(28, 27, 44, 27, fill="#C62828", width=2)
        icon_cv.create_line(37, 21, 44, 27, fill="#C62828", width=2)
        icon_cv.create_line(37, 33, 44, 27, fill="#C62828", width=2)

        # ── "Confirm Logout" two-tone heading ────────────────────────────────
        heading_frame = ctk.CTkFrame(card, fg_color="transparent")
        heading_frame.place(relx=0.5, rely=0.0, anchor="n", y=86)
        ctk.CTkLabel(
            heading_frame, text="Confirm ",
            font=("Segoe UI", 21, "bold"), text_color="#1B5E20"
        ).pack(side="left")
        ctk.CTkLabel(
            heading_frame, text="Logout",
            font=("Segoe UI", 21, "bold"), text_color="#C62828"
        ).pack(side="left")

        # ── Divider: green line — red dot — red line ──────────────────────────
        div_cv = _tk.Canvas(card, width=180, height=8,
                            bg="white", highlightthickness=0)
        div_cv.place(relx=0.5, rely=0.0, anchor="n", y=120)
        div_cv.create_line(0, 4, 78, 4,    fill="#1B5E20", width=2)
        div_cv.create_oval(80, 1, 90, 9,   fill="#C62828", outline="")
        div_cv.create_line(92, 4, 180, 4,  fill="#C62828", width=2)

        # ── Subtitle ──────────────────────────────────────────────────────────
        ctk.CTkLabel(
            card,
            text="Are you sure you want to log out of the system?",
            font=("Segoe UI", 11),
            text_color="#888888",
            fg_color="transparent",
        ).place(relx=0.5, rely=0.0, anchor="n", y=140)

        # ── Buttons ───────────────────────────────────────────────────────────
        btn_frame = ctk.CTkFrame(card, fg_color="transparent")
        btn_frame.place(relx=0.5, rely=1.0, anchor="s", y=-20)

        # "Yes, Logout" — white fill, red border
        ctk.CTkButton(
            btn_frame,
            text="➜  Yes, Logout",
            font=("Segoe UI", 12, "bold"),
            fg_color="white",
            hover_color="#FFF0F0",
            text_color="#C62828",
            border_width=2,
            border_color="#C62828",
            width=130,
            height=38,
            corner_radius=9,
            command=self.perform_logout,
        ).pack(side="left", padx=(0, 10))

        # "Cancel" — solid dark green
        ctk.CTkButton(
            btn_frame,
            text="✕  Cancel",
            font=("Segoe UI", 12, "bold"),
            fg_color="#1B5E20",
            hover_color="#2E7D32",
            text_color="white",
            border_width=0,
            width=120,
            height=38,
            corner_radius=9,
            command=self.cancel_logout,
        ).pack(side="left")

    def perform_logout(self):
        # Disconnect serial if connected
        try:
            self.app.disconnect_esp32()
        except Exception:
            pass
        
        # Revert UI state
        # Destroy main UI elements packed in self.app
        try:
            if hasattr(self.app, "header") and self.app.header:
                self.app.header.destroy()
                self.app.header = None
            if hasattr(self.app, "sidebar") and self.app.sidebar:
                self.app.sidebar.destroy()
                self.app.sidebar = None
            if hasattr(self.app, "content_frame") and self.app.content_frame:
                self.app.content_frame.destroy()
                self.app.content_frame = None
        except Exception as e:
            print("Error clearing main UI on logout:", e)
            
        # Re-show admin login page
        self.app.show_admin_login()

    def cancel_logout(self):
        # Go back to Home page
        self.app._sidebar_nav(self.app.load_home, "home")


# ─────────────────────────────────────────────────────────────────────────────
#  Home Page — Displays welcome_bg.png inside a rounded card
# ─────────────────────────────────────────────────────────────────────────────
class HomePage(tk.Frame):
    """Full-screen welcome image page shown by default after login."""

    BG_IMAGE_PATH = os.path.join("settings", "welcome_bg.png")

    def __init__(self, parent, app):
        super().__init__(parent, bg="#F4F6F8")
        self.app = app
        self._photo = None
        self._raw_img = None

        # Card container to give rounded corners and a border effect
        self.card = ctk.CTkFrame(
            self,
            fg_color="white",
            corner_radius=20,
            border_width=1,
            border_color="#E0E0E0"
        )
        self.card.pack(fill="both", expand=True, padx=15, pady=15)

        # Canvas inside the card
        self._canvas = tk.Canvas(self.card, bg="white", highlightthickness=0)
        self._canvas.pack(fill="both", expand=True, padx=10, pady=10)
        self._img_item = self._canvas.create_image(0, 0, anchor="nw")

        # Load the image once
        self._load_image()

        # Re-scale when window resizes
        self._canvas.bind("<Configure>", self._on_resize)

    def _load_image(self):
        try:
            from PIL import Image
            # ── Use preloaded image if available (loaded during splash) ──
            # The background preload thread decoded this image into RAM so
            # the first open has zero disk I/O and no decode delay.
            cached = getattr(self.app, "_cached_bg_image", None)
            if cached is not None:
                self._raw_img = cached
            else:
                img_path = resource_path(self.BG_IMAGE_PATH)
                self._raw_img = Image.open(img_path)
        except Exception as e:
            print("HomePage image load error:", e)
            self._raw_img = None

    def _on_resize(self, event=None):
        if hasattr(self, "_resize_job") and self._resize_job:
            self.after_cancel(self._resize_job)
        self._resize_job = self.after(100, self._do_resize)

    def _do_resize(self):
        self._resize_job = None
        if self._raw_img is None:
            return
        try:
            from PIL import Image, ImageTk
            w = self._canvas.winfo_width()
            h = self._canvas.winfo_height()
            if w < 10 or h < 10:
                return
            resized = self._raw_img.resize((w, h), Image.Resampling.LANCZOS)
            # Add rounded corners to matching card border
            resized = add_corners(resized, 15)
            self._photo = ImageTk.PhotoImage(resized)
            self._canvas.itemconfig(self._img_item, image=self._photo)
        except Exception as e:
            print("HomePage resize error:", e)



if __name__ == "__main__":
    # ── Create main window ────────────────────────────────────────────
    # CherryApp.__init__ now does: state("zoomed") → withdraw() → build
    # login page while hidden.  No second withdraw() needed here.
    app = CherryApp()

    # ── Full-screen splash shows immediately ──────────────────────────
    splash = SplashScreen(app)
    splash.update()

    # ════════════════════════════════════════════════════════════════
    #  REAL BACKGROUND PRE-LOADING
    #  Professional apps use splash loading time to do actual work:
    #    • Initialise all databases and run any pending migrations
    #    • Reload the live component-map from the DB
    #    • Pre-decode heavy images (welcome background, logo) into RAM
    #    • Cache frequently-used DB strings so first page is instant
    #  The splash progress bar now reflects REAL task completion.
    # ════════════════════════════════════════════════════════════════
    _preload_done  = threading.Event()
    _PRELOAD_STEPS = 5

    def _background_preload():
        """Heavy I/O on a daemon thread while the splash animation plays."""
        completed = 0
        try:
            # Step 1 ── Init / migrate all SQLite databases ────────────
            app.after(0, lambda: splash.set_status("Initialising databases…"))
            SetupDatabase.init_db()
            completed += 1
            app.after(0, lambda v=completed / _PRELOAD_STEPS:
                       splash._set_real_progress(v))

            # Step 2 ── Reload the global component map from DB ────────
            app.after(0, lambda: splash.set_status("Loading component map…"))
            try:
                global COMP_MAP
                COMP_MAP = load_component_map()
            except Exception:
                pass
            completed += 1
            app.after(0, lambda v=completed / _PRELOAD_STEPS:
                       splash._set_real_progress(v))

            # Step 3 ── Pre-decode the welcome background image ────────
            # PIL.Image.load() forces full pixel decode into RAM so the
            # first time the HomePage opens, there is zero disk I/O.
            app.after(0, lambda: splash.set_status("Loading display resources…"))
            try:
                from PIL import Image as _Im
                _bg_path = resource_path(os.path.join("settings", "welcome_bg.png"))
                if os.path.exists(_bg_path):
                    _img = _Im.open(_bg_path)
                    _img.load()   # force full decode — LANCZOS resize later is faster
                    app._cached_bg_image = _img
            except Exception:
                pass
            completed += 1
            app.after(0, lambda v=completed / _PRELOAD_STEPS:
                       splash._set_real_progress(v))

            # Step 4 ── Pre-decode the company logo ───────────────────
            try:
                from PIL import Image as _Im
                _logo_p = resource_path(os.path.join("settings", "cherry_full_logo.png"))
                if os.path.exists(_logo_p):
                    _logo = _Im.open(_logo_p)
                    _logo.load()
                    app._cached_logo_image = _logo
            except Exception:
                pass
            completed += 1
            app.after(0, lambda v=completed / _PRELOAD_STEPS:
                       splash._set_real_progress(v))

            # Step 5 ── Cache admin / company name strings ────────────
            app.after(0, lambda: splash.set_status("Almost ready…"))
            try:
                app._cached_company_name = SetupDatabase.get_company_name()
                app._cached_admin_name   = SetupDatabase.get_admin_name()
            except Exception:
                pass
            completed += 1
            app.after(0, lambda v=completed / _PRELOAD_STEPS:
                       splash._set_real_progress(v))

            app.after(0, lambda: splash.set_status("Ready  ✓"))

        except Exception as _e:
            print(f"[Preload] {_e}")
        finally:
            _preload_done.set()

    threading.Thread(target=_background_preload, daemon=True).start()

    # ── Reveal: wait for preload → fade splash → show main window ────
    def _reveal_app():
        _retries = [0]
        def _check_and_reveal():
            # Give preload up to 3 extra seconds on slow machines.
            # On fast machines _preload_done fires well before 2.5 s.
            if _preload_done.is_set() or _retries[0] >= 30:
                def _on_splash_done():
                    try:
                        app.deiconify()          # Show the window first
                        # Re-apply zoomed AFTER deiconify — on Windows,
                        # withdraw() resets the state so we must set it again
                        # here to guarantee the window opens fullscreen.
                        try:
                            app.state("zoomed")
                        except Exception:
                            try:
                                sw = app.winfo_screenwidth()
                                sh = app.winfo_screenheight()
                                app.geometry(f"{sw}x{sh}+0+0")
                            except Exception:
                                pass
                        app.lift()
                        app.focus_force()
                    except Exception:
                        pass
                try:
                    splash.fade_out(on_done=_on_splash_done)
                except Exception:
                    try:
                        app.deiconify()
                    except Exception:
                        pass
            else:
                _retries[0] += 1
                app.after(100, _check_and_reveal)
        _check_and_reveal()

    app.after(2500, _reveal_app)

    app.mainloop()